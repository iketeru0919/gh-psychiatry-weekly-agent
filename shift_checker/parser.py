"""勤務表Excel（管理用シート）の解析。HTML版チェッカーと同じ読取設定。"""

from __future__ import annotations

import datetime as dt
import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from .config import SUMMARY_ADDRESSES, Settings
from .util import (
    base_name_key,
    long_path,
    matches_exact,
    normalize,
    number_or_text,
    numeric_or_zero,
    round2,
    to_hours,
)


class ShiftFileError(Exception):
    """書式違い・破損など、この施設のファイルを解析できない場合に送出する。"""


@dataclass
class ShiftCell:
    value: str
    day: int | None
    date_label: str
    weekday: str
    cell: str


@dataclass
class Staff:
    key: str
    name: str
    row: int
    total_hours: float | None
    fulltime: str
    job: str
    day_allocation: object
    paid_count: int
    vacation_count: int
    prescribed_hours: float | None
    shifts: dict[str, ShiftCell]


@dataclass
class ShiftBook:
    path: Path
    sheet_name: str
    prescribed_hours: float | None
    facility_summary: dict
    staff: dict[str, Staff]
    duplicate_names: list[str]
    shift_symbols: set[str]
    warnings: list[str] = field(default_factory=list)
    month_label: str = ""


OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _precheck_file(path: Path) -> None:
    suffix = path.suffix.lower()
    try:
        with open(long_path(str(path)), "rb") as handle:
            head = handle.read(8)
    except OSError as error:
        raise ShiftFileError(
            f"ファイルを開けません（Dropboxで「オンラインのみ」になっていないか確認してください）: {error}"
        ) from error
    if suffix == ".xls":
        raise ShiftFileError(
            "旧形式(.xls)のため読取不可。Excelで開き「.xlsx」形式で保存し直してください。"
        )
    if head.startswith(OLE2_SIGNATURE):
        raise ShiftFileError(
            "パスワード保護されている可能性があり読取不可。保護を解除して保存し直してください。"
        )
    if not head.startswith(b"PK"):
        raise ShiftFileError("Excelファイルとして読み取れません（破損の可能性）。")


def load_workbook_safe(path: Path):
    import openpyxl

    _precheck_file(path)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            # read_only: 必要なシートだけ遅延読込するため大幅に高速
            return openpyxl.load_workbook(long_path(str(path)), data_only=True, read_only=True)
    except Exception as error:
        raise ShiftFileError(f"ファイルを開けません（破損の可能性）: {error}") from error


class SheetGrid:
    """管理用シートの必要範囲を一括読込してセル番地でアクセスするためのグリッド。

    read_onlyモードではセル単位のランダムアクセスが遅いため、
    先頭から必要行までを1回だけ走査してメモリに保持する。
    """

    def __init__(self, sheet, max_row: int, max_col: int):
        self._values: dict[tuple[int, int], object] = {}
        self._formats: dict[tuple[int, int], str] = {}
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue
                key = (cell.row, cell.column)
                self._values[key] = cell.value
                try:
                    self._formats[key] = cell.number_format or ""
                except Exception:
                    self._formats[key] = ""

    def value(self, address: str):
        column, row = coordinate_from_string(address)
        return self._values.get((row, column_index_from_string(column)))

    def value_rc(self, row: int, col: int):
        return self._values.get((row, col))

    def number_format(self, address: str) -> str:
        column, row = coordinate_from_string(address)
        return self._formats.get((row, column_index_from_string(column)), "")


def find_sheet_name(workbook, keyword: str) -> str | None:
    names = workbook.sheetnames
    if keyword:
        for name in names:
            if normalize(name) == keyword:
                return name
        for name in names:
            if keyword in normalize(name):
                return name
    for name in names:
        if "管理用" in normalize(name):
            return name
    return None


def extract_day(value) -> int | None:
    if isinstance(value, dt.datetime):
        return value.day
    if isinstance(value, dt.date):
        return value.day
    if isinstance(value, (int, float)) and not isinstance(value, bool) and 1 <= value <= 31:
        return int(value)
    match = re.match(r"^(\d{1,2})(?:日)?$", normalize(value))
    return int(match.group(1)) if match else None


def detect_date_columns(grid: SheetGrid, settings: Settings) -> list[dict]:
    start = column_index_from_string(settings.shift_start)
    end = column_index_from_string(settings.shift_end)
    columns = []
    for index in range(start, end + 1):
        letter = get_column_letter(index)
        row4 = grid.value_rc(4, index)
        row5 = grid.value_rc(5, index)
        row6 = grid.value_rc(6, index)
        day = extract_day(row4) or extract_day(row5)
        if isinstance(row5, (dt.datetime, dt.date)):
            date_label = f"{row5.month}/{row5.day}"
            key = f"DATE:{date_label}"
            date_value = row5
        elif day:
            date_label = f"{day}日"
            key = f"DAY:{day}"
            date_value = None
        else:
            continue
        columns.append({
            "letter": letter,
            "day": day,
            "date_label": date_label,
            "weekday": normalize(row6),
            "key": key,
            "date_value": date_value,
        })
    return columns


def parse_shift_book(path: Path, settings: Settings) -> ShiftBook:
    workbook = load_workbook_safe(path)
    sheet_name = find_sheet_name(workbook, settings.sheet_keyword)
    if not sheet_name:
        listed = "、".join(workbook.sheetnames[:6])
        raise ShiftFileError(
            f"「{settings.sheet_keyword}」シートが見つかりません（存在するシート: {listed}…）"
            "→ 書式が標準テンプレートと異なる可能性"
        )
    sheet = workbook[sheet_name]
    needed_row = max(
        settings.end_row,
        *(int(re.sub(r"[A-Z]", "", addr)) for addr in SUMMARY_ADDRESSES),
        int(re.sub(r"[A-Z]", "", settings.prescribed_cell)),
        6,
    )
    needed_col = max(
        *(column_index_from_string(re.sub(r"\d", "", addr)) for addr in SUMMARY_ADDRESSES),
        column_index_from_string(settings.shift_end),
        column_index_from_string(settings.total_col),
        column_index_from_string(settings.day_hours_col),
        column_index_from_string(settings.day_allocation_col),
    )
    grid = SheetGrid(sheet, needed_row, needed_col)
    workbook.close()
    book_warnings: list[str] = []

    prescribed = to_hours(
        grid.value(settings.prescribed_cell),
        grid.number_format(settings.prescribed_cell),
    )
    if prescribed is None:
        book_warnings.append(
            f"所定労働時間({settings.prescribed_cell})が数値でないため、所定不足チェックはスキップしました"
        )

    date_columns = detect_date_columns(grid, settings)
    if len(date_columns) < 27:
        raise ShiftFileError(
            f"シフト日付列({settings.shift_start}〜{settings.shift_end} 行4-5)が検出できません"
            f"（検出 {len(date_columns)}列/期待28〜31列）→ 書式が標準テンプレートと異なる可能性"
        )

    month_label = ""
    for column in date_columns:
        if isinstance(column["date_value"], (dt.datetime, dt.date)):
            month_label = f"{column['date_value'].year}年{column['date_value'].month}月"
            break

    staff: dict[str, Staff] = {}
    duplicates: list[str] = []
    occurrences: dict[str, int] = {}
    symbols: set[str] = set()

    name_col_index = column_index_from_string(settings.name_col)
    for row in range(settings.start_row, settings.end_row + 1):
        name = normalize(grid.value_rc(row, name_col_index))
        if not name:
            continue
        name_key = base_name_key(name)
        occurrence = occurrences.get(name_key, 0) + 1
        occurrences[name_key] = occurrence
        if occurrence > 1:
            duplicates.append(name)
        staff_key = f"{name_key}::{occurrence}"

        total_address = f"{settings.total_col}{row}"
        total_hours = to_hours(grid.value(total_address), grid.number_format(total_address))
        fulltime = normalize(grid.value(f"{settings.fulltime_col}{row}"))
        job = normalize(grid.value(f"{settings.job_col}{row}"))
        day_allocation = number_or_text(grid.value(f"{settings.day_allocation_col}{row}"))

        shifts: dict[str, ShiftCell] = {}
        paid_count = vacation_count = 0
        for column in date_columns:
            value = normalize(grid.value(f"{column['letter']}{row}"))
            if value:
                symbols.add(value)
            if matches_exact(value, settings.paid_words):
                paid_count += 1
            if matches_exact(value, settings.vacation_words):
                vacation_count += 1
            shifts[column["key"]] = ShiftCell(
                value=value,
                day=column["day"],
                date_label=column["date_label"],
                weekday=column["weekday"],
                cell=f"{column['letter']}{row}",
            )

        staff[staff_key] = Staff(
            key=staff_key,
            name=name,
            row=row,
            total_hours=total_hours,
            fulltime=fulltime,
            job=job,
            day_allocation=day_allocation,
            paid_count=paid_count,
            vacation_count=vacation_count,
            prescribed_hours=prescribed,
            shifts=shifts,
        )

    if not staff:
        raise ShiftFileError(
            f"職員氏名({settings.name_col}列 {settings.start_row}〜{settings.end_row}行)が"
            "1件も読めません → 書式が標準テンプレートと異なる可能性"
        )

    summary = {"DL3": number_or_text(grid.value(settings.prescribed_cell))}
    for address in SUMMARY_ADDRESSES:
        summary[address] = number_or_text(grid.value(address))
    summary["REAL_TOTAL"] = round2(
        numeric_or_zero(summary.get("DY24")) + numeric_or_zero(summary.get("DY40"))
    )

    return ShiftBook(
        path=path,
        sheet_name=sheet_name,
        prescribed_hours=prescribed,
        facility_summary=summary,
        staff=staff,
        duplicate_names=sorted(set(duplicates)),
        shift_symbols=symbols,
        warnings=book_warnings,
        month_label=month_label,
    )
