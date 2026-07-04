"""結果Excel（デスクトップ出力）の生成。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .checker import FacilityResult
from .config import SUMMARY_GROUPS, SUMMARY_LABELS, Settings
from .util import matches_exact

HEADER_FILL = PatternFill("solid", fgColor="16324F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CRITICAL_FILL = PatternFill("solid", fgColor="FFDEDE")
WARNING_FILL = PatternFill("solid", fgColor="FFF3C4")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
OK_FILL = PatternFill("solid", fgColor="DFF3E8")
NOTE_FONT = Font(color="617080", size=9)

SEVERITY_FILLS = {"critical": CRITICAL_FILL, "warning": WARNING_FILL}


def _write_table(sheet, headers: list[str], rows: list[list], row_fills=None,
                 widths: list[int] | None = None) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for index, row in enumerate(rows):
        sheet.append(row)
        fill = row_fills[index] if row_fills else None
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    sheet.freeze_panes = "A2"
    if widths:
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = width
    sheet.auto_filter.ref = sheet.dimensions


def _fmt(value):
    return "" if value is None else value


TOTAL_FILL = PatternFill("solid", fgColor="DCECFB")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="B8D5F0")
RED_FONT = Font(color="C00000", bold=True)


def _num(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return None


def write_report(results: list[FacilityResult], settings: Settings, output_dir: Path,
                 year: int, month: int, compared: bool, previous_stamp: str,
                 labor_rows=None, trend=None, external_rows=None,
                 external_trend=None) -> Path:
    workbook = Workbook()
    month_tag = f"{year % 100:02d}{month:02d}"
    now = datetime.now()

    # --- サマリー ---
    sheet = workbook.active
    sheet.title = "サマリー"
    headers = ["施設", "状態", "職員数", "重要", "要確認", "体制加算判定", "過不足結果",
               "警告・エラー内容", "使用ファイル", "候補数"]
    rows, fills = [], []
    for result in results:
        summary = result.current.facility_summary if result.current else {}
        note = result.error or " / ".join(result.warnings)
        rows.append([
            result.facility,
            result.status,
            len(result.current.staff) if result.current else "",
            result.critical_count if result.current else "",
            len(result.alerts) if result.current else "",
            _fmt(summary.get("EV78", "")),
            _fmt(summary.get("BU75", "")),
            note,
            result.source_path,
            result.duplicate_files if result.duplicate_files > 1 else "",
        ])
        if result.status != "OK":
            fills.append(ERROR_FILL)
        elif str(summary.get("EV78", "")).find("未達") >= 0 or result.critical_count:
            fills.append(WARNING_FILL)
        else:
            fills.append(OK_FILL)
    _write_table(sheet, headers, rows, fills,
                 widths=[22, 10, 8, 6, 8, 12, 10, 50, 60, 6])
    ok = sum(1 for r in results if r.status == "OK")
    footer = [
        f"実行日時: {now:%Y-%m-%d %H:%M}",
        f"対象年月: {year}年{month}月（{month_tag}）",
        f"処理結果: {ok}施設OK / {len(results) - ok}施設エラー",
        f"比較基準: {previous_stamp} 時点のスナップショット" if compared
        else "比較基準: なし（初回実行のため今回のみチェック）",
    ]
    for line in footer:
        sheet.append([line])
        sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- 要確認者 / 職員全件 ---
    staff_headers = ["施設", "判定", "職員氏名", "状態", "職種", "雇用区分", "所定",
                     "前回総労働", "今回総労働", "増減", "所定との差",
                     "有休(前)", "有休(今)", "有休差", "公休(前)", "公休(今)", "公休差",
                     "日別変更数", "内容"]
    staff_widths = [22, 8, 18, 9, 18, 9, 7, 9, 9, 7, 9, 7, 7, 7, 7, 7, 7, 9, 60]

    def staff_row_values(row):
        return [
            row.facility, row.judgment, row.name, row.status, row.job,
            row.employment_type, _fmt(row.prescribed),
            _fmt(row.before_hours), _fmt(row.after_hours), _fmt(row.hour_delta),
            _fmt(row.required_delta),
            row.before_paid, row.after_paid, row.paid_delta,
            row.before_vacation, row.after_vacation, row.vacation_delta,
            row.change_count, row.reason,
        ]

    for title, selector in (("要確認者", lambda r: r.alerts), ("職員全件", lambda r: r.staff_rows)):
        sheet = workbook.create_sheet(title)
        rows, fills = [], []
        for result in results:
            for row in selector(result):
                rows.append(staff_row_values(row))
                fills.append(SEVERITY_FILLS.get(row.severity))
        if rows:
            _write_table(sheet, staff_headers, rows, fills, staff_widths)
        else:
            _write_table(sheet, staff_headers, [], None, staff_widths)
            sheet.append(["該当なし"])

    # --- 日別変更（比較実行時のみ）---
    if compared:
        sheet = workbook.create_sheet("日別変更")
        headers = ["施設", "職員氏名", "日付", "曜日", "前回", "今回", "変更区分", "セル"]
        rows, fills = [], []
        for result in results:
            for change in result.change_rows:
                rows.append([
                    change.facility, change.name, change.date_label, change.weekday,
                    change.previous_value, change.current_value, change.change_type,
                    change.cell,
                ])
                fills.append(WARNING_FILL if change.change_type in ("休暇変更", "削除") else None)
        _write_table(sheet, headers, rows, fills, widths=[22, 18, 9, 6, 12, 12, 10, 8])
        if not rows:
            sheet.append(["変更はありませんでした"])

    # --- 配置・体制チェック（HTML版と同じ4グループ・同じ項目名＋過不足の強調）---
    sheet = workbook.create_sheet("配置・体制チェック")
    # 支援員・世話人グループに「過不足」列（法令基準と実配置の差）を追加する
    placement_groups = []
    for group_name, group_keys in SUMMARY_GROUPS:
        keys_with_diff = list(group_keys)
        if group_name == "支援員 配置詳細":
            keys_with_diff.append("SUPPORT_DIFF")
        elif group_name == "世話人 配置詳細":
            keys_with_diff.append("CARE_DIFF")
        placement_groups.append((group_name, keys_with_diff))
    diff_labels = {
        "SUPPORT_DIFF": "支援員 過不足 (DY24-FE79)",
        "CARE_DIFF": "世話人 過不足 (DY40-FE82)",
    }
    keys = [key for _, group_keys in placement_groups for key in group_keys]

    group_row = ["施設"]
    for group_name, group_keys in placement_groups:
        group_row.extend([group_name] + [""] * (len(group_keys) - 1))
    sheet.append(group_row)
    column_number = 2
    for group_name, group_keys in placement_groups:
        if len(group_keys) > 1:
            sheet.merge_cells(
                start_row=1, start_column=column_number,
                end_row=1, end_column=column_number + len(group_keys) - 1,
            )
        column_number += len(group_keys)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.append(["施設"] + [diff_labels.get(k) or SUMMARY_LABELS[k] for k in keys])
    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.merge_cells("A1:A2")

    red_targets = {"SUPPORT_DIFF", "CARE_DIFF", "BU75"}
    for result in results:
        if not result.current:
            sheet.append([result.facility, "読取エラー"] + [""] * (len(keys) - 1))
            for cell in sheet[sheet.max_row]:
                cell.fill = ERROR_FILL
            continue
        summary = result.current.facility_summary
        values = {}
        support_base = _num(summary.get("FE79"))
        support_actual = _num(summary.get("DY24"))
        care_base = _num(summary.get("FE82"))
        care_actual = _num(summary.get("DY40"))
        values["SUPPORT_DIFF"] = (
            round(support_actual - support_base, 2)
            if support_actual is not None and support_base is not None else ""
        )
        values["CARE_DIFF"] = (
            round(care_actual - care_base, 2)
            if care_actual is not None and care_base is not None else ""
        )
        sheet.append([result.facility]
                     + [values[k] if k in values else _fmt(summary.get(k, "")) for k in keys])
        row_number = sheet.max_row
        if "未達" in str(summary.get("EV78", "")):
            for cell in sheet[row_number]:
                cell.fill = WARNING_FILL
        # 不足（マイナス）は赤太字で強調
        for offset, key in enumerate(keys, start=2):
            if key not in red_targets:
                continue
            cell = sheet.cell(row_number, offset)
            number = _num(cell.value)
            if number is not None and number < 0:
                cell.font = RED_FONT
    sheet.freeze_panes = "B3"
    sheet.column_dimensions["A"].width = 22
    for index in range(len(keys)):
        sheet.column_dimensions[get_column_letter(index + 2)].width = 15
    sheet.append([])
    sheet.append(["※過不足列は実配置数−法令基準。マイナス（不足）と過不足結果(BU75)のマイナスは赤太字で表示します。"])
    sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- 労務チェック（氏名の隣に指摘カテゴリ、詳細は右端の別セル）---
    if labor_rows is not None:
        sheet = workbook.create_sheet("労務チェック")
        headers = ["施設", "判定", "職員氏名", "指摘カテゴリ",
                   "連続勤務", "夜勤明け", "夜勤超過", "公休不足", "二重割当", "時間外",
                   "社保ライン", "45h超(年)",
                   "職種", "雇用区分", "総労働時間", "週平均", "所定", "所定超過",
                   "夜勤回数", "深夜業4回以上", "公休数", "公休基準", "有休数",
                   "社保表記", "指摘詳細"]
        rows, fills = [], []
        for row in labor_rows:
            rows.append([
                row.facility, row.judgment, row.name, row.category_summary,
                row.cat_consecutive, row.cat_ake, row.cat_night, row.cat_holiday,
                row.cat_double, row.cat_overtime, row.cat_insurance,
                row.overtime_year_count,
                row.job, row.employment_type, _fmt(row.total_hours),
                _fmt(row.weekly_avg_hours), _fmt(row.prescribed), _fmt(row.overtime),
                row.night_count, row.night_check_target,
                row.holiday_count, _fmt(row.required_holidays), row.paid_count,
                row.insurance_mark,
                " / ".join(row.reasons) or "問題なし",
            ])
            fills.append(SEVERITY_FILLS.get(row.severity))
        _write_table(sheet, headers, rows, fills,
                     widths=[22, 8, 18, 24, 8, 8, 8, 9, 8, 8, 9, 9,
                             18, 9, 9, 8, 7, 8, 8, 11, 7, 8, 7, 8, 70])
        for row_number in range(2, len(rows) + 2):
            for column_number in range(5, 13):  # カテゴリ値の列は赤太字で目立たせる
                cell = sheet.cell(row_number, column_number)
                if cell.value not in (None, ""):
                    cell.font = RED_FONT
        sheet.freeze_panes = "E2"
        sheet.append([])
        sheet.append(["※連続勤務・夜勤明けは月内のみで判定（前月末からの継続は含みません）。"
                      "公休基準はconfig.jsonのrequired_holidaysで施設別に設定できます（未設定は判定なし）。"
                      "社保ラインは非常勤かつ氏名に社保表記が無い人の週平均時間の目安判定です。"
                      "45h超(年)は所定超過45時間以上だった月数（36協定の特別条項は年6回まで）。"])
        sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- 外部人材（タイミー・派遣）---
    if external_rows is not None:
        sheet = workbook.create_sheet("外部人材")
        headers = ["施設", "区分", "氏名・枠名", "勤務日数", "総労働時間", "夜勤回数",
                   "有休", "公休", "前回総労働", "時間増減", "シフト変更", "外部依存度",
                   "概算費用(円)", "備考"]
        rows, fills = [], []
        for row in external_rows:
            rows.append([
                row.facility, row.category, row.name,
                _fmt(row.work_days), _fmt(row.total_hours), _fmt(row.night_count),
                _fmt(row.paid_count), _fmt(row.vacation_count),
                _fmt(row.previous_hours), _fmt(row.hour_delta), _fmt(row.change_count),
                row.dependency, _fmt(row.cost), row.note,
            ])
            if row.is_total:
                fills.append(GRAND_TOTAL_FILL if row.facility == "全施設" else TOTAL_FILL)
            else:
                fills.append(None)
        _write_table(sheet, headers, rows, fills,
                     widths=[22, 9, 26, 9, 10, 9, 6, 6, 10, 9, 9, 10, 12, 18])
        for row_number in range(2, len(rows) + 2):
            cell = sheet.cell(row_number, 13)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
        if not rows:
            sheet.append(["外部人材（氏名にタイミー・派遣を含む職員）はいませんでした"])
        if external_trend and external_trend[1]:
            sheet.append([])
            sheet.append(["■ 外部人材の推移（直近6ヶ月）"])
            sheet.cell(sheet.max_row, 1).font = Font(bold=True, size=11)
            trend_headers, trend_rows = external_trend
            sheet.append(trend_headers)
            for cell in sheet[sheet.max_row]:
                if cell.value:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
            for trend_row in trend_rows:
                sheet.append(trend_row)
        sheet.append([])
        rates_note = ("単価はconfig.jsonのexternal_ratesで設定（未設定の区分は費用を表示しません）。"
                      if settings.external_rates else
                      "概算費用を出すにはconfig.jsonのexternal_ratesに時間単価を設定してください"
                      "（例: \"external_rates\": {\"タイミー\": 1400, \"派遣\": 2200}）。")
        sheet.append(["※氏名にキーワード（タイミー・派遣）を含む職員行を記載名ごとに集計。"
                      "括弧書きの個人名もタイミーに含めています。"
                      "外部依存度＝区分計の総労働時間÷施設全体の総労働時間。" + rates_note])
        sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- 労務推移 ---
    if trend and trend[1]:
        sheet = workbook.create_sheet("労務推移")
        trend_headers, trend_rows = trend
        _write_table(sheet, trend_headers, trend_rows, None,
                     widths=[22, 18, 10] + [10] * (len(trend_headers) - 3))
        sheet.append([])
        sheet.append(["※実行のたびに月次集計が _チェック履歴/労務推移データ.csv へ蓄積され、"
                      "直近6ヶ月分を表示します。"])
        sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- 勤務記号 ---
    sheet = workbook.create_sheet("勤務記号")
    headers = ["記号", "区分", "使用施設"]
    symbol_map: dict[str, set[str]] = {}
    for result in results:
        if result.current:
            for symbol in result.current.shift_symbols:
                symbol_map.setdefault(symbol, set()).add(result.facility)
    rows = []
    for symbol in sorted(symbol_map):
        category = ("有休" if matches_exact(symbol, settings.paid_words)
                    else "公休" if matches_exact(symbol, settings.vacation_words)
                    else "勤務")
        rows.append([symbol, category, "、".join(sorted(symbol_map[symbol]))])
    _write_table(sheet, headers, rows, None, widths=[12, 8, 70])
    sheet.append([])
    sheet.append(["※「有休」「公休」の判定文字は shift_checker/config.json の "
                  "paid_words / vacation_words で変更できます"])
    sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    # --- ダッシュボード（エリア全体を1枚で。先頭シートに配置）---
    sheet = workbook.create_sheet("ダッシュボード", 0)
    headers = ["施設", "状態", "職員数", "体制加算判定", "配置過不足(BU75)",
               "支援員過不足", "世話人過不足", "労務指摘(重要)", "労務指摘(要確認)",
               "外部依存度", "外部概算費用(円)", "入力途中(人)", "シフト変更(件)"]
    labor_by_facility: dict[str, list] = {}
    for row in (labor_rows or []):
        labor_by_facility.setdefault(row.facility, []).append(row)
    external_by_facility: dict[str, dict] = {}
    for row in (external_rows or []):
        if row.is_total and row.facility != "全施設":
            bucket = external_by_facility.setdefault(
                row.facility, {"hours": 0.0, "cost": 0.0, "has_cost": False,
                               "facility_hours": row.facility_hours})
            bucket["hours"] += row.total_hours or 0
            if row.cost is not None:
                bucket["cost"] += row.cost
                bucket["has_cost"] = True

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for result in results:
        if not result.current:
            sheet.append([result.facility, "読取エラー"] + [""] * (len(headers) - 2))
            for cell in sheet[sheet.max_row]:
                cell.fill = ERROR_FILL
            continue
        summary = result.current.facility_summary
        support_diff = care_diff = ""
        if _num(summary.get("DY24")) is not None and _num(summary.get("FE79")) is not None:
            support_diff = round(_num(summary.get("DY24")) - _num(summary.get("FE79")), 2)
        if _num(summary.get("DY40")) is not None and _num(summary.get("FE82")) is not None:
            care_diff = round(_num(summary.get("DY40")) - _num(summary.get("FE82")), 2)
        facility_labor = labor_by_facility.get(result.facility, [])
        labor_critical = sum(1 for r in facility_labor if r.severity == "critical")
        labor_warning = sum(1 for r in facility_labor if r.severity == "warning")
        incomplete = sum(1 for r in result.staff_rows if "入力途中" in r.reason)
        bucket = external_by_facility.get(result.facility)
        dependency = ""
        cost_value = ""
        if bucket and bucket.get("facility_hours"):
            dependency = f"{round(bucket['hours'] / bucket['facility_hours'] * 100, 1):g}%"
        if bucket and bucket["has_cost"]:
            cost_value = round(bucket["cost"])
        judgment = str(summary.get("EV78", ""))
        sheet.append([
            result.facility, result.status, len(result.current.staff), judgment,
            _fmt(summary.get("BU75", "")), support_diff, care_diff,
            labor_critical or "", labor_warning or "",
            dependency, cost_value, incomplete or "", len(result.change_rows) or "",
        ])
        row_number = sheet.max_row
        cells = {header: sheet.cell(row_number, index + 1)
                 for index, header in enumerate(headers)}
        if "未達" in judgment:
            cells["体制加算判定"].fill = CRITICAL_FILL
            cells["体制加算判定"].font = RED_FONT
        for header in ("配置過不足(BU75)", "支援員過不足", "世話人過不足"):
            number = _num(cells[header].value)
            if number is not None and number < 0:
                cells[header].fill = CRITICAL_FILL
                cells[header].font = RED_FONT
        if labor_critical:
            cells["労務指摘(重要)"].fill = CRITICAL_FILL
            cells["労務指摘(重要)"].font = RED_FONT
        if labor_warning:
            cells["労務指摘(要確認)"].fill = WARNING_FILL
        try:
            dep_number = float(str(dependency).rstrip("%")) if dependency else None
        except ValueError:
            dep_number = None
        if dep_number is not None:
            if dep_number >= 20:
                cells["外部依存度"].fill = CRITICAL_FILL
                cells["外部依存度"].font = RED_FONT
            elif dep_number >= 10:
                cells["外部依存度"].fill = WARNING_FILL
        if incomplete:
            cells["入力途中(人)"].fill = WARNING_FILL
        if isinstance(cells["外部概算費用(円)"].value, (int, float)):
            cells["外部概算費用(円)"].number_format = "#,##0"
    sheet.freeze_panes = "B2"
    for index, width in enumerate([22, 10, 7, 12, 12, 11, 11, 11, 12, 10, 13, 10, 11], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.append([])
    for line in (
        f"対象年月: {year}年{month}月（{month_tag}） / 実行日時: {now:%Y-%m-%d %H:%M}",
        "色の意味: 赤=不足・重要（配置不足、労務の重要指摘、外部依存20%以上）、黄=注意（要確認、依存10%以上、入力途中あり）",
        "詳細は各シート（労務チェック・配置・体制チェック・外部人材・要確認者）を参照してください。",
    ):
        sheet.append([line])
        sheet.cell(sheet.max_row, 1).font = NOTE_FONT

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"シフトチェック結果_{month_tag}_{now:%Y%m%d_%H%M}.xlsx"
    workbook.save(output_path)
    return output_path
