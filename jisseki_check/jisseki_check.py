import os
import re
import traceback
from pathlib import Path, PurePosixPath
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

# ★ここを変更: 実績記録票が入ったフォルダの親ディレクトリ
DRIVE_FOLDER_PATH = "/content/drive/MyDrive/実績"

# ★出力先フォルダ（自動作成）
OUTPUT_FOLDER_PATH = "/content/drive/MyDrive/実績/チェックレポート"

# ★対象ファイルの拡張子
TARGET_EXTENSIONS = ['.xlsm', '.xlsx']

# ============================================================
# チェックロジック
# ============================================================

# --- キーワード定義 ---
GAIBU_KW  = [
    # 医療系
    '受診','病院','HP','クリニック','神経科','眼科','歯科','整形','内科','外科',
    '耳鼻','皮膚','精神科','脳外','歯医者','透析','デイケア','リハビリ',
    # 外出・外食
    '外出','外食','自費昼食','母と外食','家族と外出','家族と受診','一人で外出',
    '相談員と外出','お見舞い',
    # 就労・生活介護・デイ系施設
    '就労','生活介護','生活訓練','多機能','B型','A型','一般就労','余暇',
    '障害者雇用',
    # 施設名（五十音順）
    'ALL宇都宮','あいあい学園','アイサポ','アルスノヴァ','アットホーム',
    'アント','いつでも','ウィズ','ウイズ','ウィング','ウーリー',
    'えがお','オルゴール',
    'きたえるーむ','きらら','ケアイロ',
    'コラソン',
    'じょあん','シンフォニー',
    'すてっぷ',
    'たいよう','たからの家','たてがみ','ツクイ','トライ',
    'ななくさ','ナルド','ニチイ',
    'にゃんこ',
    'ひくまの','ひまわり工房','ピース','ぷらねっと',
    'ふくふく','ふくろうの家','ふるさと','ふれあい作業',
    'べテルホーム','ペンタス','ベジモ',
    'ぽかぽか',
    'むつみ','みこち','メンタルサポート',
    'ユーファーム','ユニークワークス',
    'ラシクラボ','ララカフェ','リアン','リセイル','リブライフ','ロピア',
    'ワークセンター','ワークワーク',
    'だんらんの家','プティ','ステップ','トライ',
    'HANAUTA','Jump','BE HAPPY',
    '朝山','小池','招来屋',
    '自宅→','自宅へ','自宅より',
    '白梅','白萩','グリーンフィールド',
    '幸','やくわり','工房つつじ','カタクリの花',
    # 移動支援・同行援護
    '移動支援','同行援護',
    # 教会・その他
    '教会','一人暮らし体験',
]
NOT_GAIBU = ['往診','出前','腹痛','担当者会議','就労休み','デイ休み','キャンセル','昼食なし','発熱']
LUNCH_OUT = [
    # 日中サービス（昼食を施設で取る可能性あり）
    'ひくまの','シンフォニー','ウィズ','ウイズ','ステップ','トライ',
    'ワークセンター','ツクイ','デイサービス','ユニークワークス',
    'だんらんの家','プティ','ななくさ','ピース','リセイル','ふくろうの家',
    'ユーファーム','ALL宇都宮','たいよう','ふくふく','じょあん','ふるさと',
    'ナルド','HANAUTA','べテルホーム','コラソン','ぽかぽか','リブライフ',
    'たてがみ','アント','生活介護','生活訓練','B型','A型','一般就労','障害者雇用',
    '自宅→','自宅へ','自宅より',
    'すてっぷ','Jump','たからの家','ワークワーク','ロピア','ふれあい作業',
    '白梅','白萩','グリーンフィールド','ベジモ','きたえるーむ','ウーリー','むつみ','オルゴール',
    'あいあい学園','ラシクラボ','ひまわり工房','きらら','アットホーム',
    'みこち','ペンタス','やくわり','カタクリの花','工房つつじ','いつでも',
    'にゃんこ','ぷらねっと','ケアイロ','ララカフェ','招来屋','アルスノヴァ',
    'えがお','リアン','ウィング','はぐくみ','ニチイ','BE HAPPY','余暇',
]
LUNCH_IN  = ['昼発注あり', '昼食あり', '昼食発注あり', '食事発注済み', 'キャンセル不可食事発注あり']
# 短期入所の月跨ぎ利用が備考に明示されているとみなすキーワード（R17用）
TSUKI_MATAGI_KW = ['月跨ぎ', '月またぎ', '翌月', '継続', 'まで', '予定']
# 二重登録されたキーワードの除去（判定結果は変わらない）
GAIBU_KW  = list(dict.fromkeys(GAIBU_KW))
LUNCH_OUT = list(dict.fromkeys(LUNCH_OUT))
GAIHAKU_NG = {'外泊', '外泊開始'}
# F列（夜勤加配）が空欄でOKなC列値
F_EXEMPT   = {'外泊開始','外泊','入院開始','入院','退去','入院→退去','-'}
TAIKEN_PAT = re.compile(r'^体験')  # 体験①②③④等
# I列（医療連携加算）が空欄が正しいC列値（GH不在中）
I_EXEMPT   = {'入院', '外泊', '入院→退去'}  # 入院開始・外泊開始・外泊終了・退院はGH在室のためI列=1が正→除外
NYUIN_NG   = {'入院', '入院開始'}        # 入院中は夜勤加配不可
NYUIN_F_NG = {'入院', '入院開始', '退去', '入院→退去'} # F列空欄が必要な状態（退院はGH戻りのため除外）
TIME_REQ   = {'入院開始', '退院', '退去', '入院→退去'} # T列に時間記載が必要な状態
# 退去時の行先キーワード
DESTINATION_KW = ['自宅', '実家', 'GH', 'グループホーム', '施設', '病院',
                  '入院', 'アパート', '家族', '転居', '転出', '他GH']
TIME_PATTERN = re.compile(r'～?\d{1,2}[：:]\d{2}|～?\d{1,2}時\d*|\d{1,2}[：:]\d{2}～')
# 外泊予定日（月末まで外泊継続のとき備考に記載されるべき日付）
DATE_PLAN_PATTERN = re.compile(
    r'\d{1,2}/\d{1,2}まで'      # 3/2まで
    r'|\d{1,2}月\d{1,2}日まで'  # 4月1日まで
    r'|\d{1,2}日まで'             # 5日まで
    r'|翌日'                       # 翌日帰設等
    r'|\d{1,2}/\d{1,2}帰'       # 3/2帰設
    r'|\d{1,2}月\d{1,2}日帰'    # 4月1日帰設
)
# 予定日パターン: 3/31まで、4月5日予定、など
DATE_PATTERN = re.compile(
    r'\d{1,2}[/／]\d{1,2}'           # 3/31
    r'|\d{1,2}月\d{1,2}日'           # 3月31日
    r'|まで|予定'                        # ～まで、予定
)

def is_gaibu(t):
    if not t: return False
    return any(k in t for k in GAIBU_KW) and not any(k in t for k in NOT_GAIBU)

def is_lunch_out(t):
    if not t: return False
    if any(k in t for k in LUNCH_IN): return False
    return any(k in t for k in LUNCH_OUT)

def has_time(t):
    return bool(TIME_PATTERN.search(t)) if t else False

def has_destination(t):
    """退去備考に行先（自宅・実家・施設名等）が記載されているか確認"""
    if not t: return False
    return any(k in t for k in DESTINATION_KW)

def is_f_exempt(c):
    """F列（夜勤加配）が空欄でOKなC列値かどうか判定"""
    return c in F_EXEMPT or bool(TAIKEN_PAT.match(c))

def is_i_exempt(c):
    """I列（医療連携加算）が空欄でOKなC列値かどうか判定"""
    return c in I_EXEMPT or bool(TAIKEN_PAT.match(c))

def has_date_plan(t):
    """外泊継続時の備考に予定日（〇/〇まで等）が記載されているか確認"""
    return bool(DATE_PLAN_PATTERN.search(t)) if t else False

# ============================================================
# キーワード設定の外部ファイル対応
# 実績フォルダに「チェック設定.xlsx」を置くと、コードを変更せずに
# 施設名などのキーワードを現場で追加・削除できる
# ============================================================
CONFIG_FILENAME = 'チェック設定.xlsx'
KEYWORD_SHEET   = 'キーワード'
# 設定ファイルの「種別」列の値 → 上書き対象のキーワードリスト
KEYWORD_TYPE_MAP = {
    '外出キーワード':   'GAIBU_KW',
    '外出除外':         'NOT_GAIBU',
    '日中外出先':       'LUNCH_OUT',
    '昼食発注済み表記': 'LUNCH_IN',
    '退去行先':         'DESTINATION_KW',
    '月跨ぎ記載':       'TSUKI_MATAGI_KW',
}

def load_keyword_config(base_folder):
    """実績フォルダの チェック設定.xlsx を読み込み、キーワードリストを上書きする。

    ファイルが無ければ何もせず False を返す（組み込みのデフォルトを使用）。
    読み込みに成功したら True を返す。
    """
    global GAIBU_KW, NOT_GAIBU, LUNCH_OUT, LUNCH_IN, DESTINATION_KW, TSUKI_MATAGI_KW
    path = Path(base_folder) / CONFIG_FILENAME
    if not path.exists():
        return False
    wb = load_workbook(str(path), data_only=True)
    if KEYWORD_SHEET not in wb.sheetnames:
        print(f"⚠️  {CONFIG_FILENAME} に「{KEYWORD_SHEET}」シートがないため、組み込みのキーワードを使用します")
        wb.close()
        return False
    ws = wb[KEYWORD_SHEET]
    loaded = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        ktype, kw = row[0], row[1]
        if ktype is None or kw is None:
            continue
        ktype = str(ktype).strip()
        kw    = str(kw).strip()
        if ktype in KEYWORD_TYPE_MAP and kw:
            loaded[KEYWORD_TYPE_MAP[ktype]].append(kw)
    wb.close()
    g = globals()
    for var_name, kws in loaded.items():
        g[var_name] = list(dict.fromkeys(kws))
    return True

def create_config_template(base_folder):
    """現在のキーワードを チェック設定.xlsx として書き出す（現場で編集できる雛形）"""
    path = Path(base_folder) / CONFIG_FILENAME
    if path.exists():
        print(f"⚠️  既に存在するため上書きしません: {path}")
        return str(path)
    wb = Workbook()
    ws = wb.active; ws.title = KEYWORD_SHEET
    ws['A1'] = '種別'; ws['B1'] = 'キーワード'
    ws['A1'].font = Font(bold=True); ws['B1'].font = Font(bold=True)
    r = 2
    g = globals()
    for ktype, var_name in KEYWORD_TYPE_MAP.items():
        for kw in g[var_name]:
            ws.cell(row=r, column=1, value=ktype)
            ws.cell(row=r, column=2, value=kw)
            r += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 32
    ws2 = wb.create_sheet('説明')
    notes = [
        'このファイルを編集すると、コードを変更せずにチェック用キーワードを追加・削除できます。',
        '「キーワード」シートに1行1キーワードで記載してください（1列目=種別、2列目=キーワード）。',
        '',
        '種別の意味:',
        '  外出キーワード   … 備考にあると「住居外利用」とみなす語（R2で使用）',
        '  外出除外         … 上記に該当しても外出とみなさない語（往診・出前 等）',
        '  日中外出先       … 昼食発注の確認対象となる日中サービス・施設名（R5で使用）',
        '  昼食発注済み表記 … 備考にあると昼食確認から除外する語（昼発注あり 等）',
        '  退去行先         … 退去日の備考に必要な行先の語（R11で使用）',
        '  月跨ぎ記載       … 短期入所の月跨ぎ利用を明示する語（R17で使用）',
        '',
        '注意: 1文字だけのキーワード（例:「幸」）は無関係な備考にも一致しやすいため、',
        '      できるだけ2文字以上の表記で登録してください。',
    ]
    for i, t in enumerate(notes, 1):
        ws2.cell(row=i, column=1, value=t)
    ws2.column_dimensions['A'].width = 90
    wb.save(str(path))
    print(f"✅ キーワード設定の雛形を作成しました: {path}")
    return str(path)

# --- ルール定義 ---
RULES = {
    'R1': '外泊中の夜勤加配加算',
    'R2': '住居外利用チェック漏れ',
    'R3': '住居外利用の備考未記載',
    'R4': '外泊中の食事発注記録',
    'R5': '日中外出日の昼食発注確認',
    'R6': '外泊終了日の帰設時間未記載',
    'R7': '外泊開始日の出発時間未記載',
    'R8': '入院中の夜勤加配加算',
    'R9': '入院開始・退院・退去日の時間未記載',
    'R10': '入院・入院開始・退去中の夜勤加配加算',
    'R11': '退去・入院→退去の行先未記載',
    'R12': '外泊が月末まで未終了（予定日未記載）',
    'R13': '入院開始後に退院・退去なし',
    'R14': '入院中に食事発注あり',
    'R15': '夜勤加配加算の算定漏れ（同一ファイル内の相違）',
    'R16': '医療連携体制加算の誤算定（入院・外泊中）',
    'R17': '短期入所の月跨ぎ利用の可能性',
    'R18': 'シート名とI2セルの氏名が不一致（前月データ使い回しミスの可能性）',
    'R19': '同一ファイル内でD1（月）の値がシートによって異なる',
    'R20': 'ファイル名の年月とD1（月）・A1（年）の値が不一致',
    'R21': '医療連携体制加算の算定漏れ（同一ファイル内で個人のみ未算定）',
    'R22': 'C2（受給者証番号）またはP2（障害支援区分）が空白',
}
RULE_DESCS = {
    'R1': 'C列=「外泊」「外泊開始」のとき → F列（夜勤職員加配加算）は空欄が正',
    'R2': 'T列（備考）に外出・受診等の記載 → E列（住居外利用）にチェックが必要',
    'R3': 'E列（住居外利用）にチェックあり → T列（備考）に理由の記載が必要',
    'R4': 'C列=「外泊」「外泊開始」のとき → P/Q/R列（食事発注✓）は空欄が正',
    'R5': '日中外出（デイ・就労支援等）の日に昼食✓あり → 昼食不要の可能性（要確認）',
    'R6': 'C列=「外泊終了」のとき → T列（備考）に帰設時間の記載が必要',
    'R7': 'C列=「外泊開始」のとき → T列（備考）に出発時間の記載が必要',
    'R8': 'C列=「入院」「入院開始」のとき → F列（夜勤職員加配加算）は空欄が正',
    'R9': 'C列=「入院開始」「退院」「退去」のとき → T列（備考）に時間の記載が必要',
    'R10': 'C列=「入院」「入院開始」「退去」のとき → F列（夜勤加配）は空欄が正（退院はGH戻りのため除外）',
    'R11': 'C列=「退去」「入院→退去」のとき → T列（備考）に行先（自宅・実家・施設名等）の記載が必要',
    'R12': '月末時点で外泊開始後に外泊終了がない場合、外泊開始日の備考に帰設予定日の記載が必要',
    'R13': '入院開始後に退院・退去の記録が月内にない。退院日未定の可能性あり。情報として抽出のみ',
    'R14': 'C列=「入院」のとき食事✓がある。突発入院でキャンセル不可の可能性あり。情報として抽出のみ',
    'R15': '算定対象日（外泊・入院・退去・体験・-以外）なのにF列（夜勤加配）が空欄の日を検出。同一ファイル内の他の利用者と比較',
    'R16': 'C列=「入院」「外泊」のとき → I列（医療連携体制加算）は空欄が正。入院開始・外泊開始はGH在室のためI列=1が正（除外）',
    'R17': '短期入所ファイルの月末側（25日〜）でE列（退所時間）が「-」かつ備考に月跨ぎの記載がない場合、月跨ぎ利用の可能性あり',
    'R18': 'シート名とI2セルの氏名を正規化して比較。不一致は前月データ使い回しによる人物混在の可能性',
    'R19': '同一ファイル内でD1（月）の値がシートによって異なる。前月データ使い回し時の月更新忘れ',
    'R20': 'ファイル名の年月とA1（令和年）・D1（月）の値が一致しない。前月ファイルをそのまま提出するミスを防ぐ',
    'R21': '他の利用者が算定している日にI列（医療連携加算）が空欄の日を検出。個人のみ漏れている日を特定',
    'R22': 'C2セル（受給者証番号）またはP2セル（障害支援区分）が空白のまま。請求前に入力が必要',
}

# --- データ読み取り ---
EXCLUDE_SHEETS = {'実績記録票','算定項目リスト','重度障害者支援加算要件','延べ日数'}


# --- 元号変換（R20で使用）---
REIWA_BASE = 2018  # 令和元年=2019年 → 西暦-2018=令和年

def _to_int(v):
    """セル値から整数を取り出す（'令和8'→8、8.0→8。取れなければNone）"""
    if v is None or isinstance(v, bool): return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, datetime): return None
    m = re.search(r'\d{1,4}', str(v))
    return int(m.group()) if m else None

def _normalize_name(v):
    """氏名比較用の正規化（空白・「様」を除去。R18用）"""
    if not v: return ''
    s = str(v)
    for ch in (' ', '　', '\t', '様'):
        s = s.replace(ch, '')
    return s.strip()

def extract_sheet_headers(filepath):
    """各シートのヘッダー情報を一括取得（R18〜R20・R22用）
    A1=令和年 / D1=月 / I2=氏名 / C2=受給者証番号 / P2=障害支援区分"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = {}
    for sname in wb.sheetnames:
        if sname in EXCLUDE_SHEETS: continue
        ws = wb[sname]
        result[sname] = {
            'A1': ws['A1'].value, 'D1': ws['D1'].value, 'I2': ws['I2'].value,
            'C2': ws['C2'].value, 'P2': ws['P2'].value,
        }
    wb.close()
    return result

def run_header_checks(floor_label, filepath, filename):
    """R18〜R20・R22（シートのヘッダー整合性チェック）を実行してissuesを返す"""
    issues = []
    headers = extract_sheet_headers(filepath)
    if not headers:
        return issues

    # R18: シート名とI2（氏名）の不一致
    for sname, h in headers.items():
        sn = _normalize_name(sname)
        i2 = _normalize_name(h['I2'])
        if sn and i2 and sn != i2 and sn not in i2 and i2 not in sn:
            issues.append(dict(rule='R18', floor=floor_label, name=sname, date='-',
                level='error',
                detail=f"シート名=「{sname}」/ I2セル=「{h['I2']}」",
                judge='❌ シート名とI2の氏名が不一致（前月データ使い回しの可能性）'))

    # R19: 同一ファイル内でD1（月）が不統一
    d1_map = {s: _to_int(h['D1']) for s, h in headers.items()}
    d1_vals = {v for v in d1_map.values() if v is not None}
    if len(d1_vals) > 1:
        pairs = [f'{s}={v}月' for s, v in d1_map.items() if v is not None]
        issues.append(dict(rule='R19', floor=floor_label, name='（全シート）', date='-',
            level='error',
            detail=f"D1（月）がシートによって異なる: {' / '.join(pairs[:10])}",
            judge='❌ 同一ファイル内で月が不統一（前月データの月更新忘れの可能性）'))

    # R20: ファイル名の年月とA1（令和年）・D1（月）の不一致
    ym = re.search(r'(\d{4})年(\d{1,2})月', filename)
    if ym:
        seireki, month = int(ym.group(1)), int(ym.group(2))
        reiwa = seireki - REIWA_BASE
        for sname, h in headers.items():
            a1 = _to_int(h['A1']); d1 = _to_int(h['D1'])
            ng = []
            # A1は「令和年」または「西暦」のどちらかに一致すればOK
            if a1 is not None and a1 not in (reiwa, seireki):
                ng.append(f"A1=「{h['A1']}」（令和{reiwa}年={seireki}年 が正）")
            if d1 is not None and d1 != month:
                ng.append(f"D1=「{h['D1']}」（{month}月 が正）")
            if ng:
                issues.append(dict(rule='R20', floor=floor_label, name=sname, date='-',
                    level='error',
                    detail=f"ファイル名=「{filename}」/ {' / '.join(ng)}",
                    judge='❌ ファイル名の年月とシートの年月が不一致（前月ファイル提出ミスの可能性）'))

    # R22: C2（受給者証番号）・P2（障害支援区分）の空白・異常値
    for sname, h in headers.items():
        if not h['C2']:
            issues.append(dict(rule='R22', floor=floor_label, name=sname, date='-',
                level='error',
                detail='C2セル（受給者証番号）が空白',
                judge='❌ 受給者証番号が未入力（請求前に入力が必要）'))
        p2 = h['P2']
        if p2 is None or str(p2).strip() == '':
            issues.append(dict(rule='R22', floor=floor_label, name=sname, date='-',
                level='error',
                detail='P2セル（障害支援区分）が空白',
                judge='❌ 障害支援区分が未入力（請求前に入力が必要）'))
        else:
            p2n = _to_int(p2)
            if p2n is None or not (1 <= p2n <= 6):
                issues.append(dict(rule='R22', floor=floor_label, name=sname, date='-',
                    level='error',
                    detail=f'P2セル（障害支援区分）=「{p2}」（1〜6以外の値）',
                    judge='❌ 障害支援区分が1〜6以外の値です'))
    return issues

def extract_all(filepath):
    """実績記録票xlsm/xlsxから全利用者・全日のデータを読み取る

    戻り値: (result, warnings)
      result   : {シート名: {日: {列の値}}}
      warnings : 読み取れなかったシート・読み飛ばした行などの注意メッセージ
    """
    # data_only=True: 数式セルはキャッシュされた計算結果を読む
    # （従来は数式文字列がそのまま返り、チェックが素通りする恐れがあった）
    wb = load_workbook(filepath, read_only=True, data_only=True)
    result = {}
    warnings = []
    for sname in wb.sheetnames:
        if sname in EXCLUDE_SHEETS: continue
        ws = wb[sname]
        i2 = ws['I2'].value  # 簡易テンプレート検証用（氏名欄）
        days = {}
        short_rows = 0
        for i, row in enumerate(ws.iter_rows(min_row=9, max_row=39, values_only=True)):
            day = i + 1
            if len(row) < 20:
                short_rows += 1
                continue
            # E列：短期の退所時間（時刻 or 「-」「－」）
            e_raw = row[4]
            if e_raw is None or str(e_raw).strip() == '':
                e_val = ''
            elif str(e_raw).strip() in ('-', '－', '―', '−'):
                e_val = '-'
            else:
                e_val = str(e_raw).strip()
            days[day] = {
                'C': str(row[2]).strip()  if row[2]  else '',
                'E': bool(row[4]) and str(row[4]).strip() not in ('-','－','―','−'),
                'E_raw': e_val,
                'F': bool(row[5]),
                'I': row[8] in (1, '1'),
                'P': row[15] == '✓',
                'Q': row[16] == '✓',
                'R': row[17] == '✓',
                'T': str(row[19]).strip() if row[19] else '',
            }
        if days:
            result[sname] = days
            if short_rows:
                warnings.append(f'シート「{sname}」: 列数不足のため{short_rows}日分を読み飛ばしました（様式が想定と異なる可能性）')
            if not i2:
                warnings.append(f'シート「{sname}」: I2（氏名欄）が空欄です（様式が想定と異なる可能性）')
        else:
            warnings.append(f'シート「{sname}」: 9〜39行のデータを読み取れませんでした（このシートは未チェックです）')
    wb.close()
    return result, warnings

# --- チェック実行 ---
def run_checks(floor_label, data):
    """1ファイル分のチェックを実行してissuesリストを返す"""
    issues = []
    for name, days in data.items():
        for day in sorted(days.keys()):
            d = days[day]; dt = f'{day}日'

            # R1: 外泊中に夜勤加配
            if d['C'] in GAIHAKU_NG and d['F']:
                issues.append(dict(rule='R1', floor=floor_label, name=name, date=dt,
                    level='error',
                    detail=f"C列=「{d['C']}」/ F列（夜勤加配）=1",
                    judge='❌ 外泊中に夜勤職員加配加算が記録されています'))

            # R2: 住居外チェック漏れ
            if is_gaibu(d['T']) and not d['E']:
                issues.append(dict(rule='R2', floor=floor_label, name=name, date=dt,
                    level='error',
                    detail=f"備考=「{d['T']}」/ E列（住居外）=空白",
                    judge='⚠️ 住居外利用チェック漏れの可能性'))

            # R3: 住居外チェックあり・備考なし
            if d['E'] and not d['T']:
                issues.append(dict(rule='R3', floor=floor_label, name=name, date=dt,
                    level='warn',
                    detail="E列（住居外）=1 / T列（備考）=空白",
                    judge='❓ 住居外利用の理由が備考に未記載'))

            # R4: 外泊中に食事記録
            if d['C'] in GAIHAKU_NG:
                foods = [m for m, v in [('朝食',d['P']),('昼食',d['Q']),('夕食',d['R'])] if v]
                if foods:
                    issues.append(dict(rule='R4', floor=floor_label, name=name, date=dt,
                        level='error',
                        detail=f"C列=「{d['C']}」/ 食事✓あり：{'・'.join(foods)}",
                        judge=f"❌ 外泊中に食事発注記録があります（{'・'.join(foods)}）"))

            # R5: 日中外出日の昼食確認
            if is_lunch_out(d['T']) and d['Q']:
                issues.append(dict(rule='R5', floor=floor_label, name=name, date=dt,
                    level='info',
                    detail=f"備考=「{d['T']}」/ Q列（昼食）=✓",
                    judge='📌 日中外出日に昼食発注あり（意図的なら問題なし・要確認）'))

            # R6: 外泊終了の帰設時間未記載
            if d['C'] == '外泊終了' and not has_time(d['T']):
                issues.append(dict(rule='R6', floor=floor_label, name=name, date=dt,
                    level='warn',
                    detail=f"C列=「外泊終了」/ T列=「{d['T'] if d['T'] else '（空白）'}」",
                    judge='❓ 外泊終了日の備考に帰設時間の記載がありません'))

            # R7: 外泊開始の出発時間未記載
            if d['C'] == '外泊開始' and not has_time(d['T']):
                issues.append(dict(rule='R7', floor=floor_label, name=name, date=dt,
                    level='warn',
                    detail=f"C列=「外泊開始」/ T列=「{d['T'] if d['T'] else '（空白）'}」",
                    judge='❓ 外泊開始日の備考に出発時間の記載がありません'))

            # R8: 入院中の夜勤加配（既存ルール・NYUIN_NGは入院・入院開始）
            if d['C'] in NYUIN_NG and d['F']:
                issues.append(dict(rule='R8', floor=floor_label, name=name, date=dt,
                    level='error',
                    detail=f"C列=「{d['C']}」/ F列（夜勤加配）=1",
                    judge='❌ 入院中に夜勤職員加配加算が記録されています'))

            # R9: 入院開始・退院・退去日のT列時間未記載
            if d['C'] in TIME_REQ and not has_time(d['T']):
                issues.append(dict(rule='R9', floor=floor_label, name=name, date=dt,
                    level='warn',
                    detail=f"C列=「{d['C']}」/ T列=「{d['T'] if d['T'] else '（空白）'}」",
                    judge=f"❓ 「{d['C']}」の備考に時間の記載がありません"))

            # R10: 退去・入院→退去のF列夜勤加配
            # ※退院はGHに戻るため対象外、R8と対象が重複しないよう退去系のみ
            if d['C'] in {'退去', '入院→退去'} and d['F']:
                issues.append(dict(rule='R10', floor=floor_label, name=name, date=dt,
                    level='error',
                    detail=f"C列=「退去」/ F列（夜勤加配）=1",
                    judge=f"❌ 「{d['C']}」の日に夜勤職員加配加算が記録されています"))

            # R11: 退去・入院→退去の行先未記載
            if d['C'] in {'退去', '入院→退去'} and not has_destination(d['T']):
                issues.append(dict(rule='R11', floor=floor_label, name=name, date=dt,
                    level='warn',
                    detail=f"C列=「退去」/ T列=「{d['T'] if d['T'] else '（空白）'}」",
                    judge=f"❓ 「{d['C']}」の備考に行先（自宅・実家・施設名等）の記載がありません"))


            # R16: 医療連携体制加算の誤算定（入院・外泊中）
            if d['C'] in I_EXEMPT and d['I']:
                issues.append(dict(rule='R16', floor=floor_label, name=name, date=dt,
                    level='error',
                    detail=f"C列=「{d['C']}」/ I列（医療連携加算）=1",
                    judge=f"❌ 「{d['C']}」中に医療連携体制加算が算定されています"))
    # ── R12・R13・R14: 月単位の連続性チェック ──────────────
    for name, days in data.items():
        c_map = {day: days[day]['C'] for day in sorted(days.keys()) if days[day]['C']}
        day_list = sorted(c_map.keys())

        # R12: 外泊開始後に外泊終了がない（月末まで外泊中）
        gaihaku_start = None
        for day in day_list:
            c = c_map[day]
            if c == '外泊開始':
                gaihaku_start = day
            elif c in {'外泊終了', '退去', '入院→退去', '退院'}:
                gaihaku_start = None
        if gaihaku_start is not None:
            t = days[gaihaku_start]['T']
            has_plan = bool(DATE_PATTERN.search(t)) if t else False
            if not has_plan:
                issues.append(dict(
                    rule='R12', floor=floor_label, name=name,
                    date=f'{gaihaku_start}日',
                    level='warn',
                    detail=f"外泊開始後、月末まで外泊終了なし / 外泊開始日備考=「{t if t else '（空白）'}」",
                    judge='❓ 外泊が月をまたぐ場合、外泊開始日の備考に帰設予定日の記載が必要です'
                ))

        # R13: 入院開始後に退院・退去がない（月末も入院継続中）
        nyuin_start = None
        for day in day_list:
            c = c_map[day]
            if c == '入院開始':
                nyuin_start = day
            elif c in {'退院', '退去', '入院→退去'}:
                nyuin_start = None
        if nyuin_start is not None:
            issues.append(dict(
                rule='R13', floor=floor_label, name=name,
                date=f'{nyuin_start}日',
                level='info',
                detail=f"{nyuin_start}日に入院開始 / 月末時点で退院・退去の記録なし",
                judge='📋 月末時点で入院継続中（退院日未定の可能性。情報として抽出）'
            ))

        # R14: 入院中に食事発注あり（day単位だが月単位ループでまとめる）
        for day in sorted(days.keys()):
            d = days[day]
            if d['C'] == '入院':
                foods = [m for m, v in [('朝食',d['P']),('昼食',d['Q']),('夕食',d['R'])] if v]
                if foods:
                    biko = d['T']
                    if biko:
                        judge = f'📋 入院中に食事発注あり（備考:「{biko}」確認済みの可能性）'
                    else:
                        judge = '📋 入院中に食事発注あり（備考なし・要確認）'
                    issues.append(dict(
                        rule='R14', floor=floor_label, name=name,
                        date=f'{day}日',
                        level='info',
                        detail=f"C列=「入院」/ 食事✓あり：{'・'.join(foods)} / 備考=「{biko if biko else '（空白）'}」",
                        judge=judge
                    ))

        # R17: 短期入所ファイルの月跨ぎ利用チェック（月単位）
        if '短期' in floor_label:
            for day in sorted(days.keys(), reverse=True):
                d_r17 = days[day]
                if day < 25: break
                if d_r17['E_raw'] == '-':
                    t_r17 = d_r17['T']
                    has_tsuki = any(k in t_r17 for k in TSUKI_MATAGI_KW) if t_r17 else False
                    if not has_tsuki:
                        issues.append(dict(
                            rule='R17', floor=floor_label, name=name,
                            date=f'{day}日',
                            level='warn',
                            detail=(
                                f"{day}日のE列（退所時間）=「-」退所時間未記録 / "
                                f"T列備考=「{t_r17 if t_r17 else '（空白）'}」"
                            ),
                            judge='❓ 月跨ぎ利用の可能性あり（退所時間なし・備考に月跨ぎ記載なし）'
                        ))
                    break
    # 重複issue除去（安全弁：同一ルール・同一日・同一人の重複を1件にまとめる）
    seen = set()
    deduped = []
    for i in issues:
        key = (i['rule'], i['floor'], i['name'], i['date'])
        if key not in seen:
            seen.add(key)
            deduped.append(i)
    return deduped, data  # R15のためdataも返す


# ============================================================
# Excel レポート生成
# ============================================================

# スタイル定数
RED_F    = PatternFill("solid", fgColor="FFCCCC")
ORANGE_F = PatternFill("solid", fgColor="FFE0B2")
YELLOW_F = PatternFill("solid", fgColor="FFFDE7")
GREEN_F  = PatternFill("solid", fgColor="E2EFDA")
PINK_F   = PatternFill("solid", fgColor="F8BBD0")
BLUE_F   = PatternFill("solid", fgColor="DDEEFF")  # 情報・抽出のみ
HDR_F    = PatternFill("solid", fgColor="1F4E79")
HDR2_F   = PatternFill("solid", fgColor="2E75B6")

def tb():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_L = Alignment(horizontal='left',   vertical='center', wrap_text=False)
HF = Font(name='メイリオ', bold=True, color="FFFFFF", size=10)
BF = Font(name='メイリオ', size=10)
RF = Font(name='メイリオ', size=10, color="C00000", bold=True)
OF = Font(name='メイリオ', size=10, color="BF5700", bold=True)
GF = Font(name='メイリオ', size=10, color="1F7A1F", bold=True)

LEVEL_FILL  = {'error': RED_F,      'warn': ORANGE_F, 'info': YELLOW_F}
LEVEL_FONT  = {'error': RF,         'warn': OF,        'info': BF}
LEVEL_LABEL = {'error': '❌ エラー', 'warn': '⚠️ 警告',  'info': '📌 確認'}

RULE_COLORS = {
    'R1': PINK_F,   'R2': RED_F,    'R3': ORANGE_F,
    'R4': PINK_F,   'R5': YELLOW_F, 'R6': ORANGE_F,
    'R7': ORANGE_F, 'R8': PINK_F,   'R9': ORANGE_F,
    'R10': PINK_F,  'R11': ORANGE_F,
    'R12': ORANGE_F,'R13': BLUE_F,  'R14': BLUE_F,  'R15': RED_F,
    'R16': RED_F,   'R17': ORANGE_F,'R18': RED_F,   'R19': RED_F,
    'R20': RED_F,   'R21': RED_F,   'R22': RED_F,
}

# ルール略称（サマリーの列見出し用・全レポート共通）
RULE_SHORT = {
    "R1":"R1\n外泊x夜勤加配", "R2":"R2\n住居外\nChk漏れ",
    "R3":"R3\n住居外\n備考なし","R4":"R4\n外泊x食事",
    "R5":"R5\n外出x昼食",      "R6":"R6\n外泊終了\n時間なし",
    "R7":"R7\n外泊開始\n時間なし","R8":"R8\n入院x夜勤加配",
    "R9":"R9\n退院等\n時間なし","R10":"R10\n退去x\n夜勤加配",
    "R11":"R11\n退去\n行先なし","R12":"R12\n外泊\n未終了",
    "R13":"R13\n入院\n継続中", "R14":"R14\n入院x食事",
    "R15":"R15\n加配\n算定漏れ","R16":"R16\n医療連携\n誤算定",
    "R17":"R17\n短期\n月跨ぎ", "R18":"R18\n氏名\n不一致",
    "R19":"R19\nD1\n不統一",   "R20":"R20\n年月\n不一致",
    "R21":"R21\n医療連携\n算定漏れ","R22":"R22\n番号区分\n空白",
}

def auto_col_width(ws, min_w=8, max_w=60):
    """各列の内容に合わせて列幅を自動調整する（結合セル対応）"""
    from openpyxl.utils import get_column_letter as _gcl
    col_max = {}
    for row in ws.iter_rows():
        for cell in row:
            # 結合セルや数式セルはスキップ
            if cell.value is None or str(cell.value).startswith('='):
                continue
            try:
                col_idx = cell.column
            except AttributeError:
                continue
            val = str(cell.value)
            # 日本語・全角は2文字分、半角は1文字分
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            if col_idx not in col_max or length > col_max[col_idx]:
                col_max[col_idx] = length
    for col_idx, max_len in col_max.items():
        adjusted = max(min_w, min(max_len + 2, max_w))
        ws.column_dimensions[_gcl(col_idx)].width = adjusted

def set_title(ws, t, sub, n_cols):
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = t
    ws['A1'].font = Font(name='メイリオ', bold=True, size=13, color="1F4E79")
    ws['A1'].alignment = ALIGN_C; ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A2:{get_column_letter(n_cols)}2')
    ws['A2'] = sub
    ws['A2'].font = Font(name='メイリオ', size=10, color="555555")
    ws['A2'].alignment = ALIGN_L; ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 5

def set_hrow(ws, r, heads, widths, fill=None):
    for col, (h, w) in enumerate(zip(heads, widths), 1):
        c = ws.cell(row=r, column=col, value=h)
        c.fill = fill or HDR_F; c.font = HF
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        c.border = tb()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 20

def write_rows(ws, issues, start=5, detail_col=4):
    if not issues:
        ws.merge_cells(f'A{start}:F{start}')
        c = ws.cell(row=start, column=1,
                    value='✅ 問題なし：このルールに該当する不備はありませんでした')
        c.font = GF; c.alignment = ALIGN_L; ws.row_dimensions[start].height = 22
        return
    for r, i in enumerate(issues, start):
        fill = LEVEL_FILL[i['level']]
        jf   = LEVEL_FONT[i['level']]
        vals = [i['floor'], i['name'], i['date'], i['detail'], i['judge'], LEVEL_LABEL[i['level']]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val); c.fill = fill
            c.font = jf if col in (5, 6) else BF
            c.alignment = ALIGN_C if col != detail_col else ALIGN_L
            c.border = tb()
        ws.row_dimensions[r].height = 18



def add_gh_sheet(wb_out, gh_name, year_month, all_issues):
    """
    既存のWorkbookに施設ごとのシートを1枚追加する。
    シート名 = GH名（31文字以内に切り詰め）
    """
    err_n  = sum(1 for i in all_issues if i['level'] == 'error')
    warn_n = sum(1 for i in all_issues if i['level'] == 'warn')
    info_n = sum(1 for i in all_issues if i['level'] == 'info')

    # シート名はExcelの制限(31文字)に合わせて切り詰め
    sheet_title = gh_name[:28] + ('...' if len(gh_name) > 28 else '')
    # 同名シートが既にある場合は連番を付ける
    existing = [s.title for s in wb_out.worksheets]
    if sheet_title in existing:
        sheet_title = sheet_title[:25] + f'_{len(existing)}'
    ws = wb_out.create_sheet(title=sheet_title)

    # タイトル行
    set_title(ws,
        f'{gh_name}　{year_month}',
        f'❌エラー：{err_n}件　⚠️警告：{warn_n}件　📌確認：{info_n}件　合計：{len(all_issues)}件',
        7)

    # ヘッダー
    set_hrow(ws, 4,
        ['フロア/ファイル','利用者名','日付','チェックルール','チェック内容（詳細）','判定','重要度'],
        [20, 14, 14, 18, 46, 36, 10])

    # データ行
    if not all_issues:
        ws.merge_cells('A5:G5')
        c = ws.cell(row=5, column=1, value='✅ この施設のチェック結果に問題はありませんでした')
        c.font = Font(name='メイリオ', size=11, color="1F7A1F", bold=True)
        c.alignment = ALIGN_L
        ws.row_dimensions[5].height = 22
    else:
        for r, i in enumerate(all_issues, 5):
            fill = LEVEL_FILL[i['level']]
            jf   = LEVEL_FONT[i['level']]
            lv_label = LEVEL_LABEL[i['level']]
            vals = [i['floor'], i['name'], i['date'],
                    RULES[i['rule']], i['detail'], i['judge'], lv_label]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = fill
                c.font = jf if col in (6, 7) else BF
                c.alignment = ALIGN_C if col not in (4, 5) else ALIGN_L
                c.border = tb()
    ws.freeze_panes = 'A5'
    auto_col_width(ws)  # 列幅自動調整
    return ws

def build_report(gh_name, year_month, all_issues, all_data):
    """
    GH名・対象年月・全issuesから1つのExcelレポートを生成してWorkbookを返す
    all_data: {floor_label: data_dict}
    """
    wb_out = Workbook()

    err_n  = sum(1 for i in all_issues if i['level'] == 'error')
    warn_n = sum(1 for i in all_issues if i['level'] == 'warn')
    info_n = sum(1 for i in all_issues if i['level'] == 'info')

    # ── シート① 全件一覧 ─────────────────────────────────
    ws1 = wb_out.active; ws1.title = "①全チェック一覧"
    set_title(ws1,
        f'{gh_name} 実績記録票 整合性チェックレポート　{year_month}',
        f'❌エラー：{err_n}件　⚠️警告：{warn_n}件　📌確認：{info_n}件　合計：{len(all_issues)}件',
        7)
    set_hrow(ws1, 4,
        ['フロア/ファイル','利用者名','日付','チェックルール','チェック内容（詳細）','判定','重要度'],
        [16, 14, 8, 16, 46, 36, 10])
    for r, i in enumerate(all_issues, 5):
        fill = LEVEL_FILL[i['level']]; jf = LEVEL_FONT[i['level']]
        vals = [i['floor'], i['name'], i['date'],
                RULES[i['rule']], i['detail'], i['judge'], LEVEL_LABEL[i['level']]]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=r, column=col, value=val); c.fill = fill
            c.font = jf if col in (6, 7) else BF
            c.alignment = ALIGN_C if col != 5 else ALIGN_L; c.border = tb()
        ws1.row_dimensions[r].height = 18
    ws1.freeze_panes = 'A5'

    # ── シート② R1〜R22 個別タブ ──────────────────────────
    rule_sheet_names = {
        'R1':'R1_外泊×夜勤加配', 'R2':'R2_住居外チェック漏れ',
        'R3':'R3_住居外備考なし', 'R4':'R4_外泊×食事記録',
        'R5':'R5_外出日昼食確認', 'R6':'R6_外泊終了帰設時間',
        'R7':'R7_外泊開始出発時間',
        'R8':'R8_入院×夜勤加配',
        'R9':'R9_入院開始退院退去時間',
        'R10':'R10_退去×夜勤加配',
        'R11':'R11_退去行先未記載',
        'R12':'R12_外泊月末未終了',
        'R13':'R13_入院継続確認',
        'R14':'R14_入院中食事発注',
        'R15':'R15_夜勤加配算定漏れ',
        'R16':'R16_医療連携誤算定',
        'R17':'R17_短期月跨ぎ',
        'R18':'R18_氏名不一致',
        'R19':'R19_月不統一',
        'R20':'R20_年月不一致',
        'R21':'R21_医療連携算定漏れ',
        'R22':'R22_番号区分空白',
    }
    for rule_id, sheet_name in rule_sheet_names.items():
        target = [i for i in all_issues if i['rule'] == rule_id]
        ws = wb_out.create_sheet(sheet_name)
        set_title(ws,
            f'【{rule_id}】{RULES[rule_id]}　{year_month}',
            f'{RULE_DESCS[rule_id]}　　件数：{len(target)}件', 6)
        set_hrow(ws, 4,
            ['フロア/ファイル','利用者名','日付','チェック内容（詳細）','判定','重要度'],
            [16, 14, 8, 50, 38, 10])
        write_rows(ws, target)
        ws.freeze_panes = 'A5'

    # ── シート③ 利用者別サマリー ─────────────────────────
    rule_ids = [f'R{n}' for n in range(1, 23)]
    ws_s = wb_out.create_sheet("③利用者別サマリー")
    set_title(ws_s,
        f'利用者別 チェック結果サマリー　{year_month}',
        'R1〜R22 全ルールの不備件数を利用者ごとに集計', len(rule_ids) + 3)
    set_hrow(ws_s, 4,
        ['フロア/ファイル','利用者名'] + [RULE_SHORT[rid] for rid in rule_ids] + ['合計'],
        [16, 14] + [9] * len(rule_ids) + [9])

    sm = defaultdict(lambda: {'f': '', **{rid: 0 for rid in rule_ids}})
    for i in all_issues:
        k = (i['floor'], i['name']); sm[k]['f'] = i['floor']; sm[k][i['rule']] += 1
    # 不備ゼロの利用者も表示
    for floor_label, data in all_data.items():
        for nm in data:
            k = (floor_label, nm)
            if k not in sm: sm[k]['f'] = floor_label

    r = 5
    for (fl, nm), d in sorted(sm.items()):
        tot = sum(d[rid] for rid in rule_ids)
        fill = RED_F if tot >= 5 else (ORANGE_F if tot >= 2 else (YELLOW_F if tot == 1 else GREEN_F))
        for col, val in enumerate([fl, nm] + [d[rid] for rid in rule_ids] + [tot], 1):
            c = ws_s.cell(row=r, column=col, value=val)
            c.fill = fill; c.font = BF; c.alignment = ALIGN_C; c.border = tb()
        ws_s.row_dimensions[r].height = 20; r += 1
    ws_s.freeze_panes = 'A5'

    # ── シート④ ルール説明 ───────────────────────────────
    ws_d = wb_out.create_sheet("④チェックルール説明")
    for col, w in zip('ABCD', [6, 24, 20, 56]):
        ws_d.column_dimensions[col].width = w
    ws_d.merge_cells('A1:D1')
    ws_d['A1'] = 'チェックルール一覧（R1〜R22）'
    ws_d['A1'].font = Font(name='メイリオ', bold=True, size=13, color="1F4E79")
    ws_d['A1'].alignment = ALIGN_C; ws_d.row_dimensions[1].height = 28
    set_hrow(ws_d, 2, ['ID','ルール名','対象列','判定基準・説明'], [6,24,20,56], HDR2_F)

    rule_detail = [
        ('R1','外泊中の夜勤加配加算',    'C列・F列',
         'C列が「外泊」「外泊開始」のとき → F列（夜勤職員加配加算）は空欄が正。外泊中はGH不在のため加配加算は算定不可。'),
        ('R2','住居外利用チェック漏れ',  'E列・T列',
         'T列（備考）に受診・外出・デイ利用等の記載があるとき → E列（住居外利用）にチェックが必要。未チェックは加算誤りの可能性あり。'),
        ('R3','住居外利用の備考未記載',  'E列・T列',
         'E列（住居外利用）にチェックがあるとき → T列（備考）に外出先・時間等の理由を記載する必要がある。'),
        ('R4','外泊中の食事発注記録',    'C列・P/Q/R列',
         'C列が「外泊」「外泊開始」のとき → P/Q/R列（朝食・昼食・夕食の発注✓）は空欄が正。外泊中のGH食事発注は原則不要。'),
        ('R5','日中外出日の昼食発注確認','E列・Q列・T列',
         'T列（備考）にデイ・就労支援等の日中外出記載があり昼食✓がある場合は要確認。施設側で昼食提供なら不要の可能性あり。「昼発注あり」の特記があれば除外。'),
        ('R6','外泊終了日の帰設時間未記載','C列・T列',
         'C列が「外泊終了」のとき → T列（備考）に帰設時間（例：15:30帰設）の記載が必要。数字:数字 または 数字時 の形式を確認。'),
        ('R7','外泊開始日の出発時間未記載','C列・T列',
         'C列が「外泊開始」のとき → T列（備考）に出発時間（例：14:00外泊）の記載が必要。数字:数字 または 数字時 の形式を確認。'),
        ('R8','入院中の夜勤加配加算',     'C列・F列',
         'C列が「入院」「入院開始」のとき → F列（夜勤職員加配加算）は空欄が正。入院中はGH不在のため算定不可。'),
        ('R9','入院開始・退院・退去等の時間未記載','C列・T列',
         'C列が「入院開始」「退院」「退去」「入院→退去」のとき → T列（備考）に時間の記載が必要（例：15時入院開始、9時退去）。「入院」継続中は不要。'),
        ('R10','退去・入院→退去の夜勤加配', 'C列・F列',
         'C列が「退去」「入院→退去」のとき → F列（夜勤職員加配加算）は空欄が正。※退院はGH戻りのため対象外。'),
        ('R11','退去・入院→退去の行先未記載','C列・T列',
         'C列が「退去」「入院→退去」のとき → T列（備考）に行先（自宅・実家・施設名等）の記載が必要。'),
        ('R12','外泊が月末まで未終了','C列・T列（月単位）',
         '外泊開始後に外泊終了が月内にない場合、外泊開始日の備考に帰設予定日（例:3/31まで・4/5予定）の記載が必要。'),
        ('R13','入院継続中（退院・退去なし）','C列（月単位）',
         '入院開始後に退院・退去の記録が月内にない。退院日未定の可能性あり。情報として抽出のみ。'),
        ('R14','入院中に食事発注あり','C列・P/Q/R列（月単位）',
         'C列=「入院」のとき食事✓がある。突発入院でキャンセル不可の可能性あり。備考に記載があれば確認済みの可能性。情報として抽出のみ。'),
        ('R15','夜勤加配加算の算定漏れ','F列（ファイル内比較）',
         RULE_DESCS['R15']),
        ('R16','医療連携体制加算の誤算定','C列・I列',
         RULE_DESCS['R16']),
        ('R17','短期入所の月跨ぎ利用','E列・T列（短期のみ）',
         RULE_DESCS['R17']),
        ('R18','シート名とI2の氏名不一致','シート名・I2セル',
         RULE_DESCS['R18']),
        ('R19','D1（月）の不統一','D1セル（ファイル内比較）',
         RULE_DESCS['R19']),
        ('R20','ファイル名と年月の不一致','ファイル名・A1・D1セル',
         RULE_DESCS['R20']),
        ('R21','医療連携体制加算の算定漏れ','I列（ファイル内比較）',
         RULE_DESCS['R21']),
        ('R22','受給者証番号・支援区分の空白','C2・P2セル',
         RULE_DESCS['R22']),
    ]
    for r, (rid, rname, rcol, rdesc) in enumerate(rule_detail, 3):
        rfill = RULE_COLORS[rid]
        for col, val in enumerate([rid, rname, rcol, rdesc], 1):
            c = ws_d.cell(row=r, column=col, value=val)
            c.fill = rfill; c.font = BF
            c.alignment = ALIGN_C if col < 4 else ALIGN_L; c.border = tb()
        ws_d.row_dimensions[r].height = 50

    return wb_out


# ============================================================
# メイン処理: フォルダを走査して一括チェック
# ============================================================

# GH名末尾のフロア・短期表記を除去する正規表現
_FLOOR_SUFFIX = re.compile(
    r'[0-9０-９]+[FＦfｆ階]$'  # 末尾の「1F」「２F」等
    r'|短期$'                   # 末尾の「短期」
)

def _shorten_floor_label(stem):
    """ファイルstemから短いフロアラベルを生成（例: _萩丘ⅰ_1F_実績記録票_2026年3月 → 萩丘ⅰ 1F）"""
    import re as _re
    s = stem.lstrip('_')
    parts = s.split('_')
    # 「実績記録票」より前の部分を取得
    try:
        idx = parts.index('実績記録票')
        key_parts = parts[:idx]
    except ValueError:
        key_parts = parts[:2]
    # GH名とフロア（1F/2F/短期等）を空白区切りで結合
    return ' '.join(p for p in key_parts if p)

def _normalize_gh_name(raw):
    """GH名末尾のフロア番号・短期を除去して正規化する"""
    name = _FLOOR_SUFFIX.sub('', raw).rstrip('_').strip()
    return name if name else raw

def _rev_number(name):
    """「修正」表記の版数を数値で返す（修正なし=0、修正=1、修正2/修正②=2 等）"""
    m = re.search(r'修正[（(]?([0-9０-９]{1,2}|[①-⑳])?', name)
    if not m: return 0
    tok = m.group(1)
    if not tok: return 1
    if '①' <= tok <= '⑳':
        return ord(tok) - ord('①') + 1
    return int(tok.translate(str.maketrans('０１２３４５６７８９', '0123456789')))

def _pick_latest(files):
    """同一フロア・同一年月に修正版がある場合は最新の修正版を優先する
    （修正② > 修正 > 無印。版数は数値として比較する）"""
    def sort_key(f):
        return (_rev_number(f.name), f.name)
    return sorted(files, key=sort_key)[-1]

def _add_summary_sheet(wb, rows):
    """全GH横断サマリーシートをWorkbookの先頭に追加（ルール別件数も表示）"""
    ws = wb.create_sheet(title="\U0001f4ca全施設サマリー", index=0)
    R2F = PatternFill("solid", fgColor="FFCCCC")
    O2F = PatternFill("solid", fgColor="FFE0B2")
    Y2F = PatternFill("solid", fgColor="FFFDE7")
    G2F = PatternFill("solid", fgColor="E2EFDA")
    P2F = PatternFill("solid", fgColor="F8BBD0")
    B2F = PatternFill("solid", fgColor="DDEEFF")
    W2F = PatternFill("solid", fgColor="F8F8F8")
    rule_ids = [f"R{n}" for n in range(1, 23)]
    RSHORT = RULE_SHORT
    RFILL = {
        "R1":P2F,"R2":R2F,"R3":O2F,"R4":P2F,"R5":Y2F,
        "R6":O2F,"R7":O2F,"R8":P2F,"R9":O2F,"R10":P2F,
        "R11":O2F,"R12":O2F,"R13":B2F,"R14":B2F,"R15":R2F,
        "R16":R2F,"R17":O2F,"R18":R2F,"R19":R2F,"R20":R2F,
        "R21":R2F,"R22":R2F,
    }
    ERR_R = {"R1","R2","R4","R8","R10","R15","R16","R18","R19","R20","R21","R22"}
    n_cols = 6 + len(rule_ids)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = f"全施設 チェック結果サマリー\u3000実行日: {datetime.today().strftime(chr(37)+chr(89)+chr(47)+chr(37)+chr(109)+chr(47)+chr(37)+chr(100))}"
    ws["A1"].font = Font(name="\u30e1\u30a4\u30ea\u30aa", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    BH = ["GH名","対象年月","\u274c エラー","\u26a0\ufe0f 警告","\U0001f4cc 確認","合計"]
    BW = [16,10,9,9,9,8]
    RW = [8]*len(rule_ids)
    for col,(h,w) in enumerate(zip(BH+[RSHORT[r] for r in rule_ids],BW+RW),1):
        ir = col>len(BH)
        rid = rule_ids[col-len(BH)-1] if ir else None
        c = ws.cell(row=2,column=col,value=h)
        c.fill = RFILL[rid] if ir else HDR_F
        c.font = Font(name="\u30e1\u30a4\u30ea\u30aa",bold=True,
                      size=9 if ir else 10,color="000000" if ir else "FFFFFF")
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = tb()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 44
    for r,row in enumerate(rows,3):
        tot=row.get("合計",0); e=row.get("エラー",0)
        w=row.get("警告",0); i=row.get("確認",0)
        fl=(R2F if tot>=10 else O2F if tot>=3 else Y2F if tot>=1 else G2F)
        for col,val in enumerate([row.get("GH",""),row.get("年月",""),
                                   e or "",w or "",i or "",tot or ""],1):
            c=ws.cell(row=r,column=col,value=val)
            c.fill=fl
            c.font=Font(name="\u30e1\u30a4\u30ea\u30aa",bold=(col==1),size=10)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=tb()
        rc=row.get("rules",{})
        for ri,rid in enumerate(rule_ids):
            col=len(BH)+ri+1; cnt=rc.get(rid,0)
            c=ws.cell(row=r,column=col,value=cnt if cnt else "")
            c.fill=RFILL[rid] if cnt else W2F
            c.font=Font(name="\u30e1\u30a4\u30ea\u30aa",bold=bool(cnt),size=10,
                        color="C00000" if (cnt and rid in ERR_R) else "000000")
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=tb()
        ws.row_dimensions[r].height=20
    ws.freeze_panes="A3"

def _add_skipped_sheet(wb, entries):
    """チェックできなかったファイル・シートや、ファイル選別の判断を一覧するシートを追加。
    「問題なし」に見えても、ここに載っているものは検査されていないことを可視化する。"""
    ws = wb.create_sheet(title='⚠未チェック一覧')
    set_title(ws,
        'チェックできなかった項目・ファイル選別の記録',
        '※「問題なし」のレポートでも、ここに記載のあるファイル・シートは検査されていません', 3)
    set_hrow(ws, 4, ['GH名', 'ファイル名', '内容'], [18, 44, 90])
    if not entries:
        ws.merge_cells('A5:C5')
        c = ws.cell(row=5, column=1, value='✅ すべてのファイル・シートをチェックしました')
        c.font = GF; c.alignment = ALIGN_L
        ws.row_dimensions[5].height = 22
    else:
        for r, e in enumerate(entries, 5):
            for col, val in enumerate([e['gh'], e['file'], e['reason']], 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = ORANGE_F; c.font = BF
                c.alignment = ALIGN_L; c.border = tb()
            ws.row_dimensions[r].height = 18
    ws.freeze_panes = 'A5'

def _extract_floor(f):
    """ファイル名からフロア表記（1F・２Ｆ・短期 等）を抽出する"""
    parts = f.stem.lstrip('_').split('_')
    if parts:
        m = _FLOOR_SUFFIX.search(parts[0])
        if m: return m.group()
    m = re.search(r'_([0-9０-９]+[FＦfｆ階]|短期)_', f.name)
    return m.group(1) if m else ''

def find_jisseki_files(base_folder):
    """
    base_folder 直下の実績記録票ファイルをGH名ごとにグループ化する。
    ・ファイル名末尾の「1F」「2F」「短期」等を除去してGH名を統一する
    ・同一フロア・年月で修正版（修正①等）がある場合は最新の修正版を優先する

    ファイル名例:
      _宇都宮1F_実績記録票_2026年3月.xlsm    → GH名「宇都宮」
      _宇都宮２F_実績記録票_2026年3月.xlsm   → GH名「宇都宮」（同じGHとして集約）
      _RASIEL気賀1F_実績記録票_2026年3月.xlsm → GH名「RASIEL気賀」

    戻り値: ( { 'GH名': [ファイルパスのリスト] },
              [採用・除外の判断メモのリスト] )
    """
    base = Path(base_folder)
    notes = []

    # ── Step1: 全ファイルを収集してGH名を正規化 ──
    raw_groups = defaultdict(list)   # { gh_name: [Path, ...] }
    for f in sorted(base.iterdir()):
        if f.is_dir(): continue
        if f.name.startswith('~$'): continue
        if f.suffix not in TARGET_EXTENSIONS: continue
        if 'チェックレポート' in f.name: continue
        if 'チェックサマリー' in f.name: continue
        if 'チェック設定' in f.name: continue   # キーワード設定ファイルは対象外

        stem  = f.stem.lstrip('_')
        parts = stem.split('_')
        raw   = parts[0] if parts else f.stem
        gh_name = _normalize_gh_name(raw)
        raw_groups[gh_name].append(f)

    # ── Step2: 同一GH内で修正版が存在するファイルを絞り込む ──
    # 同一キー（フロア+年月）で複数ある場合は最新の修正版を優先
    gh_files = defaultdict(list)
    for gh_name, files in raw_groups.items():
        sub_groups = defaultdict(list)
        for f in files:
            ym = re.search(r'\d{4}年\d{1,2}月', f.name)
            sub_key = f"{_extract_floor(f)}_{ym.group() if ym else ''}"
            sub_groups[sub_key].append(f)

        for sub_key, sub_files in sorted(sub_groups.items()):
            chosen = _pick_latest(sub_files)
            gh_files[gh_name].append(chosen)
            if len(sub_files) > 1:
                dropped = [x.name for x in sub_files if x is not chosen]
                notes.append(
                    f'{gh_name}: 同一フロア・年月のファイルが{len(sub_files)}件 → '
                    f'「{chosen.name}」を採用 / 除外: {"、".join(dropped)}'
                )

    return gh_files, notes



def run_all(base_folder, output_folder):
    """実績フォルダ内のファイルをGH別にまとめてチェックし、レポートを出力する

    出力:
      ・【全GH】チェックサマリー_{年月}_{実行日}.xlsx
          … 全施設横断サマリー＋施設別シート＋未チェック一覧
      ・{GH名}_チェックレポート_{年月}_{実行日}.xlsx
          … 施設ごとの詳細レポート（全件一覧・ルール別タブ・利用者別サマリー付き）
    """
    Path(output_folder).mkdir(parents=True, exist_ok=True)
    today = datetime.today().strftime('%Y%m%d')

    # キーワード設定（チェック設定.xlsx）があれば読み込む
    if load_keyword_config(base_folder):
        print(f"🔧 キーワード設定を読み込みました: {Path(base_folder) / CONFIG_FILENAME}")
    else:
        print(f"🔧 組み込みのキーワードを使用します")
        print(f"   （{CONFIG_FILENAME} を実績フォルダに置くと現場で編集できます。"
              f"雛形は mod.create_config_template(実績フォルダ) で作成）")

    gh_files, scan_notes = find_jisseki_files(base_folder)
    if not gh_files:
        print(f"⚠️  対象ファイルが見つかりませんでした: {base_folder}")
        print(f"   xlsm / xlsx ファイルがフォルダ直下に置かれているか確認してください")
        return

    print(f"{'='*60}")
    print(f"  実績記録票 整合性チェック開始")
    print(f"  対象フォルダ: {base_folder}")
    print(f"  検出GH数: {len(gh_files)} / ファイル総数: {sum(len(v) for v in gh_files.values())}")
    print(f"{'='*60}\n")

    # ファイル選別（修正版採用等）の判断を表示・記録
    skipped = []   # 未チェック一覧シート用: dict(gh=, file=, reason=)
    for n in scan_notes:
        print(f"  📂 {n}")
        skipped.append(dict(gh='（ファイル選別）', file='', reason=n))
    if scan_notes:
        print()

    summary_rows = []  # 全GH横断サマリー用

    # 全施設を1ファイルにまとめるWorkbookを事前に作成
    wb_all = Workbook()
    wb_all.remove(wb_all.active)  # デフォルトシートを削除

    for gh_name, files in sorted(gh_files.items()):
        print(f"▶ {gh_name}  ({len(files)}ファイル)")

        all_issues = []
        all_data   = {}   # {floor_label: data}
        year_month = ''

        for fpath in files:
            # ファイル名からフロア/年月を推定
            floor_label = fpath.stem   # 拡張子なしのファイル名をラベルに使用

            # 年月をファイル名から取得（例: 2026年3月）
            ym_match = re.search(r'(\d{4})年(\d{1,2})月', fpath.name)
            if ym_match and not year_month:
                year_month = f"{ym_match.group(1)}年{ym_match.group(2)}月"

            # floor_labelをシンプルな表示に変換（例: 萩丘ⅰ 1F）
            display_label = _shorten_floor_label(floor_label)

            # ── 読み取り ──
            try:
                data, warnings = extract_all(str(fpath))
            except Exception as e:
                print(f"    ❌ 読み取りエラー: {fpath.name}  ({e})")
                traceback.print_exc()
                skipped.append(dict(gh=gh_name, file=fpath.name,
                                    reason=f'読み取りエラーのため未チェック: {e}'))
                continue

            for w in warnings:
                print(f"    ⚠️  {fpath.name}: {w}")
                skipped.append(dict(gh=gh_name, file=fpath.name, reason=w))

            if not data:
                print(f"    ⚠️  シートが見つかりません: {fpath.name}")
                skipped.append(dict(gh=gh_name, file=fpath.name,
                                    reason='読み取れるシートがないため未チェック'))
                continue

            # ── チェック実行 ──
            try:
                issues, _ = run_checks(display_label, data)
                issues.extend(run_header_checks(display_label, str(fpath), fpath.name))
            except Exception as e:
                # 読み取りは成功しているのにここで落ちる場合は
                # チェックロジック側の不具合の可能性が高い
                print(f"    ❌ チェック処理エラー（ツールの不具合の可能性）: {fpath.name}  ({e})")
                traceback.print_exc()
                skipped.append(dict(gh=gh_name, file=fpath.name,
                                    reason=f'チェック処理エラーのため未チェック（ツールの不具合の可能性）: {e}'))
                continue

            all_issues.extend(issues)
            all_data[display_label] = data
            err = sum(1 for i in issues if i['level']=='error')
            wrn = sum(1 for i in issues if i['level']=='warn')
            print(f"    ✓ {fpath.name}  →  ❌{err}件 ⚠️{wrn}件")

        if not all_data:
            print(f"    スキップ（有効なファイルなし）\n")
            continue

        # ── R15: 同一ファイル内 夜勤加配算定漏れチェック ──
        for floor_label, data in all_data.items():
            # 各利用者の「算定対象日」と「実際にF列=1の日」を収集
            resident_info = {}
            for name, days in data.items():
                f_days = []; tgt_days = []
                for day in sorted(days.keys()):
                    c = days[day]['C']; f = days[day]['F']
                    if is_f_exempt(c): pass
                    else:
                        tgt_days.append(day)
                        if f: f_days.append(day)
                resident_info[name] = {'f_days': set(f_days), 'tgt': tgt_days}

            valid = {n: v for n, v in resident_info.items() if v['tgt']}
            if not valid: continue
            max_f = max(len(v['f_days']) for v in valid.values())
            if max_f == 0: continue

            # 各日ごとに「何人がF列なしか」を集計
            from collections import Counter as _Counter
            day_missing_count = _Counter()
            total_valid = len(valid)
            for v in valid.values():
                for d in v['tgt']:
                    if d not in v['f_days']:
                        day_missing_count[d] += 1

            # 全員が漏れている日 = 施設全体の算定なし日（個別問題ではない）
            all_missing_days = {d for d, cnt in day_missing_count.items()
                                if cnt >= total_valid * 0.8}  # 8割以上が漏れ→全体扱い

            for name, v in valid.items():
                missing_all = [d for d in v['tgt'] if d not in v['f_days']]
                # 全体共通の漏れ日を除いた「個人だけの漏れ」
                missing_personal = [d for d in missing_all if d not in all_missing_days]

                if missing_personal:
                    all_issues.append(dict(
                        rule='R15', floor=floor_label, name=name,
                        date=', '.join(f'{d}日' for d in missing_personal),
                        level='error',
                        detail=(
                            f"この方だけF列（夜勤加配）が空欄の日あり / "
                            f"該当日: {', '.join(str(d)+'日' for d in missing_personal)} / "
                            f"同ファイルの他の利用者は算定されている"
                        ),
                        judge=f'❌ 夜勤加配加算の個人算定漏れの可能性（{len(missing_personal)}日分）'
                    ))

            # 全員共通の未算定日 = 職員配置上の理由で意図的に非算定 → スルー

        # ── R21: 医療連携体制加算の個人算定漏れチェック ──
        # 他の利用者が算定している日に自分だけ空欄 → 漏れとして検出
        # 全員が空欄の日 → スルー（職員配置等の共通理由）
        for floor_label_i, data_i in all_data.items():
            resident_i = {}
            for name_i, days_i in data_i.items():
                i_days = []; tgt_days = []
                for day_i in sorted(days_i.keys()):
                    c_i = days_i[day_i]['C']
                    i_v = days_i[day_i]['I']
                    if is_i_exempt(c_i): pass
                    else:
                        tgt_days.append(day_i)
                        if i_v: i_days.append(day_i)
                if tgt_days:
                    resident_i[name_i] = {'i_days': set(i_days), 'tgt': set(tgt_days)}

            if not resident_i: continue

            # 誰か1人でも算定している日を収集
            day_has_i = set()
            for v in resident_i.values():
                day_has_i |= v['i_days']

            if not day_has_i: continue  # 誰も算定していない → スルー

            # 算定している人がいる日に自分だけ空欄の日を検出
            for name_i, v in resident_i.items():
                missing_p = sorted(
                    d for d in v['tgt']
                    if d in day_has_i and d not in v['i_days']
                )
                if missing_p:
                    all_issues.append(dict(
                        rule='R21', floor=floor_label_i, name=name_i,
                        date=', '.join(f'{d}日' for d in missing_p),
                        level='error',
                        detail=(
                            f"他の利用者が算定している日にI列（医療連携加算）が空欄 / "
                            f"該当日: {', '.join(str(d)+'日' for d in missing_p)}"
                        ),
                        judge=f'❌ 医療連携体制加算の算定漏れの可能性（{len(missing_p)}日分）'
                    ))


        year_month = year_month or '不明'

        # 全施設ファイルにシートとして追加
        add_gh_sheet(wb_all, gh_name, year_month, all_issues)

        # 施設別の詳細レポートを出力（全件一覧・ルール別タブ・利用者別サマリー付き）
        safe_gh = re.sub(r'[\\/:*?"<>|]', '_', gh_name)
        gh_out_path = Path(output_folder) / f"{safe_gh}_チェックレポート_{year_month}_{today}.xlsx"
        wb_gh = build_report(gh_name, year_month, all_issues, all_data)
        wb_gh.save(str(gh_out_path))

        total_err  = sum(1 for i in all_issues if i['level']=='error')
        total_warn = sum(1 for i in all_issues if i['level']=='warn')
        total_info = sum(1 for i in all_issues if i['level']=='info')
        print(f"    ✓ シート追加: {gh_name}  ❌{total_err}件 ⚠️{total_warn}件 📌{total_info}件")
        print(f"    💾 施設別レポート保存: {gh_out_path.name}")

        # ルール別件数を集計
        rule_counts = {f'R{n}': sum(1 for i in all_issues if i['rule'] == f'R{n}')
                       for n in range(1, 23)}
        summary_rows.append({
            'GH': gh_name, '年月': year_month,
            'ファイル数': len(files),
            'エラー': total_err, '警告': total_warn, '確認': total_info,
            '合計': total_err + total_warn + total_info,
            'rules': rule_counts,
        })

    # 全GH横断サマリーシートを先頭に挿入
    if summary_rows:
        _add_summary_sheet(wb_all, summary_rows)

    # 未チェック一覧シートを追加（チェックされなかったものの可視化）
    _add_skipped_sheet(wb_all, skipped)

    # 全施設まとめファイルを保存
    # 年月はsummary_rowsから取得（複数年月混在の場合は最初のものを使用）
    ym_label = summary_rows[0]['年月'] if summary_rows else today
    all_out_name = f"【全GH】チェックサマリー_{ym_label}_{today}.xlsx"
    all_out_path = Path(output_folder) / all_out_name
    wb_all.save(str(all_out_path))
    print(f"\n  💾 全施設まとめ保存: {all_out_name}")
    if skipped:
        print(f"  ⚠️  未チェック項目が {len(skipped)} 件あります"
              f"（「⚠未チェック一覧」シートを確認してください）")

    print(f"{'='*60}")
    print(f"  完了！レポートの保存先: {output_folder}")
    print(f"{'='*60}")


