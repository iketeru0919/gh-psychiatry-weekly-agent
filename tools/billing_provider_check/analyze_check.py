import datetime
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl


SHORT_STAY_MARKERS = ("短期", "短期入所")
EXCLUDE_FILE_MARKERS = ("テンプレート", "旧", "コピー", "バックアップ", "テスト", "短期")
# 提供側の延べ日数シートで利用者ではない集計行を除外するための氏名
NON_USER_NAMES = ("合計", "小計", "計", "総計", "平均")


def normalize_name(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace(" ", "").replace("　", "")
    # 同一文字の異体字のみ吸収する。「菊地／菊池」のような実在する別姓の
    # 置換は別人を同一人物として統合する恐れがあるため行わない。
    replacements = {
        "髙": "高",
        "﨑": "崎",
        "邉": "邊",
        "辺": "邊",
        "栁": "柳",
        "澤": "沢",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def as_number(value):
    # 文字列で入力された数値（"31" など）も無言で0にせず数値として扱う。
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = unicodedata.normalize("NFKC", value).replace(",", "").strip()
        if not text:
            return 0
        try:
            number = float(text)
        except ValueError:
            return 0
        return int(number) if number.is_integer() else number
    return 0


def is_readable_number(value):
    # 空欄・数値・数値として解釈できる文字列なら True。
    # False の場合は as_number が0を返すため、呼び出し側で警告として記録する。
    if value is None:
        return True
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        text = unicodedata.normalize("NFKC", value).replace(",", "").strip()
        if not text:
            return True
        try:
            float(text)
        except ValueError:
            return False
        return True
    return False


def is_short_stay(service_name):
    return any(marker in str(service_name or "") for marker in SHORT_STAY_MARKERS)


def is_life_support(service_name):
    # 施設によりサービス名の表記が異なる（例：南中丸「生活援助日中Ⅰ４」、
    # おきつ「生活援助Ⅰ３・大１」、貴船「生活援助Ⅰ３」）ため、
    # 「生活援助」で始まる基本サービス行を対象とし、加算行・短期行は除外する。
    text = str(service_name or "")
    return text.startswith("生活援助") and "加算" not in text and not is_short_stay(text)


def facility_folder_tokens(facility_name):
    normalized = normalize_name(facility_name)
    return [normalized, normalize_name(f"RASIEL{facility_name}"), normalize_name(f"ラシエル{facility_name}")]


def month_dir_candidates(year, month):
    return (
        f"{year}年{month}月提供分",
        f"{year}年{month:02d}月提供分",
        f"{month}月提供分",
        f"{month:02d}月提供分",
    )


def should_exclude_excel(path):
    name = path.name
    lower = name.lower()
    if not lower.endswith((".xlsx", ".xlsm")):
        return True
    if name.startswith("~$"):
        return True
    # "test" は独立した英単語の場合のみ除外する（"latest" 等の誤除外を防ぐ）。
    if re.search(r"(?<![a-z0-9])test(?![a-z0-9])", lower):
        return True
    return any(marker in name for marker in EXCLUDE_FILE_MARKERS)


def find_provider_workbooks(provider_root, facility_name, year, month):
    root = Path(provider_root)
    records = []
    if not root.exists():
        return [], [
            {
                "facility_name": facility_name,
                "status": "未検出",
                "folder": str(root),
                "file": "",
                "note": "provider_root が存在しません。",
            }
        ]

    tokens = facility_folder_tokens(facility_name)
    facility_dirs = []
    for child in root.iterdir():
        if child.is_dir() and any(token in normalize_name(child.name) for token in tokens):
            facility_dirs.append(child)

    if not facility_dirs:
        return [], [
            {
                "facility_name": facility_name,
                "status": "未検出",
                "folder": str(root),
                "file": "",
                "note": "施設フォルダが見つかりません。",
            }
        ]

    year_text = f"{year}年"
    month_names = month_dir_candidates(year, month)
    excel_files = []
    for facility_dir in facility_dirs:
        month_dirs = []
        for subdir in facility_dir.rglob("*"):
            if not subdir.is_dir():
                continue
            normalized_name = normalize_name(subdir.name)
            if any(normalize_name(name) == normalized_name for name in month_names):
                month_dirs.append(subdir)
            elif year_text in str(subdir) and f"{month}月" in subdir.name and "提供分" in subdir.name:
                month_dirs.append(subdir)

        if not month_dirs:
            records.append(
                {
                    "facility_name": facility_name,
                    "status": "未検出",
                    "folder": str(facility_dir),
                    "file": "",
                    "note": f"{year}年{month}月提供分フォルダが見つかりません。",
                }
            )
            continue

        for month_dir in sorted(set(month_dirs)):
            candidates = [path for path in month_dir.rglob("*") if path.is_file() and not should_exclude_excel(path)]
            if not candidates:
                records.append(
                    {
                        "facility_name": facility_name,
                        "status": "未検出",
                        "folder": str(month_dir),
                        "file": "",
                        "note": "対象Excelが見つかりません。",
                    }
                )
            for candidate in sorted(candidates):
                excel_files.append(candidate)
                records.append(
                    {
                        "facility_name": facility_name,
                        "status": "候補",
                        "folder": str(month_dir),
                        "file": str(candidate),
                        "note": "読取候補として検出しました。",
                    }
                )

    return sorted(set(excel_files)), records


def find_header_map(row):
    return {value: idx for idx, value in enumerate(row) if value is not None}


def month_of(value):
    # 日付セル・Excelシリアル値・「2026/4」等の文字列を (年, 月) に変換する。
    # 解釈できない場合は None を返す。
    if isinstance(value, datetime.datetime) or isinstance(value, datetime.date):
        return (value.year, value.month)
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value > 20000:
            d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(value))
            return (d.year, d.month)
        return None
    if isinstance(value, str):
        text = unicodedata.normalize("NFKC", value).strip()
        match = re.match(r"(\d{4})[/\-.年](\d{1,2})", text)
        if match:
            return (int(match.group(1)), int(match.group(2)))
        try:
            return month_of(float(text))
        except ValueError:
            return None
    return None


def validate_db_layout(ws):
    # DBシートは列位置（M/O/P/Q/T列）を直接参照しているため、
    # 列の挿入・削除でレイアウトが変わった場合は誤読せずに停止する。
    rows = list(ws.iter_rows(min_row=5, max_row=6, values_only=True))
    row5 = rows[0] if rows else ()
    row6 = rows[1] if len(rows) > 1 else ()

    def cell_text(row, idx):
        return str(row[idx]) if len(row) > idx and row[idx] is not None else ""

    problems = []
    if "請求ベース" not in cell_text(row5, 11):
        problems.append("L5に「請求ベース」がありません")
    if "提供ベース" not in cell_text(row5, 13):
        problems.append("N5に「提供ベース」がありません")
    if "利用日数" not in cell_text(row5, 19):
        problems.append("T5に「利用日数」がありません")
    if "有無" not in cell_text(row6, 15):
        problems.append("P6に「有無」がありません")
    if "理由" not in cell_text(row6, 16):
        problems.append("Q6に「理由」がありません")
    if problems:
        raise SystemExit(
            "DBシートの列構成が想定と異なります（列の挿入・削除の可能性があります）。"
            "ツールの列位置を実際のDBシートに合わせて修正してください: " + "、".join(problems)
        )


def validate_db_target_month(ws, target_year, target_month):
    # DBシートのC4（対象月）とM6/O6（当月列の日付）を確認し、
    # 古い請求Excelを選んでしまった場合は集計せずに停止する。
    grid = list(ws.iter_rows(min_row=4, max_row=6, values_only=True))
    row4 = grid[0] if grid else ()
    row6 = grid[2] if len(grid) > 2 else ()
    candidates = []
    for row, idx in ((row4, 2), (row6, 12), (row6, 14)):
        if len(row) > idx:
            parsed = month_of(row[idx])
            if parsed:
                candidates.append(parsed)
    if candidates and (target_year, target_month) not in candidates:
        found = "、".join(f"{y}年{m}月" for y, m in sorted(set(candidates)))
        raise SystemExit(
            f"請求ExcelのDBシートの対象月（{found}）が、指定した{target_year}年{target_month}月と"
            "一致しません。請求Excelの取り違えを確認してください。"
        )


def read_db_rows(workbook_path, target_year, target_month):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "DB" not in wb.sheetnames:
        raise SystemExit(
            f"請求Excelに「DB」シートがありません。請求Excelの取り違えの可能性があります: {workbook_path}"
        )
    result = {}
    ws = wb["DB"]
    validate_db_layout(ws)
    validate_db_target_month(ws, target_year, target_month)
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        facility = row[2] if len(row) > 2 else None
        if not isinstance(facility, str):
            continue
        facility = facility.strip()
        if not facility or facility in ("施設名", "担当AM", "暦日数", "株式会社ラシエル稼働管理"):
            continue
        # M/O/T列のいずれかに数値がない行（表題行・区切り行）は施設行とみなさない。
        metrics = [row[idx] if len(row) > idx else None for idx in (12, 14, 19)]
        if not any(isinstance(v, (int, float)) and not isinstance(v, bool) for v in metrics):
            continue
        result[str(facility)] = {
            "row": row_num,
            "facility_name": str(facility),
            "j_net_change": row[9] if len(row) > 9 else None,
            "k_net_change_reason": row[10] if len(row) > 10 else None,
            "m_billed_count": row[12] if len(row) > 12 else None,
            "n_provider_prev_count": row[13] if len(row) > 13 else None,
            "o_provider_count": row[14] if len(row) > 14 else None,
            "p_gap_label": row[15] if len(row) > 15 else None,
            "q_gap_reason": row[16] if len(row) > 16 else None,
            "t_billed_days": row[19] if len(row) > 19 else None,
            "aq_home_nursing_contracts": row[42] if len(row) > 42 else None,
        }
    return result


def empty_billing():
    return {
        "present": False,
        "all_user_count": 0,
        "resident_user_count": 0,
        "life_support_days": 0,
        "carryover_days": 0,
        "users": {},
        "resident_users": {},
        "rows": [],
        "warnings": [],
    }


def read_billing_all(workbook_path, month_sheet, target_year, target_month):
    # 請求Excelは1回だけ読み、全施設分をまとめて集計する（施設ごとの再読込を避ける）。
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if month_sheet not in wb.sheetnames:
        raise SystemExit(
            f"請求Excelに「{month_sheet}」シートがありません。対象年月と請求Excelの組み合わせを"
            f"確認してください。（存在するシート: {', '.join(wb.sheetnames)}）"
        )
    ws = wb[month_sheet]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = find_header_map(header)
    required_headers = ("利用者名", "サービス内容", "回数")
    missing_headers = [name for name in required_headers if name not in header_map]
    if missing_headers:
        raise SystemExit(
            f"請求シート「{month_sheet}」の1行目に必要なヘッダーが見つかりません: "
            f"{', '.join(missing_headers)}（シートの構成が変わった可能性があります）"
        )
    has_provision_col = "サービス提供年月" in header_map
    results = {}
    collisions = {}
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        facility_name = str(row[0]).strip()
        if not facility_name:
            continue
        bucket = results.setdefault(facility_name, empty_billing())
        bucket["present"] = True
        record = {name: row[idx] if idx < len(row) else None for name, idx in header_map.items()}
        name = record.get("利用者名")
        service_name = record.get("サービス内容")
        count_raw = record.get("回数")
        if not is_readable_number(count_raw):
            bucket["warnings"].append(
                f"請求シート{excel_row}行目（{name}／{service_name}）の回数「{count_raw}」を数値として読めないため0回として扱いました。元データを確認してください。"
            )
        count = as_number(count_raw)

        # 請求シートには月遅れ請求（過去月提供分）が混在するため、
        # サービス提供年月が対象月の行だけを当月分として集計し、
        # 他月提供分は「月遅れ請求」として利用者別に別カウントする。
        provision_month = None
        is_current = True
        if has_provision_col:
            provision_raw = record.get("サービス提供年月")
            provision_month = month_of(provision_raw)
            if provision_month is None:
                if provision_raw is not None:
                    bucket["warnings"].append(
                        f"請求シート{excel_row}行目（{name}）のサービス提供年月「{provision_raw}」を解釈できないため当月分として扱いました。"
                    )
            else:
                is_current = provision_month == (target_year, target_month)

        normalized = normalize_name(name)
        user = bucket["users"].setdefault(
            normalized,
            {
                "name": name,
                "normalized_name": normalized,
                "recipient_no": record.get("受給者証番号"),
                "life_support_days": 0,
                "short_stay_days": 0,
                "is_short_stay": False,
                "carryover_days": 0,
                "carryover_months": set(),
                "services": [],
                "excel_rows": [],
            },
        )
        recipient = record.get("受給者証番号")
        if user["recipient_no"] is None:
            user["recipient_no"] = recipient
        elif recipient is not None and str(recipient) != str(user["recipient_no"]):
            # 同じ氏名で受給者証番号が異なる行がある場合は、同姓同名の別人を
            # 1人に集計している可能性があるため警告する。
            collisions.setdefault(facility_name, set()).add(
                f"利用者「{name}」に複数の受給者証番号（{user['recipient_no']}／{recipient}）があります。同姓同名の別人を1人として集計している可能性があります。"
            )
        user["excel_rows"].append(excel_row)
        user["services"].append(
            {
                "service_name": service_name,
                "unit": record.get("単位数"),
                "count": count,
                "service_units": record.get("サービス単位数"),
                "provision_month": f"{provision_month[0]}年{provision_month[1]}月" if provision_month else None,
            }
        )
        if is_current:
            if is_short_stay(service_name):
                user["is_short_stay"] = True
                user["short_stay_days"] += count
            if is_life_support(service_name):
                user["life_support_days"] += count
        elif is_life_support(service_name):
            user["carryover_days"] += count
            user["carryover_months"].add(f"{provision_month[0]}年{provision_month[1]}月")
        bucket["rows"].append(record)

    if not has_provision_col:
        for bucket in results.values():
            bucket["warnings"].append(
                "請求シートに「サービス提供年月」列が見つからないため月遅れ請求を判定できません。全行を当月分として集計しています。"
            )

    for facility_name, bucket in results.items():
        # 短期入所と生活援助日中を同月に併用する利用者がいるため、
        # 短期行の有無では除外せず、当月の生活援助日中の回数があるかどうかで入居系と判定する。
        # 短期のみの利用者は life_support_days が 0 のためここで自然に除外される。
        resident_users = {
            key: value
            for key, value in bucket["users"].items()
            if value["life_support_days"] > 0
        }
        bucket["resident_users"] = resident_users
        bucket["all_user_count"] = len(bucket["users"])
        bucket["resident_user_count"] = len(resident_users)
        bucket["life_support_days"] = sum(user["life_support_days"] for user in resident_users.values())
        bucket["carryover_days"] = sum(as_number(user["carryover_days"]) for user in bucket["users"].values())
        bucket["warnings"].extend(sorted(collisions.get(facility_name, set())))
    return results


def read_provider_file(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    if "延べ日数" not in wb.sheetnames:
        return [], []
    ws = wb["延べ日数"]
    users = []
    skipped = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[1] if len(row) > 1 else None
        if not isinstance(name, str) or not name.strip() or name.startswith("氏名"):
            continue
        if normalize_name(name) in NON_USER_NAMES:
            # 「合計」「小計」等の集計行を利用者として読まない。
            continue
        days = row[3] if len(row) > 3 else None
        if days is None or not is_readable_number(days):
            # 延べ日数が空欄・数値以外（数式の計算結果未保存など）の行を
            # 無言で落とすと差異の見落としにつながるため、スキップ行として記録する。
            skipped.append(
                f"{row_num}行目「{name}」の延べ日数「{days}」を数値として読めないためスキップしました。"
            )
            continue
        days = as_number(days)
        users.append(
            {
                "name": name,
                "normalized_name": normalize_name(name),
                "category": row[2] if len(row) > 2 else None,
                "days": days,
                "single_days": row[5] if len(row) > 5 else None,
                "combined_days": row[6] if len(row) > 6 else None,
                "half_days": row[7] if len(row) > 7 else None,
                "direct_billing": row[8] if len(row) > 8 else None,
                "source_file": str(path),
                "source_row": row_num,
            }
        )
    return users, skipped


def read_provider(paths):
    merged = {}
    duplicates = []
    file_records = []
    for path in paths:
        try:
            users, skipped = read_provider_file(path)
        except Exception as exc:
            file_records.append(
                {
                    "status": "読取エラー",
                    "file": str(path),
                    "folder": str(Path(path).parent),
                    "note": str(exc),
                }
            )
            continue
        file_records.append(
            {
                "status": "読取OK" if users else "対象外",
                "file": str(path),
                "folder": str(Path(path).parent),
                "note": "延べ日数シートを読み取りました。" if users else "延べ日数シートまたは利用者行が見つかりません。",
            }
        )
        for message in skipped:
            file_records.append(
                {
                    "status": "警告",
                    "file": str(path),
                    "folder": str(Path(path).parent),
                    "note": message,
                }
            )
        for user in users:
            key = user["normalized_name"]
            if key in merged:
                duplicates.append({"normalized_name": key, "name": user["name"], "source_file": user["source_file"]})
                merged[key]["days"] += as_number(user["days"])
                merged[key]["single_days"] = as_number(merged[key]["single_days"]) + as_number(user["single_days"])
                merged[key]["combined_days"] = as_number(merged[key]["combined_days"]) + as_number(user["combined_days"])
                merged[key]["half_days"] = as_number(merged[key]["half_days"]) + as_number(user["half_days"])
                merged[key]["source_file"] += "\n" + user["source_file"]
            else:
                merged[key] = user
    for dup in duplicates:
        # 1F/2Fの合算は正常な動きだが、同姓同名の別人を合算している可能性も
        # あるため、どの利用者を合算したかを読取ファイル一覧で確認できるようにする。
        file_records.append(
            {
                "status": "情報",
                "file": dup["source_file"],
                "folder": str(Path(dup["source_file"]).parent),
                "note": f"利用者「{dup['name']}」が複数ファイルに存在するため延べ日数を合算しました。同姓同名の別人でないか確認してください。",
            }
        )
    return {
        "users": merged,
        "sheet_user_count": len(merged),
        "active_user_count": sum(1 for user in merged.values() if as_number(user["days"]) > 0),
        "total_days": sum(as_number(user["days"]) for user in merged.values()),
        "duplicates": duplicates,
        "file_records": file_records,
    }


def build_reason(diff, summary, provider_user=None):
    item = diff["item"]
    billing = as_number(diff.get("billing_value"))
    provider = as_number(diff.get("provider_value"))
    user_name = diff.get("user_name") or "対象者"
    direct_billing = provider_user.get("direct_billing") if provider_user else None
    evidence_base = (
        f"請求延べ{summary['billing_life_support_days']}日、"
        f"提供延べ{summary['provider_total_days']}日。"
    )

    if item == "提供あり請求なし" and provider == 0 and direct_billing:
        return {
            "reason_confidence": "高",
            "reason_candidate": "直接請求（他事業所・自費等）の利用者で、この請求Excelの対象外の可能性があります。",
            "evidence": f"{evidence_base}{user_name}さんは提供側にシートがあり延べ0日、直接請求フラグ（{direct_billing}）が付いています。",
            "check_point": "直接請求の請求経路と、DB O列の提供人数に直接請求利用者を含める定義か確認してください。",
            "db_q_draft": "理由：直接請求の利用者を提供ベース人数に含めているため人数乖離。請求延べ日数と提供延べ日数は一致しており、請求漏れなし。対策：直接請求利用者の計上定義を確認。",
        }

    if item == "提供あり請求なし" and provider == 0:
        return {
            "reason_confidence": "高",
            "reason_candidate": "延べ0日の利用者が提供ベース人数に含まれている可能性があります。",
            "evidence": f"{evidence_base}{user_name}さんは提供側にシートがありますが延べ0日です。",
            "check_point": "DB O列の提供人数に延べ0日利用者を含める定義か確認してください。",
            "db_q_draft": "理由：延べ0日の利用者を提供ベース人数に含めているため人数乖離。請求延べ日数と提供延べ日数は一致しており、請求漏れなし。対策：提供人数の計上定義を確認し、必要に応じて延べ0日利用者を除外。",
        }

    if item == "提供あり請求なし" and direct_billing:
        return {
            "reason_confidence": "中",
            "reason_candidate": "直接請求（他事業所・自費等）のため、この請求Excelに載らない利用者の可能性があります。未請求とは区別して確認してください。",
            "evidence": f"{user_name}さんは提供側に{provider}日あり直接請求フラグ（{direct_billing}）が付いていますが、この請求Excelの生活援助日中系に見当たりません。",
            "check_point": "直接請求の請求経路で当月分が請求済みか確認してください。",
            "db_q_draft": f"理由：{user_name}さんは直接請求のため当請求Excel対象外の可能性。提供実績{provider}日。対策：直接請求側の請求状況を確認。",
        }

    if item == "提供あり請求なし":
        return {
            "reason_confidence": "中",
            "reason_candidate": "未請求、月遅れ請求、受給者証更新待ちの可能性があります。",
            "evidence": f"{user_name}さんは提供側に{provider}日ありますが、請求側の生活援助日中系に見当たりません。",
            "check_point": "請求データの対象月、月遅れ予定、受給者証更新状況を確認してください。",
            "db_q_draft": f"理由：{user_name}さんの提供実績{provider}日が請求側に未反映のため乖離。対策：月遅れ・受給者証状況を確認し、必要に応じて翌月請求または請求修正。",
        }

    if item == "請求あり提供なし":
        return {
            "reason_confidence": "中",
            "reason_candidate": "過剰請求、提供記録漏れ、施設違いの可能性があります。",
            "evidence": f"{user_name}さんは請求側に{billing}日ありますが、提供側の延べ日数に見当たりません。",
            "check_point": "提供実績記録票の日別実績、対象施設、請求明細の施設名を確認してください。",
            "db_q_draft": f"理由：{user_name}さんの請求{billing}日が提供実績で確認できないため乖離。対策：提供記録と請求明細を確認し、過剰請求の場合は修正対応。",
        }

    if item == "利用者別延べ日数" and provider > billing:
        gap = provider - billing
        return {
            "reason_confidence": "中",
            "reason_candidate": "未請求、月遅れ、受給者証更新待ちの可能性があります。",
            "evidence": f"{user_name}さんは提供側が請求側より{gap}日多いです。",
            "check_point": "差異日数に該当する日別実績と、請求データの掲載月を確認してください。",
            "db_q_draft": f"理由：{user_name}さんの提供実績が請求より{gap}日多く、月遅れまたは未請求の可能性。対策：請求予定月を確認し、必要に応じて翌月請求。",
        }

    if item == "利用者別延べ日数" and billing > provider:
        gap = billing - provider
        return {
            "reason_confidence": "中",
            "reason_candidate": "過剰請求、提供記録漏れ、入院・外泊反映漏れの可能性があります。",
            "evidence": f"{user_name}さんは請求側が提供側より{gap}日多いです。",
            "check_point": "実績記録票の日別実績、入院・外泊、請求回数を確認してください。",
            "db_q_draft": f"理由：{user_name}さんの請求が提供実績より{gap}日多く、過剰請求または提供記録漏れの可能性。対策：実績記録票と請求明細を照合し、必要に応じて修正。",
        }

    return {
        "reason_confidence": "低",
        "reason_candidate": "差異の原因を追加確認してください。",
        "evidence": evidence_base,
        "check_point": "請求側・提供側の元データを確認してください。",
        "db_q_draft": "理由：差異原因を確認中。対策：請求データと提供実績を照合し、必要に応じて修正。",
    }


def attach_reason(diff, summary, provider_user=None):
    enriched = dict(diff)
    enriched.update(build_reason(diff, summary, provider_user))
    carryover = as_number(diff.get("billing_carryover_days"))
    if carryover:
        # 当月シートに他月提供分の請求がある利用者は月遅れ請求サイクルの
        # 可能性が高いため、その事実を根拠に追記する。
        enriched["evidence"] += f" なお当月の請求シートには、この利用者の他月提供分{carryover}日の月遅れ請求が含まれています。"
    return enriched


def classify_facility(db_row, billing, provider, diffs):
    billing_days = as_number(billing.get("life_support_days"))
    provider_days = as_number(provider.get("total_days"))
    resident_count = as_number(billing.get("resident_user_count"))
    active_count = as_number(provider.get("active_user_count"))

    if any(item["severity"] == "重大" for item in diffs):
        return "重大"
    if any(item["severity"] == "要確認" for item in diffs) or billing_days != provider_days:
        return "要確認"
    if not db_row:
        # DB行が見つからない場合は0との比較で誤判定せず、未照合として注意を返す。
        return "注意"
    db_m = as_number(db_row.get("m_billed_count"))
    db_o = as_number(db_row.get("o_provider_count"))
    db_t = as_number(db_row.get("t_billed_days"))
    if db_t != billing_days:
        # DB T列（利用日数・請求ベース）と請求Excelの延べ日数の不一致も判定に含める。
        return "注意"
    if db_m != resident_count or db_o != active_count:
        return "注意"
    return "OK"


def summarize_main_reason(summary):
    if summary["judgement"] == "OK":
        return "主要な人数・延べ日数は一致しています。"
    if summary["judgement"] == "注意":
        if summary.get("db_row") is None:
            return "DBシートに施設行が見つからないため、DB列との照合ができていません。施設名の表記を確認してください。"
        parts = []
        if as_number(summary.get("db_t_billed_days")) != as_number(summary.get("billing_life_support_days")):
            parts.append("DB T列の利用日数と請求Excelの延べ日数が一致していません。DBへの転記を確認してください。")
        counts_mismatch = (
            as_number(summary.get("db_m_billed_count")) != as_number(summary.get("billing_resident_user_count"))
            or as_number(summary.get("db_o_provider_count")) != as_number(summary.get("provider_active_user_count"))
        )
        if counts_mismatch:
            if summary.get("billing_life_support_days") == summary.get("provider_total_days"):
                parts.append("人数に乖離がありますが、請求延べ日数と提供延べ日数は一致しています。延べ0日利用者や人数定義を確認してください。")
            else:
                parts.append("人数に乖離があります。差異一覧を確認してください。")
        if parts:
            return "".join(parts)
    return "差異一覧の理由候補・根拠・確認ポイントを確認してください。"


def analyze_facility(config, db_rows, billing_by_facility):
    facility_name = config["facility_name"]
    db_row = db_rows.get(facility_name, {})
    billing = billing_by_facility.get(facility_name) or empty_billing()
    discovery_records = []
    provider_paths = [Path(path) for path in config.get("provider_workbooks", [])]
    if not provider_paths and CONFIG.get("provider_root"):
        provider_paths, discovery_records = find_provider_workbooks(
            CONFIG["provider_root"],
            facility_name,
            int(CONFIG["target_year"]),
            int(CONFIG["target_month"]),
        )
    provider = read_provider(provider_paths)
    file_records = []
    for record in discovery_records:
        file_records.append(record)
    for record in provider.get("file_records", []):
        item = dict(record)
        item["facility_name"] = facility_name
        file_records.append(item)
    for message in billing.get("warnings", []):
        file_records.append(
            {
                "facility_name": facility_name,
                "status": "警告",
                "folder": "",
                "file": str(CONFIG["billing_workbook"]),
                "note": message,
            }
        )
    if not billing.get("present"):
        file_records.append(
            {
                "facility_name": facility_name,
                "status": "警告",
                "folder": "",
                "file": str(CONFIG["billing_workbook"]),
                "note": f"請求シートに施設「{facility_name}」の行が見つかりません。施設名の表記（config・請求シートの一致）を確認してください。",
            }
        )
    if not db_row:
        file_records.append(
            {
                "facility_name": facility_name,
                "status": "警告",
                "folder": "",
                "file": str(CONFIG["billing_workbook"]),
                "note": f"DBシートに施設「{facility_name}」の行が見つかりません。DB列（M/O/P/Q/T）との照合は行われていません。",
            }
        )

    billing_users = billing.get("resident_users", {})
    provider_users = provider.get("users", {})
    summary = {
        "facility_name": facility_name,
        "db_row": db_row.get("row"),
        "db_m_billed_count": db_row.get("m_billed_count"),
        "db_o_provider_count": db_row.get("o_provider_count"),
        "db_p_gap_label": db_row.get("p_gap_label"),
        "db_q_gap_reason": db_row.get("q_gap_reason"),
        "db_t_billed_days": db_row.get("t_billed_days"),
        "billing_sheet_present": billing.get("present"),
        "billing_all_user_count": billing.get("all_user_count"),
        "billing_resident_user_count": billing.get("resident_user_count"),
        "billing_life_support_days": billing.get("life_support_days"),
        "billing_carryover_days": billing.get("carryover_days"),
        "provider_sheet_user_count": provider.get("sheet_user_count"),
        "provider_active_user_count": provider.get("active_user_count"),
        "provider_total_days": provider.get("total_days"),
    }

    all_billing_users = billing.get("users", {})
    provider_readable = any(record.get("status") == "読取OK" for record in provider.get("file_records", []))

    diffs = []
    if provider_readable:
        for key in sorted(set(provider_users) - set(billing_users)):
            user = provider_users[key]
            days = as_number(user.get("days"))
            diff = {
                "facility_name": facility_name,
                "user_name": user.get("name"),
                "item": "提供あり請求なし",
                "billing_value": 0,
                "provider_value": days,
                "difference": -days,
                "severity": "注意" if days == 0 else "重大",
                "billing_carryover_days": as_number(all_billing_users.get(key, {}).get("carryover_days")),
            }
            diffs.append(attach_reason(diff, summary, user))

        for key in sorted(set(billing_users) - set(provider_users)):
            user = billing_users[key]
            days = as_number(user.get("life_support_days"))
            diff = {
                "facility_name": facility_name,
                "user_name": user.get("name"),
                "item": "請求あり提供なし",
                "billing_value": days,
                "provider_value": 0,
                "difference": days,
                "severity": "重大",
                "billing_carryover_days": as_number(user.get("carryover_days")),
            }
            diffs.append(attach_reason(diff, summary))

        for key in sorted(set(billing_users) & set(provider_users)):
            billing_user = billing_users[key]
            provider_user = provider_users[key]
            billing_days = as_number(billing_user.get("life_support_days"))
            provider_days = as_number(provider_user.get("days"))
            if billing_days != provider_days:
                diff = {
                    "facility_name": facility_name,
                    "user_name": provider_user.get("name"),
                    "item": "利用者別延べ日数",
                    "billing_value": billing_days,
                    "provider_value": provider_days,
                    "difference": billing_days - provider_days,
                    "severity": "要確認",
                    "billing_carryover_days": as_number(billing_user.get("carryover_days")),
                }
                diffs.append(attach_reason(diff, summary, provider_user))

    if not provider_readable:
        # 提供実績Excelが1つも読めていない施設は、利用者別の差異を出すと
        # 全員が「請求あり提供なし（重大）」になり誤報の嵐になるため、
        # 比較不能として判定を分ける。
        summary["judgement"] = "提供未読"
        summary["main_reason"] = "提供実績Excelを読み取れていないため、請求との比較ができていません。読取ファイル一覧で対象フォルダ・ファイルを確認してください。"
    else:
        summary["judgement"] = classify_facility(db_row, billing, provider, diffs)
        summary["main_reason"] = summarize_main_reason(summary)

    # DB P列（乖離の有無）・Q列（理由記載）とツールの検出結果の整合を確認する。
    tool_gap = bool(diffs) or summary["judgement"] in ("重大", "要確認", "注意")
    if summary["judgement"] == "提供未読":
        summary["db_p_check"] = "提供実績未読のため照合不可"
        summary["db_q_check"] = "提供実績未読のため照合不可"
    elif not db_row:
        summary["db_p_check"] = "DB行未検出のため照合不可"
        summary["db_q_check"] = "DB行未検出のため照合不可"
    else:
        p_label = str(db_row.get("p_gap_label") or "").strip()
        db_says_gap = "乖離" in p_label or p_label == "有"
        if db_says_gap and tool_gap:
            summary["db_p_check"] = "整合（双方乖離あり）"
        elif db_says_gap:
            summary["db_p_check"] = "DBは乖離ありだがツールでは差異未検出。解消済みか、DB記載を確認してください"
        elif tool_gap:
            summary["db_p_check"] = "ツールで差異検出だがDB P列に乖離記載なし。DB記入漏れの可能性があります"
        else:
            summary["db_p_check"] = "整合（双方乖離なし）"
        q_text = str(db_row.get("q_gap_reason") or "").strip()
        if tool_gap and not q_text:
            summary["db_q_check"] = "差異があるのにDB Q列が未記載。記載案の転記を検討してください"
        elif tool_gap:
            summary["db_q_check"] = "DB Q列に記載あり。ツールの見立てと内容を照合してください"
        elif q_text:
            summary["db_q_check"] = "ツール差異なしだがQ列に記載あり。解消済みの記載か確認してください"
        else:
            summary["db_q_check"] = "整合（差異・記載なし）"

    user_rows = []
    diff_by_name = {normalize_name(diff.get("user_name")): diff for diff in diffs}
    # 月遅れ請求のみの利用者（当月提供分の請求がない利用者）も一覧に出す。
    carryover_keys = {
        key for key, user in all_billing_users.items() if as_number(user.get("carryover_days"))
    }
    for key in sorted(set(billing_users) | set(provider_users) | carryover_keys):
        billing_user = billing_users.get(key) or all_billing_users.get(key, {})
        provider_user = provider_users.get(key, {})
        billing_days = as_number(billing_user.get("life_support_days"))
        provider_days = as_number(provider_user.get("days"))
        carryover_days = as_number(billing_user.get("carryover_days"))
        carryover_months = "、".join(sorted(billing_user.get("carryover_months", set())))
        related_diff = diff_by_name.get(key, {})
        user_rows.append(
            {
                "facility_name": facility_name,
                "user_name": provider_user.get("name") or billing_user.get("name"),
                "normalized_name": key,
                "billing_life_support_days": billing_days,
                "provider_days": provider_days,
                "difference": billing_days - provider_days,
                "billing_carryover_days": carryover_days,
                "carryover_note": f"{carryover_months}提供分{carryover_days}日の月遅れ請求あり" if carryover_days else "",
                "direct_billing": provider_user.get("direct_billing"),
                "status": "OK" if billing_days == provider_days and billing_days > 0 else "確認",
                "reason_candidate": related_diff.get("reason_candidate"),
                "reason_confidence": related_diff.get("reason_confidence"),
                "evidence": related_diff.get("evidence"),
                "check_point": related_diff.get("check_point"),
                "db_q_draft": related_diff.get("db_q_draft"),
            }
        )

    return {"summary": summary, "diffs": diffs, "user_rows": user_rows, "file_records": file_records}


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: analyze_check.py CONFIG_JSON OUTPUT_JSON")
    global CONFIG
    config_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    with config_path.open("r", encoding="utf-8-sig") as f:
        CONFIG = json.load(f)
    if "target_month_label" not in CONFIG and "target_month" in CONFIG:
        CONFIG["target_month_label"] = f"{int(CONFIG['target_month'])}月"

    target_year = int(CONFIG["target_year"])
    target_month = int(CONFIG["target_month"])
    db_rows = read_db_rows(CONFIG["billing_workbook"], target_year, target_month)
    billing_by_facility = read_billing_all(
        CONFIG["billing_workbook"], CONFIG["target_month_label"], target_year, target_month
    )
    raw_facilities = CONFIG.get("facilities") or []
    if not raw_facilities or (
        len(raw_facilities) == 1
        and isinstance(raw_facilities[0], str)
        and raw_facilities[0].strip().lower() in ("all", "全施設")
    ):
        # facilities が未指定・空・"all"・"全施設" の場合はDBシートの全施設を対象にする。
        facilities_config = [{"facility_name": name} for name in db_rows]
    else:
        facilities_config = [
            {"facility_name": item} if isinstance(item, str) else item
            for item in raw_facilities
        ]
    facilities = [analyze_facility(item, db_rows, billing_by_facility) for item in facilities_config]
    payload = {
        "report_title": CONFIG.get("report_title", "請求提供差異チェック"),
        "target_month_label": CONFIG.get("target_month_label"),
        "facilities": facilities,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
