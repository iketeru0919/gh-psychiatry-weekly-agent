import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def n(value):
    return value if value is not None else ""


def flatten_summaries(payload):
    rows = []
    for facility in payload.get("facilities", []):
        s = facility["summary"]
        rows.append(
            {
                "facility_name": s.get("facility_name"),
                "judgement": s.get("judgement"),
                "db_m_billed_count": s.get("db_m_billed_count"),
                "billing_resident_user_count": s.get("billing_resident_user_count"),
                "db_o_provider_count": s.get("db_o_provider_count"),
                "provider_active_user_count": s.get("provider_active_user_count"),
                "provider_sheet_user_count": s.get("provider_sheet_user_count"),
                "db_t_billed_days": s.get("db_t_billed_days"),
                "billing_life_support_days": s.get("billing_life_support_days"),
                "provider_total_days": s.get("provider_total_days"),
                "db_p_gap_label": s.get("db_p_gap_label"),
                "main_reason": s.get("main_reason"),
            }
        )
    return rows


def flatten_db_checks(payload):
    rows = []
    for facility in payload.get("facilities", []):
        s = facility["summary"]
        billed_match = (s.get("db_m_billed_count") or 0) == (s.get("billing_resident_user_count") or 0)
        provider_match_active = (s.get("db_o_provider_count") or 0) == (s.get("provider_active_user_count") or 0)
        provider_match_sheet = (s.get("db_o_provider_count") or 0) == (s.get("provider_sheet_user_count") or 0)
        days_match = (
            (s.get("db_t_billed_days") or 0)
            == (s.get("billing_life_support_days") or 0)
            == (s.get("provider_total_days") or 0)
        )
        rows.append(
            {
                "facility_name": s.get("facility_name"),
                "db_row": s.get("db_row"),
                "db_m_billed_count": s.get("db_m_billed_count"),
                "billing_resident_user_count": s.get("billing_resident_user_count"),
                "billed_count_check": "OK" if billed_match else "確認",
                "db_o_provider_count": s.get("db_o_provider_count"),
                "provider_active_user_count": s.get("provider_active_user_count"),
                "provider_sheet_user_count": s.get("provider_sheet_user_count"),
                "provider_count_check": "OK" if provider_match_active else ("シート人数一致" if provider_match_sheet else "確認"),
                "db_t_billed_days": s.get("db_t_billed_days"),
                "billing_life_support_days": s.get("billing_life_support_days"),
                "provider_total_days": s.get("provider_total_days"),
                "days_check": "OK" if days_match else "確認",
                "db_q_gap_reason": s.get("db_q_gap_reason"),
            }
        )
    return rows


def flatten_users(payload):
    rows = []
    for facility in payload.get("facilities", []):
        for row in facility.get("user_rows", []):
            billing = row.get("billing_life_support_days") or 0
            provider = row.get("provider_days") or 0
            result = "一致"
            explanation = "請求側の生活援助日中系回数と、提供側の延べ日数が一致しています。"
            check_point = ""
            if billing > provider:
                result = "請求が多い"
                explanation = f"請求側が提供側より{billing - provider}日多い状態です。過剰請求または提供側記録漏れの可能性があります。"
                check_point = "提供実績記録票の日別実績、入院・外泊、月途中入退居を確認してください。"
            elif provider > billing:
                result = "提供が多い"
                explanation = f"提供側が請求側より{provider - billing}日多い状態です。未請求、月遅れ、受給者証待ちの可能性があります。"
                check_point = "請求データの対象月、月遅れ、受給者証更新状況を確認してください。"
            elif billing == 0 and provider == 0:
                result = "延べ0日"
                explanation = "提供側に利用者シートはありますが、延べ日数は0日です。請求漏れではない可能性が高いです。"
                check_point = "DBの提供人数に含める定義か確認してください。"
            rows.append(
                {
                    **row,
                    "user_result": result,
                    "readable_difference": explanation,
                    "check_point_display": row.get("check_point") or check_point,
                }
            )
    return rows


def flatten_diffs(payload):
    rows = []
    for facility in payload.get("facilities", []):
        for row in facility.get("diffs", []):
            billing = row.get("billing_value") or 0
            provider = row.get("provider_value") or 0
            billing_text = f"{billing}日"
            provider_text = f"{provider}日"
            meaning = "請求側と提供側の状態を確認してください。"
            if row.get("item") == "提供あり請求なし" and provider == 0:
                billing_text = "請求なし"
                provider_text = "提供シートあり、延べ0日"
                meaning = "提供側には利用者がいますが延べ日数は0日のため、通常の請求漏れとは分けて確認します。"
            elif row.get("item") == "提供あり請求なし":
                billing_text = "請求なし"
                provider_text = f"提供{provider}日"
                meaning = f"提供実績が{provider}日ありますが、請求側の生活援助日中系に見当たりません。未請求または月遅れの可能性があります。"
            elif row.get("item") == "請求あり提供なし":
                billing_text = f"請求{billing}日"
                provider_text = "提供なし"
                meaning = f"請求が{billing}日ありますが、提供側の延べ日数に見当たりません。過剰請求または提供記録漏れの可能性があります。"
            elif billing > provider:
                meaning = f"請求側が提供側より{billing - provider}日多い状態です。"
            elif provider > billing:
                meaning = f"提供側が請求側より{provider - billing}日多い状態です。"
            rows.append(
                {
                    **row,
                    "billing_text": billing_text,
                    "provider_text": provider_text,
                    "difference_meaning": meaning,
                }
            )
    if not rows:
        rows.append(
            {
                "facility_name": "",
                "user_name": "",
                "item": "差異なし",
                "billing_text": "-",
                "provider_text": "-",
                "difference_meaning": "対象範囲では差異がありません。",
                "severity": "OK",
                "reason_candidate": "",
                "reason_confidence": "",
                "evidence": "",
                "check_point": "",
                "db_q_draft": "",
            }
        )
    return rows


def flatten_file_records(payload):
    rows = []
    for facility in payload.get("facilities", []):
        rows.extend(facility.get("file_records", []))
    if not rows:
        rows.append(
            {
                "facility_name": "",
                "status": "記録なし",
                "folder": "",
                "file": "",
                "note": "読取記録がありません。",
            }
        )
    return rows


def write_sheet(wb, title, rows, headers, widths):
    ws = wb.create_sheet(title)
    ws.append([label for _, label in headers])
    for row in rows:
        ws.append([n(row.get(key)) for key, _ in headers])

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22 if row_idx == 1 else 45
    return ws


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: build_report.py INPUT_JSON OUTPUT_XLSX")
    input_json = Path(sys.argv[1])
    output_xlsx = Path(sys.argv[2])
    payload = json.loads(input_json.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb,
        "サマリー",
        flatten_summaries(payload),
        [
            ("facility_name", "施設名"),
            ("judgement", "判定"),
            ("db_m_billed_count", "DB M列 請求人数"),
            ("billing_resident_user_count", "請求 入居系人数"),
            ("db_o_provider_count", "DB O列 提供人数"),
            ("provider_active_user_count", "提供 実利用者数"),
            ("provider_sheet_user_count", "提供 シート人数"),
            ("db_t_billed_days", "DB T列 利用日数"),
            ("billing_life_support_days", "請求 延べ日数"),
            ("provider_total_days", "提供 延べ日数"),
            ("db_p_gap_label", "DB P列"),
            ("main_reason", "主な見立て"),
        ],
        [12, 12, 16, 16, 16, 16, 16, 16, 14, 14, 12, 60],
    )
    write_sheet(
        wb,
        "DBチェック",
        flatten_db_checks(payload),
        [
            ("facility_name", "施設名"),
            ("db_row", "DB行"),
            ("db_m_billed_count", "DB M列 請求人数"),
            ("billing_resident_user_count", "請求 入居系人数"),
            ("billed_count_check", "請求人数判定"),
            ("db_o_provider_count", "DB O列 提供人数"),
            ("provider_active_user_count", "提供 実利用者数"),
            ("provider_sheet_user_count", "提供 シート人数"),
            ("provider_count_check", "提供人数判定"),
            ("db_t_billed_days", "DB T列 利用日数"),
            ("billing_life_support_days", "請求 延べ日数"),
            ("provider_total_days", "提供 延べ日数"),
            ("days_check", "延べ日数判定"),
            ("db_q_gap_reason", "DB Q列 理由・対策"),
        ],
        [12, 8, 16, 16, 14, 16, 16, 16, 16, 16, 14, 14, 14, 60],
    )
    write_sheet(
        wb,
        "利用者別チェック",
        flatten_users(payload),
        [
            ("facility_name", "施設名"),
            ("user_name", "利用者名"),
            ("user_result", "照合結果"),
            ("billing_life_support_days", "請求側：生活援助日中系の回数"),
            ("provider_days", "提供側：延べ日数"),
            ("difference", "差異（請求－提供）"),
            ("readable_difference", "差異の読み方"),
            ("reason_candidate", "理由候補"),
            ("reason_confidence", "理由確度"),
            ("evidence", "根拠"),
            ("check_point_display", "確認ポイント"),
            ("db_q_draft", "DB Q列への記載案"),
            ("direct_billing", "直接請求フラグ"),
        ],
        [12, 16, 14, 18, 16, 14, 58, 44, 12, 54, 44, 72, 14],
    )
    write_sheet(
        wb,
        "差異一覧",
        flatten_diffs(payload),
        [
            ("facility_name", "施設名"),
            ("user_name", "利用者名"),
            ("item", "差異区分"),
            ("billing_text", "請求側の状態"),
            ("provider_text", "提供側の状態"),
            ("difference_meaning", "差異の意味"),
            ("severity", "判定"),
            ("reason_candidate", "理由候補"),
            ("reason_confidence", "理由確度"),
            ("evidence", "根拠"),
            ("check_point", "次に確認すること"),
            ("db_q_draft", "DB Q列への記載案"),
        ],
        [12, 16, 20, 20, 24, 58, 12, 44, 12, 54, 44, 72],
    )
    write_sheet(
        wb,
        "読取ファイル一覧",
        flatten_file_records(payload),
        [
            ("facility_name", "施設名"),
            ("status", "読取結果"),
            ("folder", "対象フォルダ"),
            ("file", "読取ファイル"),
            ("note", "備考"),
        ],
        [14, 14, 72, 72, 44],
    )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


if __name__ == "__main__":
    main()
