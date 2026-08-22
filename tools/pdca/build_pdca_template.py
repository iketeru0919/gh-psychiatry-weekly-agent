"""週間PDCAシート改善版テンプレートを生成するスクリプト。

既存の「PDCAシート（施設単位・週4ブロック横並び）」の問題点を解消し、
添付の施設担当一覧をマスタとして組み込んだテンプレートを出力する。

使い方:
    python tools/pdca/build_pdca_template.py \
        --facility-master tools/pdca/facility_master_source.xlsx \
        --out dist/週間PDCA_改善版テンプレート.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ---------------------------------------------------------------- 見た目定義

NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT = "DDEBF7"
GRAY = "F2F2F2"
YELLOW = "FFF2CC"
GREEN = "E2EFDA"
RED = "FFC7CE"

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORIES = [
    "①施設稼働",
    "②短期入所",
    "③加算詳細",
    "④訪看契約",
    "⑤SQ利用",
    "⑥人材・労務",
    "⑦その他",
]
STATUSES = ["未着手", "対応中", "完了", "保留", "中止"]
WEEKS = ["第1週", "第2週", "第3週", "第4週", "第5週"]

# カテゴリごとの標準KPI（週次でログを取る指標）
STANDARD_KPIS = [
    ("①施設稼働", "GH稼働率", "%"),
    ("①施設稼働", "営業接点件数", "件"),
    ("①施設稼働", "見学・体験件数", "件"),
    ("①施設稼働", "入居契約件数", "件"),
    ("②短期入所", "SS稼働率", "%"),
    ("②短期入所", "SS案件発掘件数", "件"),
    ("③加算詳細", "夜勤加配 取得率", "%"),
    ("④訪看契約", "訪看契約件数", "件"),
    ("⑤SQ利用", "SQ契約件数", "件"),
]

PREF_RE = re.compile(r"^(.{2,3}?[都道府県])")


def title_cell(ws, cell: str, text: str, size: int = 14) -> None:
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=NAVY)


def header_row(ws, row: int, headers: list[str], start_col: int = 1, fill: str = BLUE) -> None:
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def box_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in row:
            c.border = BOX
            if c.alignment is None or not c.alignment.wrap_text:
                c.alignment = Alignment(vertical="top", wrap_text=True)


# ---------------------------------------------------------------- 施設マスタ


def read_facilities(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["施設一覧"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        kentei, name, gyoshu, tanto, addr = (r + (None,) * 5)[:5]
        main, sub = split_owner(tanto)
        pref = PREF_RE.match(addr or "")
        rows.append(
            {
                "指定権者": kentei,
                "施設名": name,
                "施設業種": gyoshu,
                "担当AM": main,
                "兼務・副担当": sub,
                "都道府県": pref.group(1) if pref else "",
                "住所": addr,
            }
        )
    return rows


def split_owner(value: str | None) -> tuple[str, str]:
    """'東 将学（尾方 範夫）' -> ('東 将学', '尾方 範夫')"""
    if not value:
        return "", ""
    m = re.match(r"^(.*?)\s*[（(](.*?)[）)]\s*$", value)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return value.strip(), ""


# ---------------------------------------------------------------- 各シート


def sheet_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("00_はじめに")
    set_widths(ws, {"A": 4, "B": 26, "C": 96})
    title_cell(ws, "B2", "週間PDCAシート 改善版テンプレート", 16)
    ws["B3"] = "既存PDCAシートの構造課題を解消し、施設担当一覧（40施設）をマスタとして内蔵した版です。"
    ws["B3"].font = Font(color="404040")

    items = [
        ("設計の狙い", "①1施設1ファイルの属人運用から、施設マスタ×週次ログの共通フォーマットへ。②文章だけのPDCAから、数値（目標／実績／達成率）で判定できるPDCAへ。"),
        ("使う順番", "03_設定 で対象施設と対象月を選ぶ → 05_週次KPIログ に毎週の実績を入力 → 04_課題管理表 でP/D/C/Aと期限を管理 → 06_ダッシュボード で会議前に確認。"),
        ("入力してよい欄", "黄色セルのみが入力欄です。白・グレーのセルは数式なので触らないでください。"),
        ("施設・担当者の選択", "プルダウンは 01_施設マスタ / 02_担当者マスタ を参照しています。施設異動や担当変更はマスタ側だけを直せば全シートに反映されます。"),
        ("週の定義", "03_設定 の「第1週の基準日」を入れると第2〜第5週は自動採番されます（旧シートの『+365日』方式は廃止）。"),
        ("横展開", "05_週次KPIログ は施設名を持つ縦持ちデータなので、40施設分を1ファイルに束ねてもピボットで集計できます。"),
        ("継続運用", "月ごとにシートを増やしません。月末に締め処理を行うと、当月ブックが実績として保管され、週次KPIは 15_KPI累積ログ に値で積み上がり、日別入力欄が空の翌月ブックが作られます。"),
        ("過去の参照", "月をまたいだ推移は 16_年間推移 で見ます。個別の日別実績は、締めのときに保管された『〇〇_実績.xlsx』を開いてください。"),
        ("改善の根拠", "07_改善提案一覧 に、旧シートで検出した具体的な不具合と改善内容を記載しています。"),
    ]
    r = 5
    for k, v in items:
        ws.cell(row=r, column=2, value=k).font = Font(bold=True, color=NAVY)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=r, column=2).alignment = Alignment(vertical="top")
        c = ws.cell(row=r, column=3, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 34
        r += 1
    box_range(ws, 5, r - 1, 2, 3)


def sheet_facility_master(wb: Workbook, facilities: list[dict]) -> None:
    ws = wb.create_sheet("01_施設マスタ")
    headers = ["施設ID", "施設名", "施設業種", "指定権者", "都道府県", "担当AM", "兼務・副担当", "住所", "PDCA対象"]
    header_row(ws, 1, headers)
    for i, f in enumerate(facilities, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=f"F{i:03d}")
        ws.cell(row=row, column=2, value=f["施設名"])
        ws.cell(row=row, column=3, value=f["施設業種"])
        ws.cell(row=row, column=4, value=f["指定権者"])
        ws.cell(row=row, column=5, value=f["都道府県"])
        ws.cell(row=row, column=6, value=f["担当AM"])
        ws.cell(row=row, column=7, value=f["兼務・副担当"])
        ws.cell(row=row, column=8, value=f["住所"])
        ws.cell(row=row, column=9, value="対象")
    last = len(facilities) + 1
    set_widths(ws, {"A": 9, "B": 34, "C": 14, "D": 10, "E": 10, "F": 20, "G": 16, "H": 42, "I": 10})
    box_range(ws, 1, last, 1, 9)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:I{last}"
    dv = DataValidation(type="list", formula1='"対象,対象外"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I{last}")
    return last


def sheet_owner_master(wb: Workbook, facilities: list[dict], fac_last: int) -> int:
    ws = wb.create_sheet("02_担当者マスタ")
    owners: list[str] = []
    for f in facilities:
        if f["担当AM"] and f["担当AM"] not in owners:
            owners.append(f["担当AM"])
    owners.sort()
    header_row(ws, 1, ["担当AM", "担当施設数", "GH（日）", "生活介護系", "主な担当エリア"])
    for i, o in enumerate(owners, start=2):
        ws.cell(row=i, column=1, value=o)
        ws.cell(row=i, column=2, value=f"=COUNTIF('01_施設マスタ'!$F$2:$F${fac_last},$A{i})")
        ws.cell(row=i, column=3, value=f'=COUNTIFS(\'01_施設マスタ\'!$F$2:$F${fac_last},$A{i},\'01_施設マスタ\'!$C$2:$C${fac_last},"GH（日）")')
        ws.cell(row=i, column=4, value=f"=$B{i}-$C{i}")
        areas = sorted({f["都道府県"] or f["指定権者"] for f in facilities if f["担当AM"] == o})
        ws.cell(row=i, column=5, value="／".join(areas))
    last = len(owners) + 1
    ws.cell(row=last + 2, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=last + 2, column=2, value=f"=SUM(B2:B{last})").font = Font(bold=True)
    set_widths(ws, {"A": 22, "B": 12, "C": 12, "D": 12, "E": 34})
    box_range(ws, 1, last, 1, 5)
    ws.freeze_panes = "A2"
    return last


def sheet_settings(wb: Workbook, fac_last: int) -> None:
    ws = wb.create_sheet("03_設定")
    set_widths(ws, {"A": 4, "B": 22, "C": 34, "D": 4, "E": 20, "F": 34})
    title_cell(ws, "B2", "対象設定（このシートを最初に入力）", 14)

    labels = [
        ("B4", "対象施設", "B5", None),
        ("B6", "対象年", "B7", 2026),
        ("B8", "対象月", "B9", 8),
        ("B10", "第1週の基準日", "B11", None),
        ("B12", "作成者", "B13", None),
    ]
    for lab_cell, lab, in_cell, default in labels:
        ws[lab_cell] = lab
        ws[lab_cell].font = Font(bold=True, color=NAVY)
        ws[lab_cell].fill = PatternFill("solid", fgColor=LIGHT)
        ws[lab_cell].border = BOX
        c = ws[in_cell]
        if default is not None:
            c.value = default
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.border = BOX
    ws["B11"].number_format = "yyyy/mm/dd"
    ws["C10"] = "※ 第1週の月曜日など、運用上の起点日を入力してください。"
    ws["C10"].font = Font(size=9, color="808080")

    dv_fac = DataValidation(type="list", formula1="施設名リスト", allow_blank=False)
    ws.add_data_validation(dv_fac)
    dv_fac.add("B5")

    title_cell(ws, "E4", "選択施設の情報（自動表示）", 12)
    info = [
        ("業種", "C"),
        ("指定権者", "D"),
        ("都道府県", "E"),
        ("担当AM", "F"),
        ("兼務・副担当", "G"),
        ("住所", "H"),
    ]
    r = 5
    for lab, col in info:
        ws.cell(row=r, column=5, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(row=r, column=5).border = BOX
        f = (
            f"=IFERROR(INDEX('01_施設マスタ'!${col}$2:${col}${fac_last},"
            f"MATCH($B$5,'01_施設マスタ'!$B$2:$B${fac_last},0)),\"\")"
        )
        c = ws.cell(row=r, column=6, value=f)
        c.border = BOX
        r += 1

    title_cell(ws, "B15", "週の定義（自動計算）", 12)
    ws["B16"] = "※ 集計期間は旧シートの『7日／14日／21日／28日時点』と同じ、月内日付での区切りです。"
    ws["B16"].font = Font(size=9, color="808080")
    header_row(ws, 17, ["週", "集計開始日", "集計終了日", "PDCA会議日", "入力締切"], start_col=2)
    for i, w in enumerate(WEEKS):
        row = 18 + i
        ws.cell(row=row, column=2, value=w)
        ws.cell(
            row=row, column=3,
            value=f'=IF(DAY(EOMONTH(DATE($B$7,$B$9,1),0))<{i * 7 + 1},"",DATE($B$7,$B$9,{i * 7 + 1}))',
        )
        ws.cell(
            row=row, column=4,
            value=f'=IF($C{row}="","",MIN(EOMONTH(DATE($B$7,$B$9,1),0),DATE($B$7,$B$9,{(i + 1) * 7})))',
        )
        if i == 0:
            ws.cell(row=row, column=5, value='=IF($B$11="","",$B$11)')
        else:
            ws.cell(row=row, column=5, value=f'=IF($E{row - 1}="","",$E{row - 1}+7)')
        ws.cell(row=row, column=6, value=f'=IF($E{row}="","",$E{row}-1)')
        for col in (3, 4, 5, 6):
            ws.cell(row=row, column=col).number_format = "yyyy/mm/dd"
    box_range(ws, 17, 22, 2, 6)


def sheet_issues(wb: Workbook, fac_last: int) -> int:
    ws = wb.create_sheet("04_課題管理表")
    headers = [
        "No.",
        "対象施設",
        "カテゴリ（KGI領域）",
        "課題（KPI）",
        "KGI目標",
        "KPI目標値",
        "単位",
        "現状値",
        "達成率",
        "対応方針（Plan）",
        "実行内容（Do）",
        "検証・所感（Check）",
        "次アクション（Act）",
        "担当者",
        "起票日",
        "期限",
        "ステータス",
        "残日数",
        "遅延判定",
        "完了日",
        "備考",
    ]
    header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 40
    rows = 60
    for i in range(2, rows + 2):
        ws.cell(row=i, column=1, value=f'=IF($D{i}="","",ROW()-1)')
        ws.cell(row=i, column=2, value='=IF($D%d="","",設定対象施設)' % i)
        ws.cell(row=i, column=9, value=f'=IFERROR(IF($F{i}=0,"",$H{i}/$F{i}),"")')
        ws.cell(row=i, column=9).number_format = "0.0%"
        ws.cell(row=i, column=15).number_format = "yyyy/mm/dd"
        ws.cell(row=i, column=16).number_format = "yyyy/mm/dd"
        ws.cell(row=i, column=20).number_format = "yyyy/mm/dd"
        ws.cell(row=i, column=18, value=f'=IF(OR($P{i}="",$Q{i}="完了",$Q{i}="中止"),"",$P{i}-TODAY())')
        ws.cell(
            row=i,
            column=19,
            value=(
                f'=IF($P{i}="","",IF(OR($Q{i}="完了",$Q{i}="中止"),"—",'
                f'IF($P{i}<TODAY(),"遅延",IF($P{i}-TODAY()<=2,"要注意","")))) '
            ).strip(),
        )
    last = rows + 1
    set_widths(
        ws,
        {
            "A": 6, "B": 26, "C": 16, "D": 30, "E": 22, "F": 12, "G": 8, "H": 12, "I": 10,
            "J": 34, "K": 30, "L": 30, "M": 30, "N": 16, "O": 12, "P": 12, "Q": 12,
            "R": 10, "S": 10, "T": 12, "U": 24,
        },
    )
    box_range(ws, 2, last, 1, 21)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:U{last}"

    # 入力欄（黄色）は D〜H, J〜Q, T, U
    for col in list("DEFGH") + list("JKLMNOPQ") + ["T", "U"]:
        for r in range(2, last + 1):
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=YELLOW)

    dv_cat = DataValidation(type="list", formula1="カテゴリリスト", allow_blank=True)
    dv_sta = DataValidation(type="list", formula1="ステータスリスト", allow_blank=True)
    dv_own = DataValidation(type="list", formula1="担当者リスト", allow_blank=True)
    for dv, ref in ((dv_cat, f"C2:C{last}"), (dv_sta, f"Q2:Q{last}"), (dv_own, f"N2:N{last}")):
        ws.add_data_validation(dv)
        dv.add(ref)

    ws.conditional_formatting.add(
        f"A2:U{last}",
        FormulaRule(formula=[f'$S2="遅延"'], fill=PatternFill("solid", fgColor=RED), stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        f"A2:U{last}",
        FormulaRule(formula=[f'$Q2="完了"'], fill=PatternFill("solid", fgColor=GRAY), stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        f"I2:I{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    return last



# 週次KPIログの「実績」を、月次入力シート（10〜14）から自動集計するための対応表
PIPE_LAST = 61  # 14_案件管理台帳のデータ最終行


def _week_start(row: int) -> str:
    return f"INDEX('03_設定'!$C$18:$C$22,MATCH($E{row},'03_設定'!$B$18:$B$22,0))"


def _week_end(row: int) -> str:
    return f"INDEX('03_設定'!$D$18:$D$22,MATCH($E{row},'03_設定'!$B$18:$B$22,0))"


def _pipeline_count(col: str, row: int, extra: str = "") -> str:
    """14_案件管理台帳の日付列が、その週の期間に入る件数を数える。"""
    rng = f"'14_案件管理台帳'!${col}$2:${col}${PIPE_LAST}"
    return (
        f'COUNTIFS({rng},">="&{_week_start(row)},{rng},"<="&{_week_end(row)}{extra})'
    )


def _lookup_week(sheet: str, value_range: str, label_range: str, row: int) -> str:
    return (
        f"IFERROR(INDEX('{sheet}'!{value_range},"
        f"MATCH($E{row},'{sheet}'!{label_range},0)),\"\")"
    )


def _auto_source(kpi: str, row: int) -> str:
    if kpi == "GH稼働率":
        return _lookup_week("10_月次_GH稼働", "$AQ$4:$AQ$8", "$AN$4:$AN$8", row)
    if kpi == "SS稼働率":
        return _lookup_week("11_月次_SS稼働", "$AQ$4:$AQ$8", "$AN$4:$AN$8", row)
    if kpi == "営業接点件数":
        return _lookup_week("12_月次_営業行動", "$AO$4:$AO$8", "$AN$4:$AN$8", row)
    if kpi == "夜勤加配 取得率":
        return _lookup_week("13_月次_加算要件", "$AQ$5:$AQ$9", "$AN$5:$AN$9", row)
    if kpi == "見学・体験件数":
        return f"{_pipeline_count('H', row)}+{_pipeline_count('I', row)}"
    if kpi == "入居契約件数":
        return _pipeline_count("L", row)
    if kpi == "SS案件発掘件数":
        return _pipeline_count("G", row, ",'14_案件管理台帳'!$D$2:$D$%d,\"SS\"" % PIPE_LAST)
    return ""


def sheet_kpi_log(wb: Workbook) -> int:
    ws = wb.create_sheet("05_週次KPIログ")
    headers = [
        "記録日", "対象施設", "年", "月", "週", "カテゴリ", "指標名", "単位",
        "週次目標", "実績", "達成率", "差異", "担当者", "未達要因（Check）", "次週アクション（Act）",
        "実績（手入力上書き）",
    ]
    header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 32

    row = 2
    for w in WEEKS:
        for cat, kpi, unit in STANDARD_KPIS:
            ws.cell(row=row, column=2, value="=設定対象施設")
            ws.cell(row=row, column=3, value="=設定対象年")
            ws.cell(row=row, column=4, value="=設定対象月")
            ws.cell(row=row, column=5, value=w)
            ws.cell(row=row, column=6, value=cat)
            ws.cell(row=row, column=7, value=kpi)
            ws.cell(row=row, column=8, value=unit)
            auto = _auto_source(kpi, row)
            if auto:
                ws.cell(row=row, column=10, value=f'=IF($P{row}<>"",$P{row},{auto})')
            else:
                ws.cell(row=row, column=10, value=f'=IF($P{row}<>"",$P{row},"")')
            ws.cell(row=row, column=11, value=f'=IFERROR(IF($I{row}=0,"",$J{row}/$I{row}),"")')
            ws.cell(row=row, column=11).number_format = "0.0%"
            ws.cell(row=row, column=12, value=f'=IF(OR($I{row}="",$J{row}=""),"",$J{row}-$I{row})')
            ws.cell(row=row, column=1).number_format = "yyyy/mm/dd"
            if unit == "%":
                ws.cell(row=row, column=9).number_format = "0.0%"
                ws.cell(row=row, column=10).number_format = "0.0%"
            row += 1
    last = row - 1
    set_widths(
        ws,
        {"A": 12, "B": 26, "C": 7, "D": 7, "E": 9, "F": 14, "G": 20, "H": 7,
         "I": 11, "J": 11, "K": 10, "L": 9, "M": 16, "N": 34, "O": 34, "P": 14},
    )
    box_range(ws, 2, last, 1, 16)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:P{last}"
    for col in ["A", "I", "M", "N", "O", "P"]:
        for r in range(2, last + 1):
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=YELLOW)
    ws.cell(row=1, column=16).comment = None
    ws.cell(row=last + 2, column=1, value=(
        "※ J列『実績』は 10〜14 の入力シートから自動集計されます。"
        "自動値を使わない場合のみ P列に数値を入れると、そちらが優先されます。"
    )).font = Font(size=9, color="808080")
    dv_own = DataValidation(type="list", formula1="担当者リスト", allow_blank=True)
    ws.add_data_validation(dv_own)
    dv_own.add(f"M2:M{last}")
    ws.conditional_formatting.add(
        f"K2:K{last}",
        CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        f"K2:K{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    return last


def sheet_dashboard(wb: Workbook, kpi_last: int, issue_last: int, owner_last: int) -> None:
    ws = wb.create_sheet("06_ダッシュボード")
    set_widths(
        ws,
        {"A": 4, "B": 20, "C": 14, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10,
         "I": 12, "J": 12, "K": 10, "L": 18, "M": 12, "N": 12},
    )
    title_cell(ws, "B2", "週間PDCA ダッシュボード", 16)
    ws["B3"] = '=設定対象施設&"／"&設定対象年&"年"&設定対象月&"月"'
    ws["B3"].font = Font(bold=True, color=BLUE, size=12)

    title_cell(ws, "B5", "■ KPI週次推移（05_週次KPIログの集計）", 12)
    header_row(ws, 6, ["指標名", "単位"] + WEEKS + ["月間目標", "月間実績", "達成率", "判定"], start_col=2)
    r = 7
    for cat, kpi, unit in STANDARD_KPIS:
        ws.cell(row=r, column=2, value=kpi)
        ws.cell(row=r, column=3, value=unit)
        # 率（%）は週平均、件数は合計で月間値を出す
        agg = "AVERAGEIFS" if unit == "%" else "SUMIFS"
        for i, w in enumerate(WEEKS):
            col = 4 + i
            ws.cell(
                row=r,
                column=col,
                value=(
                    f"=IFERROR({agg}('05_週次KPIログ'!$J$2:$J${kpi_last},"
                    f"'05_週次KPIログ'!$G$2:$G${kpi_last},$B{r},"
                    f"'05_週次KPIログ'!$E$2:$E${kpi_last},\"{w}\"),\"\")"
                ),
            )
            if unit == "%":
                ws.cell(row=r, column=col).number_format = "0.0%"
        ws.cell(
            row=r, column=9,
            value=f"=IFERROR({agg}('05_週次KPIログ'!$I$2:$I${kpi_last},'05_週次KPIログ'!$G$2:$G${kpi_last},$B{r}),\"\")",
        )
        ws.cell(
            row=r, column=10,
            value=f"=IFERROR({agg}('05_週次KPIログ'!$J$2:$J${kpi_last},'05_週次KPIログ'!$G$2:$G${kpi_last},$B{r}),\"\")",
        )
        if unit == "%":
            ws.cell(row=r, column=9).number_format = "0.0%"
            ws.cell(row=r, column=10).number_format = "0.0%"
        ws.cell(row=r, column=11, value=f'=IFERROR(IF($I{r}=0,"",$J{r}/$I{r}),"")')
        ws.cell(row=r, column=11, value=f'=IFERROR(IF($I{r}=0,"",$J{r}/$I{r}),"")')
        ws.cell(row=r, column=11).number_format = "0.0%"
        ws.cell(row=r, column=12, value=f'=IF($K{r}="","未入力",IF($K{r}>=1,"〇 達成",IF($K{r}>=0.9,"△ 要注意","✕ 未達")))')
        r += 1
    kpi_end = r - 1
    box_range(ws, 6, kpi_end, 2, 12)
    ws.conditional_formatting.add(
        f"L7:L{kpi_end}",
        FormulaRule(formula=[f'$L7="✕ 未達"'], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        f"L7:L{kpi_end}",
        FormulaRule(formula=[f'$L7="〇 達成"'], fill=PatternFill("solid", fgColor=GREEN)),
    )

    top = kpi_end + 2
    title_cell(ws, f"B{top}", "■ 課題ステータス（04_課題管理表の集計）", 12)
    header_row(ws, top + 1, ["ステータス", "件数"], start_col=2)
    for i, s in enumerate(STATUSES):
        rr = top + 2 + i
        ws.cell(row=rr, column=2, value=s)
        ws.cell(row=rr, column=3, value=f"=COUNTIF('04_課題管理表'!$Q$2:$Q${issue_last},$B{rr})")
    rr = top + 2 + len(STATUSES)
    ws.cell(row=rr, column=2, value="うち遅延").font = Font(bold=True, color="C00000")
    ws.cell(row=rr, column=3, value=f'=COUNTIF(\'04_課題管理表\'!$S$2:$S${issue_last},"遅延")')
    ws.cell(row=rr + 1, column=2, value="うち要注意（期限2日以内）")
    ws.cell(row=rr + 1, column=3, value=f'=COUNTIF(\'04_課題管理表\'!$S$2:$S${issue_last},"要注意")')
    box_range(ws, top + 1, rr + 1, 2, 3)

    title_cell(ws, f"E{top}", "■ 担当AM別 課題状況（02_担当者マスタ連動）", 12)
    header_row(ws, top + 1, ["担当AM", "担当施設数", "課題件数", "対応中", "遅延件数"], start_col=5)
    for i in range(2, owner_last + 1):
        rr = top + i
        ws.cell(row=rr, column=5, value=f"='02_担当者マスタ'!$A{i}")
        ws.cell(row=rr, column=6, value=f"='02_担当者マスタ'!$B{i}")
        ws.cell(row=rr, column=7, value=f"=COUNTIF('04_課題管理表'!$N$2:$N${issue_last},$E{rr})")
        ws.cell(
            row=rr, column=8,
            value=f'=COUNTIFS(\'04_課題管理表\'!$N$2:$N${issue_last},$E{rr},\'04_課題管理表\'!$Q$2:$Q${issue_last},"対応中")',
        )
        ws.cell(
            row=rr, column=9,
            value=f'=COUNTIFS(\'04_課題管理表\'!$N$2:$N${issue_last},$E{rr},\'04_課題管理表\'!$S$2:$S${issue_last},"遅延")',
        )
    box_range(ws, top + 1, top + owner_last, 5, 9)
    ws.conditional_formatting.add(
        f"I{top + 2}:I{top + owner_last}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
    )


def sheet_proposals(wb: Workbook) -> None:
    ws = wb.create_sheet("07_改善提案一覧")
    headers = ["No.", "区分", "現行シートの状態", "何が問題か（リスク）", "改善案", "本テンプレートでの対応", "優先度"]
    header_row(ws, 1, headers)
    data = [
        ("不具合", "週間PDCA AE5/AE7/AE9/AE11（第4週ブロック）が第1週の =D5/=D7/=D9/=D11 を参照。第2→第3週は連鎖しているのに第4週だけ第1週に戻っている。",
         "第4週の行動設定が第3週の続きにならず、月末週のPDCAが第1週の焼き直しになる。改善サイクルが1周分失われる。",
         "週ブロックを横並びコピーで作らず、週を『行』として縦持ちにする。",
         "05_週次KPIログ で週を行に持たせ、参照連鎖そのものを廃止。", "高"),
        ("不具合", "週間PDCA C17 が ='2月'!BF30 と特定月シートに直リンク。他のKGI/KPIは AM列のHLOOKUP経由。",
         "対象月を切り替えても④訪看契約のKPIだけ2月の値が出続ける。誤った達成判断につながる。",
         "参照経路を1本化し、対象月は1か所の設定セルで制御する。",
         "03_設定 の対象年・対象月を全シートが参照。個別月シートへの直リンクなし。", "高"),
        ("不具合", "AN65:AY65 などの翌月系が =AO51〜=AZ51 と1列ずらし。最終列 AY65 は空のAZ51を参照。",
         "3月（年度末）の『翌月稼働』『リピート』『予算比』『前月比』が常に0になる。",
         "ずらし参照ではなく、月をキーにした縦持ち＋集計で翌月値を取る。",
         "05_週次KPIログ＋06_ダッシュボードのSUMIFS集計に置換。", "高"),
        ("不具合", "G1 = AM63+365（月シートのN3日付に365日を足して基準日を作る）。以降は +7 の連鎖。",
         "うるう年で1日ずれる。元の日付の意味も追えず、年度をまたぐと破綻する。",
         "基準日は人が入れる1セルに集約し、週は+7で自動採番する。",
         "03_設定 B11『第1週の基準日』→ 第1〜第5週を自動計算。", "高"),
        ("構造", "月シートが22枚あるのに、HLOOKUP表 AN24:AY63 は12か月分しか列を持たない。'4月 '（末尾スペース）、'26年６月'など命名も不統一。",
         "26年6月以降のシートは集計対象外。シート名を1文字直しただけで全参照が #REF! になる。",
         "月ごとにシートを増やさず、1枚のログに年・月列を持たせる。",
         "05_週次KPIログ に『年』『月』列。シート追加不要で何年でも蓄積可能。", "高"),
        ("構造", "1施設1ファイル運用で、ファイル内のどこにも施設名・担当AMが入っていない。担当者欄（F/O/X/AG列）は空欄の自由入力。",
         "40施設を横断した集計・比較ができない。ファイル名だけが識別子なので取り違えが起きる。担当者の表記ゆれで名寄せできない。",
         "施設マスタ／担当者マスタを内蔵し、施設・担当はプルダウン選択にする。",
         "01_施設マスタ（40施設）・02_担当者マスタ（16名）＋03_設定で施設を選択。担当AM・指定権者・住所は自動表示。", "高"),
        ("構造", "第1〜第4週が横に4ブロック並び、各ブロックが前週セルを参照するコピー構造。1枚に51列。",
         "行の挿入・削除でブロックがずれる。第5週がある月に対応できない。印刷・レビューがしづらい。",
         "縦持ちのテーブルに変更し、週は行方向へ増やす。",
         "04_課題管理表・05_週次KPIログはいずれも縦持ちテーブル。第5週まで標準対応。", "高"),
        ("構造", "継続運用の方法が『月シートを1枚ずつ増やす』こと。実ファイルでは22枚まで増え、命名も不統一になっている。",
         "シートが増えるほどブックが重くなり、集計式（12か月固定のHLOOKUP表）が追従できない。年度をまたぐと参照が壊れ、過去の修正が現在の集計に波及する。",
         "月ごとにシートを増やさず、月末に締めて『実績ブックの保管＋累積ログへの追記＋翌月ブックの生成』に分ける。",
         "15_KPI累積ログ（値で保管）＋16_年間推移＋tools/pdca/close_month.py（締め処理）。ブックの枚数は月に関係なく一定。", "高"),
        ("運用", "対応方針・所感が1セルに改行で詰め込まれた長文（【営業行動】…▶2/25 追客 等）。",
         "目標値と実績が数式で取り出せず、達成／未達の自動判定ができない。読む人によって解釈が変わる。",
         "『数値で置く欄』と『文章で残す欄』を分離する。",
         "週次目標・実績・達成率・差異を数値列に分離し、所感は別列に残す。", "高"),
        ("運用", "ステータス列・期限列（G/H, P/Q…）は結合セルで、入力規則も条件付き書式もなく空欄のまま。",
         "期限超過が誰にも見えない。ステータスの表記ゆれ（完了／済／OK）で集計不能。",
         "ステータスはプルダウン、期限は日付、遅延は自動判定＋色付け。",
         "04_課題管理表にステータスDV・残日数・遅延判定・行の自動着色を実装。", "高"),
        ("運用", "『※未対象施設』とだけ書かれた行（④訪看契約・⑤SQ利用）が毎週コピーされ続けている。",
         "対象外の項目が常時表示され、シートのノイズになる。対象化した時に誰が判断するか決まっていない。",
         "対象／対象外を施設マスタ側の属性で持ち、非対象は非表示にする。",
         "01_施設マスタ I列『PDCA対象』で管理。カテゴリは04で必要な行だけ起票する運用。", "中"),
        ("運用", "『週間PDCA (見本)』が本体と別に存在し、内容が本体とずれている。",
         "見本の更新漏れで、書き方の基準が人によって変わる。",
         "見本を別シートで持たず、テンプレート自体に入力例と注記を持たせる。",
         "00_はじめに に運用手順、各シートの黄色セルで入力箇所を明示。", "中"),
        ("運用", "予算達成状況の判定が =IF(予算比>=0,\"〇\",\"✕\") の2値のみ。前月比の一部は文字列 '-'。",
         "あと一歩の未達と大幅未達が同じ『✕』。文字列混在で平均・グラフが崩れる。",
         "達成／要注意／未達の3段階にし、空欄は空欄のまま扱う。",
         "06_ダッシュボードで 100%以上=〇、90%以上=△、それ未満=✕、未入力は『未入力』と判定。", "中"),
        ("データ", "Sheet2 に居室番号・区分・年齢の生データが残置。個人に紐づく情報がシート内に混在。",
         "配布時に意図しない情報が同梱される。マスタとしての位置づけも不明。",
         "個人データは配布用テンプレートから分離する。",
         "本テンプレートには個人データを含めない。必要なら別ファイルで管理。", "中"),
        ("拡張", "施設一覧側に『グループホームイノベル信中島／公式名 小信中島』など名称差3件、住所照合注記あり。",
         "施設名がキーなのに正式名称と社内呼称が混在すると、集計時に名寄せできない。",
         "施設IDを主キーにし、名称は表示用の属性として持つ。",
         "01_施設マスタに施設ID（F001〜F040）を付与。名称変更はマスタ1か所で吸収。", "中"),
        ("構造", "月シート（4月〜26年9月の22枚）が、利用者ごとの日別稼働カレンダー・SS予約・営業行動・案件管理・加算要件の入力面を兼ねている。週間PDCAの文章は BF16 等の長大な文字列連結式で自動生成されている。",
         "入力面（原簿）と管理面（PDCA）が同じシートに同居し、月が変わるたびシートを増やす必要がある。連結式で作った文章は、あとから数値として再利用できない。",
         "原簿は月次入力シートとして分離し、集計値だけをPDCA側へ渡す。",
         "10〜13の月次入力シートに分離。週別集計値は05_週次KPIログが自動で参照する。", "高"),
        ("構造", "案件管理が『kintone名×日別セルに進捗記号』のマトリクス（55〜78行）。相談元・区分・希望種別が別セルに散る。",
         "同じ案件の履歴が月シートをまたいで分断され、発掘→見学→体験→契約のリードタイムが測れない。",
         "案件は1件1行の台帳にし、各フェーズの日付を列で持つ。",
         "14_案件管理台帳。日付列から見学・体験・入居件数を週次で自動集計し、停滞14日超は赤表示。", "高"),
        ("運用", "重度障害者支援体制・福祉専門職員配置等加算の要件確認が、月シートの隅にチェック欄として置かれている（BR88 常勤比率0.368→✕ 等）。",
         "算定可否の根拠が月シートを開かないと見えず、要件を割った時に誰がいつ気づくかが運用任せ。",
         "要件判定を独立させ、割り込み時に『要対応』として表示する。",
         "13_月次_加算要件に集約。夜勤加配は日別実績から取得率を自動計算、体制加算は✕が1つでもあれば『要対応』。", "中"),
        ("拡張", "40施設分のファイルが個別に存在し、AMは自分の担当分だけを見る運用。",
         "エリア長・本部が全体傾向（未達KPI・遅延課題の偏り）を把握するのに手作業の集約が必要。",
         "縦持ちログを束ねてピボット集計できる形にする。",
         "05_週次KPIログは施設名・担当者を各行に持つため、40施設分を連結してそのまま集計可能。06に担当AM別の遅延件数を実装。", "中"),
    ]
    for i, row in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        for j, v in enumerate(row, start=2):
            ws.cell(row=i, column=j, value=v)
        ws.row_dimensions[i].height = 58
    last = len(data) + 1
    set_widths(ws, {"A": 6, "B": 10, "C": 46, "D": 42, "E": 38, "F": 40, "G": 8})
    box_range(ws, 2, last, 1, 7)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:G{last}"
    ws.conditional_formatting.add(
        f"G2:G{last}", FormulaRule(formula=['$G2="高"'], fill=PatternFill("solid", fgColor=RED))
    )


def sheet_rules(wb: Workbook) -> None:
    ws = wb.create_sheet("08_運用ルール")
    set_widths(ws, {"A": 4, "B": 14, "C": 20, "D": 30, "E": 60})
    title_cell(ws, "B2", "週次運用ルール（案）", 14)
    header_row(ws, 4, ["タイミング", "誰が", "何を", "判断基準"], start_col=2)
    rows = [
        ("週次（月）", "管理者", "05_週次KPIログに前週実績を入力", "金曜締め・月曜正午までに入力完了"),
        ("週次（月）", "担当AM", "06_ダッシュボードで未達KPIを確認", "達成率90%未満の指標は必ず課題化する"),
        ("週次（火）", "担当AM＋管理者", "04_課題管理表でPlan/Doを更新し、期限を再設定", "遅延判定が『遅延』の課題はその場で期限を引き直す"),
        ("週次（火）", "担当AM", "Check（未達要因）とAct（次週アクション）を1行で記載", "要因は事実、アクションは動詞で書く"),
        ("月次（月初）", "エリア長", "40施設分のログを連結し担当AM別に確認", "遅延件数が2件以上のAMはヒアリング対象"),
        ("月次（月末）", "担当AM", "Excelで保存後、月締め処理を実行して翌月ブックを作成", "締める前に必ずExcelで保存する（数式の値が確定するため）"),
        ("月次（翌月初）", "担当AM", "16_年間推移で前月までの推移を確認", "2か月連続で未達のKPIは対応方針を作り直す"),
        ("随時", "本部", "01_施設マスタの担当変更・新規開設を反映", "マスタ更新は本部のみ。現場は編集しない"),
    ]
    for i, r in enumerate(rows, start=5):
        for j, v in enumerate(r, start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 30
    box_range(ws, 4, 4 + len(rows), 2, 5)


def add_defined_names(wb: Workbook, fac_last: int, owner_last: int) -> None:
    names = {
        "施設名リスト": f"'01_施設マスタ'!$B$2:$B${fac_last}",
        "担当者リスト": f"'02_担当者マスタ'!$A$2:$A${owner_last}",
        "カテゴリリスト": "'09_リスト'!$A$2:$A$8",
        "ステータスリスト": "'09_リスト'!$B$2:$B$6",
        "週リスト": "'09_リスト'!$C$2:$C$6",
        "設定対象施設": "'03_設定'!$B$5",
        "設定対象年": "'03_設定'!$B$7",
        "設定対象月": "'03_設定'!$B$9",
    }
    for n, ref in names.items():
        wb.defined_names.add(DefinedName(n, attr_text=ref))


def sheet_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("09_リスト")
    header_row(ws, 1, ["カテゴリ", "ステータス", "週"])
    for i, v in enumerate(CATEGORIES, start=2):
        ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(STATUSES, start=2):
        ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(WEEKS, start=2):
        ws.cell(row=i, column=3, value=v)
    set_widths(ws, {"A": 16, "B": 12, "C": 10})
    ws.sheet_state = "hidden"


# ------------------------------------------------------------ 月次入力（原簿）

DAY_COL0 = 5  # E列から日付が始まる
DAYS = 31
MARKS_GH = '"●,入,外"'  # ●=在籍稼働 / 入=入院 / 外=外泊


def _day_header(ws, row: int, first_col: int = DAY_COL0) -> None:
    """1〜31日の日付ヘッダと、その下に週区分行を敷く。"""
    for d in range(1, DAYS + 1):
        col = first_col + d - 1
        c = ws.cell(
            row=row,
            column=col,
            value=f'=IF(DAY(EOMONTH(DATE(設定対象年,設定対象月,1),0))<{d},"",DATE(設定対象年,設定対象月,{d}))',
        )
        c.number_format = "d"
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = BOX
        w = ws.cell(row=row + 1, column=col, value=f'=IF({get_column_letter(col)}{row}="","","第"&ROUNDUP({d}/7,0)&"週")')
        w.font = Font(size=8, color="808080")
        w.alignment = Alignment(horizontal="center")
        w.border = BOX
        ws.column_dimensions[get_column_letter(col)].width = 4.2
    ws.cell(row=row, column=first_col - 1, value="日").font = Font(bold=True, size=9)
    ws.cell(row=row + 1, column=first_col - 1, value="週").font = Font(size=8, color="808080")


def _week_block(ws, top: int, headers: list[str], col: int = 40) -> int:
    """右側に第1〜第5週の集計ブロックを置き、先頭行番号を返す。"""
    header_row(ws, top, ["週"] + headers, start_col=col)
    for i, w in enumerate(WEEKS):
        ws.cell(row=top + 1 + i, column=col, value=w)
    box_range(ws, top, top + len(WEEKS), col, col + len(headers))
    for i in range(len(headers) + 1):
        ws.column_dimensions[get_column_letter(col + i)].width = 13
    return top + 1


def sheet_gh_occupancy(wb: Workbook) -> None:
    ws = wb.create_sheet("10_月次_GH稼働")
    title_cell(ws, "A1", "共同生活援助 稼働実績（●=稼働 / 入=入院 / 外=外泊。入院・外泊は未算定日）", 12)
    header_row(ws, 3, ["No.", "利用者ID", "利用者名（任意）", "区分"])
    _day_header(ws, 3)
    last_day_col = DAY_COL0 + DAYS - 1
    dl = get_column_letter(last_day_col)
    first = 5
    rows = 30
    for i in range(rows):
        r = first + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=DAY_COL0 + DAYS + 1, value=f'=COUNTIF($E{r}:${dl}{r},"●")')
        ws.cell(row=r, column=DAY_COL0 + DAYS + 2, value=f'=COUNTIF($E{r}:${dl}{r},"入")')
        ws.cell(row=r, column=DAY_COL0 + DAYS + 3, value=f'=COUNTIF($E{r}:${dl}{r},"外")')
    last = first + rows - 1
    header_row(ws, 4, ["稼働日数", "入院", "外泊"], start_col=DAY_COL0 + DAYS + 1)
    for c in range(DAY_COL0 + DAYS + 1, DAY_COL0 + DAYS + 4):
        ws.column_dimensions[get_column_letter(c)].width = 9
    set_widths(ws, {"A": 5, "B": 11, "C": 20, "D": 6})
    box_range(ws, 3, last, 1, DAY_COL0 + DAYS + 3)
    ws.freeze_panes = "E5"
    dv = DataValidation(type="list", formula1=MARKS_GH, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{first}:{dl}{last}")
    for r in range(first, last + 1):
        for c in range(2, DAY_COL0 + DAYS):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=YELLOW)

    # 日別合計行（週別集計はこの行を参照する：配列の形を1行に揃えるため）
    tot = {}
    for k, (label, mark) in enumerate({"稼働": "●", "入院": "入", "外泊": "外"}.items()):
        rr = last + 1 + k
        tot[label] = rr
        ws.cell(row=rr, column=2, value=f"日別{label}数").font = Font(bold=True, size=9)
        for d in range(DAYS):
            cl = get_column_letter(DAY_COL0 + d)
            ws.cell(row=rr, column=DAY_COL0 + d, value=f'=COUNTIF({cl}${first}:{cl}${last},"{mark}")')
        box_range(ws, rr, rr, 2, DAY_COL0 + DAYS - 1)

    cap_row = last + 5
    ws.cell(row=cap_row, column=1, value="定員").font = Font(bold=True)
    cap = ws.cell(row=cap_row, column=4, value=20)
    cap.fill = PatternFill("solid", fgColor=YELLOW)
    cap.border = BOX
    ws.cell(row=cap_row + 1, column=1, value="※ 定員を入力すると週別稼働率が自動計算されます").font = Font(size=9, color="808080")

    top = _week_block(ws, 3, ["日数", "稼働回数", "稼働率", "入院", "外泊"])
    for i, w in enumerate(WEEKS):
        r = top + i
        ws.cell(row=r, column=41, value=f'=COUNTIF($E$4:${dl}$4,$AN{r})')
        ws.cell(row=r, column=42, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*$E${tot["稼働"]}:${dl}${tot["稼働"]})')
        ws.cell(row=r, column=43, value=f'=IFERROR($AP{r}/($AO{r}*$D${cap_row}),"")')
        ws.cell(row=r, column=43).number_format = "0.0%"
        ws.cell(row=r, column=44, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*$E${tot["入院"]}:${dl}${tot["入院"]})')
        ws.cell(row=r, column=45, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*$E${tot["外泊"]}:${dl}${tot["外泊"]})')


def sheet_ss_occupancy(wb: Workbook) -> None:
    ws = wb.create_sheet("11_月次_SS稼働")
    title_cell(ws, "A1", "短期入所 稼働実績（●=利用 / 予=予約 / ×=キャンセル）", 12)
    header_row(ws, 3, ["No.", "利用者ID", "利用者名（任意）", "区分"])
    _day_header(ws, 3)
    dl = get_column_letter(DAY_COL0 + DAYS - 1)
    first, rows = 5, 15
    for i in range(rows):
        r = first + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=DAY_COL0 + DAYS + 1, value=f'=COUNTIF($E{r}:${dl}{r},"●")')
        ws.cell(row=r, column=DAY_COL0 + DAYS + 2, value=f'=COUNTIF($E{r}:${dl}{r},"×")')
    last = first + rows - 1
    header_row(ws, 4, ["利用回数", "キャンセル"], start_col=DAY_COL0 + DAYS + 1)
    for c in range(DAY_COL0 + DAYS + 1, DAY_COL0 + DAYS + 3):
        ws.column_dimensions[get_column_letter(c)].width = 10
    set_widths(ws, {"A": 5, "B": 11, "C": 20, "D": 6})
    box_range(ws, 3, last, 1, DAY_COL0 + DAYS + 2)
    ws.freeze_panes = "E5"
    dv = DataValidation(type="list", formula1='"●,予,×"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{first}:{dl}{last}")
    for r in range(first, last + 1):
        for c in range(2, DAY_COL0 + DAYS):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=YELLOW)

    tot_use, tot_cxl = last + 1, last + 2
    for rr, label, mark in ((tot_use, "利用", "●"), (tot_cxl, "キャンセル", "×")):
        ws.cell(row=rr, column=2, value=f"日別{label}数").font = Font(bold=True, size=9)
        for d in range(DAYS):
            cl = get_column_letter(DAY_COL0 + d)
            ws.cell(row=rr, column=DAY_COL0 + d, value=f'=COUNTIF({cl}${first}:{cl}${last},"{mark}")')
        box_range(ws, rr, rr, 2, DAY_COL0 + DAYS - 1)

    bed_row = last + 4
    ws.cell(row=bed_row, column=1, value="申請床数").font = Font(bold=True)
    bed = ws.cell(row=bed_row, column=4, value=2)
    bed.fill = PatternFill("solid", fgColor=YELLOW)
    bed.border = BOX

    ws.cell(row=2, column=47, value="▼作業列（週別 利用有無）").font = Font(size=8, color="808080")
    top = _week_block(ws, 3, ["日数", "稼働回数", "稼働率", "キャンセル", "利用人数"])
    for i, w in enumerate(WEEKS):
        r = top + i
        ws.cell(row=r, column=41, value=f'=COUNTIF($E$4:${dl}$4,$AN{r})')
        ws.cell(row=r, column=42, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*$E${tot_use}:${dl}${tot_use})')
        ws.cell(row=r, column=43, value=f'=IFERROR($AP{r}/($AO{r}*$D${bed_row}),"")')
        ws.cell(row=r, column=43).number_format = "0.0%"
        ws.cell(row=r, column=44, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*$E${tot_cxl}:${dl}${tot_cxl})')
        # 利用人数＝その週に1回でも●がある利用者数（AU:AY の作業列で判定）
        helper = get_column_letter(47 + i)
        ws.cell(row=r, column=45, value=f"=SUM(${helper}${first}:${helper}${last})")
        hc = ws.cell(row=3, column=47 + i, value=w)
        hc.font = Font(size=8, color="808080")
        ws.column_dimensions[helper].width = 7
        for rr in range(first, last + 1):
            ws.cell(
                row=rr,
                column=47 + i,
                value=f'=IF(SUMPRODUCT(($E$4:${dl}$4=$AN{r})*($E{rr}:${dl}{rr}="●"))>0,1,0)',
            )


def sheet_sales_actions(wb: Workbook) -> None:
    ws = wb.create_sheet("12_月次_営業行動")
    title_cell(ws, "A1", "営業行動（訪問・郵送・FAX等の接点件数を日別に入力）", 12)
    header_row(ws, 3, ["No.", "営業チャネル", "月間目標", "備考"])
    _day_header(ws, 3)
    dl = get_column_letter(DAY_COL0 + DAYS - 1)
    channels = [
        "相談支援事業所",
        "基幹相談支援事業所",
        "障がい福祉課、生活保護課",
        "病棟のある病院",
        "就労支援、生活介護、宿泊訓練",
    ]
    first = 5
    for i, ch in enumerate(channels):
        r = first + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=ch)
        ws.cell(row=r, column=DAY_COL0 + DAYS + 1, value=f"=SUM($E{r}:${dl}{r})")
    last = first + len(channels) - 1
    total = last + 1
    ws.cell(row=total, column=2, value="合計").font = Font(bold=True)
    ws.cell(row=total, column=3, value=f"=SUM($C{first}:$C{last})")
    for d in range(DAYS):
        col = get_column_letter(DAY_COL0 + d)
        ws.cell(row=total, column=DAY_COL0 + d, value=f"=SUM({col}{first}:{col}{last})")
    ws.cell(row=total, column=DAY_COL0 + DAYS + 1, value=f"=SUM($E{total}:${dl}{total})")
    header_row(ws, 4, ["月間合計"], start_col=DAY_COL0 + DAYS + 1)
    ws.column_dimensions[get_column_letter(DAY_COL0 + DAYS + 1)].width = 11
    set_widths(ws, {"A": 5, "B": 26, "C": 10, "D": 14})
    box_range(ws, 3, total, 1, DAY_COL0 + DAYS + 1)
    ws.freeze_panes = "E5"
    for r in range(first, last + 1):
        for c in list(range(3, 5)) + list(range(DAY_COL0, DAY_COL0 + DAYS)):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=YELLOW)

    top = _week_block(ws, 3, ["接点件数"])
    for i, w in enumerate(WEEKS):
        r = top + i
        ws.cell(row=r, column=41, value=f'=SUMPRODUCT(($E$4:${dl}$4=$AN{r})*($E${total}:${dl}${total}))')


def sheet_addition_requirements(wb: Workbook) -> None:
    ws = wb.create_sheet("13_月次_加算要件")
    title_cell(ws, "A1", "加算要件の実績管理（夜勤職員加配・重度障害者支援体制・福祉専門職員配置）", 12)

    ws["A3"] = "■夜勤職員加配加算（●=加配配置日）"
    ws["A3"].font = Font(bold=True, color=NAVY)
    header_row(ws, 4, ["No.", "配置単位", "対象", "備考"])
    _day_header(ws, 4)
    dl = get_column_letter(DAY_COL0 + DAYS - 1)
    first = 6
    units = ["1階部", "2階部"]
    for i, u in enumerate(units):
        r = first + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=u)
        ws.cell(row=r, column=3, value="対象")
        ws.cell(row=r, column=DAY_COL0 + DAYS + 1, value=f'=COUNTIF($E{r}:${dl}{r},"●")')
    last = first + len(units) - 1
    header_row(ws, 5, ["算定日数"], start_col=DAY_COL0 + DAYS + 1)
    dv = DataValidation(type="list", formula1='"●,✕"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E{first}:{dl}{last}")
    for r in range(first, last + 1):
        for c in list(range(3, 5)) + list(range(DAY_COL0, DAY_COL0 + DAYS)):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=YELLOW)
    box_range(ws, 4, last, 1, DAY_COL0 + DAYS + 1)

    tot_row = last + 1
    ws.cell(row=tot_row, column=2, value="日別算定数").font = Font(bold=True, size=9)
    for d in range(DAYS):
        cl = get_column_letter(DAY_COL0 + d)
        ws.cell(row=tot_row, column=DAY_COL0 + d, value=f'=COUNTIF({cl}${first}:{cl}${last},"●")')
    box_range(ws, tot_row, tot_row, 2, DAY_COL0 + DAYS - 1)

    top = _week_block(ws, 4, ["日数", "算定日数", "取得率"])
    for i, w in enumerate(WEEKS):
        r = top + i
        ws.cell(row=r, column=41, value=f'=COUNTIF($E$5:${dl}$5,$AN{r})')
        ws.cell(row=r, column=42, value=f'=SUMPRODUCT(($E$5:${dl}$5=$AN{r})*$E${tot_row}:${dl}${tot_row})')
        ws.cell(row=r, column=43, value=f'=IFERROR($AP{r}/($AO{r}*COUNTA($B${first}:$B${last})),"")')
        ws.cell(row=r, column=43).number_format = "0.0%"

    r0 = tot_row + 3
    ws.cell(row=r0, column=1, value="■重度障害者支援体制加算（対象者ごとの計画・記録）").font = Font(bold=True, color=NAVY)
    header_row(ws, r0 + 1, ["No.", "利用者ID", "利用者名（任意）", "計画作成", "記録作成", "確認日", "確認者", "備考"])
    for i in range(10):
        rr = r0 + 2 + i
        ws.cell(row=rr, column=1, value=i + 1)
        for c in range(2, 9):
            ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row=rr, column=6).number_format = "yyyy/mm/dd"
    dv2 = DataValidation(type="list", formula1='"〇,✕,対象外"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"D{r0 + 2}:E{r0 + 11}")
    box_range(ws, r0 + 1, r0 + 11, 1, 8)
    ws.cell(row=r0 + 13, column=1, value="判定：計画・記録がすべて〇なら算定可").font = Font(size=9, color="808080")
    ws.cell(row=r0 + 13, column=4, value=f'=IF(COUNTIF($D{r0 + 2}:$E{r0 + 11},"✕")>0,"要対応","問題なし")')

    r1 = r0 + 15
    ws.cell(row=r1, column=1, value="■福祉専門職員配置等加算").font = Font(bold=True, color=NAVY)
    header_row(ws, r1 + 1, ["No.", "職員名", "常勤/非常勤", "社福", "介福", "精保", "公認心理師", "勤続3年以上"])
    for i in range(20):
        rr = r1 + 2 + i
        ws.cell(row=rr, column=1, value=i + 1)
        for c in range(2, 9):
            ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=YELLOW)
    box_range(ws, r1 + 1, r1 + 21, 1, 8)
    dv3 = DataValidation(type="list", formula1='"常勤,非常勤"', allow_blank=True)
    dv4 = DataValidation(type="list", formula1='"〇"', allow_blank=True)
    ws.add_data_validation(dv3)
    ws.add_data_validation(dv4)
    dv3.add(f"C{r1 + 2}:C{r1 + 21}")
    dv4.add(f"D{r1 + 2}:H{r1 + 21}")
    rj = r1 + 23
    ws.cell(row=rj, column=1, value="常勤数").font = Font(bold=True)
    ws.cell(row=rj, column=2, value=f'=COUNTIF($C{r1 + 2}:$C{r1 + 21},"常勤")')
    ws.cell(row=rj + 1, column=1, value="常勤比率").font = Font(bold=True)
    ws.cell(row=rj + 1, column=2, value=f'=IFERROR($B{rj}/COUNTA($B{r1 + 2}:$B{r1 + 21}),"")')
    ws.cell(row=rj + 1, column=2).number_format = "0.0%"
    # 有資格判定は作業列 I（1人1行）で持ち、比率はその合計で出す
    ws.cell(row=r1 + 1, column=9, value="有資格").font = Font(bold=True, size=9)
    for i in range(20):
        rr = r1 + 2 + i
        ws.cell(row=rr, column=9, value=f'=IF($B{rr}="","",IF(COUNTIF($D{rr}:$G{rr},"〇")>0,1,0))')
    ws.cell(row=rj + 2, column=1, value="有資格者比率（Ⅰ要件35%）").font = Font(bold=True)
    ws.cell(
        row=rj + 2, column=2,
        value=f'=IFERROR(SUM($I{r1 + 2}:$I{r1 + 21})/COUNTA($B{r1 + 2}:$B{r1 + 21}),"")',
    )
    ws.cell(row=rj + 2, column=2).number_format = "0.0%"
    set_widths(ws, {"A": 5, "B": 18, "C": 12, "D": 10, "E": 8, "F": 8, "G": 12, "H": 12})


def sheet_pipeline(wb: Workbook) -> int:
    ws = wb.create_sheet("14_案件管理台帳")
    headers = [
        "案件ID", "対象施設", "kintone名/イニシャル", "希望種別", "区分", "相談元",
        "発掘日", "見学日", "体験日", "実調日", "契約日", "入居/利用開始日",
        "ステータス", "次回追客日", "担当者", "停滞日数", "メモ",
    ]
    header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 32
    rows = 60
    for i in range(2, rows + 2):
        ws.cell(row=i, column=1, value=f'=IF($C{i}="","","C"&TEXT(ROW()-1,"000"))')
        ws.cell(row=i, column=2, value=f'=IF($C{i}="","",設定対象施設)')
        for c in range(7, 13):
            ws.cell(row=i, column=c).number_format = "yyyy/mm/dd"
        ws.cell(row=i, column=14).number_format = "yyyy/mm/dd"
        ws.cell(
            row=i, column=16,
            value=f'=IF($C{i}="","",IF($M{i}="入居","—",TODAY()-MAX($G{i}:$L{i})))',
        )
    last = rows + 1
    set_widths(
        ws,
        {"A": 9, "B": 24, "C": 20, "D": 10, "E": 6, "F": 14, "G": 11, "H": 11, "I": 11,
         "J": 11, "K": 11, "L": 14, "M": 12, "N": 12, "O": 16, "P": 10, "Q": 28},
    )
    box_range(ws, 2, last, 1, 17)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:Q{last}"
    for col in list("CDEF") + list("GHIJKL") + ["M", "N", "O", "Q"]:
        for r in range(2, last + 1):
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=YELLOW)
    dvs = [
        (DataValidation(type="list", formula1='"SS,入居,将来参考,入居待機"', allow_blank=True), f"D2:D{last}"),
        (DataValidation(type="list", formula1='"相談支援,基幹相談,病院,行政,家族,本人,他事業所"', allow_blank=True), f"F2:F{last}"),
        (DataValidation(type="list", formula1='"発掘,追客,見学,体験,実調,契約,入居,見送り"', allow_blank=True), f"M2:M{last}"),
        (DataValidation(type="list", formula1="担当者リスト", allow_blank=True), f"O2:O{last}"),
    ]
    for dv, ref in dvs:
        ws.add_data_validation(dv)
        dv.add(ref)
    ws.conditional_formatting.add(
        f"P2:P{last}",
        CellIsRule(operator="greaterThan", formula=["14"], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.cell(row=last + 2, column=1, value="※ 停滞日数が14日を超えた案件は赤表示。週次で追客日を再設定してください。").font = Font(
        size=9, color="808080"
    )
    return last


# ------------------------------------------------------- 累積ログ／年間推移

ARCHIVE_LAST = 2000  # 15_KPI累積ログの想定最終行（約4年分）

ARCHIVE_HEADERS = [
    "記録日", "対象施設", "年", "月", "週", "カテゴリ", "指標名", "単位",
    "週次目標", "実績", "達成率", "差異", "担当者", "未達要因（Check）",
    "次週アクション（Act）", "締め処理日",
]


def sheet_archive(wb: Workbook) -> None:
    ws = wb.create_sheet("15_KPI累積ログ")
    header_row(ws, 1, ARCHIVE_HEADERS)
    ws.row_dimensions[1].height = 32
    set_widths(
        ws,
        {"A": 12, "B": 26, "C": 7, "D": 7, "E": 9, "F": 14, "G": 20, "H": 7,
         "I": 11, "J": 11, "K": 10, "L": 9, "M": 16, "N": 30, "O": 30, "P": 12},
    )
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:P{ARCHIVE_LAST}"
    ws.cell(row=3, column=18, value=(
        "このシートは月締め処理（tools/pdca/close_month.py）で自動追記される保管庫です。"
    )).font = Font(bold=True, color=NAVY)
    ws.cell(row=4, column=18, value=(
        "・数式ではなく『値』が入ります。対象月を切り替えても過去の実績が消えません。"
    )).font = Font(size=9, color="808080")
    ws.cell(row=5, column=18, value=(
        "・手で編集しないでください。修正が必要な場合は該当行を直し、締め日を入れ直します。"
    )).font = Font(size=9, color="808080")
    ws.cell(row=6, column=18, value=(
        "・40施設分をこのシートに縦に積めば、そのままピボットで全社集計できます。"
    )).font = Font(size=9, color="808080")


def sheet_yearly_trend(wb: Workbook) -> None:
    ws = wb.create_sheet("16_年間推移")
    title_cell(ws, "B2", "年間推移（15_KPI累積ログの集計）", 14)
    ws["B3"] = '=設定対象年&"年度／"&設定対象施設'
    ws["B3"].font = Font(bold=True, color=BLUE, size=12)
    ws["B4"] = "※ 月締めが済んだ月だけ数値が入ります。率は週平均、件数は月合計です。"
    ws["B4"].font = Font(size=9, color="808080")

    months = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    header_row(ws, 6, ["指標名", "単位"] + [f"{m}月" for m in months] + ["年度計/平均"], start_col=2)
    rng_val = f"'15_KPI累積ログ'!$J$2:$J${ARCHIVE_LAST}"
    rng_kpi = f"'15_KPI累積ログ'!$G$2:$G${ARCHIVE_LAST}"
    rng_year = f"'15_KPI累積ログ'!$C$2:$C${ARCHIVE_LAST}"
    rng_month = f"'15_KPI累積ログ'!$D$2:$D${ARCHIVE_LAST}"

    r = 7
    for cat, kpi, unit in STANDARD_KPIS:
        ws.cell(row=r, column=2, value=kpi)
        ws.cell(row=r, column=3, value=unit)
        agg = "AVERAGEIFS" if unit == "%" else "SUMIFS"
        for i, m in enumerate(months):
            col = 4 + i
            year = "設定対象年" if m >= 4 else "設定対象年+1"
            ws.cell(
                row=r, column=col,
                value=(
                    f'=IFERROR({agg}({rng_val},{rng_kpi},$B{r},'
                    f'{rng_year},{year},{rng_month},{m}),"")'
                ),
            )
            if unit == "%":
                ws.cell(row=r, column=col).number_format = "0.0%"
        first_c, last_c = get_column_letter(4), get_column_letter(15)
        fn = "AVERAGE" if unit == "%" else "SUM"
        ws.cell(row=r, column=16, value=f'=IFERROR({fn}(${first_c}{r}:${last_c}{r}),"")')
        if unit == "%":
            ws.cell(row=r, column=16).number_format = "0.0%"
        r += 1
    last = r - 1
    box_range(ws, 6, last, 2, 16)
    set_widths(ws, {"A": 3, "B": 20, "C": 6, "P": 13})
    for i in range(12):
        ws.column_dimensions[get_column_letter(4 + i)].width = 9
    ws.freeze_panes = "D7"


def build(facility_master: Path, out: Path) -> None:
    facilities = read_facilities(facility_master)
    wb = Workbook()
    wb.remove(wb.active)

    sheet_readme(wb)
    fac_last = sheet_facility_master(wb, facilities)
    owner_last = sheet_owner_master(wb, facilities, fac_last)
    sheet_settings(wb, fac_last)
    issue_last = sheet_issues(wb, fac_last)
    kpi_last = sheet_kpi_log(wb)
    sheet_gh_occupancy(wb)
    sheet_ss_occupancy(wb)
    sheet_sales_actions(wb)
    sheet_addition_requirements(wb)
    sheet_pipeline(wb)
    sheet_archive(wb)
    sheet_yearly_trend(wb)
    sheet_dashboard(wb, kpi_last, issue_last, owner_last)
    sheet_proposals(wb)
    sheet_rules(wb)
    sheet_lists(wb)
    add_defined_names(wb, fac_last, owner_last)
    wb._sheets.sort(key=lambda w: w.title)

    wb["03_設定"]["B5"] = facilities[0]["施設名"]
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"生成しました: {out}（施設 {len(facilities)} 件 / 担当AM {owner_last - 1} 名）")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--facility-master", type=Path, default=Path("tools/pdca/facility_master_source.xlsx"))
    p.add_argument("--out", type=Path, default=Path("dist/週間PDCA_改善版テンプレート.xlsx"))
    a = p.parse_args()
    build(a.facility_master, a.out)


if __name__ == "__main__":
    main()
