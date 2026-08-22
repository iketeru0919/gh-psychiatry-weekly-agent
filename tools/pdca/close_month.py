"""月締め処理：当月の週次KPIを累積ログへ値で追記し、翌月用ブックを作る。

毎月シートを増やす代わりに、
  ①当月ブックをそのまま実績アーカイブとして保存し、
  ②KPIの週次実績だけを 15_KPI累積ログ に値で追記し、
  ③日別の入力欄を空にして対象月を1つ進めた翌月用ブックを出力する
という流れで継続運用します。

前提:
    Excelで開いて保存されたブックであること（数式のキャッシュ値を読むため）。
    Excelを一度も通していないブックでは実績が空のまま追記されます。

使い方:
    python tools/pdca/close_month.py 週間PDCA_2026年08月.xlsx
        -> 週間PDCA_2026年08月_実績.xlsx（当月の保管用・編集しない）
        -> 週間PDCA_2026年09月.xlsx      （翌月用・こちらを使う）
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

KPI_SHEET = "05_週次KPIログ"
ARCHIVE_SHEET = "15_KPI累積ログ"
SETTINGS_SHEET = "03_設定"

# 締めのたびに空にする日別入力レンジ（利用者名や職員名は翌月へ引き継ぐ）
CLEAR_RANGES: dict[str, list[tuple[int, int, int, int]]] = {
    # sheet: [(min_row, max_row, min_col, max_col), ...]
    "10_月次_GH稼働": [(5, 34, 5, 35)],
    "11_月次_SS稼働": [(5, 19, 5, 35)],
    "12_月次_営業行動": [(5, 9, 5, 35)],
    "13_月次_加算要件": [(6, 7, 5, 35), (13, 22, 4, 8)],
}


def read_month_rows(values: Worksheet, facility: str, year, month) -> list[list]:
    """05_週次KPIログから、実績または目標が入っている行だけを取り出す。"""
    rows: list[list] = []
    for r in values.iter_rows(min_row=2, max_row=values.max_row, values_only=True):
        if not r or not r[6]:  # G列: 指標名
            continue
        target, actual = r[8], r[9]
        if target in (None, "") and actual in (None, ""):
            continue
        row = list(r[:15]) + [dt.date.today()]
        row[1] = facility
        row[2] = year
        row[3] = month
        rows.append(row)
    return rows


def first_empty_row(ws: Worksheet) -> int:
    r = 2
    while ws.cell(row=r, column=7).value not in (None, ""):
        r += 1
    return r


def next_month(year: int, month: int) -> tuple[int, int]:
    return (year + 1, 1) if month == 12 else (year, month + 1)


def close(path: Path, out_dir: Path | None = None) -> tuple[Path, Path]:
    out_dir = out_dir or path.parent
    wb_val = load_workbook(path, data_only=True)
    wb = load_workbook(path)

    st = wb_val[SETTINGS_SHEET]
    facility = st["B5"].value
    year = st["B7"].value
    month = st["B9"].value
    if not (facility and year and month):
        raise SystemExit("03_設定 の 対象施設／対象年／対象月 が空です。先に設定してください。")

    rows = read_month_rows(wb_val[KPI_SHEET], facility, year, month)
    if not rows:
        print("警告: 週次KPIログに実績・目標がありません。Excelで保存済みのブックか確認してください。")

    archive = wb[ARCHIVE_SHEET]
    start = first_empty_row(archive)
    # 同一施設・同一年月が既に入っていれば二重登録
    for r in archive.iter_rows(min_row=2, max_row=start - 1, values_only=True):
        if r[1] == facility and r[2] == year and r[3] == month:
            raise SystemExit(
                f"{facility} の {year}年{month}月 は既に 15_KPI累積ログ に登録済みです。"
                "再締めする場合は該当行を削除してから実行してください。"
            )
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            archive.cell(row=start + i, column=j, value=v)
        archive.cell(row=start + i, column=16).number_format = "yyyy/mm/dd"
        archive.cell(row=start + i, column=1).number_format = "yyyy/mm/dd"

    # ① 当月ブックを実績として保管（数式のまま、入力も残す）
    archived = out_dir / f"{path.stem}_実績{path.suffix}"
    shutil.copy2(path, archived)

    # ② 翌月用：日別入力を消し、対象月を進める
    for sheet, ranges in CLEAR_RANGES.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r0, r1, c0, c1 in ranges:
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    ws.cell(row=r, column=c).value = None

    ny, nm = next_month(int(year), int(month))
    wb[SETTINGS_SHEET]["B7"] = ny
    wb[SETTINGS_SHEET]["B9"] = nm
    wb[SETTINGS_SHEET]["B11"] = dt.date(ny, nm, 1)

    nxt = out_dir / f"週間PDCA_{facility}_{ny}年{nm:02d}月.xlsx"
    wb.save(nxt)
    print(f"累積ログへ {len(rows)} 行を追記しました（{year}年{month}月）")
    print(f"当月の保管用: {archived}")
    print(f"翌月用ブック: {nxt}")
    return archived, nxt


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path, help="締める対象のブック")
    ap.add_argument("--out-dir", type=Path, default=None)
    a = ap.parse_args()
    close(a.workbook, a.out_dir)


if __name__ == "__main__":
    main()
