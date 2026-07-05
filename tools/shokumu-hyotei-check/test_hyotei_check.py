# -*- coding: utf-8 -*-
"""実ファイルで観測したレイアウトを再現してロジックを検証するテスト。"""
import datetime as dt
import shutil
from pathlib import Path

import openpyxl
from hyotei_check import (check_folder, parse_month_token, month_from_sheet_title,
                          write_output)

TMP = Path(__file__).parent / "testdata"


def build_sheet(ws, eval_date, people):
    """観測した様式を再現: B2=評価月ラベル, B3=日付, 6行目左ヘッダー,
    7-8行目は例(8行目のH/I/Jにラベル), 9行目からデータ, 残骸FALSE行。
    people: (氏名, 職種, 雇用形態, 入社日, H氏名, 順位, 評価)"""
    ws["B1"] = "ラシエル 従業員勤務評定表"
    ws["B2"] = "評価月"; ws["C2"] = "事業所名"; ws["E2"] = "評価者"
    ws["B3"] = eval_date; ws["C3"] = "ラシエルテスト"
    ws["B6"] = "氏　　名"; ws["C6"] = "職　　種"; ws["D6"] = "総　　評"
    ws["E6"] = "雇用形態"; ws["F6"] = "入社日"; ws["H6"] = "↓↓入力変更禁止↓↓"
    ws["B7"] = "○○　○○（例）"; ws["C7"] = "生活支援員"; ws["E7"] = "契社"
    ws["H7"] = "評定計算式"
    ws["B8"] = "○○　○○（例）"; ws["C8"] = "世話人"; ws["E8"] = "派遣"
    ws["H8"] = "氏名"; ws["I8"] = "順位"; ws["J8"] = "評価"
    r = 9
    for p in people:
        ws.cell(r, 2, p[0]); ws.cell(r, 3, p[1]); ws.cell(r, 4, "総評テキスト")
        ws.cell(r, 5, p[2]); ws.cell(r, 6, p[3])
        ws.cell(r, 8, p[4]); ws.cell(r, 9, p[5]); ws.cell(r, 10, p[6])
        r += 1
    for _ in range(5):  # 数式残骸
        ws.cell(r, 10, "FALSE")
        r += 1


def make_files():
    if TMP.exists():
        shutil.rmtree(TMP)
    TMP.mkdir()
    d = dt.date
    # 浄水型: 単独シート、シート名 "2602" だが評価月セルは 2026/6 → セルを優先すべき
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "2602"
    build_sheet(ws, d(2026, 6, 20), [
        ("瀬口　公子", "管理者", "契社", d(2026, 6, 1), "瀬口　公子", 1, "A"),
        ("井上美加", "生活支援員", "時給", d(2025, 12, 1), "井上美加", 2, "B"),
    ])
    wb.save(TMP / "【浄水】職務評定表（植松）2606.xlsx")

    # 名取型: シート名 "Sheet1"、順位の重複(4)と欠番(3)、空欄・値異常・氏名なし行あり
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sheet1"
    build_sheet(ws, d(2026, 6, 22), [
        ("浜塚　淳", "生活支援員", "正社", d(2022, 6, 1), "浜塚　淳", 1, "A"),
        ("今野　千佳", "管理者候補", "時給", d(2022, 5, 8), "今野　千佳", 2, "A"),
        ("曽我　ゆみ", "世話人", "", d(2022, 3, 1), "曽我　ゆみ", 4, "B"),   # 雇用形態空欄
        ("瀬野　美智代", "世話人", "時給", "不明", "瀬野　美智代", 4, "B"),  # 入社日異常+順位重複
        ("南舘　恵理子", "世話人", "時給", d(2022, 10, 25), "南舘　恵理子", 5, "G"),  # 評価異常
        ("追加　太郎", "世話人", "時給", d(2023, 1, 1), "", "", ""),  # H未反映
        ("", "", "", "", "", 6, "E"),  # 氏名なし行(レコード化せず警告1件のみ)
    ])
    wb.save(TMP / "【名取】職務評定表（吉川）2026・6.xlsx")

    # 石巻型: 複数月シート。同一人物の評価推移(A→B)を含む
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for (y, m, g2) in [(2026, 4, "A"), (2026, 6, "B")]:
        ws = wb.create_sheet(f"{y}.{m}")
        build_sheet(ws, d(y, m, 15), [
            ("菅原　一郎", "施設長", "正社", d(2020, 4, 1), "菅原　一郎", 1, "A"),
            ("鈴木　二郎", "世話人", "時給", d(2021, 5, 1), "鈴木　二郎", 2, g2),
        ])
    wb.save(TMP / "【石巻】勤務評定表（菅原）2026.4.xlsx")

    # 上溝型: 月別シートの連番だが、2026.4 シートの評価月セルが 2026/2 のまま
    # → シート名を優先して 2026-04 として扱い、食い違いを警告するべき
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for title, cell_date in [("上溝2026.2", d(2026, 2, 1)),
                             ("上溝2026.4", d(2026, 2, 1)),   # 更新忘れ
                             ("上溝2026.6", d(2026, 6, 1))]:
        ws = wb.create_sheet(title)
        build_sheet(ws, cell_date, [
            ("田中　泰枝", "サビ管", "契社", d(2026, 2, 1), "田中　泰枝", 1, "A"),
        ])
    wb.save(TMP / "【上溝】勤務評定表（保科）2026.6.xlsx")

    # 下手野型: 6月ファイルのはずがシートは4月のみ → 対象月シート不在になるはず
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "202604"
    build_sheet(ws, d(2026, 4, 10), [
        ("米田　花子", "世話人", "時給", d(2024, 4, 1), "米田　花子", 1, "B"),
    ])
    wb.save(TMP / "【下手野】職務評定表（米田）202606.xlsx")

    # 宮下型: シート名 "宮下6月"、評価月セルは日付あり
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "宮下6月"
    build_sheet(ws, d(2026, 6, 1), [
        ("堀切　三郎", "世話人", "時給", d(2023, 6, 1), "堀切　三郎", 1, "C"),
    ])
    wb.save(TMP / "【宮下】職務評定（堀切）202606.xlsx")

    # 一時ファイルはスキップされること
    (TMP / "~$【浄水】職務評定表（植松）2606.xlsx").write_bytes(b"junk")


def main():
    make_files()

    # 月指定パーサ
    assert parse_month_token("2026-06") == "2026-06"
    assert parse_month_token("2026.6") == "2026-06"
    assert parse_month_token("202606") == "2026-06"
    assert parse_month_token("2026/6") == "2026-06"
    assert parse_month_token("2026年6月") == "2026-06"

    # シート名パーサ
    assert month_from_sheet_title("2025.4") == (2025, 4)
    assert month_from_sheet_title("R8.6") == (2026, 6)
    assert month_from_sheet_title("202604") == (2026, 4)
    assert month_from_sheet_title("2602") == (2026, 2)
    assert month_from_sheet_title("上溝2026.4") == (2026, 4)
    assert month_from_sheet_title("宮下6月") == (None, 6)
    assert month_from_sheet_title("Sheet1") == (None, None)

    months = ["2026-06"]
    paths = list(TMP.glob("*.xlsx"))
    records, issues = check_folder(paths, months)

    facs = {r.facility for r in records}
    assert facs == {"浄水", "名取", "石巻", "宮下", "上溝"}, facs

    # 浄水: 単独シート+食い違い → 評価月セル(2026-06)を採用
    josui = [r for r in records if r.facility == "浄水"]
    assert len(josui) == 2 and all(r.month == "2026-06" for r in josui)

    # 上溝: 連番シート → シート名を優先。6月シートの1名だけが対象
    kamimizo = [r for r in records if r.facility == "上溝"]
    assert len(kamimizo) == 1 and kamimizo[0].sheet_name == "上溝2026.6"

    # 石巻: 2026-06 のシートだけ読まれる
    ishi = [r for r in records if r.facility == "石巻"]
    assert len(ishi) == 2 and all(r.sheet_name == "2026.6" for r in ishi)

    # 氏名なし行はレコード化されない
    assert not any(r.facility == "名取" and not r.name for r in records)

    kinds = {}
    for i in issues:
        kinds.setdefault(i.kind, []).append(i)

    assert any("雇用形態" in i.detail for i in kinds.get("空欄", []))
    assert any("順位 [4]" in i.detail for i in kinds.get("順位の重複", []))
    assert any("[3" in i.detail for i in kinds.get("順位の欠番", []))
    assert any("評価 'G'" in i.detail for i in kinds.get("値の異常", []))
    assert any("入社日 '不明'" in i.detail for i in kinds.get("値の異常", []))
    assert any("追加　太郎" in i.detail for i in kinds.get("不整合", []))
    assert any(i.facility == "下手野" for i in kinds.get("対象月シート不在", []))
    assert all(i.facility == "下手野" for i in kinds.get("対象月シート不在", []))
    assert any(i.facility == "浄水" and "2026-06 を採用" in i.detail
               for i in kinds.get("月表記の食い違い", [])), kinds.get("月表記の食い違い")
    assert any(i.facility == "名取" for i in kinds.get("氏名なし", []))
    # 氏名なし行から空欄エラーが量産されていないこと
    assert not any("不明" in i.detail for i in kinds.get("空欄", []))

    # --- 複数月指定 + マトリクス出力 ---
    months3 = ["2026-02", "2026-04", "2026-06"]
    records3, issues3 = check_folder(paths, months3)

    # 上溝: 食い違いシートは 2026-04 として採用され、2月と重複しない
    kami3 = [r for r in records3 if r.facility == "上溝"]
    assert {r.month for r in kami3} == {"2026-02", "2026-04", "2026-06"}, {r.month for r in kami3}
    assert any(i.kind == "月表記の食い違い" and i.facility == "上溝" and "2026-04 を採用" in i.detail
               for i in issues3)
    # 同一人物×同一月の重複なし → write_output 後も「重複」エラーが出ない
    out = TMP / "output.xlsx"
    write_output(records3, issues3, months3, out)
    assert not any(i.kind == "重複" for i in issues3)

    wb = openpyxl.load_workbook(out)
    expect = {"エラーサマリー", "全職員一覧", "エラー詳細",
              "浄水", "名取", "石巻", "宮下", "上溝", "下手野"}
    assert set(wb.sheetnames) == expect, wb.sheetnames
    assert wb.sheetnames[0] == "エラーサマリー"
    assert wb.sheetnames[-1] == "エラー詳細"

    # 石巻マトリクス: 行=職員、月が横に並び、鈴木二郎は A→B で推移が「↓」
    ws = wb["石巻"]
    assert ws.cell(1, 5).value == "2026-02" and ws.cell(1, 7).value == "2026-04" \
        and ws.cell(1, 9).value == "2026-06", [ws.cell(1, c).value for c in range(1, 12)]
    names = {ws.cell(r, 1).value: r for r in (3, 4)}
    r = names["鈴木　二郎"]
    assert ws.cell(r, 5).value == "－"          # 2月はデータなし
    assert ws.cell(r, 7).value == "A" and ws.cell(r, 9).value == "B"
    assert ws.cell(r, 10).value == 2            # 6月の順位
    assert ws.cell(r, 11).value == "↓"

    # 全職員一覧: 施設列つき
    wa = wb["全職員一覧"]
    assert wa.cell(2, 1).value is None or wa.cell(1, 1).value == "施設"
    assert wa.cell(1, 1).value == "施設" and wa.cell(1, 2).value == "氏名"
    fac_col = [wa.cell(r, 1).value for r in range(3, wa.max_row + 1)]
    assert "石巻" in fac_col and "上溝" in fac_col

    # エラーサマリー: 種別ごとの件数表になっている
    es = wb["エラーサマリー"]
    header = [es.cell(1, c).value for c in range(1, es.max_column + 1)]
    assert header[0] == "施設" and header[-1] == "合計"
    assert "空欄" in header

    print("全テスト成功:", len(records3), "件のレコード,", len(issues3), "件のエラー検出")
    for i in issues3:
        print(f"  [{i.kind}] {i.facility} {i.month} 行{i.row}: {i.detail}")


if __name__ == "__main__":
    main()
