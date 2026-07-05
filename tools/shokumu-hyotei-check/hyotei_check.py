# -*- coding: utf-8 -*-
"""職務評定表チェック・集約ロジック

Google Drive の「職務評定」フォルダ内の Excel(施設別・月別シート)から
氏名/順位/評価/雇用形態/入社日/職種 を抽出し、施設別に1つの Excel に
まとめる。あわせて入力エラーを検出してエラー一覧を作る。
"""

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
VALID_GRADES = {"S", "A", "B", "C", "D", "E", "F"}

# 列の位置(1始まり): C=職種, E=雇用形態, F=入社日, H=氏名, I=順位, J=評価
COL_SHOKUSHU = 3
COL_KOYO = 5
COL_NYUSHA = 6
COL_NAME_LEFT = 2   # B列の氏名(手入力側)
COL_NAME = 8
COL_RANK = 9
COL_GRADE = 10


@dataclass
class Record:
    facility: str
    file_name: str
    sheet_name: str
    month: str          # "YYYY-MM"
    row: int
    name: str = ""
    rank: object = None
    grade: object = None
    koyo: str = ""
    nyusha: object = None
    shokushu: str = ""


@dataclass
class Issue:
    facility: str
    file_name: str
    sheet_name: str
    month: str
    row: object         # 行番号 or "-"
    kind: str
    detail: str


def normalize(s):
    """判定・比較用の正規化(全角→半角等)。表示用の値には使わない。"""
    if s is None:
        return ""
    return unicodedata.normalize("NFKC", str(s)).strip()


def clean(s):
    """表示用: 前後の空白だけ除去し、元の表記(全角スペース等)は保つ。"""
    if s is None:
        return ""
    return str(s).strip()


def facility_from_filename(name):
    m = re.match(r"[【\[](.+?)[】\]]", name)
    return m.group(1) if m else Path(name).stem


def parse_month_token(token):
    """ユーザー指定の月 ("2026-06", "2026.6", "202606", "2026/6") → "YYYY-MM"."""
    t = normalize(token).replace("年", "-").replace("月", "")
    m = re.fullmatch(r"(\d{4})[.\-/](\d{1,2})", t)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    m = re.fullmatch(r"(\d{4})(\d{2})", t)
    if m and 1 <= int(m.group(2)) <= 12:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    raise ValueError(f"月の指定 '{token}' を解釈できません。例: 2026-06 / 2026.6 / 202606")


def month_from_sheet_title(title):
    """シート名から (year, month) を推定。判らない要素は None。"""
    t = normalize(title)
    m = re.search(r"[RrRr]\s*(\d{1,2})[.\-年]?\s*(\d{1,2})", t)  # R8.6 (令和)
    if m:
        return 2018 + int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d{4})\s*[.．\-/年]\s*(\d{1,2})", t)  # 2026.6 / 2026年6月
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
    m = re.fullmatch(r"(\d{4})(\d{2})", t)  # 202606
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
    m = re.fullmatch(r"(\d{2})(\d{2})", t)  # 2606 → 2026-06
    if m and 1 <= int(m.group(2)) <= 12:
        return 2000 + int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d{1,2})\s*月", t)  # 宮下6月 → 月のみ
    if m and 1 <= int(m.group(1)) <= 12:
        return None, int(m.group(1))
    return None, None


def month_from_hyokazuki_cell(ws):
    """シート内の「評価月」ラベルの下のセルから (year, month) を取る。"""
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=12):
        for cell in row:
            if normalize(cell.value) == "評価月":
                below = ws.cell(row=cell.row + 1, column=cell.column)
                v = below.value
                if isinstance(v, (datetime, date)):
                    return v.year, v.month
                m = re.search(r"(\d{4})\s*[./\-年]\s*(\d{1,2})", normalize(v))
                if m and 1 <= int(m.group(2)) <= 12:
                    return int(m.group(1)), int(m.group(2))
                return None, None
    return None, None


def _ym(y, m):
    return f"{y:04d}-{m:02d}" if (y and m) else None


def sheet_month_info(ws):
    """シートの月の手がかりを返す: (シート名からの月, シート名の月のみ, 評価月セルからの月)"""
    ty, tm = month_from_sheet_title(ws.title)
    cy, cm = month_from_hyokazuki_cell(ws)
    return _ym(ty, tm), tm, _ym(cy, cm)


def decide_sheet_month(title_month, title_m_only, cell_month, is_series):
    """シートの対象月を決める。
    シート名と評価月セルが食い違う場合:
      - 月別シートが並ぶファイル(is_series)ではシート名を採用(セルの更新忘れが多い)
      - 単独シートのファイルでは評価月セルを採用(シート名がテンプレの古い名前のことが多い)
    返り値: (month, conflict_detail or None)"""
    if title_month and cell_month:
        if title_month == cell_month:
            return title_month, None
        chosen = title_month if is_series else cell_month
        other = cell_month if is_series else title_month
        src = "シート名" if is_series else "評価月セル"
        return chosen, (f"シート名からは {title_month}、評価月セルからは {cell_month} と読め、"
                        f"食い違っています。{src}の {chosen} を採用しました({other} ではない点に注意)")
    if title_month:
        return title_month, None
    if cell_month:
        detail = None
        if title_m_only and int(cell_month[-2:]) != title_m_only:
            detail = (f"シート名の月({title_m_only}月)と評価月セル({cell_month})が食い違っています。"
                      f"評価月セルの {cell_month} を採用しました")
        return cell_month, detail
    return None, None


def find_data_start(ws):
    """H/I/J のラベル行(氏名・順位・評価)を探し、その次の行番号を返す。"""
    for row in ws.iter_rows(min_row=1, max_row=15, min_col=COL_NAME, max_col=COL_GRADE):
        vals = [normalize(c.value) for c in row]
        if vals[0] == "氏名" and vals[COL_RANK - COL_NAME] == "順位":
            return row[0].row + 1
    return None


def is_formula_error(v):
    return normalize(v) in FORMULA_ERRORS


def parse_sheet(ws, facility, file_name, month_disp):
    """1シートを読み、(records, issues) を返す。"""
    records, issues = [], []
    start = find_data_start(ws)
    if start is None:
        issues.append(Issue(facility, file_name, ws.title, month_disp, "-",
                            "様式不一致", "氏名・順位・評価のヘッダー行が見つかりません"))
        return records, issues

    for r in range(start, ws.max_row + 1):
        name_l = clean(ws.cell(r, COL_NAME_LEFT).value)
        name = clean(ws.cell(r, COL_NAME).value)
        rank = ws.cell(r, COL_RANK).value
        grade = ws.cell(r, COL_GRADE).value
        koyo = clean(ws.cell(r, COL_KOYO).value)
        nyusha = ws.cell(r, COL_NYUSHA).value
        shokushu = clean(ws.cell(r, COL_SHOKUSHU).value)

        if "（例）" in name_l or "(例)" in name_l:
            continue
        grade_n = normalize(grade)
        # 氏名のない行はデータとして扱わない(数式残骸の FALSE/0 等)。
        # ただし評価や順位らしき値が入っていれば1件だけ警告する。
        if not name_l and not name:
            has_rank = isinstance(rank, (int, float)) or normalize(rank).isdigit()
            if grade_n in VALID_GRADES or has_rank:
                issues.append(Issue(facility, file_name, ws.title, month_disp, r, "氏名なし",
                                    f"氏名が空欄の行に順位『{clean(rank)}』/評価『{grade_n}』が入っています"))
            continue

        rec = Record(facility, file_name, ws.title, month_disp, r,
                     name=name or name_l, rank=rank, grade=grade_n,
                     koyo=koyo, nyusha=nyusha, shokushu=shokushu)
        records.append(rec)

        # --- エラーチェック(行単位) ---
        def add(kind, detail):
            issues.append(Issue(facility, file_name, ws.title, month_disp, r, kind, detail))

        for label, v in (("氏名(H列)", name), ("順位(I列)", rank), ("評価(J列)", grade),
                         ("雇用形態(E列)", koyo), ("入社日(F列)", nyusha), ("職種(C列)", shokushu)):
            if is_formula_error(v):
                add("数式エラー", f"{label} が {normalize(v)} になっています")
            elif normalize(v) == "":
                add("空欄", f"{label} が空欄です(氏名: {rec.name or '不明'})")

        if name_l and name and name_l != name:
            add("不整合", f"B列の氏名『{name_l}』とH列の氏名『{name}』が一致しません")
        if not name and name_l:
            add("不整合", f"B列に『{name_l}』がありますがH列(評定計算式)に反映されていません")

        if rank is not None and not is_formula_error(rank):
            if not (isinstance(rank, (int, float)) and float(rank).is_integer()):
                add("値の異常", f"順位 '{rank}' が整数ではありません(氏名: {rec.name})")
        if grade_n and grade_n not in VALID_GRADES and not is_formula_error(grade):
            add("値の異常", f"評価 '{grade_n}' が想定値({'/'.join(sorted(VALID_GRADES))})ではありません(氏名: {rec.name})")
        if nyusha is not None and normalize(nyusha) != "" and not is_formula_error(nyusha):
            if not isinstance(nyusha, (datetime, date)):
                if not re.fullmatch(r"\d{4}[./\-]\d{1,2}[./\-]\d{1,2}", normalize(nyusha)):
                    add("値の異常", f"入社日 '{nyusha}' が日付として読めません(氏名: {rec.name})")

    # --- 順位の重複・欠番(シート単位) ---
    ranks = [int(r.rank) for r in records
             if isinstance(r.rank, (int, float)) and float(r.rank).is_integer()]
    if ranks:
        seen, dups = set(), set()
        for v in ranks:
            (dups if v in seen else seen).add(v)
        if dups:
            issues.append(Issue(facility, file_name, ws.title, month_disp, "-", "順位の重複",
                                f"順位 {sorted(dups)} が複数の職員に付いています"))
        missing = sorted(set(range(1, max(ranks) + 1)) - set(ranks))
        if missing:
            issues.append(Issue(facility, file_name, ws.title, month_disp, "-", "順位の欠番",
                                f"順位 {missing} が抜けています(1〜{max(ranks)} のうち)"))
    return records, issues


def process_workbook(path, target_months):
    """1ファイルを処理。target_months に一致するシートのみ読む。
    返り値: (records, issues, found_months)"""
    file_name = Path(path).name
    facility = facility_from_filename(file_name)
    wb = openpyxl.load_workbook(path, data_only=True)
    records, issues, found = [], [], set()

    infos = [(ws, *sheet_month_info(ws)) for ws in wb.worksheets]
    # シート名から異なる複数の月が読めるファイルは「月別シートの連番」とみなす
    title_months = {tmon for _, tmon, _, _ in infos if tmon}
    is_series = len(title_months) >= 2

    for ws, title_month, title_m_only, cell_month in infos:
        month, conflict = decide_sheet_month(title_month, title_m_only, cell_month, is_series)
        if month:
            found.add(month)
        if month is None:
            # 月が判定できないシート: 対象月が1つでシートも1つなら読んでみる
            if len(wb.worksheets) == 1 and len(target_months) == 1:
                recs, iss = parse_sheet(ws, facility, file_name, target_months[0])
                issues.append(Issue(facility, file_name, ws.title, target_months[0], "-",
                                    "月判定不可", "シート名・評価月セルから月を判定できないため、"
                                    f"指定月 {target_months[0]} のデータとして扱いました"))
                for r_ in recs:
                    r_.month = target_months[0]
                records += recs
                issues += iss
                found.add(target_months[0])
            continue
        if month in target_months:
            if conflict:
                issues.append(Issue(facility, file_name, ws.title, month, "-",
                                    "月表記の食い違い", conflict))
            recs, iss = parse_sheet(ws, facility, file_name, month)
            for r_ in recs:
                r_.month = month
            records += recs
            issues += iss
    wb.close()
    return records, issues, found


def check_folder(xlsx_paths, target_months):
    """複数ファイルを処理し、施設単位で対象月シートの不在も検出する。"""
    all_records, all_issues = [], []
    facility_months = {}
    for p in sorted(xlsx_paths):
        name = Path(p).name
        if name.startswith("~$"):
            continue
        facility = facility_from_filename(name)
        try:
            recs, iss, found = process_workbook(p, target_months)
        except Exception as e:  # 壊れたファイル等
            all_issues.append(Issue(facility, name, "-", "-", "-", "読込エラー", str(e)))
            continue
        all_records += recs
        all_issues += iss
        facility_months.setdefault(facility, set()).update(found)

    for facility, months in sorted(facility_months.items()):
        for m in target_months:
            if m not in months:
                all_issues.append(Issue(facility, "-", "-", m, "-", "対象月シート不在",
                                        f"{m} のシートが施設『{facility}』のどのファイルにも見つかりません"))
    return all_records, all_issues


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ERROR_FILL = PatternFill("solid", fgColor="FCE4EC")
MISSING_FILL = PatternFill("solid", fgColor="E7E6E6")
GRADE_FILLS = {
    "S": PatternFill("solid", fgColor="548235"),
    "A": PatternFill("solid", fgColor="70AD47"),
    "B": PatternFill("solid", fgColor="C6E0B4"),
    "C": PatternFill("solid", fgColor="FFE699"),
    "D": PatternFill("solid", fgColor="FFC000"),
    "E": PatternFill("solid", fgColor="F4B183"),
    "F": PatternFill("solid", fgColor="FF7C80"),
}
GRADE_ORDER = {"S": 0, "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
CENTER = Alignment(horizontal="center")


def _safe_sheet_name(name, used):
    s = re.sub(r"[\[\]:*?/\\]", "_", name)[:31] or "施設"
    base, i = s, 2
    while s in used:
        s = f"{base[:28]}_{i}"
        i += 1
    used.add(s)
    return s


def _person_key(name):
    """同一人物の判定用: 空白を除去し全角半角を揃える。"""
    return re.sub(r"\s+", "", normalize(name))


def _build_people(records):
    """施設ごとに 職員×月 のマトリクスデータを作る。
    返り値: {facility: [person, ...]}  person = dict(name, shokushu, koyo, nyusha, months={m: rec})"""
    by_fac = {}
    for r in records:
        fac = by_fac.setdefault(r.facility, {})
        p = fac.setdefault(_person_key(r.name), {"name": r.name, "months": {}})
        # 同一人物×同一月が複数あれば最初のものを残す(重複は別途エラー報告)
        p["months"].setdefault(r.month, r)
    result = {}
    for fac, people in by_fac.items():
        rows = []
        for p in people.values():
            latest = p["months"][max(p["months"])]
            p["name"] = latest.name
            p["shokushu"] = latest.shokushu
            p["koyo"] = latest.koyo
            nyusha = latest.nyusha
            p["nyusha"] = nyusha.date() if isinstance(nyusha, datetime) else nyusha
            rows.append(p)
        result[fac] = rows
    return result


def _trend(p, months):
    """データのある直近2つの月の評価を比べて ↑/↓/→ を返す。"""
    grades = [GRADE_ORDER.get(p["months"][m].grade) for m in months if m in p["months"]]
    grades = [g for g in grades if g is not None]
    if len(grades) < 2:
        return ""
    prev, last = grades[-2], grades[-1]
    if last < prev:
        return "↑"
    if last > prev:
        return "↓"
    return "→"


def _sort_people(rows, months):
    """直近月の順位→氏名 の順に並べる。直近月にいない人は後ろ。"""
    last = months[-1]
    def key(p):
        rec = p["months"].get(last)
        rank = rec.rank if rec and isinstance(rec.rank, (int, float)) else 10 ** 6
        return (rank, p["name"])
    return sorted(rows, key=key)


def _write_matrix_sheet(ws, rows, months, with_facility=None):
    """職員×月のマトリクスを書き込む。with_facility: (人→施設名) を先頭列に出す場合の dict。"""
    fixed = (["施設"] if with_facility else []) + ["氏名", "職種", "雇用形態", "入社日"]
    nfix = len(fixed)
    # 1行目: 月見出し(評価+順位の2列を結合)、2行目: 列名
    for i, h in enumerate(fixed, 1):
        c = ws.cell(1, i, h)
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
    for j, m in enumerate(months):
        col = nfix + 1 + j * 2
        ws.cell(1, col, m)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(2, col, "評価")
        ws.cell(2, col + 1, "順位")
    tcol = nfix + 1 + len(months) * 2
    ws.cell(1, tcol, "推移")
    ws.merge_cells(start_row=1, start_column=tcol, end_row=2, end_column=tcol)
    for row in ws.iter_rows(min_row=1, max_row=2, max_col=tcol):
        for c in row:
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            c.alignment = CENTER

    r = 3
    for p in rows:
        col = 1
        if with_facility:
            ws.cell(r, col, with_facility[id(p)]); col += 1
        ws.cell(r, col, p["name"])
        ws.cell(r, col + 1, p["shokushu"])
        ws.cell(r, col + 2, p["koyo"])
        c = ws.cell(r, col + 3, p["nyusha"])
        c.number_format = "yyyy/m/d"
        for j, m in enumerate(months):
            gc = ws.cell(r, nfix + 1 + j * 2)
            rc = ws.cell(r, nfix + 2 + j * 2)
            gc.alignment = CENTER
            rc.alignment = CENTER
            rec = p["months"].get(m)
            if rec is None:
                gc.value = "－"; rc.value = "－"
                gc.fill = MISSING_FILL; rc.fill = MISSING_FILL
                continue
            gc.value = rec.grade or ""
            if rec.grade in GRADE_FILLS:
                gc.fill = GRADE_FILLS[rec.grade]
            elif not rec.grade or is_formula_error(rec.grade):
                gc.fill = ERROR_FILL
            rank = rec.rank
            if isinstance(rank, float) and rank.is_integer():
                rank = int(rank)
            rc.value = rank if rank is not None else ""
            if rank is None or is_formula_error(rank):
                rc.fill = ERROR_FILL
        tc = ws.cell(r, tcol, _trend(p, months))
        tc.alignment = CENTER
        if tc.value == "↑":
            tc.font = Font(color="2E7D32", bold=True)
        elif tc.value == "↓":
            tc.font = Font(color="C00000", bold=True)
        r += 1

    widths = ([14] if with_facility else []) + [18, 14, 10, 12] + [7, 7] * len(months) + [7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(3, nfix + 1).coordinate


def write_output(records, issues, target_months, out_path):
    months = sorted(target_months)
    wb = openpyxl.Workbook()
    people_by_fac = _build_people(records)

    # 同一人物×同一月の重複を検出してエラーに追加
    dup_seen = {}
    for rec in records:
        k = (rec.facility, _person_key(rec.name), rec.month)
        if k in dup_seen:
            issues.append(Issue(rec.facility, rec.file_name, rec.sheet_name, rec.month, rec.row,
                                "重複", f"『{rec.name}』の {rec.month} の評価が複数の行/シートにあります"
                                f"(先の行: {dup_seen[k]})"))
        else:
            dup_seen[k] = f"{rec.sheet_name} 行{rec.row}"

    # --- エラーサマリー(先頭シート) ---
    ws = wb.active
    ws.title = "エラーサマリー"
    kinds = sorted({i.kind for i in issues})
    counts = {}
    for i in issues:
        counts[(i.facility, i.kind)] = counts.get((i.facility, i.kind), 0) + 1
    ws.append(["施設"] + kinds + ["合計"])
    for fac in sorted({i.facility for i in issues}):
        row = [counts.get((fac, k), 0) for k in kinds]
        ws.append([fac] + row + [sum(row)])
    total_row = ["合計"] + [sum(counts.get((f, k), 0) for f, kk in counts if kk == k)
                          for k in kinds]
    total_row.append(sum(total_row[1:]))
    ws.append(total_row)
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = HEADER_FILL
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    if not issues:
        ws.append([]); ws.append(["エラーはありませんでした"])
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(kinds) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "A2"

    # --- 全職員一覧(施設横断) ---
    used = {"エラーサマリー", "全職員一覧", "エラー詳細"}
    ws_all = wb.create_sheet("全職員一覧")
    all_rows, fac_of = [], {}
    for fac in sorted(people_by_fac):
        for p in _sort_people(people_by_fac[fac], months):
            all_rows.append(p)
            fac_of[id(p)] = fac
    _write_matrix_sheet(ws_all, all_rows, months, with_facility=fac_of)

    # --- 施設別シート ---
    for fac in sorted(people_by_fac):
        wsf = wb.create_sheet(_safe_sheet_name(fac, used))
        _write_matrix_sheet(wsf, _sort_people(people_by_fac[fac], months), months)

    # --- エラー詳細(末尾) ---
    wse = wb.create_sheet("エラー詳細")
    wse.append(["施設", "ファイル名", "シート名", "対象月", "行", "種別", "内容"])
    for i in sorted(issues, key=lambda x: (x.facility, x.kind, str(x.month), str(x.row))):
        wse.append([i.facility, i.file_name, i.sheet_name, i.month, i.row, i.kind, i.detail])
    for c in wse[1]:
        c.font = Font(bold=True); c.fill = HEADER_FILL
    for i, w in enumerate([14, 40, 14, 10, 6, 16, 70], 1):
        wse.column_dimensions[get_column_letter(i)].width = w
    wse.freeze_panes = "A2"
    wse.auto_filter.ref = f"A1:G{max(wse.max_row, 1)}"

    wb.save(out_path)
    return out_path
