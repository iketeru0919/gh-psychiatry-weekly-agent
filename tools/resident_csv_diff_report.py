# -*- coding: utf-8 -*-
"""
入居者情報管理CSV 差分レポート生成（Colab 1セル実行版 / 修正版）

旧版からの主な変更点
  1. 期限追跡の対象列を DATE_COLS に統一（個別支援計画更新日などの脱落を解消）
  2. 退去者フィルタを差分計算の「後」に適用し、🚪退去 と 🗑️削除 を分離
  3. 重複シートの廃止（旧⑦期限切れ・対応なし / 旧⑧改善確認 → ②③の列に統合）
  4. 同一人物・同一期限日の項目をまとめ、行数を圧縮
  5. 「入居者別対応リスト」を新設（1人1行の実務用ワークリスト）
  6. 期限日を日付型で出力／エンコーディング自動判定の堅牢化／データ不備シート追加
"""

# =========================================================
# 0. 実行環境の準備
# =========================================================
import os

try:
    from google.colab import drive
    drive.mount('/content/drive')
    print('✅ Google Drive のマウント完了')
    BASE_DIR = '/content/drive/MyDrive/入居者情報管理'
except ImportError:
    BASE_DIR = './入居者情報管理'
    print(f'ℹ️ Colab以外で実行しています（BASE_DIR={BASE_DIR}）')

# ▼ 旧CSVのファイル名（inputフォルダ内）
OLD_FILENAME = 'ご入居者情報管理_20260728T080703+0900.csv'

# ▼ 新CSVのファイル名（inputフォルダ内）
NEW_FILENAME = 'ご入居者情報管理_20260818T140216+0900.csv'

INPUT_DIR  = f'{BASE_DIR}/input'
OUTPUT_DIR = f'{BASE_DIR}/output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

OLD_PATH = f'{INPUT_DIR}/{OLD_FILENAME}'
NEW_PATH = f'{INPUT_DIR}/{NEW_FILENAME}'

print(f'📂 旧ファイル: {OLD_FILENAME}')
print(f'📂 新ファイル: {NEW_FILENAME}')
print(f'📂 出力先   : {OUTPUT_DIR}')

import re
import unicodedata
from collections import defaultdict
from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =========================================================
# 1. 定数定義
# =========================================================
KEY_COL     = 'レコード番号'
NAME_COL    = 'ご入居者名（受給者証統一）（※姓名の間全角スペース）'
SITE_COL    = '施設名'
EXIT_COL    = '退去日'
SERVICE_COL = 'サービス種別（or退去）'
EXIT_LABEL  = '退去済'

REQUIRED_COLS = [KEY_COL, NAME_COL, SITE_COL, EXIT_COL, SERVICE_COL]

# 完全一致で除外する列（'Unnamed: xx' は自動除外）
IGNORE_COLS = ['作成者']

# 期限として追跡する日付列
DATE_COLS = [
    '区分認定期間（終了）',
    'サービス期間（終了）',
    '特定給付期間（終了）',
    '負担上限期間（終了）',
    '特別加算期間（終了）',
    '更新日（その他制度）',
    '個別支援計画更新日',
]

# 【修正①】旧版は ALERT_COLS が DATE_COLS と食い違っており、
#          個別支援計画更新日（最多）が対応状況の追跡から丸ごと漏れていた。
#          また日付列でない「特別加算項目」が含まれ、常に評価対象外だった。
ALERT_COLS = list(DATE_COLS)

ALERT_DAYS = 60          # 何日先まで予告するか
WORST_SITE_TOP = 10      # 施設別ワーストの表示件数

COLOR_HEADER   = '2F5496'
COLOR_RED      = 'FFCCCC'
COLOR_ORANGE   = 'FFE5CC'
COLOR_YELLOW   = 'FFFACC'
COLOR_GREEN    = 'E2EFDA'
COLOR_GRAY     = 'D9D9D9'
COLOR_ROW_ALT  = 'F2F2F2'
COLOR_OLD_VAL  = 'FFE0E0'
COLOR_NEW_VAL  = 'E2EFDA'
COLOR_TOTAL    = 'D9E1F2'
COLOR_STRONG   = 'FF9999'

FONT_NAME = 'Meiryo'
today     = date.today()

# =========================================================
# 2. 読み込みユーティリティ
# =========================================================
ENCODING_CANDIDATES = ['utf-8-sig', 'cp932', 'utf-8', 'euc-jp']

DATE_FORMATS = ('%Y/%m/%d', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%d %H:%M:%S', '%Y年%m月%d日', '%Y.%m.%d')


def load_csv(path):
    """文字コード候補を順に試す。chardetの誤判定（cp932/shift_jis）を避ける。"""
    last_error = None
    for enc in ENCODING_CANDIDATES:
        try:
            return pd.read_csv(path, encoding=enc, dtype=str), enc
        except UnicodeDecodeError as e:
            last_error = e
    try:
        import chardet
        with open(path, 'rb') as f:
            enc = chardet.detect(f.read())['encoding']
        return pd.read_csv(path, encoding=enc, dtype=str), f'{enc}(chardet)'
    except Exception:
        raise last_error


def parse_date(value):
    """日付として解釈できれば date、できなければ None。"""
    text = str(value).strip()
    if text in ('', 'nan', 'NaT', 'None'):
        return None
    text = unicodedata.normalize('NFKC', text)
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_date(value):
    """日付列の表記ゆれ（2026/8/1 と 2026/08/01 等）を吸収し、偽の差分を防ぐ。"""
    d = parse_date(value)
    if d:
        return d.strftime('%Y/%m/%d')
    return str(value).strip() if str(value).strip().lower() != 'nan' else ''


def days_until(value):
    d = parse_date(value)
    return (d - today).days if d else None


def read_source(path, label):
    df, enc = load_csv(path)
    df.columns = df.columns.str.strip()

    drop_cols = [c for c in df.columns
                 if c in IGNORE_COLS or c.startswith('Unnamed')]
    df = df.drop(columns=drop_cols)

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise KeyError(f'{label}CSVに必須列がありません: {missing}')

    # 全列を「前後空白を除いた文字列」に揃え、NaN と '' の比較ゆれをなくす
    for col in df.columns:
        df[col] = df[col].fillna('').astype(str).str.strip()

    for col in DATE_COLS:
        if col in df.columns:
            df[col] = df[col].map(normalize_date)

    print(f'📥 {label}CSV: {len(df)}行 / {len(df.columns)}列 '
          f'（encoding={enc}, 除外列={drop_cols or "なし"}）')
    return df


def active_only(df):
    """退去日が入っている、またはサービス種別が退去済の方を除外。"""
    return df[(df[EXIT_COL] == '') & (df[SERVICE_COL] != EXIT_LABEL)]


raw_old = read_source(OLD_PATH, '旧')
raw_new = read_source(NEW_PATH, '新')

# 【修正②】退去者フィルタは差分計算「後」に使う。
#          旧版はフィルタ後に差分を取っていたため、今回退去した方が
#          「🗑️削除レコード」として出力され、実削除と区別できなかった。
act_old = active_only(raw_old)
act_new = active_only(raw_new)

print(f'✅ 在籍者: 旧 {len(act_old)}名 → 新 {len(act_new)}名')

# =========================================================
# 3. データ不備チェック
# =========================================================
issue_rows = []


def add_issue(kind, key, site, name, detail):
    issue_rows.append({'種別': kind, 'レコード番号': key, '施設名': site,
                       '入居者名': name, '内容': detail})


# 3-1. レコード番号の重複
for label, df in (('旧CSV', raw_old), ('新CSV', raw_new)):
    dup_keys = df[KEY_COL][df[KEY_COL].duplicated()].unique()
    for key in dup_keys:
        row = df[df[KEY_COL] == key].iloc[0]
        add_issue('レコード番号の重複', key, row[SITE_COL], row[NAME_COL],
                  f'{label}に同じレコード番号が複数存在します（先頭行のみ採用）')

# 3-2. 同一施設・同姓同名（別レコード番号）＝二重登録の疑い
dup_person = act_new[act_new.duplicated([SITE_COL, NAME_COL], keep=False)]
for (site, name), grp in dup_person.groupby([SITE_COL, NAME_COL]):
    keys = '、'.join(sorted(grp[KEY_COL], key=lambda k: (len(k), k)))
    add_issue('二重登録の疑い', keys, site, name,
              f'同一施設に同姓同名が{len(grp)}件（レコード番号 {keys}）')

# 3-3. 氏名の区切り（規定は姓名の間に全角スペース1つ）
NAME_OK = re.compile(r'^[^\s　]+　[^\s　]+$')
for _, row in act_new.iterrows():
    name = row[NAME_COL]
    if name and not NAME_OK.match(name):
        add_issue('氏名の表記ゆれ', row[KEY_COL], row[SITE_COL], name,
                  '姓名の区切りが全角スペース1つになっていません')

# 3-4. 日付として解釈できない値
for _, row in act_new.iterrows():
    for col in DATE_COLS:
        value = row.get(col, '')
        if value and parse_date(value) is None:
            add_issue('日付書式エラー', row[KEY_COL], row[SITE_COL], row[NAME_COL],
                      f'{col} = 「{value}」')

# 重複キーは先頭行を採用してインデックス化
IDX_OLD = act_old.drop_duplicates(KEY_COL).set_index(KEY_COL, drop=False)
IDX_NEW = act_new.drop_duplicates(KEY_COL).set_index(KEY_COL, drop=False)
ALL_OLD_KEYS = set(raw_old[KEY_COL])
ALL_NEW_KEYS = set(raw_new[KEY_COL])

# 【修正③】set をそのまま回すと実行ごとに行順が変わるため常にソートする
old_keys    = sorted(IDX_OLD.index)
new_keys    = sorted(IDX_NEW.index)
common_keys = sorted(set(old_keys) & set(new_keys))

# =========================================================
# 4. 差分の抽出
# =========================================================
compare_cols = [c for c in act_new.columns
                if c in act_old.columns and c != KEY_COL]

changed_pairs = set()      # (レコード番号, 列名) 変更があった組み合わせ
change_items  = []         # 1変更1件

for key in common_keys:
    row_old = IDX_OLD.loc[key]     # 事前インデックス化でO(n^2)の全行走査を解消
    row_new = IDX_NEW.loc[key]
    for col in compare_cols:
        v_old = row_old.get(col, '')
        v_new = row_new.get(col, '')
        if v_old == v_new:
            continue
        changed_pairs.add((key, col))
        change_items.append({
            'レコード番号': key,
            '施設名'      : row_new[SITE_COL],
            '入居者名'    : row_new[NAME_COL],
            '変更項目'    : col,
            '旧の値'      : v_old or '（空欄）',
            '新の値'      : v_new or '（空欄）',
            '_old_raw'    : v_old,
            '_col'        : col,
        })

# --- 4-1. 更新レコード：同一人物・同一の「旧値→新値」を1行にまとめる ---
update_groups = defaultdict(list)
for item in change_items:
    update_groups[(item['レコード番号'], item['旧の値'], item['新の値'])].append(item)

update_rows = []
for (key, v_old, v_new), items in update_groups.items():
    head = items[0]
    # 期限切れだった日付が更新されたかどうかのフラグ（旧⑧改善確認の代替）
    resolved = ''
    for item in items:
        if item['_col'] in ALERT_COLS:
            d_old = days_until(item['_old_raw'])
            if d_old is not None and d_old < 0:
                resolved = '✅ 期限切れ→更新'
                break
    update_rows.append({
        'レコード番号': key,
        '施設名'      : head['施設名'],
        '入居者名'    : head['入居者名'],
        '変更項目'    : '、'.join(i['変更項目'] for i in items),
        '項目数'      : len(items),
        '旧の値'      : v_old,
        '新の値'      : v_new,
        '期限切れ対応': resolved,
    })
update_rows.sort(key=lambda r: (r['施設名'], r['レコード番号']))

# --- 4-2. 新規／再登録・退去／削除 ---
added_rows = []
for key in sorted(set(new_keys) - set(old_keys)):
    row = IDX_NEW.loc[key]
    kind = '🔄 再登録' if key in ALL_OLD_KEYS else '🆕 新規'
    added_rows.append({'種別': kind, 'row': row, 'key': key})

removed_rows = []
for key in sorted(set(old_keys) - set(new_keys)):
    old_row = IDX_OLD.loc[key]
    if key in ALL_NEW_KEYS:
        # 新CSVには存在するが在籍者から外れた＝退去（新CSV側の値を表示する）
        new_row = raw_new[raw_new[KEY_COL] == key].iloc[0]
        removed_rows.append({
            '種別': '🚪 退去', 'レコード番号': key,
            '施設名': new_row[SITE_COL], '入居者名': new_row[NAME_COL],
            'サービス種別': new_row[SERVICE_COL],
            '入居日': new_row.get('入居日', ''), '退去日': new_row.get(EXIT_COL, ''),
            '備考': '新CSVで退去処理済み',
        })
    else:
        removed_rows.append({
            '種別': '🗑️ 削除', 'レコード番号': key,
            '施設名': old_row[SITE_COL], '入居者名': old_row[NAME_COL],
            'サービス種別': old_row[SERVICE_COL],
            '入居日': old_row.get('入居日', ''), '退去日': old_row.get(EXIT_COL, ''),
            '備考': '新CSVにレコードが存在しません（要確認）',
        })
removed_rows.sort(key=lambda r: (r['種別'], r['施設名']))

# =========================================================
# 5. 期限アラート
# =========================================================
def status_of(days):
    if days < 0:
        return '🔴 期限切れ'
    if days <= 30:
        return '🟠 30日以内'
    return '🟡 60日以内'


def overdue_bucket(days):
    if days >= 0:
        return '－'
    if days >= -30:
        return '① 30日以内'
    if days >= -90:
        return '② 31〜90日'
    return '③ 91日以上'


STATUS_ORDER = {'🔴 期限切れ': 0, '🟠 30日以内': 1, '🟡 60日以内': 2}

alert_items = []
for key in new_keys:
    row = IDX_NEW.loc[key]
    for col in ALERT_COLS:
        value = row.get(col, '')
        days  = days_until(value)
        if days is None or days > ALERT_DAYS:
            continue
        alert_items.append({
            'レコード番号': key,
            '施設名'      : row[SITE_COL],
            '入居者名'    : row[NAME_COL],
            '項目名'      : col,
            '期限日'      : parse_date(value),
            '残り日数'    : days,
            'ステータス'  : status_of(days),
            '経過区分'    : overdue_bucket(days),
            '更新有無'    : (key, col) in changed_pairs,
        })

# --- 5-1. 明細：同一人物・同一期限日の項目を1行にまとめる ---
alert_groups = defaultdict(list)
for item in alert_items:
    alert_groups[(item['レコード番号'], item['期限日'])].append(item)

alert_rows = []
for (key, limit_date), items in alert_groups.items():
    head    = items[0]
    changed = sum(1 for i in items if i['更新有無'])
    if changed == 0:
        action = '❌ 未対応'
    elif changed == len(items):
        action = '✅ 前回から更新'
    else:
        action = '⚠️ 一部のみ更新'
    alert_rows.append({
        'ステータス'  : head['ステータス'],
        '経過区分'    : head['経過区分'],
        '施設名'      : head['施設名'],
        '入居者名'    : head['入居者名'],
        '項目名'      : '、'.join(i['項目名'] for i in items),
        '項目数'      : len(items),
        '期限日'      : limit_date,
        '残り日数'    : head['残り日数'],
        '前回対応'    : action,
        'レコード番号': key,
    })
alert_rows.sort(key=lambda r: (r['残り日数'], r['施設名'], r['入居者名']))

# --- 5-2. 入居者別対応リスト（1人1行） ---
person_groups = defaultdict(list)
for item in alert_items:
    person_groups[item['レコード番号']].append(item)

person_rows = []
for key, items in person_groups.items():
    head    = items[0]
    top     = min(items, key=lambda i: i['残り日数'])
    changed = sum(1 for i in items if i['更新有無'])
    person_rows.append({
        '対応区分'    : top['ステータス'],
        'レコード番号': key,
        '施設名'      : head['施設名'],
        '入居者名'    : head['入居者名'],
        '🔴期限切れ'  : sum(1 for i in items if i['ステータス'].startswith('🔴')),
        '🟠30日以内'  : sum(1 for i in items if i['ステータス'].startswith('🟠')),
        '🟡60日以内'  : sum(1 for i in items if i['ステータス'].startswith('🟡')),
        '最短期限日'  : top['期限日'],
        '残り日数'    : top['残り日数'],
        '最優先項目'  : top['項目名'],
        '対象項目'    : '、'.join(sorted({i['項目名'] for i in items})),
        '前回対応'    : '❌ 未対応' if changed == 0 else f'✅ {changed}項目更新',
    })
# 未対応かつ期限が切迫している人を上に
person_rows.sort(key=lambda r: (r['前回対応'].startswith('✅'), r['残り日数']))

# =========================================================
# 6. 集計
# =========================================================
alert_persons = {i['レコード番号'] for i in alert_items}
expired_items = [i for i in alert_items if i['残り日数'] < 0]
expired_persons = {i['レコード番号'] for i in expired_items}
untouched_expired_persons = {
    key for key in expired_persons
    if not any(i['更新有無'] for i in person_groups[key] if i['残り日数'] < 0)
}


def count_status(prefix):
    return sum(1 for i in alert_items if i['ステータス'].startswith(prefix))


site_stats = defaultdict(lambda: {'人数': set(), '件数': 0, '期限切れ': 0, '未対応': 0})
for item in alert_items:
    s = site_stats[item['施設名']]
    s['人数'].add(item['レコード番号'])
    s['件数'] += 1
    if item['残り日数'] < 0:
        s['期限切れ'] += 1
        if not item['更新有無']:
            s['未対応'] += 1

print('✅ 集計完了')
print(f'   期限アラート     : {len(alert_items)}件 / {len(alert_persons)}名'
      f'（明細{len(alert_rows)}行に集約）')
print(f'   　🔴期限切れ     : {count_status("🔴")}件 / {len(expired_persons)}名'
      f'（うち前回から未対応 {len(untouched_expired_persons)}名）')
print(f'   更新レコード     : {len(change_items)}項目（{len(update_rows)}行に集約）'
      f' / {len({i["レコード番号"] for i in change_items})}名')
print(f'   新規・再登録     : {len(added_rows)}件')
print(f'   退去・削除       : {len(removed_rows)}件')
print(f'   データ不備       : {len(issue_rows)}件')

# =========================================================
# 7. Excel出力
# =========================================================
def fill(color):
    return PatternFill('solid', start_color=color, end_color=color)


def border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)


def put(ws, row, col, value, *, bold=False, bg=None, align='left', size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name=FONT_NAME, size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = border()
    if bg:
        cell.fill = fill(bg)
    if isinstance(value, date):
        cell.number_format = 'yyyy/mm/dd'   # 日付型で書き出し、期間フィルタを使えるようにする
    return cell


def write_header(ws, headers, color=COLOR_HEADER, row=1):
    for i, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=name)
        cell.font      = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
        cell.fill      = fill(color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border()


def write_table(ws, headers, records, row_color=None, header_color=COLOR_HEADER,
                empty_message='該当データはありません'):
    assert len(set(headers)) == len(headers), f'見出しが重複しています: {headers}'
    ws.sheet_view.showGridLines = False
    write_header(ws, headers, header_color)
    if records:
        for r_idx, record in enumerate(records, start=2):
            bg = row_color(record) if row_color else (
                COLOR_ROW_ALT if r_idx % 2 == 0 else None)
            for c_idx, key in enumerate(headers, start=1):
                value = record.get(key, '')
                put(ws, r_idx, c_idx, value, bg=bg,
                    align='center' if isinstance(value, (int, date)) else 'left')
    else:
        put(ws, 2, 1, empty_message)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)


def auto_width(ws, min_w=10, max_w=42):
    for col in ws.columns:
        longest = 0
        for cell in col:
            if cell.value is None:
                continue
            text = cell.value.strftime('%Y/%m/%d') if isinstance(cell.value, date) else str(cell.value)
            # 全角は2文字幅として概算
            width = sum(2 if unicodedata.east_asian_width(ch) in 'WFA' else 1 for ch in text)
            longest = max(longest, width)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(longest + 2, min_w), max_w)


def status_color(status):
    if status.startswith('🔴'):
        return COLOR_RED
    if status.startswith('🟠'):
        return COLOR_ORANGE
    if status.startswith('🟡'):
        return COLOR_YELLOW
    return None


wb = Workbook()
wb.remove(wb.active)
now_str = datetime.now().strftime('%Y/%m/%d %H:%M')

# ---------- ①サマリー ----------
ws = wb.create_sheet('①サマリー')
ws.sheet_view.showGridLines = False
for col, width in zip('ABCDEF', (30, 18, 14, 14, 14, 14)):
    ws.column_dimensions[col].width = width

summary_lines = [
    ('【実行情報】', '', None),
    ('実行日時', now_str, None),
    ('基準日', today, None),
    ('旧ファイル', OLD_FILENAME, None),
    ('新ファイル', NEW_FILENAME, None),
    ('対象人数（在籍）', f'旧 {len(act_old)}名 → 新 {len(act_new)}名', None),
    ('', '', None),
    ('【期限アラート】', '', None),
    ('🔴 期限切れ', f'{count_status("🔴")} 件', COLOR_RED),
    ('🟠 30日以内', f'{count_status("🟠")} 件', COLOR_ORANGE),
    ('🟡 60日以内', f'{count_status("🟡")} 件', COLOR_YELLOW),
    ('対象人数', f'{len(alert_persons)} 名', None),
    ('', '', None),
    ('【期限切れの経過日数】', '', None),
    ('① 30日以内', f'{sum(1 for i in expired_items if i["経過区分"].startswith("①"))} 件', COLOR_ORANGE),
    ('② 31〜90日', f'{sum(1 for i in expired_items if i["経過区分"].startswith("②"))} 件', COLOR_RED),
    ('③ 91日以上（棚卸し対象）', f'{sum(1 for i in expired_items if i["経過区分"].startswith("③"))} 件', COLOR_STRONG),
    ('', '', None),
    ('【対応状況】', '', None),
    ('🚨 期限切れ・前回から未対応', f'{len(untouched_expired_persons)} 名', COLOR_STRONG),
    ('✅ 期限切れを更新した人数',
     f'{len({r["レコード番号"] for r in update_rows if r["期限切れ対応"]})} 名', COLOR_GREEN),
    ('', '', None),
    ('【前回からの動き】', '', None),
    ('✏️ 更新のあった人数',
     f'{len({i["レコード番号"] for i in change_items})} 名 / {len(change_items)} 項目', None),
    ('🆕 新規入居', f'{sum(1 for r in added_rows if r["種別"].endswith("新規"))} 件', COLOR_GREEN),
    ('🔄 再登録', f'{sum(1 for r in added_rows if r["種別"].endswith("再登録"))} 件', None),
    ('🚪 退去', f'{sum(1 for r in removed_rows if r["種別"].endswith("退去"))} 件', COLOR_GRAY),
    ('🗑️ 削除（要確認）', f'{sum(1 for r in removed_rows if r["種別"].endswith("削除"))} 件', COLOR_STRONG),
    ('', '', None),
    ('【データ品質】', '', None),
    ('⚠️ データ不備', f'{len(issue_rows)} 件', COLOR_YELLOW if issue_rows else None),
]

r = 1
for label, value, bg in summary_lines:
    if label:
        put(ws, r, 1, label, bold=label.startswith('【'))
        if value != '':
            put(ws, r, 2, value, bold=bool(bg), bg=bg,
                align='center' if isinstance(value, date) else 'left')
    r += 1

# 項目別・期限アラート集計
r += 1
put(ws, r, 1, '【項目別・期限アラート集計】', bold=True)
r += 1
write_header(ws, ['項目名', '🔴 期限切れ', '🟠 30日以内', '🟡 60日以内', '合計'], row=r)
for col_name in DATE_COLS:
    r += 1
    items = [i for i in alert_items if i['項目名'] == col_name]
    counts = [sum(1 for i in items if i['ステータス'].startswith(p)) for p in ('🔴', '🟠', '🟡')]
    put(ws, r, 1, col_name)
    for offset, (count, color) in enumerate(
            zip(counts, (COLOR_RED, COLOR_ORANGE, COLOR_YELLOW))):
        put(ws, r, 2 + offset, count, bg=color if count else None, align='center')
    put(ws, r, 5, sum(counts), bg=COLOR_GREEN if sum(counts) else None, align='center')
r += 1
put(ws, r, 1, '合計', bold=True, bg=COLOR_TOTAL)
for offset, prefix in enumerate(('🔴', '🟠', '🟡')):
    put(ws, r, 2 + offset, count_status(prefix), bold=True, bg=COLOR_TOTAL, align='center')
put(ws, r, 5, len(alert_items), bold=True, bg=COLOR_TOTAL, align='center')

# 施設別ワースト
r += 2
put(ws, r, 1, f'【施設別ワースト{WORST_SITE_TOP}（期限切れ・未対応の多い順）】', bold=True)
r += 1
write_header(ws, ['施設名', '対象人数', 'アラート件数', '🔴 期限切れ', '🚨 未対応'], row=r)
ranked = sorted(site_stats.items(), key=lambda kv: (-kv[1]['未対応'], -kv[1]['期限切れ']))
for site, stat in ranked[:WORST_SITE_TOP]:
    r += 1
    put(ws, r, 1, site)
    put(ws, r, 2, len(stat['人数']), align='center')
    put(ws, r, 3, stat['件数'], align='center')
    put(ws, r, 4, stat['期限切れ'], bg=COLOR_RED if stat['期限切れ'] else None, align='center')
    put(ws, r, 5, stat['未対応'], bg=COLOR_STRONG if stat['未対応'] else None, align='center')

# ---------- ②入居者別対応リスト ----------
ws = wb.create_sheet('②入居者別対応リスト')
write_table(
    ws,
    ['対応区分', '施設名', '入居者名', '🔴期限切れ', '🟠30日以内', '🟡60日以内',
     '最短期限日', '残り日数', '最優先項目', '対象項目', '前回対応', 'レコード番号'],
    person_rows,
    row_color=lambda rec: COLOR_STRONG if rec['前回対応'].startswith('❌') and rec['残り日数'] < 0
    else status_color(rec['対応区分']),
    empty_message='期限アラートはありません',
)

# ---------- ③期限アラート明細 ----------
ws = wb.create_sheet('③期限アラート明細')
write_table(
    ws,
    ['ステータス', '経過区分', '施設名', '入居者名', '項目名', '項目数',
     '期限日', '残り日数', '前回対応', 'レコード番号'],
    alert_rows,
    row_color=lambda rec: status_color(rec['ステータス']),
    empty_message='期限アラートはありません',
)

# ---------- ④更新レコード ----------
ws = wb.create_sheet('④更新レコード')
write_table(
    ws,
    ['レコード番号', '施設名', '入居者名', '変更項目', '項目数',
     '旧の値', '新の値', '期限切れ対応'],
    update_rows,
    row_color=lambda rec: COLOR_GREEN if rec['期限切れ対応'] else None,
    empty_message='変更はありません',
)
for row_cells in ws.iter_rows(min_row=2, min_col=6, max_col=7):
    for offset, cell in enumerate(row_cells):
        if cell.value not in (None, ''):
            cell.fill = fill(COLOR_OLD_VAL if offset == 0 else COLOR_NEW_VAL)
            if offset == 1:
                cell.font = Font(name=FONT_NAME, size=10, bold=True)

# ---------- ⑤新規・再登録 ----------
show_cols = [c for c in ['サービス種別（or退去）', '入居日', '区分',
                         'サービス期間（終了）', '負担上限期間（終了）']
             if c in act_new.columns]
added_records = []
for entry in added_rows:
    row, key = entry['row'], entry['key']
    record = {'種別': entry['種別'], 'レコード番号': key,
              '施設名': row[SITE_COL], '入居者名': row[NAME_COL]}
    record.update({c: row.get(c, '') for c in show_cols})
    hits = person_groups.get(key, [])
    record['期限アラート'] = (
        f'⚠️ {len(hits)}件（最短 {min(h["残り日数"] for h in hits)}日）' if hits else '－')
    added_records.append(record)
added_records.sort(key=lambda r_: (r_['種別'], r_['施設名']))

ws = wb.create_sheet('⑤新規・再登録')
write_table(
    ws,
    ['種別', 'レコード番号', '施設名', '入居者名'] + show_cols + ['期限アラート'],
    added_records,
    row_color=lambda rec: COLOR_YELLOW if rec['期限アラート'] != '－' else COLOR_GREEN,
    header_color='1F7A4D',
    empty_message='新規追加はありません',
)

# ---------- ⑥退去・削除 ----------
ws = wb.create_sheet('⑥退去・削除')
write_table(
    ws,
    ['種別', 'レコード番号', '施設名', '入居者名', 'サービス種別',
     '入居日', '退去日', '備考'],
    removed_rows,
    row_color=lambda rec: COLOR_GRAY if rec['種別'].endswith('退去') else COLOR_STRONG,
    header_color='7F7F7F',
    empty_message='退去・削除はありません',
)

# ---------- ⑦データ不備 ----------
ws = wb.create_sheet('⑦データ不備')
issue_rows.sort(key=lambda r_: (r_['種別'], r_['施設名']))
write_table(
    ws,
    ['種別', 'レコード番号', '施設名', '入居者名', '内容'],
    issue_rows,
    row_color=lambda rec: COLOR_YELLOW,
    header_color='B8860B',
    empty_message='データ不備は検出されませんでした',
)

# ---------- 保存 ----------
out_name = f'入居者情報_差分レポート_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
out_path = f'{OUTPUT_DIR}/{out_name}'
wb.save(out_path)

print('✅ レポート出力完了')
print(f'📄 ファイル名: {out_name}')
print(f'📂 保存先   : {OUTPUT_DIR}')
