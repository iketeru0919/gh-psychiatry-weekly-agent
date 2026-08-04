"""実績記録表ツールの回帰テスト。

    py -m unittest test_jisseki -v
"""

from pathlib import Path
import shutil
import tempfile
import unittest

import pandas as pd
from openpyxl import Workbook

import jisseki_common as common
import jisseki_cell_extract as extract


class MonthMatchTest(unittest.TestCase):
    """1月指定で11月、2月指定で12月を拾ってしまう不具合の回帰テスト。"""

    def match(self, path, year, month):
        return common.is_target_month_file(Path(path), year, month, Path(r"C:\base"))

    def test_matches_correct_month(self):
        self.assertTrue(self.match(r"C:\base\施設\2026年4月\実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\施設\2026年04月\実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\施設\R8.4\実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\施設\令和8年4月\実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\施設\202604_実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\2026\4月\実績1F.xlsx", 2026, 4))
        self.assertTrue(self.match(r"C:\base\施設\２０２６年４月\実績1F.xlsx", 2026, 4))

    def test_does_not_match_other_month(self):
        self.assertFalse(self.match(r"C:\base\施設\2026年5月\実績1F.xlsx", 2026, 4))
        self.assertFalse(self.match(r"C:\base\施設\2025年4月\実績1F.xlsx", 2026, 4))

    def test_january_does_not_match_october_to_december(self):
        for path in [
            r"C:\base\施設\2026年10月\実績1F.xlsx",
            r"C:\base\施設\2026年11月\実績1F.xlsx",
            r"C:\base\施設\2026年12月\実績1F.xlsx",
            r"C:\base\施設\2026-10\実績1F.xlsx",
            r"C:\base\施設\2026-11\実績1F.xlsx",
            r"C:\base\施設\R8.10\実績1F.xlsx",
            r"C:\base\2026\11月\実績1F.xlsx",
        ]:
            with self.subTest(path=path):
                self.assertFalse(self.match(path, 2026, 1))

    def test_february_does_not_match_december(self):
        self.assertFalse(self.match(r"C:\base\施設\2026年12月\実績1F.xlsx", 2026, 2))
        self.assertFalse(self.match(r"C:\base\2026\12月\実績1F.xlsx", 2026, 2))

    def test_extract_year_month(self):
        cases = [
            (r"C:\base\施設\2026年4月\実績1F.xlsx", (2026, 4)),
            (r"C:\base\施設\R8.12\実績1F.xlsx", (2026, 12)),
            (r"C:\base\施設\令和8年1月\実績1F.xlsx", (2026, 1)),
            (r"C:\base\施設\実績1F.xlsx", None),
        ]

        for path, expected in cases:
            with self.subTest(path=path):
                self.assertEqual(
                    common.extract_year_month_from_path(Path(path), Path(r"C:\base")),
                    expected,
                )


class FileTypeTest(unittest.TestCase):
    """全角・半角の組み合わせをすべて拾えること。"""

    def test_all_width_variants(self):
        for name in ["実績1F.xlsx", "実績１F.xlsx", "実績1Ｆ.xlsx", "実績１Ｆ.xlsx"]:
            with self.subTest(name=name):
                self.assertEqual(common.get_file_type(name), "1F")

        for name in ["実績2F.xlsx", "実績２F.xlsx", "実績2Ｆ.xlsx", "実績２Ｆ.xlsx"]:
            with self.subTest(name=name):
                self.assertEqual(common.get_file_type(name), "2F")

    def test_shortstay_wins(self):
        self.assertEqual(common.get_file_type("短期実績1F.xlsx"), "短期")

    def test_unknown(self):
        self.assertIsNone(common.get_file_type("実績記録表.xlsx"))


class SheetFilterTest(unittest.TestCase):
    def test_non_user_sheets_are_excluded(self):
        for name in [
            "実績記録票",
            "実績記録表",
            "延べ日数",
            "算定項目リスト",
            "重度障害者支援加算要件",
            "空き部屋２０１",
            "原本",
            "記入例",
        ]:
            with self.subTest(name=name):
                self.assertFalse(common.is_target_sheet(name))

    def test_user_sheets_are_kept(self):
        for name in ["山本哲也", "101田中", "石原浩・区4", "ヒロオカ義也"]:
            with self.subTest(name=name):
                self.assertTrue(common.is_target_sheet(name))


class FolderFilterTest(unittest.TestCase):
    def test_test_folder_excluded(self):
        self.assertTrue(common.is_excluded_folder("99.テストフォルダ　※担当以外使用禁止"))
        self.assertTrue(common.is_excluded_folder("過去分"))

    def test_real_facility_is_not_excluded(self):
        # 施設名に「テスト」が含まれるだけでは除外しない
        for name in ["RASIEL三蔵子⑤　★", "RASIELテスト施設⑤　★", "RASIEL旧河内⑤　★"]:
            with self.subTest(name=name):
                self.assertFalse(common.is_excluded_folder(name))


class ValueTest(unittest.TestCase):
    def test_is_applicable_treats_not_applicable_as_zero(self):
        for value in ["非該当", "なし", "－", "0", "０", "", None]:
            with self.subTest(value=value):
                self.assertEqual(extract.is_applicable(value), 0)

        for value in ["Ⅰ", "Ⅱ", "Ⅴ", "該当", "条件①", "●"]:
            with self.subTest(value=value):
                self.assertEqual(extract.is_applicable(value), 1)

    def test_to_number(self):
        self.assertEqual(extract.to_number("1,234"), 1234)
        self.assertEqual(extract.to_number("３０"), 30)
        self.assertEqual(extract.to_number(True), 1)
        self.assertEqual(extract.to_number(None), 0)
        self.assertEqual(extract.to_number(""), 0)

    def test_to_number_records_error(self):
        errors = []
        context = {"対象月": "2026年4月", "施設名": "A", "区分": "1F", "利用者名": "山田"}

        self.assertEqual(extract.to_number("あああ", context, errors), 0)
        self.assertEqual(len(errors), 1)
        self.assertIn("数値に変換できません", errors[0]["内容"])

    def test_extract_day_count(self):
        cases = [
            ("短期入所　１４日/月", 14),
            ("短期入所　　　５日/月\t", 5),
            ("10日/月", 10),
            ("短期入所10日／日", 10),
            (14, 14),
            (None, None),
        ]

        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(extract.extract_day_count(value), expected)


# ===== 疑似ファイルを作って通しで検証する =====


def build_gh_workbook(path: Path, users, with_template=True):
    """実ファイル（RASIEL 1F/2F 2026年7月）と同じ配置で疑似シートを作る。

    users: (シート名, 合計算定日数, C列の状況, 状況を書く日数, 住居外利用回数)
    """
    wb = Workbook()
    wb.remove(wb.active)

    def fill_sheet(ws, day_count, status, status_days, outside_use, name=None):
        if name:
            ws["I2"] = name          # 支給決定障害者等氏名
            ws["P2"] = 5             # 支援区分

        ws["E3"] = "非該当"
        ws["I3"] = "非該当"
        ws["M3"] = "Ⅴ"
        ws["Q3"] = "非該当"
        ws["E4"] = "非該当"
        ws["I4"] = "非該当"
        ws["M4"] = "非該当"
        ws["Q4"] = "非該当"

        ws["C8"] = "サービス提供の状況"
        ws["E8"] = "住居外利用"

        if status:
            for offset in range(status_days):
                ws.cell(row=extract.GH_DATA_ROW_START + offset, column=3, value=status)

        # 住居外利用は日次に 1 を立て、E40 がその合計になっている
        for offset in range(outside_use):
            ws.cell(row=extract.GH_DATA_ROW_START + offset, column=5, value=1)

        ws.cell(row=extract.GH_TOTAL_ROW, column=1, value="合計")
        ws[f"C{extract.GH_TOTAL_ROW}"] = day_count
        ws[f"E{extract.GH_TOTAL_ROW}"] = outside_use
        ws[f"F{extract.GH_TOTAL_ROW}"] = day_count
        ws[f"I{extract.GH_TOTAL_ROW}"] = day_count

        # 表の下の見出し行。行範囲を限定していないと誤カウントされる。
        ws["C42"] = "基本算定日数"
        ws["E42"] = "住居外利用"
        ws["C45"] = "入院"
        ws["G45"] = "○"

    for name, day_count, status, status_days, outside_use in users:
        fill_sheet(wb.create_sheet(name), day_count, status, status_days, outside_use, name)

    if with_template:
        # 実ファイルのテンプレートシート。31日分が埋まっていて I2 は空欄。
        fill_sheet(wb.create_sheet("実績記録票"), 31, "", 0, 0)
        wb.create_sheet("延べ日数")
        wb.create_sheet("算定項目リスト")
        wb.create_sheet("重度障害者支援加算要件")

    wb.save(path)


def build_ss_workbook(path: Path, users, with_template=True):
    wb = Workbook()
    wb.remove(wb.active)

    for name, count, normal, combined, band in users:
        ws = wb.create_sheet(name)
        ws["B3"] = "短期入所　１４日/月"
        ws["K3"] = band
        ws[f"C{extract.SS_TOTAL_ROW}"] = count

        row = extract.SS_DATA_ROW_START
        for _ in range(normal):
            ws.cell(row=row, column=4, value="通常利用")
            row += 1
        for _ in range(combined):
            ws.cell(row=row, column=4, value="併給利用")
            row += 1

    if with_template:
        ws = wb.create_sheet("算定項目リスト")
        # 選択肢一覧。列全体を走査すると併給が1件多く数えられる。
        ws["D2"] = "通常利用"
        ws["D3"] = "併給利用"
        ws["B3"] = "サービス提供の状況"
        wb.create_sheet("実績記録票")
        wb.create_sheet("延べ日数")

    wb.save(path)


class EndToEndTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.base = self.tmp / "保存先"
        self.out = self.tmp / "out"
        self.base.mkdir()

        facility = self.base / "RASIELテスト施設⑤　★" / "2026年4月"
        facility.mkdir(parents=True)

        build_gh_workbook(
            facility / "実績記録表1F.xlsx",
            [("山田太郎", 30, "", 0, 5), ("鈴木花子", 0, "入院", 30, 0)],
        )
        build_gh_workbook(
            facility / "実績記録表2Ｆ.xlsx",   # 半角2＋全角Ｆ
            [("佐藤一郎", 30, "", 0, 3)],
        )
        build_ss_workbook(
            facility / "短期実績.xlsx",
            [("田中次郎", 8, 4, 4, 3)],
        )

        excluded = self.base / "99.テストフォルダ　※担当以外使用禁止"
        excluded.mkdir()
        build_gh_workbook(excluded / "実績記録表1F.xlsx", [("除外太郎", 30, "", 0, 0)])

        self._orig_base = common.BASE_DIR
        self._orig_out = extract.OUTPUT_DIR
        common.BASE_DIR = self.base
        extract.OUTPUT_DIR = self.out
        extract.get_run_mode = lambda: ("month", 2026, 4)

    def tearDown(self):
        common.BASE_DIR = self._orig_base
        extract.OUTPUT_DIR = self._orig_out
        shutil.rmtree(self.tmp, ignore_errors=True)

    def read_output(self):
        import pandas as pd

        files = list(self.out.glob("*.xlsx"))
        self.assertEqual(len(files), 1)

        with pd.ExcelFile(files[0]) as xl:
            return {name: xl.parse(name) for name in xl.sheet_names}

    def test_extraction(self):
        extract.main()
        sheets = self.read_output()

        gh = sheets["GH_利用者別明細"]
        ss = sheets["短期_利用者別明細"]
        skipped = sheets["スキップシート一覧"]

        # A-1: テンプレート/非利用者シートが利用者として入らない
        self.assertEqual(
            sorted(gh["利用者名"].tolist()), ["佐藤一郎", "山田太郎", "鈴木花子"]
        )
        self.assertEqual(ss["利用者名"].tolist(), ["田中次郎"])
        self.assertIn("実績記録票", skipped["シート名"].tolist())

        # B-4: テストフォルダが走査されない
        self.assertNotIn("除外太郎", gh["利用者名"].tolist())

        # C-1: 半角1＋全角Ｆ のファイルも 2F として読める
        self.assertIn("2F", gh["区分"].tolist())

        # A-1: 合計にテンプレートの31日が混ざらない
        self.assertEqual(gh["基本算定日数"].sum(), 60)

        # A-3: 表の下の注記（C45の「入院」/ G45の「○」）を数えない
        self.assertEqual(gh["C列_入院"].sum(), 30)
        self.assertEqual(gh["入院時支援加算回数"].sum(), 0)

        # A-3: 算定項目リストの「併給」を数えない
        self.assertEqual(ss["併給回数"].sum(), 4)
        self.assertEqual(ss["通常利用回数"].sum(), 4)

        # C-5: 支給量の文字列から日数を取り出す
        self.assertEqual(ss["短期支給量(日)"].tolist(), [14])

    def test_addon_summary_counts_only_applicable(self):
        extract.main()
        sheets = self.read_output()

        addon = sheets["GH_加算集計"]
        detail = sheets["GH_加算内訳"]

        # A-2: 「非該当」は該当人数に入らない
        self.assertEqual(addon["重度支援加算(Ⅰ)_該当人数"].sum(), 0)
        self.assertEqual(addon["強行地域移行加算_該当人数"].sum(), 0)

        # 人員配置体制加算は Ⅴ なので該当
        self.assertEqual(addon["人員配置体制加算_該当人数"].sum(), 3)
        self.assertEqual(addon["利用者数"].sum(), 3)

        row = detail[
            (detail["加算項目"] == "人員配置体制加算") & (detail["区分値"] == "Ⅴ")
        ]
        self.assertEqual(row["人数"].iloc[0], 3)

    def test_outside_use(self):
        extract.main()
        sheets = self.read_output()

        gh = sheets["GH_利用者別明細"]
        detail = sheets["GH_住居外利用_利用者別"]
        summary = sheets["GH_住居外利用_施設別"]
        values = sheets["GH_E列値内訳"]

        # E40 が利用者ごとの住居外利用の総数
        by_user = dict(zip(gh["利用者名"], gh["住居外利用"]))
        self.assertEqual(by_user["山田太郎"], 5)
        self.assertEqual(by_user["佐藤一郎"], 3)
        self.assertEqual(by_user["鈴木花子"], 0)

        # 0回の利用者は「誰に何回」の一覧に出さない
        self.assertEqual(sorted(detail["利用者名"].tolist()), ["佐藤一郎", "山田太郎"])

        # 施設ごとの合計回数と該当人数
        self.assertEqual(summary["住居外利用_合計回数"].sum(), 8)
        self.assertEqual(summary["住居外利用のあった人数"].sum(), 2)

        # 合計セルと日次入力数が一致していれば「要確認」は付かない
        self.assertEqual(detail["合計と日次の照合"].tolist(), ["一致", "一致"])

        # E列に実際に入っている値
        self.assertEqual(values["件数"].sum(), 8)

    def test_outside_use_flags_mismatch(self):
        extract.main()
        sheets = self.read_output()
        gh = sheets["GH_利用者別明細"]

        row = gh[gh["利用者名"] == "山田太郎"].iloc[0]
        self.assertEqual(row["住居外利用_日次入力数"], 5)
        self.assertEqual(row["氏名"], "山田太郎")
        self.assertEqual(row["支援区分"], 5)

    def test_calculated_cells_are_highlighted(self):
        from openpyxl import load_workbook

        extract.main()
        path = next(self.out.glob("*.xlsx"))
        wb = load_workbook(path)
        ws = wb["GH_利用者別明細"]

        headers = {str(c.value): c.column for c in ws[1]}
        col = headers["住居外利用"]
        name_col = headers["利用者名"]

        filled = {}
        for row in ws.iter_rows(min_row=2):
            name = ws.cell(row=row[0].row, column=name_col).value
            cell = ws.cell(row=row[0].row, column=col)
            filled[name] = (cell.value, cell.fill.fgColor.rgb)

        # 0ではない住居外利用は塗られる
        self.assertEqual(filled["山田太郎"][0], 5)
        self.assertIn(extract.HIGHLIGHT_FILL_COLOR, str(filled["山田太郎"][1]))
        self.assertEqual(filled["佐藤一郎"][0], 3)
        self.assertIn(extract.HIGHLIGHT_FILL_COLOR, str(filled["佐藤一郎"][1]))

        # 0 は塗らない
        self.assertEqual(filled["鈴木花子"][0], 0)
        self.assertNotIn(extract.HIGHLIGHT_FILL_COLOR, str(filled["鈴木花子"][1]))

        # 対象外の列（基本算定日数）は 0 でなくても塗らない
        base_col = headers["基本算定日数"]
        base_cell = ws.cell(row=2, column=base_col)
        self.assertNotIn(extract.HIGHLIGHT_FILL_COLOR, str(base_cell.fill.fgColor.rgb))

        wb.close()

    def test_outside_use_rate(self):
        from openpyxl import load_workbook

        extract.main()
        sheets = self.read_output()

        gh = sheets["GH_利用者別明細"]
        rates = dict(zip(gh["利用者名"], gh["住居外利用率"]))

        # 山田太郎: 住居外利用5 / 基本算定日数30
        self.assertAlmostEqual(rates["山田太郎"], 5 / 30)
        self.assertAlmostEqual(rates["佐藤一郎"], 3 / 30)

        # 基本算定日数が0（入院中）の人は空欄。0% にはしない
        self.assertTrue(pd.isna(rates["鈴木花子"]))

        # 住居外利用の右隣に置く
        cols = list(gh.columns)
        self.assertEqual(cols[cols.index("住居外利用") + 1], "住居外利用率")

        # 施設別は 合計÷合計
        facility = sheets["GH_施設別集計"]
        self.assertAlmostEqual(facility["住居外利用率"].iloc[0], 8 / 60)

        # Excel 上でパーセント表示になっている
        wb = load_workbook(next(self.out.glob("*.xlsx")))
        ws = wb["GH_利用者別明細"]
        headers = {str(c.value): c.column for c in ws[1]}
        cell = ws.cell(row=2, column=headers["住居外利用率"])
        self.assertEqual(cell.number_format, extract.PERCENT_FORMAT)
        wb.close()

    def test_should_highlight(self):
        self.assertTrue(extract.should_highlight(3))
        self.assertTrue(extract.should_highlight("Ⅱ"))
        self.assertTrue(extract.should_highlight("該当"))
        self.assertFalse(extract.should_highlight(0))
        self.assertFalse(extract.should_highlight(None))
        self.assertFalse(extract.should_highlight("非該当"))
        self.assertFalse(extract.should_highlight(""))

    def test_error_columns_are_stable(self):
        extract.main()
        sheets = self.read_output()

        errors = sheets["エラー一覧"]
        self.assertEqual(list(errors.columns), extract.ERROR_COLUMNS)


class StaleFileNameTest(unittest.TestCase):
    """7月のフォルダに「2026年6月」という名前のファイルが入っているケース。

    実運用（和歌山）で、ファイル名の年月から当月日数を30日と誤判定し、
    31日算定の利用者19人が丸ごと集計から落ちた事故の回帰テスト。
    """

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.base = self.tmp / "保存先"
        self.out = self.tmp / "out"

        # フォルダは2026年7月、ファイル名は2026年6月のまま
        facility = self.base / "RASIEL和歌山" / "2026年7月"
        facility.mkdir(parents=True)

        build_gh_workbook(
            facility / "【和歌山1F】実績記録票　2026年6月.xlsm",
            [("森本聖也", 31, "", 0, 0), ("浦井浩志", 29, "", 0, 25)],
            with_template=False,
        )
        build_gh_workbook(
            facility / "【和歌山2F】実績記録票　2026年6月.xlsm",
            [("山川満", 31, "", 0, 0)],
            with_template=False,
        )
        build_ss_workbook(
            facility / "【和歌山 短期】実績記録票　2026年6月.xlsm",
            [("慈幸裕矢", 5, 1, 4, 3)],
            with_template=False,
        )

        self._orig_base = common.BASE_DIR
        self._orig_out = extract.OUTPUT_DIR
        common.BASE_DIR = self.base
        extract.OUTPUT_DIR = self.out
        extract.get_run_mode = lambda: ("month", 2026, 7)

    def tearDown(self):
        common.BASE_DIR = self._orig_base
        extract.OUTPUT_DIR = self._orig_out
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_31day_users_are_not_dropped(self):
        extract.main()

        with pd.ExcelFile(next(self.out.glob("*.xlsx"))) as xl:
            gh = xl.parse("GH_利用者別明細")
            errors = xl.parse("エラー一覧")

        # 7月は31日なので、31日算定の利用者を落としてはいけない
        self.assertEqual(
            sorted(gh["利用者名"].tolist()), ["山川満", "森本聖也", "浦井浩志"]
        )
        self.assertEqual(gh["基本算定日数"].sum(), 91)

        # ただしファイル名の年月が違うことは警告する
        mismatch = errors[errors["内容"].str.contains("指定月と一致しません", na=False)]
        self.assertEqual(len(mismatch), 3)
        self.assertTrue((mismatch["重要度"] == "要対応").all())


class NameComparisonTest(unittest.TestCase):
    def test_full_width_space_is_ignored(self):
        self.assertTrue(extract.same_person_name("清水　雅智", "清水雅智"))
        self.assertTrue(extract.same_person_name("長谷川 勇也", "長谷川勇也"))
        self.assertFalse(extract.same_person_name("清水雅智", "清水雅之"))


class ImpossibleDayCountTest(unittest.TestCase):
    """シート名の除外をすり抜けても、当月日数を超える値は弾く。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.base = self.tmp / "保存先"
        self.out = self.tmp / "out"
        facility = self.base / "RASIELテスト施設⑤　★" / "2026年4月"
        facility.mkdir(parents=True)

        build_gh_workbook(
            facility / "実績記録表1F.xlsx",
            # 「空き部屋」のようなシート名ではなく、名前では判別できないケースを試す
            [("山田太郎", 30, "", 0, 0), ("山田花子", 31, "", 0, 0)],
            with_template=False,
        )
        build_ss_workbook(facility / "短期実績.xlsx", [("田中次郎", 8, 4, 4, 3)],
                          with_template=False)
        build_gh_workbook(facility / "実績記録表2F.xlsx", [("佐藤一郎", 30, "", 0, 0)],
                          with_template=False)

        self._orig_base = common.BASE_DIR
        self._orig_out = extract.OUTPUT_DIR
        common.BASE_DIR = self.base
        extract.OUTPUT_DIR = self.out
        extract.get_run_mode = lambda: ("month", 2026, 4)

    def tearDown(self):
        common.BASE_DIR = self._orig_base
        extract.OUTPUT_DIR = self._orig_out
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_over_month_days_is_excluded(self):
        import pandas as pd

        extract.main()

        with pd.ExcelFile(next(self.out.glob("*.xlsx"))) as xl:
            gh = xl.parse("GH_利用者別明細")
            errors = xl.parse("エラー一覧")

        self.assertNotIn("山田花子", gh["利用者名"].tolist())
        self.assertIn("山田太郎", gh["利用者名"].tolist())
        self.assertTrue(errors["内容"].str.contains("当月日数").any())


if __name__ == "__main__":
    unittest.main()
