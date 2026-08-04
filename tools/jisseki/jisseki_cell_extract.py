"""実績記録表からセルを抽出して集計する。

使い方:
    py jisseki_cell_extract.py

想定レイアウト（様式が変わったら LAYOUT の定数を直す）:
    GH  : 3〜4行目が加算区分、9〜39行目が日次、40行目が合計
    短期 : 3行目が支援区分・支給量、7〜37行目が日次、38行目が合計
"""

from pathlib import Path
from datetime import datetime
import re
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from jisseki_common import (
    OUTPUT_DIR,
    collect_target_files,
    days_in_month,
    is_target_sheet,
)

# ===== レイアウト設定 =====

# 日次データの行数（31日分）。合計行から逆算して開始行を決める。
DAY_ROW_COUNT = 31

GH_TOTAL_ROW = 40
GH_DATA_ROW_START = GH_TOTAL_ROW - DAY_ROW_COUNT   # 9
GH_DATA_ROW_END = GH_TOTAL_ROW - 1                 # 39

SS_TOTAL_ROW = 38
SS_DATA_ROW_START = SS_TOTAL_ROW - DAY_ROW_COUNT   # 7
SS_DATA_ROW_END = SS_TOTAL_ROW - 1                 # 37

# 利用者名（支給決定障害者等氏名）のセル。空欄のシートは利用者シートではないと判定する。
# GH は実ファイルで I2 と確認済み。短期は未確認のため未設定。
GH_USER_NAME_CELL = "I2"
SS_USER_NAME_CELL = None

# 支援区分のセル（GH は O2 が見出し、P2 が値）
GH_BAND_CELL = "P2"

# 短期の支援区分セル。様式が変わって位置が動いた場合はここを直す。
SS_BAND_CELL = "K3"
SS_SUPPLY_CELL = "B3"

# 当月日数を超える算定日数のシートはテンプレート等とみなして除外する
EXCLUDE_IMPOSSIBLE_DAY_COUNT = True

# ===== 住居外利用（E列）の設定 =====

# E40 がそのシートの住居外利用の総数。C40 や F40 と同じ合計行にある。
GH_OUTSIDE_USE_COLUMN = "E"

# 合計セルが数式のまま保存されていて値が取れないと 0 になってしまうため、
# 日次の入力数を別に数えて突き合わせられるようにする。
GH_OUTSIDE_USE_CROSS_CHECK = True

# 日次の入力値を一覧に出す列。様式が変わって表記が変わったときに気づけるようにする。
# （○前提で数えていたために入院時支援・帰宅時支援が全社0件になった経緯がある）
GH_VALUE_AUDIT_COLUMNS = {
    "E": "住居外利用",
    "G": "入院時支援加算",
    "H": "帰宅時支援加算",
}

# ===== 項目設定 =====

GH_STATUS_ITEMS = [
    "体験①",
    "体験②",
    "体験③",
    "体験④",
    "本入居",
    "退去",
    "（退去）",
    "入院開始",
    "入院",
    "退院",
    "入院→退去",
    "入院→外泊",
    "入院→帰所→外泊",
    "入院開始→退去",
    "外泊開始",
    "外泊",
    "外泊終了",
    "外泊→退去",
    "外泊→入院",
    "実費のみ",
]

GH_BASIC_SUM_COLUMNS = [
    "基本算定日数",
    "住居外利用",
    "夜勤加配",
    "医療連携",
    "入院時支援加算回数",
    "帰宅時支援加算回数",
    "精神障害地域移行加算",
    "地域生活移行支援加算",
    "自立生活支援加算(Ⅱ)",
    "自立生活支援加算(Ⅲ)",
    "集中的支援加算(Ⅰ)",
    "集中的支援加算(Ⅱ)",
]

GH_ADDON_FLAG_COLUMNS = [
    "重度支援加算(Ⅰ)",
    "重度支援加算(Ⅱ)",
    "人員配置体制加算",
    "福祉専門職等配置加算",
    "強行地域移行加算",
    "強行体験利用加算",
    "ピアサポート加算",
    "感染対策向上加算",
]

SS_SUM_COLUMNS = [
    "SS算定回数",
    "通常利用回数",
    "併給回数",
    "緊急短期受入加算(Ⅰ)",
    "集中的支援加算(Ⅰ)",
    "集中的支援加算(Ⅱ)",
    "重度支援加算(Ⅰ)",
    "重度支援加算(Ⅱ)",
    "定員超過特例",
]

# 「非該当」を意味する表記。加算の該当人数を数えるときに 0 として扱う。
NOT_APPLICABLE_TEXTS = {
    "非該当",
    "なし",
    "無",
    "無し",
    "×",
    "✕",
    "－",
    "ー",
    "-",
    "―",
    "0",
    "０",
}

ERROR_COLUMNS = ["重要度", "対象月", "施設名", "区分", "利用者名", "内容"]

# エラー一覧の並び順（要対応を先頭に）
ERROR_LEVEL_ORDER = {"要対応": 0, "確認": 1}

# ===== 算定できているセルの強調表示 =====

# ここに挙げた列で「0ではない／該当している」セルを塗りつぶす。
# 算定できている箇所を目立たせて、算定漏れとの差を見えるようにするため。
HIGHLIGHT_COLUMNS = {
    "入院時支援加算回数",
    "帰宅時支援加算回数",
    "精神障害地域移行加算",
    "地域生活移行支援加算",
    "自立生活支援加算(Ⅱ)",
    "自立生活支援加算(Ⅲ)",
    "集中的支援加算(Ⅰ)",
    "集中的支援加算(Ⅱ)",
    "住居外利用",
    "緊急短期受入加算(Ⅰ)",
    "重度支援加算(Ⅰ)",
    "重度支援加算(Ⅱ)",
    "定員超過特例",
    "強行地域移行加算",
    "強行体験利用加算",
    "ピアサポート加算",
    "感染対策向上加算",
    "福祉専門職等配置加算",
}

# 末尾がこれで終わる列も対象にする（GH_加算集計 の「◯◯_該当人数」など）
HIGHLIGHT_COLUMN_SUFFIXES = ("_該当人数", "_合計回数")

# 塗りつぶしの色。派手にしたい場合は "FFFF00"（純粋な黄色）にする。
HIGHLIGHT_FILL_COLOR = "FFEB9C"
HIGHLIGHT_FONT_COLOR = "9C6500"

# 「◯◯率」の列の表示形式
PERCENT_FORMAT = "0.0%"

CIRCLE_TEXTS = {"○", "〇", "◯", "●"}

_ZEN_TO_HAN = str.maketrans("０１２３４５６７８９", "0123456789")


# ===== 小さなユーティリティ =====


def make_error(target_month_label, facility_name, file_type, user_name, message,
               level="要対応") -> dict:
    """エラー行のキーを常に同じにする（列順・欠損を安定させるため）。"""
    return {
        "重要度": level,
        "対象月": target_month_label,
        "施設名": facility_name,
        "区分": file_type,
        "利用者名": user_name,
        "内容": message,
    }


def strip_spaces(text: str) -> str:
    """全角・半角の空白をすべて除く。"""
    return re.sub(r"[\s\u3000]+", "", text)


def same_person_name(name_a: str, name_b: str) -> bool:
    """「清水　雅智」と「清水雅智」を同一人物として扱う。

    シート名は姓名の間に空白を入れる運用があり、氏名セルには入っていないため。
    """
    return strip_spaces(name_a) == strip_spaces(name_b)


def normalize_cell_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def to_number(value, context=None, errors=None):
    """数値化する。変換できなかった場合はエラーとして記録する。

    黙って 0 にすると「本当に0」と「読めなかった」の区別がつかなくなるため。
    """
    if value is None:
        return 0

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        return value

    text = normalize_cell_text(value)

    if text == "":
        return 0

    text = text.translate(_ZEN_TO_HAN).replace(",", "").replace("，", "")

    try:
        return float(text)
    except ValueError:
        if errors is not None and context is not None:
            errors.append(make_error(
                context["対象月"],
                context["施設名"],
                context["区分"],
                context["利用者名"],
                f"数値に変換できません: {value!r}",
            ))
        return 0


def extract_day_count(value):
    """「短期入所　１４日/月」のような文字列から日数を取り出す。"""
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return value

    text = normalize_cell_text(value).translate(_ZEN_TO_HAN)

    match = re.search(r"(\d+)\s*日", text)
    if match:
        return int(match.group(1))

    match = re.search(r"\d+", text)
    if match:
        return int(match.group(0))

    return None


def is_applicable(value) -> int:
    """加算が「該当」しているかを 1/0 で返す。

    「非該当」「なし」も入力済みなので、空欄判定では該当人数にならない。
    """
    text = normalize_cell_text(value)

    if text == "":
        return 0

    if text in NOT_APPLICABLE_TEXTS:
        return 0

    return 1


# ===== シートの読み取り =====


def read_grid(ws, max_row: int):
    """必要な範囲だけを一度に読み込んで {(行, 列): 値} にする。

    セル単位のランダムアクセスを繰り返すより速く、read_only でも動く。
    戻り値: (grid, 実際に読めた行数)
    """
    grid = {}
    row_count = 0

    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1
    ):
        row_count = row_index

        for col_index, value in enumerate(row, start=1):
            if value is not None:
                grid[(row_index, col_index)] = value

    return grid, row_count


_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")


def cell_value(grid, ref: str):
    match = _CELL_REF.match(ref)
    column = column_index_from_string(match.group(1))
    row = int(match.group(2))

    return grid.get((row, column))


def column_texts(grid, column_letter: str, row_start: int, row_end: int):
    """指定列の、日次データ範囲だけのテキストを返す。

    列全体を走査するとヘッダー・凡例・表下の注記まで数えてしまうため、
    必ず行範囲を限定する。
    """
    column = column_index_from_string(column_letter)

    for row in range(row_start, row_end + 1):
        yield normalize_cell_text(grid.get((row, column)))


def count_circle(grid, column_letter, row_start, row_end) -> int:
    return sum(
        1 for text in column_texts(grid, column_letter, row_start, row_end)
        if text in CIRCLE_TEXTS
    )


def count_exact(grid, column_letter, row_start, row_end, target_text) -> int:
    return sum(
        1 for text in column_texts(grid, column_letter, row_start, row_end)
        if text == target_text
    )


def count_contains(grid, column_letter, row_start, row_end, target_text) -> int:
    return sum(
        1 for text in column_texts(grid, column_letter, row_start, row_end)
        if target_text in text
    )


def count_filled(grid, column_letter, row_start, row_end) -> int:
    """日次範囲で何か入力されているセルの数。合計セルとの突き合わせに使う。"""
    return sum(
        1 for text in column_texts(grid, column_letter, row_start, row_end)
        if is_marked(text)
    )


def is_marked(text: str) -> bool:
    """日次セルに「算定した」印が入っているか。

    様式・施設によって ○ / 〇 / ● / ✓ / 1 と表記が分かれるため、
    空欄と 0 以外はすべて算定ありとして扱う。
    """
    if text == "":
        return False

    if text in {"0", "０", "0.0", "-", "－", "―", "ー", "×", "✕"}:
        return False

    return True


def count_marked(grid, column_letter, row_start, row_end) -> int:
    return sum(
        1 for text in column_texts(grid, column_letter, row_start, row_end)
        if is_marked(text)
    )


def collect_column_values(grid, column_letter, row_start, row_end) -> dict:
    """日次範囲に実際に入っている値と、その件数。"""
    values = {}

    for text in column_texts(grid, column_letter, row_start, row_end):
        if text == "":
            continue

        values[text] = values.get(text, 0) + 1

    return values


_STATUS_SET = set(GH_STATUS_ITEMS)


def count_gh_status_items(grid) -> dict:
    result = {f"C列_{item}": 0 for item in GH_STATUS_ITEMS}

    for text in column_texts(grid, "C", GH_DATA_ROW_START, GH_DATA_ROW_END):
        if text in _STATUS_SET:
            result[f"C列_{text}"] += 1

    return result


def extract_gh_sheet(grid, common_info: dict, errors: list) -> dict:
    def num(ref):
        return to_number(cell_value(grid, ref), common_info, errors)

    total = GH_TOTAL_ROW

    return {
        **common_info,

        "氏名": cell_value(grid, GH_USER_NAME_CELL) if GH_USER_NAME_CELL else None,
        "支援区分": cell_value(grid, GH_BAND_CELL) if GH_BAND_CELL else None,

        "基本算定日数": num(f"C{total}"),
        "住居外利用": num(f"{GH_OUTSIDE_USE_COLUMN}{total}"),
        "住居外利用_日次入力数": count_filled(
            grid, GH_OUTSIDE_USE_COLUMN, GH_DATA_ROW_START, GH_DATA_ROW_END
        ),
        "夜勤加配": num(f"F{total}"),
        "医療連携": num(f"I{total}"),
        # G40/H40 に合計セルが無いため日次を数える。
        # 様式によって「○」と「1」が混在するので、空欄・0以外をすべて1回とする。
        "入院時支援加算回数": count_marked(grid, "G", GH_DATA_ROW_START, GH_DATA_ROW_END),
        "帰宅時支援加算回数": count_marked(grid, "H", GH_DATA_ROW_START, GH_DATA_ROW_END),

        "重度支援加算(Ⅰ)": cell_value(grid, "E3"),
        "重度支援加算(Ⅱ)": cell_value(grid, "I3"),
        "人員配置体制加算": cell_value(grid, "M3"),
        "福祉専門職等配置加算": cell_value(grid, "Q3"),
        "強行地域移行加算": cell_value(grid, "E4"),
        "強行体験利用加算": cell_value(grid, "I4"),
        "ピアサポート加算": cell_value(grid, "M4"),
        "感染対策向上加算": cell_value(grid, "Q4"),

        "精神障害地域移行加算": num(f"J{total}"),
        "地域生活移行支援加算": num(f"K{total}"),
        "自立生活支援加算(Ⅱ)": num(f"L{total}"),
        "自立生活支援加算(Ⅲ)": num(f"M{total}"),
        "集中的支援加算(Ⅰ)": num(f"N{total}"),
        "集中的支援加算(Ⅱ)": num(f"O{total}"),

        **count_gh_status_items(grid),
    }


def extract_shortstay_sheet(grid, common_info: dict, errors: list) -> dict:
    def num(ref):
        return to_number(cell_value(grid, ref), common_info, errors)

    total = SS_TOTAL_ROW
    supply_raw = cell_value(grid, SS_SUPPLY_CELL)

    return {
        **common_info,

        "支援区分": cell_value(grid, SS_BAND_CELL),
        "短期支給量": supply_raw,
        "短期支給量(日)": extract_day_count(supply_raw),
        "SS算定回数": num(f"C{total}"),

        "通常利用回数": count_exact(grid, "D", SS_DATA_ROW_START, SS_DATA_ROW_END, "通常利用"),
        "併給回数": count_contains(grid, "D", SS_DATA_ROW_START, SS_DATA_ROW_END, "併給"),

        "緊急短期受入加算(Ⅰ)": num(f"F{total}"),
        "集中的支援加算(Ⅰ)": num(f"G{total}"),
        "集中的支援加算(Ⅱ)": num(f"H{total}"),
        "重度支援加算(Ⅰ)": num(f"I{total}"),
        "重度支援加算(Ⅱ)": num(f"J{total}"),
        "定員超過特例": num(f"K{total}"),
    }


# ===== 集計 =====


def make_gh_facility_summary(df_gh: pd.DataFrame) -> pd.DataFrame:
    if df_gh.empty:
        return pd.DataFrame()

    sum_cols = [col for col in GH_BASIC_SUM_COLUMNS if col in df_gh.columns]
    status_cols = [f"C列_{item}" for item in GH_STATUS_ITEMS if f"C列_{item}" in df_gh.columns]

    summary = df_gh[["施設名"] + sum_cols + status_cols].groupby(
        "施設名", as_index=False
    ).sum(numeric_only=True)

    counts = df_gh.groupby("施設名", as_index=False).size().rename(columns={"size": "利用者数"})
    summary = counts.merge(summary, on="施設名", how="right")

    if {"住居外利用", "基本算定日数"} <= set(summary.columns):
        summary.insert(
            summary.columns.get_loc("住居外利用") + 1,
            "住居外利用率",
            [
                calc_rate(use, base)
                for use, base in zip(summary["住居外利用"], summary["基本算定日数"])
            ],
        )

    return summary


def make_gh_status_summary(df_gh: pd.DataFrame) -> pd.DataFrame:
    if df_gh.empty:
        return pd.DataFrame()

    status_cols = [f"C列_{item}" for item in GH_STATUS_ITEMS if f"C列_{item}" in df_gh.columns]

    if not status_cols:
        return pd.DataFrame()

    summary = df_gh[["施設名"] + status_cols].groupby("施設名", as_index=False).sum(numeric_only=True)

    return summary.rename(columns={f"C列_{item}": item for item in GH_STATUS_ITEMS})


def make_gh_addon_summary(df_gh: pd.DataFrame) -> pd.DataFrame:
    """加算の「該当人数」を施設ごとに集計する。

    「非該当」「なし」は該当に含めない。
    """
    if df_gh.empty:
        return pd.DataFrame()

    rows = []

    for facility_name, group in df_gh.groupby("施設名"):
        row = {"施設名": facility_name, "利用者数": len(group)}

        for col in GH_ADDON_FLAG_COLUMNS:
            if col in group.columns:
                row[f"{col}_該当人数"] = int(group[col].apply(is_applicable).sum())

        rows.append(row)

    return pd.DataFrame(rows)


def make_gh_addon_detail(df_gh: pd.DataFrame) -> pd.DataFrame:
    """加算の区分値ごとの人数（例: 人員配置体制加算のⅠ/Ⅱ/Ⅴ別人数）。"""
    if df_gh.empty:
        return pd.DataFrame()

    rows = []

    for facility_name, group in df_gh.groupby("施設名"):
        for col in GH_ADDON_FLAG_COLUMNS:
            if col not in group.columns:
                continue

            texts = group[col].apply(normalize_cell_text).replace("", "（空欄）")

            for value, count in texts.value_counts().items():
                rows.append({
                    "施設名": facility_name,
                    "加算項目": col,
                    "区分値": value,
                    "人数": int(count),
                    "該当扱い": "○" if is_applicable(value) else "",
                })

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows).sort_values(["施設名", "加算項目", "区分値"])


def calc_rate(numerator, denominator):
    """割合を返す。分母が0のときは空欄（0%ではない）にする。

    入院などで基本算定日数が0の利用者に 0% と出すと、
    「住居外利用がない人」と区別がつかなくなるため。
    """
    if not denominator:
        return None

    return numerator / denominator


def add_outside_use_rate(df: pd.DataFrame) -> pd.DataFrame:
    """基本算定日数を分母にした住居外利用の割合を追加する。"""
    if df.empty:
        return df

    if not {"住居外利用", "基本算定日数"} <= set(df.columns):
        return df

    df = df.copy()
    df["住居外利用率"] = [
        calc_rate(use, base)
        for use, base in zip(df["住居外利用"], df["基本算定日数"])
    ]

    return df


def make_outside_use_detail(df_gh: pd.DataFrame) -> pd.DataFrame:
    """住居外利用が誰に何回ついているか（0回の利用者は載せない）。"""
    if df_gh.empty or "住居外利用" not in df_gh.columns:
        return pd.DataFrame()

    cols = [c for c in ["施設名", "区分", "利用者名", "基本算定日数", "住居外利用"]
            if c in df_gh.columns]

    if "住居外利用率" in df_gh.columns:
        cols.append("住居外利用率")

    if "住居外利用_日次入力数" in df_gh.columns:
        cols.append("住居外利用_日次入力数")

    detail = df_gh[cols].copy()
    detail = detail[
        (detail["住居外利用"] > 0)
        | (detail.get("住居外利用_日次入力数", 0) > 0)
    ]

    if detail.empty:
        return detail

    if "住居外利用_日次入力数" in detail.columns:
        detail["合計と日次の照合"] = [
            "一致" if total == daily else "要確認"
            for total, daily in zip(detail["住居外利用"], detail["住居外利用_日次入力数"])
        ]

    return detail.sort_values(["施設名", "住居外利用"], ascending=[True, False])


def make_outside_use_summary(df_gh: pd.DataFrame) -> pd.DataFrame:
    """住居外利用が施設ごとに何回あったか。"""
    if df_gh.empty or "住居外利用" not in df_gh.columns:
        return pd.DataFrame()

    grouped = df_gh.groupby("施設名")

    rows = []

    for facility_name, group in grouped:
        total_use = group["住居外利用"].sum()
        total_base = (
            group["基本算定日数"].sum() if "基本算定日数" in group.columns else 0
        )

        rows.append({
            "施設名": facility_name,
            "基本算定日数_合計": total_base,
            "住居外利用_合計回数": total_use,
            "住居外利用率": calc_rate(total_use, total_base),
            "住居外利用のあった人数": int((group["住居外利用"] > 0).sum()),
            "利用者数": len(group),
        })

    return pd.DataFrame(rows).sort_values("住居外利用_合計回数", ascending=False)


VALUE_AUDIT_COLUMNS = ["施設名", "列", "項目", "値", "件数"]


def make_outside_use_values(value_rows: list) -> pd.DataFrame:
    """日次範囲に実際に入っている値の一覧（表記の確認用）。

    「○で数えていたら実際は1だった」といった様式の食い違いに気づくため。
    """
    if not value_rows:
        return pd.DataFrame(columns=VALUE_AUDIT_COLUMNS)

    df = pd.DataFrame(value_rows)

    return (
        df.groupby(["施設名", "列", "項目", "値"], as_index=False)["件数"].sum()
        .sort_values(["列", "施設名", "件数"], ascending=[True, True, False])
    )


def make_ss_facility_summary(df_ss: pd.DataFrame) -> pd.DataFrame:
    if df_ss.empty:
        return pd.DataFrame()

    sum_cols = [col for col in SS_SUM_COLUMNS if col in df_ss.columns]

    summary = df_ss[["施設名"] + sum_cols].groupby("施設名", as_index=False).sum(numeric_only=True)
    counts = df_ss.groupby("施設名", as_index=False).size().rename(columns={"size": "利用者数"})

    return counts.merge(summary, on="施設名", how="right")


def reorder_columns(df: pd.DataFrame, first_cols: list) -> pd.DataFrame:
    if df.empty:
        return df

    existing = [col for col in first_cols if col in df.columns]
    remaining = [col for col in df.columns if col not in existing]

    return df[existing + remaining]


def sort_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    sort_cols = [col for col in ["対象月", "施設名", "区分", "利用者名"] if col in df.columns]

    if sort_cols:
        return df.sort_values(sort_cols)

    return df


# ===== 実行モード =====


def get_run_mode():
    print("")
    print("集計モードを選択してください")
    print("1: 最新ファイルを集計")
    print("2: 月を指定して集計")
    print("")

    while True:
        try:
            mode = input("番号を入力してください（1 または 2）: ").strip()
        except EOFError:
            print("入力が取得できませんでした。処理を中止します。")
            sys.exit(1)

        if mode == "1":
            return "latest", None, None

        if mode == "2":
            return ("month",) + _ask_year_month()

        print("1 または 2 を入力してください。")


def _ask_year_month():
    while True:
        try:
            year_text = input("対象年を入力してください（例：2026）: ").strip()
            month_text = input("対象月を入力してください（例：6）: ").strip()
        except EOFError:
            print("入力が取得できませんでした。処理を中止します。")
            sys.exit(1)

        try:
            target_year = int(year_text)
            target_month = int(month_text)
        except ValueError:
            print("年と月は数字で入力してください。")
            continue

        if not (2000 <= target_year <= 2100):
            print("対象年は2000〜2100で入力してください。")
            continue

        if not (1 <= target_month <= 12):
            print("対象月は1〜12で入力してください。")
            continue

        return target_year, target_month


def create_output_file_path(mode: str, target_year=None, target_month=None) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now_text = datetime.now().strftime("%Y%m%d_%H%M%S")

    if mode == "latest":
        file_name = f"実績記録表_セル抽出結果_最新_{now_text}.xlsx"
    else:
        file_name = f"実績記録表_セル抽出結果_{target_year}年{target_month:02d}月_{now_text}.xlsx"

    return OUTPUT_DIR / file_name


def open_workbook(file_path: Path):
    """速度優先で read_only、失敗したら通常モードで開き直す。"""
    try:
        return load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return load_workbook(file_path, data_only=True, read_only=False)


# ===== 出力の整形 =====


def is_highlight_column(header: str) -> bool:
    if header in HIGHLIGHT_COLUMNS:
        return True

    return header.endswith(HIGHLIGHT_COLUMN_SUFFIXES)


def should_highlight(value) -> bool:
    """算定できている（0ではない／該当している）か。"""
    if value is None:
        return False

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    return bool(is_applicable(value))


def highlight_calculated_cells(ws, highlight_fill, highlight_font):
    """加算の列で、算定できているセルを塗りつぶす。"""
    target_columns = {}

    for cell in ws[1]:
        if cell.value is None:
            continue

        if is_highlight_column(str(cell.value).strip()):
            target_columns[cell.column] = True

    if not target_columns:
        return 0

    count = 0

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column not in target_columns:
                continue

            if should_highlight(cell.value):
                cell.fill = highlight_fill
                cell.font = highlight_font
                count += 1

    return count


def apply_percent_format(ws):
    """「◯◯率」の列をパーセント表示にする。"""
    target_columns = [
        cell.column for cell in ws[1]
        if cell.value is not None and str(cell.value).strip().endswith("率")
    ]

    if not target_columns:
        return

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in target_columns:
                cell.number_format = PERCENT_FORMAT


def format_excel_file(file_path: Path):
    wb = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warn_header_fill = PatternFill("solid", fgColor="F4CCCC")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    warn_sheets = {"エラー一覧", "スキップシート一覧"}

    highlight_fill = PatternFill("solid", fgColor=HIGHLIGHT_FILL_COLOR)
    highlight_font = Font(bold=True, color=HIGHLIGHT_FONT_COLOR)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = warn_header_fill if ws.title in warn_sheets else header_fill

            highlight_calculated_cells(ws, highlight_fill, highlight_font)
            apply_percent_format(ws)

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 35)

    wb.save(file_path)


def save_output(output_file: Path, sheets: dict):
    """出力先が Excel で開かれている場合は別名で保存する。"""
    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except PermissionError:
        alt = output_file.with_name(f"{output_file.stem}_1{output_file.suffix}")
        print(f"出力先を開けませんでした。別名で保存します: {alt.name}")

        with pd.ExcelWriter(alt, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return alt

    return output_file


# ===== メイン =====


def main():
    mode, target_year, target_month = get_run_mode()

    if mode == "latest":
        target_month_label = "最新"
    else:
        target_month_label = f"{target_year}年{target_month}月"

    output_file = create_output_file_path(mode, target_year, target_month)

    gh_results = []
    ss_results = []
    error_results = []
    skipped_sheets = []
    outside_use_values = []

    target_files, file_errors, excluded_folders = collect_target_files(
        mode, target_year, target_month
    )

    for error in file_errors:
        error_results.append(make_error(
            target_month_label,
            error.get("施設名", ""),
            error.get("区分", ""),
            error.get("利用者名", ""),
            error.get("内容", ""),
            level=error.get("重要度", "要対応"),
        ))

    for folder_name in excluded_folders:
        skipped_sheets.append({
            "施設名": folder_name,
            "区分": "",
            "シート名": "",
            "理由": "除外対象フォルダのため走査していません",
        })

    detected_months = set()

    for item in target_files:
        facility_name = item["施設名"]
        file_type = item["区分"]
        file_path = item["ファイル"]
        detected = item["検出年月"]

        if detected:
            detected_months.add(detected)

        print(f"処理中: {facility_name} / {file_type}")

        detected_label = f"{detected[0]}年{detected[1]}月" if detected else ""

        # 当月日数は「利用者が指定した月」を正とする。
        # ファイル名の年月が古いまま（7月のデータが6月名で保存されている等）でも、
        # 実在の利用者を集計から落とさないため。
        if mode == "month":
            month_days = days_in_month(target_year, target_month)
        elif detected:
            month_days = days_in_month(*detected)
        else:
            month_days = 31

        if mode == "month" and detected and detected != (target_year, target_month):
            error_results.append(make_error(
                target_month_label, facility_name, file_type, "",
                f"ファイル名の年月（{detected_label}）が指定月と一致しません。"
                f"中身が対象月のものか確認してください: {file_path.name}",
            ))

        if file_path.suffix.lower() == ".xls":
            error_results.append(make_error(
                target_month_label, facility_name, file_type, "",
                ".xls形式は読み取り対象外です（.xlsx / .xlsm に変換してください）",
            ))
            continue

        try:
            wb = open_workbook(file_path)
        except Exception as e:
            error_results.append(make_error(
                target_month_label, facility_name, file_type, "",
                f"ファイルを開けません: {e}",
            ))
            continue

        try:
            total_row = GH_TOTAL_ROW if file_type in ["1F", "2F"] else SS_TOTAL_ROW

            # チャートシート等を除くため wb.worksheets を使う
            for ws in wb.worksheets:
                sheet_name = ws.title

                if not is_target_sheet(sheet_name):
                    skipped_sheets.append({
                        "施設名": facility_name,
                        "区分": file_type,
                        "シート名": sheet_name,
                        "理由": "利用者シートではないシート名のため除外",
                    })
                    continue

                common_info = {
                    "対象月": target_month_label,
                    "検出年月": detected_label,
                    "施設名": facility_name,
                    "区分": file_type,
                    "利用者名": sheet_name,
                    "ファイル名": file_path.name,
                    "候補ファイル数": item["候補ファイル数"],
                    "最新更新日時": item["最新更新日時"],
                }

                try:
                    grid, row_count = read_grid(ws, total_row)
                except Exception as e:
                    error_results.append(make_error(
                        target_month_label, facility_name, file_type, sheet_name,
                        f"シート読み取りエラー: {e}",
                    ))
                    continue

                if row_count < total_row:
                    skipped_sheets.append({
                        "施設名": facility_name,
                        "区分": file_type,
                        "シート名": sheet_name,
                        "理由": f"想定の合計行（{total_row}行目）が存在しません（{row_count}行）",
                    })
                    error_results.append(make_error(
                        target_month_label, facility_name, file_type, sheet_name,
                        f"様式が想定と異なります（{total_row}行目が無い）。集計から除外しました",
                    ))
                    continue

                name_cell = GH_USER_NAME_CELL if file_type in ["1F", "2F"] else SS_USER_NAME_CELL

                if name_cell and normalize_cell_text(cell_value(grid, name_cell)) == "":
                    skipped_sheets.append({
                        "施設名": facility_name,
                        "区分": file_type,
                        "シート名": sheet_name,
                        "理由": f"利用者名セル({name_cell})が空欄のため除外",
                    })
                    continue

                try:
                    if file_type in ["1F", "2F"]:
                        row = extract_gh_sheet(grid, common_info, error_results)
                        day_count = row["基本算定日数"]
                    else:
                        row = extract_shortstay_sheet(grid, common_info, error_results)
                        day_count = row["SS算定回数"]

                except Exception as e:
                    error_results.append(make_error(
                        target_month_label, facility_name, file_type, sheet_name,
                        f"シート抽出エラー: {e}",
                    ))
                    continue

                if day_count > month_days:
                    message = (
                        f"算定日数({day_count})が当月日数({month_days})を超えています。"
                        "テンプレート等の可能性があります"
                    )

                    if EXCLUDE_IMPOSSIBLE_DAY_COUNT:
                        skipped_sheets.append({
                            "施設名": facility_name,
                            "区分": file_type,
                            "シート名": sheet_name,
                            "理由": message + "（集計から除外）",
                        })
                        error_results.append(make_error(
                            target_month_label, facility_name, file_type, sheet_name,
                            message + "。集計から除外しました",
                        ))
                        continue

                    error_results.append(make_error(
                        target_month_label, facility_name, file_type, sheet_name, message,
                    ))

                if file_type in ["1F", "2F"]:
                    real_name = normalize_cell_text(row.get("氏名"))

                    if real_name and not same_person_name(real_name, sheet_name):
                        error_results.append(make_error(
                            target_month_label, facility_name, file_type, sheet_name,
                            f"シート名と氏名({GH_USER_NAME_CELL})が一致しません: {real_name}",
                            level="確認",
                        ))

                    if GH_OUTSIDE_USE_CROSS_CHECK:
                        for column, label in GH_VALUE_AUDIT_COLUMNS.items():
                            for value, count in collect_column_values(
                                grid, column,
                                GH_DATA_ROW_START, GH_DATA_ROW_END,
                            ).items():
                                outside_use_values.append({
                                    "施設名": facility_name,
                                    "列": column,
                                    "項目": label,
                                    "値": value,
                                    "件数": count,
                                })

                    gh_results.append(row)
                else:
                    if normalize_cell_text(row["支援区分"]) == "":
                        error_results.append(make_error(
                            target_month_label, facility_name, file_type, sheet_name,
                            f"支援区分({SS_BAND_CELL})が空欄です",
                            level="確認",
                        ))

                    ss_results.append(row)

        finally:
            wb.close()

    if mode == "latest" and len(detected_months) > 1:
        months = " / ".join(f"{y}年{m}月" for y, m in sorted(detected_months))
        error_results.append(make_error(
            target_month_label, "", "", "",
            f"最新モードで複数の年月が混在しています: {months}",
        ))

    df_gh = add_outside_use_rate(sort_dataframe(pd.DataFrame(gh_results)))
    df_ss = sort_dataframe(pd.DataFrame(ss_results))
    df_errors = pd.DataFrame(error_results, columns=ERROR_COLUMNS)

    if not df_errors.empty:
        df_errors = df_errors.sort_values(
            by="重要度",
            key=lambda col: col.map(lambda v: ERROR_LEVEL_ORDER.get(v, 9)),
            kind="stable",
        )
    df_skipped = pd.DataFrame(
        skipped_sheets, columns=["施設名", "区分", "シート名", "理由"]
    )

    df_gh = reorder_columns(df_gh, [
        "対象月",
        "検出年月",
        "施設名",
        "区分",
        "利用者名",
        "氏名",
        "支援区分",
        "基本算定日数",
        "住居外利用",
        "住居外利用率",
        "住居外利用_日次入力数",
        "夜勤加配",
        "医療連携",
        "入院時支援加算回数",
        "帰宅時支援加算回数",
        "ファイル名",
        "候補ファイル数",
        "最新更新日時",
    ])

    df_ss = reorder_columns(df_ss, [
        "対象月",
        "検出年月",
        "施設名",
        "区分",
        "利用者名",
        "支援区分",
        "短期支給量",
        "短期支給量(日)",
        "SS算定回数",
        "通常利用回数",
        "併給回数",
        "ファイル名",
        "候補ファイル数",
        "最新更新日時",
    ])

    sheets = {
        "GH_利用者別明細": df_gh,
        "GH_施設別集計": make_gh_facility_summary(df_gh),
        "GH_入退居状況集計": make_gh_status_summary(df_gh),
        "GH_加算集計": make_gh_addon_summary(df_gh),
        "GH_加算内訳": make_gh_addon_detail(df_gh),
        "GH_住居外利用_利用者別": make_outside_use_detail(df_gh),
        "GH_住居外利用_施設別": make_outside_use_summary(df_gh),
        "GH_日次入力値内訳": make_outside_use_values(outside_use_values),
        "短期_利用者別明細": df_ss,
        "短期_施設別集計": make_ss_facility_summary(df_ss),
        "エラー一覧": df_errors,
        "スキップシート一覧": df_skipped,
    }

    output_file = save_output(output_file, sheets)
    format_excel_file(output_file)

    print("")
    print("抽出完了")
    print(f"GH明細: {len(df_gh)}行 / 短期明細: {len(df_ss)}行")
    if "住居外利用" in df_gh.columns:
        print(
            f"住居外利用: 合計{int(df_gh['住居外利用'].sum())}回 / "
            f"該当{int((df_gh['住居外利用'] > 0).sum())}人"
        )

    if df_errors.empty:
        print("エラー・警告: 0件 / スキップシート: 0件")
    else:
        counts = df_errors["重要度"].value_counts()
        print(
            f"エラー・警告: 要対応{counts.get('要対応', 0)}件 / "
            f"確認{counts.get('確認', 0)}件 / "
            f"スキップシート: {len(df_skipped)}件"
        )
    print(f"出力先: {output_file}")


if __name__ == "__main__":
    main()
