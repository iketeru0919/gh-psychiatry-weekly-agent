"""抽出対象になるファイルを事前確認する。

判定ロジックは jisseki_common.py を共有しているので、
このツールで「対象ファイルあり」と出たものは jisseki_cell_extract.py でも読まれる。
"""

from pathlib import Path
from datetime import datetime
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from jisseki_common import BASE_DIR, OUTPUT_DIR, collect_target_files


def get_run_mode():
    print("")
    print("確認モードを選択してください")
    print("1: 最新ファイルを確認")
    print("2: 月を指定して確認")
    print("")

    while True:
        try:
            mode = input("番号を入力してください（1 または 2）: ").strip()
        except EOFError:
            print("入力が取得できませんでした。処理を中止します。")
            sys.exit(1)

        if mode == "1":
            return "latest", None, None

        if mode == "2":
            return ("month",) + _ask_year_month()

        print("1 または 2 を入力してください。")


def _ask_year_month():
    while True:
        try:
            year_text = input("対象年を入力してください（例：2026）: ").strip()
            month_text = input("対象月を入力してください（例：6）: ").strip()
        except EOFError:
            print("入力が取得できませんでした。処理を中止します。")
            sys.exit(1)

        try:
            target_year = int(year_text)
            target_month = int(month_text)
        except ValueError:
            print("年と月は数字で入力してください。")
            continue

        if not (2000 <= target_year <= 2100):
            print("対象年は2000〜2100で入力してください。")
            continue

        if not (1 <= target_month <= 12):
            print("対象月は1〜12で入力してください。")
            continue

        return target_year, target_month


def format_excel_file(file_path: Path):
    wb = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warn_header_fill = PatternFill("solid", fgColor="F4CCCC")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = (
                    warn_header_fill
                    if ws.title in {"確認事項", "未使用ファイル一覧"}
                    else header_fill
                )

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)

    wb.save(file_path)


def main():
    mode, target_year, target_month = get_run_mode()

    if not BASE_DIR.exists():
        print("保存先フォルダが見つかりません。")
        print(BASE_DIR)
        return

    target_files, errors, excluded_folders, unused_files = collect_target_files(
        mode, target_year, target_month
    )

    results = []

    for item in target_files:
        detected = item["検出年月"]

        results.append({
            "施設名": item["施設名"],
            "元フォルダ名": item["元フォルダ名"],
            "区分": item["区分"],
            "最新ファイル名": item["ファイル"].name,
            "検出年月": f"{detected[0]}年{detected[1]}月" if detected else "",
            "最新更新日時": item["最新更新日時"],
            "同区分の候補ファイル数": item["候補ファイル数"],
            "フルパス": str(item["ファイル"]),
        })

    for folder_name in excluded_folders:
        errors.append({
            "施設名": folder_name,
            "区分": "",
            "利用者名": "",
            "内容": "除外対象フォルダのため走査していません",
        })

    df_results = pd.DataFrame(
        results,
        columns=[
            "施設名", "元フォルダ名", "区分", "最新ファイル名", "検出年月",
            "最新更新日時", "同区分の候補ファイル数", "フルパス",
        ],
    )
    df_errors = pd.DataFrame(errors, columns=["施設名", "区分", "利用者名", "内容"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now_text = datetime.now().strftime("%Y%m%d_%H%M%S")

    if mode == "latest":
        output_file = OUTPUT_DIR / f"実績記録表_対象ファイル確認_最新_{now_text}.xlsx"
    else:
        output_file = OUTPUT_DIR / (
            f"実績記録表_対象ファイル確認_{target_year}年{target_month:02d}月_{now_text}.xlsx"
        )

    df_unused = pd.DataFrame(
        unused_files,
        columns=["施設名", "区分", "ファイル名", "理由", "サブフォルダ", "フルパス"],
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_results.to_excel(writer, sheet_name="対象ファイル一覧", index=False)
        df_errors.to_excel(writer, sheet_name="確認事項", index=False)
        df_unused.to_excel(writer, sheet_name="未使用ファイル一覧", index=False)

    format_excel_file(output_file)

    print("")
    print("確認完了")
    print(
        f"対象ファイル: {len(df_results)}件 / 確認事項: {len(df_errors)}件 / "
        f"未使用ファイル: {len(df_unused)}件"
    )
    print(f"出力先: {output_file}")


if __name__ == "__main__":
    main()
