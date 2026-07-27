"""シフトExcel（勤務表・実績）→ 職員名一覧 プロトタイプ抽出器."""
import re
import sys
import unicodedata
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import column_index_from_string as ci

SHEET = "管理用シート"
EXCLUDE_NAME = re.compile(r"タイミー|勤務時間|営業活動|合計|派遣$")
# NFKCで ① → "1" に変換されるため、正規化「後」に数字/英字の連番を落とす
SUFFIX = re.compile(r"[0-9AB]$")


def norm(s):
    """氏名の正規化キー：全半角統一・空白除去・①②等の連番除去・（派遣）除去."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"[（(].*?[）)]", "", s)          # (派遣) など
    s = re.sub(r"[\s　]+", "", s)
    s = SUFFIX.sub("", s)
    return s


@dataclass
class Staff:
    name: str
    shokushu: str
    keitai: str = ""
    hours: float = 0.0          # 総労働時間 (CU)
    daytime: float = 0.0        # 日中勤務時間数 (DE) = 調査対象サービス分
    kansan: float = 0.0         # 日中配置換算 人工 (DL)
    rows: list = field(default_factory=list)
    note: str = ""


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]

    merged = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = v

    def g(r, col):
        c = ci(col)
        v = ws.cell(r, c).value
        if v in (None, ""):
            v = merged.get((r, c))
        return v

    # --- ヘッダー行をアンカー検索（決め打ちしない）
    head = None
    for r in range(1, 30):
        if str(g(r, "B") or "").strip() == "職種" and "氏名" in str(g(r, "M") or ""):
            head = r
            break
    if head is None:
        raise RuntimeError(f"{path}: ヘッダー行（B=職種 / M=氏名）が見つかりません")

    # --- マスタ設定から常勤/非常勤のフォールバック辞書を作る
    fallback = {}
    if "マスタ設定" in wb.sheetnames:
        ms = wb["マスタ設定"]
        mm = {}
        for rng in ms.merged_cells.ranges:
            v = ms.cell(rng.min_row, rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    mm[(r, c)] = v

        def mg(r, col):
            c = ci(col)
            v = ms.cell(r, c).value
            if v in (None, ""):
                v = mm.get((r, c))
            return v

        for r in range(1, ms.max_row + 1):
            nm, kt = mg(r, "AE"), mg(r, "BK")
            if nm and kt in ("常勤", "非常勤"):
                fallback.setdefault(norm(nm), kt)

    # --- 職員行の走査
    people = {}
    skipped = []
    timee = {"hours": 0.0, "daytime": 0.0, "rows": 0, "by_job": {}}

    for r in range(head + 1, head + 200):
        name = g(r, "M")
        shokushu = g(r, "DW")
        if not name or not shokushu:
            continue
        name, shokushu = str(name).strip(), str(shokushu).strip()
        if shokushu not in ("生活支援員", "世話人"):
            continue

        num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
        hours, daytime, kansan = num(g(r, "CU")), num(g(r, "DE")), num(g(r, "DL"))

        if EXCLUDE_NAME.search(name):
            if name.startswith("タイミー"):
                timee["hours"] += hours
                timee["daytime"] += daytime
                timee["rows"] += 1
                acc = timee["by_job"].setdefault(shokushu, [0.0, 0.0])
                acc[0] += hours
                acc[1] += daytime
            skipped.append((r, name, "集計行/タイミー等のため名簿から除外"))
            continue

        keitai = str(g(r, "AE") or "").strip()
        if keitai not in ("常勤", "非常勤"):
            keitai = fallback.get(norm(name), "")

        key = (shokushu, norm(name))
        p = people.get(key)
        if p is None:
            p = people[key] = Staff(name=re.sub(r"[①②③④⑤]$", "", name).strip(),
                                    shokushu=shokushu, keitai=keitai)
        else:
            # 常勤・非常勤が食い違う場合は常勤を優先（週40h側に寄せる運用ルール）
            if keitai == "常勤" and p.keitai != "常勤":
                p.keitai = "常勤"
                p.note = "常勤/非常勤の表記が行によって異なるため常勤側を採用"
            elif keitai and not p.keitai:
                p.keitai = keitai
        p.hours += hours
        p.daytime += daytime
        p.kansan += kansan
        p.rows.append(r)

    # --- シート自身の合計行を拾って照合材料にする
    sheet_totals = {}
    num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
    for r in range(head + 1, head + 200):
        label = str(g(r, "B") or "")
        for sh in ("生活支援員", "世話人"):
            if label.startswith(f"{sh}の勤務時間"):
                sheet_totals.setdefault(
                    sh, (num(g(r, "CU")), num(g(r, "DE")), num(g(r, "DL"))))

    facility = str(ws["A2"].value or "").strip()
    ym = f"{ws['G1'].value}年{ws['O1'].value}月"
    return facility, ym, list(people.values()), skipped, timee, sheet_totals


def report(path):
    facility, ym, people, skipped, timee, sheet_totals = load(path)
    print("=" * 78)
    print(f"{facility}　{ym}")
    print("=" * 78)
    for sh in ("生活支援員", "世話人"):
        block = [p for p in people if p.shokushu == sh]
        for kt in ("常勤", "非常勤"):
            grp = [p for p in block if p.keitai == kt]
            print(f"\n【{sh}・{kt}】{len(grp)}名")
            for p in grp:
                mark = f"  ※{p.note}" if p.note else ""
                print(f"   {p.name:16} 総労働{p.hours:6.1f}h  日中{p.daytime:6.1f}h  "
                      f"換算{p.kansan:5.2f}  行{p.rows}{mark}")
        unk = [p for p in block if p.keitai not in ("常勤", "非常勤")]
        if unk:
            print(f"\n【{sh}・区分不明】{len(unk)}名 ← 要確認")
            for p in unk:
                print(f"   {p.name:16} 行{p.rows}")
        mine = (sum(p.hours for p in block), sum(p.daytime for p in block))
        print(f"\n  ── {sh} 小計: {len(block)}名 / 総労働{mine[0]:.1f}h "
              f"/ 日中{mine[1]:.1f}h / 換算{sum(p.kansan for p in block):.2f}")
        if sh in sheet_totals:
            cu, de, _ = sheet_totals[sh]
            # タイミー分はシートの合計には含まれるので足し戻して照合する
            got = (mine[0] + timee["by_job"].get(sh, [0.0, 0.0])[0],
                   mine[1] + timee["by_job"].get(sh, [0.0, 0.0])[1])
            ok = abs(got[0] - cu) < 0.51 and abs(got[1] - de) < 0.51
            print(f"     照合: シート合計 総労働{cu:.1f}h/日中{de:.1f}h  vs  "
                  f"抽出{got[0]:.1f}h/{got[1]:.1f}h  → {'一致' if ok else '★不一致（要確認）'}")
    if timee["rows"]:
        print(f"\n[タイミー] {timee['rows']}行 総労働{timee['hours']:.1f}h 日中{timee['daytime']:.1f}h（氏名不明のため名簿除外）")
    print(f"[除外行] {len(skipped)}件")


if __name__ == "__main__":
    for p in sys.argv[1:]:
        report(p)
