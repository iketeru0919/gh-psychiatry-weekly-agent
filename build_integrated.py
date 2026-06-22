"""
内部監査 × 実地指導対応 統合ツール 生成スクリプト
シート構成:
  ① 内部監査チェックリスト  ← 改善版ベース + 実地指導頻度列
  ② 実地指導 準備チェックリスト ← RASIELシート④ + 改善
  ③ 確認資料リスト          ← 改善版そのまま
  ④ 実施・指導履歴          ← RASIELシート①
  ⑤ 自治体別マトリクス      ← RASIELシート⑦
"""
import copy, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC_AUDIT  = '/home/user/gh-psychiatry-weekly-agent/内部監査チェックリスト_改善版.xlsx'
SRC_RASIEL = '/root/.claude/uploads/1d33ad44-761d-53a1-bc4c-9772bc5d1911/999b5dfe-RASIEL_____________2.xlsx'
DST        = '/home/user/gh-psychiatry-weekly-agent/内部監査_実地指導_統合ツール.xlsx'

# ── 色定数 ─────────────────────────────────────────────────────
C_TITLE    = 'FF1F4E79'
C_HEADER   = 'FF2E75B6'
C_S        = 'FFFCE4D6'
C_A        = 'FFE2EFDA'
C_B        = 'FFFFF2CC'
C_NEW      = 'FFF2F2F2'
C_LEGEND   = 'FFDAE3F3'
C_SUMMARY  = 'FFEAF4FB'
C_INSP3    = 'FFFF0000'   # ★★★ 文字色（赤）
C_INSP2    = 'FFBF8F00'   # ★★  文字色（濃黄）
C_INSP1    = 'FF7F7F7F'   # ★   文字色（グレー）
C_ORANGE   = 'FFFCE4D6'
C_GREEN    = 'FFE2EFDA'
C_PURPLE   = 'FFD9D2E9'
C_BLUE_LT  = 'FFD9EAD3'
WHITE      = 'FFFFFFFF'
BLACK      = 'FF000000'
GRAY_BD    = 'FFB0B0B0'

PRIO_COLOR = {'S': C_S, 'A': C_A, 'B': C_B}

def fill(rgb): return PatternFill('solid', fgColor=rgb)
def fnt(bold=False, color=BLACK, size=10, name='游ゴシック'):
    return Font(bold=bold, color=color, size=size, name=name)
def aln(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bd():
    s = Side(style='thin', color=GRAY_BD)
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value='', bold=False, bg=None, h='left', v='center',
        wrap=True, sz=10, fc=BLACK):
    c = ws.cell(row=row, column=col, value=value)
    c.font = fnt(bold, fc, sz)
    if bg: c.fill = fill(bg)
    c.alignment = aln(h, v, wrap)
    c.border = bd()
    return c

# ── 内部監査 × 実地指導 対応マップ ──────────────────────────────
# (内部監査項目の一部キーワード) → 実地指導頻度
INSP_MAP = {
    '管理者の配置':      '★★★',
    '管理者の雇用契約書': '★★★',
    '管理者の勤務実態':   '★★★',
    '勤務形態一覧表':     '★★★',
    'サビ管の専従':       '★★★',
    '退職者・不在者':     '★★★',
    '所定労働時間超過':   '★★★',
    '生活支援員':        '★★★',
    '世話人':           '★★★',
    'サビ管の配置数':     '★★★',
    '夜勤実績':          '★★★',
    '体制等状況一覧表':   '★★★',
    '算定加算と届出':     '★★★',
    '算定要件':          '★★★',
    '減算対象':          '★★★',
    '大規模住居':        '★★★',
    '日中所在':          '★★★',
    '誤請求':           '★★★',
    '個別支援計画未作成': '★★★',
    '食材料費':          '★★',
    '光熱水費':          '★★',
    '費用徴収の書面同意': '★★★',
    '領収証':           '★★★',
    'BCP研修':          '★★★',
    '感染予防委員会':     '★★★',
    '虐待防止':          '★★★',
    '身体拘束':          '★★★',
    '契約書・重説':       '★★★',
    '短期入所重説':       '★★★',
    '医療連携':          '★★',
    '入退居記録':        '★★',
    'SS支給量':          '★★',
    '秘密保持誓約書':     '★★★',
    '職務専従':          '★★★',
    '居室面積':          '★★',
    'ユニット定員':       '★★',
    'SS居室':           '★★',
    '掲示':             '★★',
    '記録の5年保存':      '★★',
    '地域連携推進会議':   '★★',
    'WAM NET':          '★★',
    '年間研修計画':       '★★★',
    '個人情報保管':       '★★',
    'サテライト型':       '★★',
    '定員類型':          '★★',
    '処遇改善加算':       '★★★',
    '管理者の知識':       '★★',
    '自立生活支援加算':   '★★',
    'SS入退所':          '★★',
    'サービス提供証明書': '★',
    '苦情記録':          '★★★',
    '人員欠如':          '★★★',
    '事故報告書':        '★★★',
    '緊急連絡体制':       '★★★',
    '健康診断':          '★★★',
    '退職者の個人情報':   '★★',
    '2024年改定加算':    '★★',
    '緊急短期入所加算':   '★★',
    '重要事項説明書の改訂': '★★★',
    '前回監査':          '—',
    '消防設備':          '★★★',
}

def get_insp_freq(item_name):
    for kw, freq in INSP_MAP.items():
        if kw in str(item_name):
            return freq
    return '★'

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_sheet1_audit(wb_dst, wb_audit):
    """① 内部監査チェックリスト"""
    ws_src = wb_audit['内部監査チェックリスト']
    ws = wb_dst.create_sheet('①内部監査チェックリスト')

    # ── 行1 タイトル ──────────────────────────────────────────
    ws.merge_cells('A1:N1')
    c = ws.cell(row=1, column=1,
        value='内部監査チェックリスト　【施設名・監査日・サービス種別を入力して使用】')
    c.font = fnt(True, WHITE, 13); c.fill = fill(C_TITLE)
    c.alignment = aln('center', 'center'); ws.row_dimensions[1].height = 22

    # ── 行2〜4 入力フィールド ─────────────────────────────────
    def info_row(r, label_a, label_h):
        ws.merge_cells(f'A{r}:D{r}')
        sc(ws, r, 1, label_a, True, C_HEADER, 'center', fc=WHITE)
        ws.merge_cells(f'E{r}:G{r}')
        sc(ws, r, 5, '')
        ws.merge_cells(f'H{r}:I{r}')
        sc(ws, r, 8, label_h, True, C_HEADER, 'center', fc=WHITE)
        ws.merge_cells(f'J{r}:N{r}')
        sc(ws, r, 10, '')
        ws.row_dimensions[r].height = 20

    info_row(2, '施設名', '監査日')
    info_row(3, '監査実施者', '前回監査日')

    ws.merge_cells('A4:D4')
    sc(ws, 4, 1, '前回指摘事項の是正状況', True, C_HEADER, 'center', fc=WHITE)
    ws.merge_cells('E4:N4')
    sc(ws, 4, 5, '□ 全て完了　□ 一部未了（詳細は指摘内容欄参照）　□ 未対応（理由：　　　　　　　　）')
    ws.row_dimensions[4].height = 20

    # ── 行5 凡例 ───────────────────────────────────────────────
    ws.merge_cells('A5:N5')
    legend = ('【判定】OK＝適合　NG＝不適合（是正要）　N/A＝対象外　是正中＝対応中　未＝未確認'
              '　　【実地指導頻度】★★★＝ほぼ全件必須　★★＝多数　★＝一部')
    c = ws.cell(row=5, column=1, value=legend)
    c.font = Font(bold=False, color='FF1F4E79', size=9, name='游ゴシック')
    c.fill = fill(C_LEGEND); c.alignment = aln('left', 'center')
    ws.row_dimensions[5].height = 18

    # ── 行6 列ヘッダー ─────────────────────────────────────────
    headers = ['優先度','サービス','分類','監査項目','確認内容','確認資料',
               '確認方法','判定','実地指導\n頻度','指摘内容','是正期限','担当者','是正確認日','Dropbox\n格納状況']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=ci, value=h)
        c.font = fnt(True, WHITE, 9); c.fill = fill(C_HEADER)
        c.alignment = aln('center', 'center'); c.border = bd()
    ws.row_dimensions[6].height = 28

    # ── 行7以降 データ読み込み＆書き込み ─────────────────────
    SKIP_ROWS = {1,2,3,4,5,6}  # ヘッダー系
    r_dst = 7
    for r_src in range(7, ws_src.max_row + 1):
        row_vals = [ws_src.cell(r_src, c).value for c in range(1, 13)]
        if not any(v is not None for v in row_vals):
            continue
        # サマリーセクション開始行はスキップ
        if str(row_vals[0]).startswith('■'):
            break
        prio = str(row_vals[0]).strip() if row_vals[0] else ''
        bg = PRIO_COLOR.get(prio, WHITE)
        item_name = row_vals[3] or ''
        insp_freq = get_insp_freq(item_name)
        # 実地指導頻度の文字色
        freq_color = (C_INSP3 if insp_freq == '★★★'
                      else C_INSP2 if insp_freq == '★★'
                      else C_INSP1)

        for ci, val in enumerate(row_vals[:8], 1):
            bold = ci in (1, 2, 3, 4)
            c = ws.cell(row=r_dst, column=ci, value=val)
            c.font = fnt(bold, BLACK, 9)
            c.fill = fill(bg)
            c.alignment = aln('center' if ci in (1,2,7,8) else 'left', 'center', True)
            c.border = bd()

        # 列9: 実地指導頻度
        c9 = ws.cell(row=r_dst, column=9, value=insp_freq)
        c9.font = fnt(True, freq_color, 10); c9.fill = fill(C_NEW)
        c9.alignment = aln('center', 'center', False); c9.border = bd()

        # 列10〜14: 空欄
        for ci in range(10, 15):
            c = ws.cell(row=r_dst, column=ci, value='')
            c.fill = fill(C_NEW); c.border = bd()
            c.alignment = aln('center' if ci in (11,12,13,14) else 'left', 'center', True)

        ws.row_dimensions[r_dst].height = 34
        r_dst += 1

    # ── 判定ドロップダウン ─────────────────────────────────────
    last_r = r_dst - 1
    dv = DataValidation(type='list', formula1='"未,OK,NG,N/A,是正中"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = f'H7:H{last_r}'
    ws.add_data_validation(dv)

    # ── Dropbox格納状況ドロップダウン ──────────────────────────
    dv2 = DataValidation(type='list', formula1='"済,未,N/A"',
                         allow_blank=True, showDropDown=False)
    dv2.sqref = f'N7:N{last_r}'
    ws.add_data_validation(dv2)

    # ── サマリー ────────────────────────────────────────────────
    sr = last_r + 2
    ws.merge_cells(f'A{sr}:N{sr}')
    c = ws.cell(row=sr, column=1, value='■ 監査結果サマリー')
    c.font = fnt(True, WHITE, 11); c.fill = fill(C_TITLE)
    c.alignment = aln('left', 'center'); ws.row_dimensions[sr].height = 20
    sr += 1
    sum_heads = ['区分','OK','NG','N/A','是正中','未確認','合計']
    ws.merge_cells(f'H{sr}:N{sr}')
    for ci, h in enumerate(sum_heads, 1):
        c = ws.cell(row=sr, column=ci, value=h)
        c.font = fnt(True, WHITE, 9); c.fill = fill(C_HEADER)
        c.alignment = aln('center'); c.border = bd()
    ws.row_dimensions[sr].height = 18
    for lbl in ['S優先度', 'A優先度', 'B優先度', '合計']:
        sr += 1
        ws.merge_cells(f'H{sr}:N{sr}')
        sc(ws, sr, 1, lbl, True, C_SUMMARY)
        for ci in range(2, 8):
            c = ws.cell(row=sr, column=ci, value='')
            c.fill = fill(C_SUMMARY); c.border = bd(); c.alignment = aln('center')
        ws.row_dimensions[sr].height = 18

    # ── 列幅 ────────────────────────────────────────────────────
    widths = {'A':5.5,'B':8,'C':13,'D':28,'E':48,'F':34,'G':12,'H':8,
              'I':8,'J':26,'K':11,'L':10,'M':12,'N':10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_sheet2_prep(wb_dst, wb_rasiel):
    """② 実地指導 準備チェックリスト"""
    ws_src = wb_rasiel['④準備チェックリスト']
    ws = wb_dst.create_sheet('②実地指導準備チェックリスト')

    # タイトル
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1,
        value='運営指導・実地指導　準備チェックリスト（全22通知統合版）')
    c.font = fnt(True, WHITE, 13); c.fill = fill(C_TITLE)
    c.alignment = aln('center', 'center'); ws.row_dimensions[1].height = 22

    # 入力欄
    ws.merge_cells('A2:C2')
    sc(ws, 2, 1, '事業所名', True, C_HEADER, 'center', fc=WHITE)
    ws.merge_cells('D2:G2')
    sc(ws, 2, 4, ''); ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:C3')
    sc(ws, 3, 1, '指導種別', True, C_HEADER, 'center', fc=WHITE)
    ws.merge_cells('D3:E3')
    sc(ws, 3, 4, '')
    sc(ws, 3, 6, '実施日', True, C_HEADER, 'center', fc=WHITE)
    sc(ws, 3, 7, ''); ws.row_dimensions[3].height = 20

    # 凡例
    ws.merge_cells('A4:G4')
    c = ws.cell(row=4, column=1,
        value='【★★★】全自治体ほぼ必須　【★★】多数の自治体で要求　【★】一部自治体のみ')
    c.font = Font(bold=False, color='FF1F4E79', size=9, name='游ゴシック')
    c.fill = fill(C_LEGEND); c.alignment = aln('left', 'center')
    ws.row_dimensions[4].height = 16

    # 列ヘッダー
    headers = ['チェック', '書類名', '種別', '優先度（22通知）', '備考・準備ポイント', '格納場所', '準備完了日']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=ci, value=h)
        c.font = fnt(True, WHITE, 9); c.fill = fill(C_HEADER)
        c.alignment = aln('center', 'center'); c.border = bd()
    ws.row_dimensions[5].height = 24

    r_dst = 6
    cur_section = ''
    for r_src in range(3, ws_src.max_row + 1):
        vals = [ws_src.cell(r_src, c).value for c in range(1, 6)]
        if not any(v is not None for v in vals):
            continue
        chk, name, typ, prio, note = vals

        # セクションヘッダー（【事前提出】【当日準備】◆ ～）
        if str(chk).startswith('【') or str(chk).startswith('◆'):
            ws.merge_cells(f'A{r_dst}:G{r_dst}')
            label = chk if str(chk).startswith('【') else f'  {chk}'
            c = ws.cell(row=r_dst, column=1, value=label)
            is_major = str(chk).startswith('【')
            c.font = fnt(True, WHITE if is_major else 'FF1F4E79', 10 if is_major else 9)
            c.fill = fill(C_HEADER if is_major else C_LEGEND)
            c.alignment = aln('left', 'center', False); c.border = bd()
            ws.row_dimensions[r_dst].height = 18
            r_dst += 1
            continue

        # 優先度別背景色
        prio_str = str(prio) if prio else ''
        bg = (C_S if '★★★' in prio_str else
              C_A if '★★' in prio_str else C_B)
        freq_color = (C_INSP3 if '★★★' in prio_str else
                      C_INSP2 if '★★' in prio_str else C_INSP1)

        sc(ws, r_dst, 1, '□', False, bg, 'center', 'center', False)
        sc(ws, r_dst, 2, name, False, bg, 'left', 'center', True, 9)
        sc(ws, r_dst, 3, typ, False, bg, 'center', 'center', False, 9)
        c4 = ws.cell(row=r_dst, column=4, value=prio_str)
        c4.font = fnt(True, freq_color, 10); c4.fill = fill(bg)
        c4.alignment = aln('center', 'center', False); c4.border = bd()
        sc(ws, r_dst, 5, note or '', False, bg, 'left', 'center', True, 9)
        sc(ws, r_dst, 6, 'Dropbox', False, C_NEW, 'center', 'center', False, 9)
        sc(ws, r_dst, 7, '', False, C_NEW, 'center', 'center', False, 9)
        ws.row_dimensions[r_dst].height = 22
        r_dst += 1

    # チェック列ドロップダウン
    dv = DataValidation(type='list', formula1='"□,☑,N/A"',
                        allow_blank=True, showDropDown=False)
    dv.sqref = f'A6:A{r_dst-1}'
    ws.add_data_validation(dv)

    widths = {'A':6, 'B':42, 'C':8, 'D':14, 'E':36, 'F':12, 'G':12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_sheet3_docs(wb_dst, wb_audit):
    """③ 確認資料リスト"""
    ws_src = wb_audit['確認資料リスト']
    ws = wb_dst.create_sheet('③確認資料リスト')

    for r in range(1, ws_src.max_row + 1):
        for c_idx in range(1, ws_src.max_column + 1):
            src_cell = ws_src.cell(r, c_idx)
            dst_cell = ws.cell(r, c_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font      = copy.copy(src_cell.font)
                dst_cell.fill      = copy.copy(src_cell.fill)
                dst_cell.border    = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
        ws.row_dimensions[r].height = ws_src.row_dimensions[r].height or 15

    for col, dim in ws_src.column_dimensions.items():
        ws.column_dimensions[col].width = dim.width or 10

    for mc in ws_src.merged_cells.ranges:
        ws.merge_cells(str(mc))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_sheet4_history(wb_dst, wb_rasiel):
    """④ 実施・指導履歴"""
    ws_src = wb_rasiel['①実施一覧']
    ws = wb_dst.create_sheet('④実施・指導履歴')

    # タイトル
    ws.merge_cells('A1:J1')
    c = ws.cell(row=1, column=1, value='運営指導・実地指導　実施履歴管理')
    c.font = fnt(True, WHITE, 13); c.fill = fill(C_TITLE)
    c.alignment = aln('center', 'center'); ws.row_dimensions[1].height = 22

    # 凡例
    ws.merge_cells('A2:J2')
    c = ws.cell(row=2, column=1,
        value='【色分け】黄＝実地指導　緑＝運営指導　橙＝運営指導（監査）　紫＝実地指導+業務管理体制一般検査')
    c.font = Font(bold=False, color='FF1F4E79', size=9, name='游ゴシック')
    c.fill = fill(C_LEGEND); c.alignment = aln('left', 'center')
    ws.row_dimensions[2].height = 16

    # 列ヘッダー
    headers = ['事業所名','自治体','指導種別','対象サービス','実施日',
               '事前提出期限','根拠文書番号','指摘件数','主な指摘内容','是正完了日']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = fnt(True, WHITE, 9); c.fill = fill(C_HEADER)
        c.alignment = aln('center', 'center'); c.border = bd()
    ws.row_dimensions[3].height = 22

    # 指導種別→色マップ
    type_color = {
        '実地指導':                         'FFFFF2CC',
        '運営指導':                         'FFE2EFDA',
        '運営指導（監査）':                  'FFFCE4D6',
        '実地指導・業務管理体制一般検査':      'FFD9D2E9',
        '運営指導・業務管理体制一般検査':      'FFD9EAD3',
    }
    r_dst = 4
    for r_src in range(3, ws_src.max_row + 1):
        vals = [ws_src.cell(r_src, c).value for c in range(1, 8)]
        if not any(v is not None for v in vals):
            continue
        if str(vals[0]).startswith('【') or str(vals[0]).startswith('集計'):
            continue
        name, pref, typ, svc, date, deadline, doc = vals
        bg = type_color.get(str(typ).strip(), WHITE)
        row_data = [name, pref, typ, svc, date, deadline, doc, '', '', '']
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r_dst, column=ci, value=val)
            c.font = fnt(False, BLACK, 9); c.fill = fill(bg)
            c.alignment = aln('center' if ci in (5,6,8,10) else 'left', 'center', True)
            c.border = bd()
        ws.row_dimensions[r_dst].height = 22
        r_dst += 1

    widths = {'A':20,'B':14,'C':18,'D':22,'E':10,'F':12,'G':22,'H':8,'I':30,'J':12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_sheet5_matrix(wb_dst, wb_rasiel):
    """⑤ 自治体別マトリクス"""
    ws_src = wb_rasiel['⑦自治体別比較マトリクス']
    ws = wb_dst.create_sheet('⑤自治体別マトリクス')

    # タイトル
    ws.merge_cells(f'A1:{get_column_letter(ws_src.max_column)}1')
    c = ws.cell(row=1, column=1, value='自治体別　事前提出要求書類マトリクス（全22通知）')
    c.font = fnt(True, WHITE, 13); c.fill = fill(C_TITLE)
    c.alignment = aln('center', 'center'); ws.row_dimensions[1].height = 22

    for r_src in range(1, ws_src.max_row + 1):
        r_dst = r_src + 1  # タイトル行分ずらす
        for c_src in range(1, ws_src.max_column + 1):
            src_cell = ws_src.cell(r_src, c_src)
            dst_cell = ws.cell(r_dst, c_src)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font      = copy.copy(src_cell.font)
                dst_cell.fill      = copy.copy(src_cell.fill)
                dst_cell.border    = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
            else:
                dst_cell.border = bd()
                dst_cell.alignment = aln('center', 'center', True)
        ws.row_dimensions[r_dst].height = ws_src.row_dimensions[r_src].height or 30

    # ヘッダー行（r_dst=2）を強調
    for ci in range(1, ws_src.max_column + 1):
        c = ws.cell(row=2, column=ci)
        if c.value is not None:
            c.font = fnt(True, WHITE, 8)
            c.fill = fill(C_HEADER)
            c.alignment = aln('center', 'center', True)

    ws.column_dimensions['A'].width = 28
    for i in range(2, ws_src.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 7

    # ● セルを色付け（視認性向上）
    for r in range(3, ws_src.max_row + 2):
        for ci in range(2, ws_src.max_column + 1):
            c = ws.cell(row=r, column=ci)
            if c.value == '●':
                c.font = fnt(True, 'FF1F4E79', 11)
                c.fill = fill('FFD9EAD3')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    wb_audit  = openpyxl.load_workbook(SRC_AUDIT)
    wb_rasiel = openpyxl.load_workbook(SRC_RASIEL)
    wb_dst    = openpyxl.Workbook()
    wb_dst.remove(wb_dst.active)  # デフォルトシート削除

    print('① 内部監査チェックリスト 生成中...')
    build_sheet1_audit(wb_dst, wb_audit)

    print('② 実地指導準備チェックリスト 生成中...')
    build_sheet2_prep(wb_dst, wb_rasiel)

    print('③ 確認資料リスト 生成中...')
    build_sheet3_docs(wb_dst, wb_audit)

    print('④ 実施・指導履歴 生成中...')
    build_sheet4_history(wb_dst, wb_rasiel)

    print('⑤ 自治体別マトリクス 生成中...')
    build_sheet5_matrix(wb_dst, wb_rasiel)

    wb_dst.save(DST)
    print(f'\n完了: {DST}')
    print(f'シート数: {len(wb_dst.sheetnames)}')
    for s in wb_dst.sheetnames:
        print(f'  - {s}')

if __name__ == '__main__':
    main()
