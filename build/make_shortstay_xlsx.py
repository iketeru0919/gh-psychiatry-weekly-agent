# -*- coding: utf-8 -*-
"""ショートステイ・体験利用 管理台帳（Excel）を生成する。

設計の要:
  - 入力は「1滞在 = 1行」(01_予約入力) のみ。他シートは全て数式で導出する。
  - 見る側の主役はカレンダー。02_カレンダー(月グリッド)と03_帯表(ガント)で
    「だれが・いつから・いつまで」を視覚的に示す。
  - 部屋数は M_部屋 の行数で決まる。数式は最大4室ぶん用意し、
    未登録の室は空欄を返すので、1室施設・2室施設で同じファイルが使える。
  - 空室・稼働は「その日の夜に泊まるか」で数える(退所日の夜は空き)。
    ただしカレンダーの帯は「利用日(入所日〜退所日)」で描き、退所日は ◀ を付ける。
  - 重い配列数式は 04_月間表 の31行に集約し、カレンダーと帯表はそこを
    MATCH で引くだけにする。日付を全部展開しない。
  - 色分けは「隠しミラー列」に区分/状態コードを持たせ、条件付き書式が
    一定のオフセットでそれを参照する方式。数式と見た目を分離できる。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment
import datetime

OUT = "/home/user/gh-psychiatry-weekly-agent/build/ショートステイ管理台帳.xlsx"

FONT = "Meiryo"
MAX_ROOMS = 4
BK_LAST = 404          # 01_予約入力 のデータ最終行(400件 ≒ 3年半ぶん)
USER_ROWS = 30
BAND_ROWS = 20         # 03_帯表 に出す利用者数
CAT_ROWS = 10
MONTH0 = datetime.date(2026, 4, 1)
CAL_START = 9                 # 02_カレンダー のグリッド開始行
CAL_STRIDE = 8                # 1週ぶんの行数（日付+部屋4+空き+食事+空行）
ADJ_TOP, ADJ_END = 59, 78     # 02_カレンダー 下部「食事の個別調整」表の行範囲

F_INPUT  = PatternFill("solid", fgColor="FFF7DD", bgColor="FFF7DD")
F_AUTO   = PatternFill("solid", fgColor="F2F2F2", bgColor="F2F2F2")
F_HEAD   = PatternFill("solid", fgColor="2F6690", bgColor="2F6690")
F_SUB    = PatternFill("solid", fgColor="DCE6EF", bgColor="DCE6EF")
F_WARN   = PatternFill("solid", fgColor="FBD3D3", bgColor="FBD3D3")
F_WARN2  = PatternFill("solid", fgColor="FCE4C8", bgColor="FCE4C8")
F_OK     = PatternFill("solid", fgColor="D9EDDC", bgColor="D9EDDC")
F_TENT   = PatternFill("solid", fgColor="EDE3F5", bgColor="EDE3F5")
F_WKND   = PatternFill("solid", fgColor="E8EEF4", bgColor="E8EEF4")
F_WHITE  = PatternFill("solid", fgColor="FFFFFF", bgColor="FFFFFF")
F_OUT    = PatternFill("solid", fgColor="FAFAFA", bgColor="FAFAFA")   # 月外の日
# 区分の色（M_区分 の1〜4行目に対応）。元アプリの配色に合わせる。
CAT_FILLS = [PatternFill("solid", fgColor="CFE0F3", bgColor="CFE0F3"),   # SS   青
             PatternFill("solid", fgColor="D3EBD8", bgColor="D3EBD8"),   # 無料 緑
             PatternFill("solid", fgColor="E5D8F0", bgColor="E5D8F0"),   # 契約 紫
             PatternFill("solid", fgColor="FAE0C3", bgColor="FAE0C3")]   # 予備 橙
CAT_FONTS = ["1B4F72", "1D6F42", "5B2C8D", "9C4A06"]   # 区分ごとの文字色（濃色）
F_BAND   = PatternFill("solid", fgColor="9DC3E6", bgColor="9DC3E6")      # 帯表の確定バー
F_BANDT  = PatternFill("solid", fgColor="E4E4EC", bgColor="E4E4EC")      # 帯表の仮バー
F_TENT2  = PatternFill("solid", fgColor="EFEFEF", bgColor="EFEFEF")      # カレンダーの仮予約

C_HEAD  = Font(name=FONT, size=10, bold=True, color="FFFFFF")
C_SUB   = Font(name=FONT, size=10, bold=True, color="1F3B52")
C_BODY  = Font(name=FONT, size=10)
C_AUTO  = Font(name=FONT, size=10, color="666666", italic=True)
C_TITLE = Font(name=FONT, size=13, bold=True, color="1F3B52")
C_NOTE  = Font(name=FONT, size=9, color="666666")
C_BOLD  = Font(name=FONT, size=10, bold=True)
C_DAY   = Font(name=FONT, size=11, bold=True, color="333333")
C_SUN   = Font(name=FONT, size=11, bold=True, color="C0392B")
C_SAT   = Font(name=FONT, size=11, bold=True, color="2471A3")
C_LANE  = Font(name=FONT, size=9)
C_TENTF = Font(name=FONT, size=9, color="777777", italic=True)
C_BANDF = Font(name=FONT, size=9, color="3A6EA5")

THIN = Side(style="thin", color="BFBFBF")
HAIR = Side(style="hair", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_L = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws

def head_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = C_HEAD; c.fill = F_HEAD; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = C_TITLE
    if sub:
        ws["A2"] = sub; ws["A2"].font = C_NOTE

def label_input(ws, row, label, value, note="", fmt=None, col=1):
    L = get_column_letter(col); V = get_column_letter(col + 1); N = get_column_letter(col + 2)
    ws[f"{L}{row}"] = label; ws[f"{L}{row}"].font = C_SUB
    ws[f"{L}{row}"].fill = F_SUB; ws[f"{L}{row}"].border = BORDER
    c = ws[f"{V}{row}"]; c.value = value; c.fill = F_INPUT; c.font = C_BODY; c.border = BORDER
    if fmt: c.number_format = fmt
    if note:
        ws[f"{N}{row}"] = note; ws[f"{N}{row}"].font = C_NOTE

def label_auto(ws, row, label, formula, note="", fmt=None, col=1):
    L = get_column_letter(col); V = get_column_letter(col + 1); N = get_column_letter(col + 2)
    ws[f"{L}{row}"] = label; ws[f"{L}{row}"].font = C_SUB
    ws[f"{L}{row}"].fill = F_SUB; ws[f"{L}{row}"].border = BORDER
    c = ws[f"{V}{row}"]; c.value = formula; c.fill = F_AUTO; c.font = C_AUTO; c.border = BORDER
    if fmt: c.number_format = fmt
    if note:
        ws[f"{N}{row}"] = note; ws[f"{N}{row}"].font = C_NOTE

# ── シート参照（全てクォート） ────────────────────────────────────
BK   = "'01_予約入力'"
CAL  = "'02_カレンダー'"
BAND = "'03_帯表'"
MON  = "'04_月間表'"
MSET = "'M_施設設定'"
MROOM= "'M_部屋'"
MUSER= "'利用者マスタ'"
MCAT = "'M_区分'"

R = lambda col: f"{BK}!${col}$5:${col}${BK_LAST}"
# 01_予約入力 の列（入力：B〜F,H,I,K,N〜U ／ 自動：A,G,J,L,M,V,W,X ／ 作業:Y〜AB）
NAME_, CAT_, ROOM_ = R("B"), R("C"), R("D")
IN_, OUT_          = R("G"), R("J")      # 自動計算された入所日・退所予定日
BF_, LF_, DF_      = R("L"), R("M"), R("N")   # 食事の指定（空欄=標準 / ○=全日 / ×=なし）
ST_                = R("Q")              # 予約状態
NOTE_              = R("V")              # 備考
ALW_               = R("AA")             # 支給量対象日数
SORTK_             = R("AC")             # 当月の並び順キー（帯表の明細用）
TIN_, TOUT_        = R("AD"), R("AE")    # 時刻の表示用文字列
INN_, OUTN_        = R("AF"), R("AG")    # 日付の数値版（空欄は0）。四則演算はこちらを使う
VALID = f'({IN_}<>"")*({ST_}<>"キャンセル")'
TENT  = f'(({ST_}="仮予約")+({ST_}="調整中"))'
ROWID = f'(ROW({IN_})-4)'

CAP   = f"{MSET}!$B$6"
UROW  = f"{MUSER}!$A$4:$A$60"
UALW  = f"{MUSER}!$B$4:$B$60"
CROW  = f"{MCAT}!$A$4:$A$20"
CCNT  = f"{MCAT}!$B$4:$B$20"
RROW  = f"{MROOM}!$A$4:$A$20"

M1 = f"{CAL}!$B$4"        # 対象月の月初（唯一の基準）
YM = f'YEAR({CAL}!$B$4)&"年"&MONTH({CAL}!$B$4)&"月"'   # 「2026年8月」
SHOWTIME = f"{CAL}!$B$5"  # 「時刻を表示」の切替（○ で表示）


# ══════════════════════════════════════════════════════════════════
# 00_使い方
# ══════════════════════════════════════════════════════════════════
ws = sheet("00_使い方")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 100
title(ws, "ショートステイ・体験利用 管理台帳",
      "入力するのは 01_予約入力 と、M_ で始まるマスタだけです。02〜09 は全て自動計算です。")
rows = [
    ("SEC", "画面の役割", ""),
    ("02_カレンダー", "★毎日見るのはここ。ふつうの月カレンダーの形で、日ごとに「どの部屋に誰がいるか」と"
                 "入室・退室の時刻を表示します（「時刻を表示」を空欄にすると時刻を消せます）。"
                 "連泊は ▶入所日 → 継続 → 退所日◀ と続けて出るので、いつからいつまでかが目で追えます。"
                 "色は区分ごと（SS・無料・契約）。仮予約はグレーの斜体です。★対象月の変更もこのシートで行います。"),
    ("03_帯表", "1人1行・横軸が1〜31日のガントチャートです。誰がどの期間を押さえているかを1か月分まとめて俯瞰できます。"
              "▶入所 ■連泊 ◀退所 ◆日帰り。白抜き（▷□◁◇）は仮予約です。"
              "★その下に「当月の予約明細」があり、01_予約入力 の内容（部屋・区分・入退所の日時・泊数・状態・備考）が"
              "入所日順に自動で並びます。"),
    ("04_月間表", "日付を縦に並べた一覧。空室数や食数の根拠を数字で確認したいときに使います。"
              "カレンダーと帯表はこのシートを引いているので、内容は必ず一致します。"),
    ("05_空室照会", "「この期間空いてますか」と電話で聞かれたとき用。日付を2つ入れるだけです。"),
    ("06_食数表", "厨房へ渡す日別の朝・昼・夕の食数。★個別の調整は 02_カレンダー の下部で日付ごとに入れます。"),
    ("07_月次集計", "上司報告・請求突合用。期間を入れると★稼働率・区分別・利用者別が出ます。"
                 "月単位でも年度単位でも指定できます。"),
    ("08_支給量管理", "受給者証の月あたり上限に対する残日数。"),
    ("09_FAX空き表", "★このまま印刷してFAX送信できる1枚ものの用紙です。送信票と空き状況カレンダーが一体に"
                 "なっており、送信先・送信元・挨拶文・凡例・注意書き・連絡事項欄まで入っています。"
                 "利用者名は載りません。A4縦1枚に収まるよう設定済みです。"),
    ("SEC", "★ 日付の入れ方（年月は最初の1回だけ）", ""),
    ("年月は1回だけ", "01_予約入力 の「年月」列に、最初の行で 2026/8 のように一度だけ入れてください。"
                "以降の行は空欄のままでかまいません。上の行の年月を自動で引き継ぎます。"),
    ("あとは日にちだけ", "「入所(日)」「退所(日)」には日にちの数字（1〜31）だけを入れます。"
                  "その右の「入所日(自動)」「退所予定日(自動)」に、年月と組み合わせた日付が自動で入ります。"),
    ("月をまたぐとき", "退所の日にちが入所より前になる場合は、自動で「翌月のその日」と判断します"
                 "（例: 入所31 → 退所1 なら翌月1日）。"),
    ("月が変わったら", "新しい月の最初の行に、その月の年月（例 2026/9）を入れてください。そこから下が新しい月になります。"),
    ("直接日付でも可", "1か月を超える滞在など、特別なときは日にち欄に日付そのもの（2026/9/3 など）を入れても動きます。"),
    ("★時刻は「時」だけ", "「入所(時)」「退所(時)」には時の数字だけを入れてください（16 と入れれば 16:00）。"
                  "9時なら 9 だけで 09:00 になります。分は入力不要です。"),
    ("分が必要なとき", "16:30 のように「時:分」で入れれば、そのまま分まで表示されます。"),
    ("時刻は空欄でも可", "時刻を入れなくても予約は成立します。カレンダーに時刻が出ないだけです。"),
    ("★注意", "行の並べ替え（ソート）はしないでください。年月の引き継ぎが崩れます。"
             "絞り込みたいときは、並べ替えではなくフィルタを使ってください。"),
    ("SEC", "はじめに設定すること", ""),
    ("① M_施設設定", "施設名・棟・標準の入所/退所時刻・住所・TEL/FAX を入れます。"),
    ("② M_部屋", "部屋名を1行1室で登録します。★ここの行数が、そのまま定員（部屋数）になります。"),
    ("③ 利用者マスタ", "氏名と支給量（受給者証の月あたり上限日数）。支給量が空欄の方は残日数を管理しません。"),
    ("⑤ M_FAX送付先", "相談支援事業所などFAXの送り先を登録します。09_FAX空き表 の宛先が一覧から選べるようになり、"
                  "番号の打ち間違いによる誤送信を防げます。"),
    ("④ M_区分・M_送迎", "区分（SS・無料・契約など）と、その区分を支給量に数えるかどうか。"
                     "★区分名はカレンダーの色分けに使います。M_区分 の上から4つまでが色に対応します。"),
    ("SEC", "★ 部屋数が施設ごとに違う場合（1室の施設と2室の施設）", ""),
    ("同じファイルで使えます", "1室の施設は M_部屋 に1行だけ、2室の施設は2行だけ登録してください。"
                        "定員・空室数・空き記号（○△×）は全て M_部屋 の行数から計算しているので、"
                        "シートも数式も一切変えずに両方に対応します。最大4室まで。"),
    ("使わない部屋の行を隠す", "02_カレンダー は各週に「部屋1〜部屋4」の4行があります。1〜2室の施設では、"
                        "使わない部屋の行（部屋名が空欄の行）をまとめて選択し、右クリック→非表示 にしてください。"
                        "6週ぶんまとめて選べます。03_帯表・04_月間表 も同様に列を非表示にできます。"),
    ("部屋を減らしたとき", "2室から1室に変更したのに古い予約が「部屋2」のまま残っていると、"
                        "01_予約入力 のS列に「⚠部屋が未登録」と赤で出ます。その予約の部屋を選び直してください。"),
    ("1室のときの表示", "1室の施設では「△（残りわずか）」は構造上発生せず、必ず ○ か × になります。"),
    ("★ファイルは施設ごとに分ける", "このファイルを施設の数だけコピーして「施設A_管理台帳.xlsx」「施設B_管理台帳.xlsx」"
                        "のように分けてください。1つのファイルに複数施設を混ぜると、定員が施設ごとに違うため"
                        "空き判定が成立せず、全ての集計に施設の絞り込みが必要になって必ず崩れます。"),
    ("SEC", "★ 食事の決め方（3段階）", ""),
    ("① 標準（何もしない）", "日帰り=昼のみ ／ 初日=夕 ／ 中日=朝昼夕 ／ 最終日=朝。"
                     "ほとんどの予約はこれで足ります。"),
    ("② 予約ごとに決める", "01_予約入力 の「朝食」「昼食」「夕食」列で選びます。"
                    "空欄=標準、○=その滞在の全ての利用日で出す、×=その滞在では出さない。"
                    "「この方は毎日昼食もつける」「この方は朝食なし」といった決め方に使います。"),
    ("③ 特定の日だけ直す", "02_カレンダー 下部の「食事の個別調整」に日付と食数を入れます。"
                    "★①②より優先されます。「この日だけ来客で1食多い」といった例外に使います。"),
    ("反映先", "どれで決めても、02_カレンダー の食事行と 06_食数表 に自動で反映されます。"),
    ("SEC", "★ 食事をカレンダーで見る・直す", ""),
    ("いつもは自動", "各週のいちばん下の「食事」行に、その日の 朝○ 昼○ 夕○ が自動で出ます。"
               "予約ごとの指定（01_予約入力 の朝食・昼食・夕食）もここに反映されます。"),
    ("違う日だけ入れる", "02_カレンダー のカレンダーの下に「食事の個別調整」表があります。"
                  "日付と、その日の実際の食数（朝・昼・夕）を入れてください。空欄の欄は自動のままです。"
                  "0 を入れれば「その食事なし」になります。"),
    ("入れたらどうなる", "カレンダーの食事行に「※調整」と付き、06_食数表 の食数もその値に変わります。"
                  "厨房へ渡す数字も自動で合います。"),
    ("★月をまたいでも安全", "調整は日付で結び付けているので、対象月を切り替えても他の月に影響しません。"
                    "同じ日付を2回入れると「⚠日付が重複」と赤く出ます。"),
    ("表示を消したいとき", "02_カレンダー の「食事を表示」を空欄にすると、食事行が非表示になります。"),
    ("SEC", "カレンダーの記号の読み方", ""),
    ("▶16:00 利用者A", "その日が入所日。数字は入室時刻です（01_予約入力 に 16 と入れたもの）。"),
    ("利用者A", "記号なしは連泊の中日。前の日から続いています。"),
    ("利用者A ◀09:00", "その日が退所日。数字は退室時刻です。朝に退所するので、★その日の夜はもう次の方が入れます。"),
    ("◆ 利用者A 10:00-16:00", "日帰り（体験利用）。宿泊しません。"),
    ("2人並ぶとき", "「利用者A ◀09:00　▶16:30 利用者B」のように、退所と入所が同じ日に重なると1つのマスに並びます。"),
    ("空き ○ △ × 仮", "○=空室あり／△=残りわずか／×=満室／仮=仮予約で調整中。これは「その日の夜」の空き状況です。"),
    ("色分け", "名前のマスは区分ごとに色が付きます（M_区分 の上から4つ＝青・緑・紫・橙）。"
             "仮予約・調整中はグレーの斜体になります。"),
    ("SEC", "数え方の定義（全シート共通のルール）", ""),
    ("利用日数", "入所日から退所日までの暦日数。両端を含みます（4/1入所・4/3退所 = 3日）。"),
    ("宿泊数", "退所日 − 入所日（4/1入所・4/3退所 = 2泊）。"),
    ("日帰り（体験利用）", "入所日 = 退所日。1利用日・0泊。食事は昼のみが既定です。"),
    ("空室の判定", "「その日の夜に泊まるか」で見ます。★退所日の当日の夜は空き扱いです（次の方が入れます）。"
                "カレンダーで退所日に名前が出ていても、その夜は空いています。"),
    ("集計に入る予約", "確定・利用済み・仮予約・調整中は集計に入り、キャンセルだけが除外されます。"
                    "キャンセルは行を削除せず、状態を「キャンセル」に変えて記録として残してください。"),
    ("食事の既定", "日帰り=昼のみ ／ 初日=夕 ／ 中日=朝昼夕 ／ 最終日=朝。"
                 "例外だけ 06_食数表 の「調整」列に実数を入れると、そちらが優先されます。"),
    ("SEC", "セルの色の意味", ""),
    ("うすい黄色", "手で入力するセルです。"),
    ("うすい灰色（斜体）", "自動計算です。上書きすると壊れます。触らないでください。"),
    ("赤・オレンジ", "警告です。ダブルブッキング（⚠重複）・支給量オーバー・未登録の部屋を知らせます。"),
    ("SEC", "毎日の使い方", ""),
    ("予約が入ったら", "01_予約入力 の空いている行に、1滞在=1行で入れます。日にちも時刻も数字だけ。"
                  "食事が標準と違うときだけ 朝食・昼食・夕食 列で ○/× を選びます。"
                  "Y列・Z列が赤くならないことだけ確認してください。"),
    ("★部屋は必ず選ぶ", "部屋が空欄のままだと、その予約はカレンダーの部屋の行に出ません"
                  "（「空き」の判定には数えるので×にはなります）。V列に「⚠部屋が未選択」と赤で出るので直してください。"),
    ("入れたあと確認", "02_カレンダー を見て、意図した期間に帯が出ているかを目で確かめます。"),
    ("「空いてますか」", "05_空室照会 に日付を2つ入れるだけです。"),
    ("月を切り替える", "★02_カレンダー の「対象月」を変えると、03_帯表・06_食数表・09_FAX空き表 の月も一緒に変わります。"),
    ("FAXで空き状況を送る", "09_FAX空き表 の「送信先」を選び、必要なら連絡事項を書いて印刷するだけです。"
                    "送信先は M_FAX送付先 に登録しておくと一覧から選べます（番号の打ち間違いによる誤送信を防げます）。"
                    "未登録・未入力の欄は手書き用の下線が印刷されるので、そのままでも使えます。"),
    ("SEC", "★ 月が替わったとき（ファイルは作り直しません）", ""),
    ("1つのファイルを使い続けます", "月ごとに新しいファイルを作る必要も、シートを増やす必要もありません。"
                       "予約は 01_予約入力 に下へ追記していくだけ。見るときは 02_カレンダー の"
                       "「対象月」を変えれば、カレンダー・帯表・食数表・FAX表がまとめて切り替わります。"),
    ("新しい月の1行目だけ", "月が変わる最初の行で「年月」に新しい月（例 2026/9）を入れてください。"
                    "そこから下の行はその月を引き継ぎます。あとは日にちの数字だけです。"),
    ("★シートを月ごとに増やさないでください", "数式が月の数だけ複製され、直すときに全部を直すことになります。"
                    "また 8/31→9/1 のように月をまたぐ滞在をどちらのシートに置くか決められず、"
                    "年度集計や支給量管理が成り立たなくなります。"
                    "「入力は1か所、表示は対象月で切り替え」が、この台帳のいちばんの利点です。"),
    ("入る件数", "予約は400件まで入ります。月に10件なら3年以上ぶんです。"
              "足りなくなったら、その時点でファイルをコピーして年度で分けてください。"),
    ("年度でまとめて見る", "07_月次集計 の期間に 2026/4/1〜2027/3/31 のように入れれば、"
                   "年度ぶんの稼働率・区分別・利用者別がそのまま出ます。"),
    ("年度で分けるとき", "年度末にファイルをコピーして「2026年度_管理台帳.xlsx」として保存し、"
                  "新しい方は古い年度の行を消して使い始めると軽く保てます。消す前に必ず控えを残してください。"),
    ("SEC", "★ 稼働率の見かた", ""),
    ("計算式", "稼働率 ＝ 延べ宿泊数 ÷（定員 × 期間の日数）。07_月次集計 の上部に自動で出ます。"),
    ("「泊」で数えます", "退所日の夜は次の方が使えるので、稼働率には入りません。"
                 "日帰り（0泊）も宿泊ではないため稼働率には入りません。"),
    ("確定と見込み", "「稼働率（確定）」は確定・利用済み・状態未記入のぶんだけ。"
               "「稼働率（仮予約を含む）」は仮予約・調整中も足した見込みです。報告には確定を使ってください。"),
    ("区分別の稼働率", "SS・無料・契約それぞれの稼働率も区分別の表に出ます（こちらは仮予約も含みます）。"),
    ("SEC", "バックアップ", ""),
    ("月に一度はコピーを", "このファイル自体が台帳です。月初に「YYYYMM_管理台帳.xlsx」の名前でコピーを別フォルダに残してください。"),
]
r = 4
for item in rows:
    if item[0] == "SEC":
        c = ws.cell(row=r, column=1, value=item[1]); c.font = C_SUB; c.fill = F_SUB
        ws.cell(row=r, column=2).fill = F_SUB
        r += 1; continue
    a, b = item
    ca = ws.cell(row=r, column=1, value=a); ca.font = C_BOLD
    ca.alignment = Alignment(vertical="top", wrap_text=True)
    cb = ws.cell(row=r, column=2, value=b); cb.font = C_BODY
    cb.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 32
    r += 1


# ══════════════════════════════════════════════════════════════════
# M_施設設定 ほかマスタ
# ══════════════════════════════════════════════════════════════════
ws = sheet("M_施設設定")
title(ws, "施設設定", "黄色のセルを施設に合わせて書き換えてください。")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 62
label_input(ws, 4, "施設名", "○○事業所", "帳票の見出しに使います")
label_input(ws, 5, "棟・フロア", "本棟 1F")
label_auto (ws, 6, "部屋数（定員）", f"=COUNTA({RROW})", "自動。M_部屋 の行数がそのまま定員になります（1部屋につき1名）")
label_input(ws, 7, "標準の入所時刻", "16:00", "予約入力の目安。実際の時刻は予約ごとに入れます")
label_input(ws, 8, "標準の退所時刻", "09:00")
label_input(ws, 9, "入力者名", "", "更新履歴に残す名前")
label_input(ws, 10, "郵便番号", "")
label_input(ws, 11, "住所", "", "FAX帳票の送信元として印字されます")
label_input(ws, 12, "電話番号", "")
label_input(ws, 13, "FAX番号", "")
label_input(ws, 14, "担当者名", "")
ADDR, TEL, FAXNO = f"{MSET}!$B$11", f"{MSET}!$B$12", f"{MSET}!$B$13"

def master_grid(ws, ncols, first, last, input_cols):
    for i in range(first, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER; cell.font = C_BODY
            if c in input_cols: cell.fill = F_INPUT

ws = sheet("M_部屋")
title(ws, "部屋マスタ", "1行 = 1室。★ここの行数が定員になります。1室の施設は1行だけ、2室の施設は2行だけ残してください。")
head_row(ws, 3, ["部屋名", "備考"], [18, 60])
master_grid(ws, 2, 4, 20, {1})
ws["A4"] = "部屋1"; ws["A5"] = "部屋2"
ws["B5"] = "1室のみの施設では、この行を削除してください"; ws["B5"].font = C_NOTE

ws = sheet("利用者マスタ")
title(ws, "利用者マスタ", "支給量＝受給者証の月あたり上限日数。空欄ならその方の残日数は管理しません。")
head_row(ws, 3, ["氏名", "支給量（日／月）", "備考"], [24, 16, 52])
master_grid(ws, 3, 4, 60, {1, 2})
for i, (nm, alw, note) in enumerate([("利用者A", 7, "サンプルです。実際の氏名に書き換えてください"),
                                     ("利用者B", 5, ""), ("利用者C", None, "支給量を管理しない例")], start=4):
    ws.cell(row=i, column=1, value=nm)
    if alw is not None: ws.cell(row=i, column=2, value=alw)
    ws.cell(row=i, column=3, value=note).font = C_NOTE
ws["A3"].comment = Comment("氏名を書き換えると 01_予約入力 のプルダウンにも反映されます。\n"
                           "ただし入力済みの予約の氏名は自動では変わりません。\n"
                           "03_帯表 に出るのは、このマスタの上から20名です。", "設計メモ", height=110, width=330)

ws = sheet("M_区分")
title(ws, "区分マスタ",
      "「支給量に含める」に ○ を入れた区分だけ 08_支給量管理 の日数に数えます。★上から4つがカレンダーの色に対応します。")
head_row(ws, 3, ["区分", "支給量に含める（○／空欄）", "カレンダーの色", "備考"], [16, 24, 18, 40])
master_grid(ws, 4, 4, 20, {1, 2})
for i, (nm, cnt) in enumerate([("SS", "○"), ("無料", "○"), ("契約", "○")], start=4):
    ws.cell(row=i, column=1, value=nm); ws.cell(row=i, column=2, value=cnt)
for i, lbl in enumerate(["青", "緑", "紫", "橙"]):
    c = ws.cell(row=4 + i, column=3, value=lbl)
    c.fill = CAT_FILLS[i]; c.font = C_BODY; c.border = BORDER
    c.alignment = Alignment(horizontal="center")
ws["D8"] = "5つめ以降の区分は色が付きません（白のまま）"; ws["D8"].font = C_NOTE

ws = sheet("M_送迎")
title(ws, "送迎ステータス マスタ", "")
head_row(ws, 3, ["送迎区分", "備考"], [18, 52])
master_grid(ws, 2, 4, 15, {1})
for i, nm in enumerate(["あり", "なし", "調整中"], start=4):
    ws.cell(row=i, column=1, value=nm)

ws = sheet("M_状態")
title(ws, "予約状態マスタ", "★「キャンセル」「仮予約」「調整中」の文字は集計と表示が参照しています。変更しないでください。")
head_row(ws, 3, ["予約状態", "意味"], [16, 74])
for i, (nm, note) in enumerate([("確定", "集計・稼働に入り、カレンダーでは区分の色で表示します"),
                                ("仮予約", "集計には入りますが、カレンダーではグレーの斜体、帯表では白抜き、FAX表では「仮」"),
                                ("調整中", "仮予約と同じ扱いです"),
                                ("利用済み", "実績。確定と同じ扱いです"),
                                ("キャンセル", "記録として残しますが、全ての集計・表示・重複判定から外れます")], start=4):
    ws.cell(row=i, column=1, value=nm).font = C_BODY
    ws.cell(row=i, column=2, value=note).font = C_NOTE
    for c in (1, 2): ws.cell(row=i, column=c).border = BORDER

ws = sheet("M_FAX送付先")
title(ws, "FAX送付先マスタ", "番号を都度手入力せず一覧から選ぶことで、誤送信を防ぎます。")
head_row(ws, 3, ["事業所名", "担当者", "FAX番号", "電話番号", "備考"], [30, 16, 18, 18, 34])
master_grid(ws, 5, 4, 24, {1, 2, 3, 4, 5})

def defname(name, ref): wb.defined_names.add(DefinedName(name, attr_text=ref))
defname("氏名一覧", f"OFFSET({MUSER}!$A$4,0,0,MAX(1,COUNTA({UROW})),1)")
defname("部屋一覧", f"OFFSET({MROOM}!$A$4,0,0,MAX(1,COUNTA({RROW})),1)")
defname("区分一覧", f"OFFSET({MCAT}!$A$4,0,0,MAX(1,COUNTA({CROW})),1)")
defname("送迎一覧", "OFFSET('M_送迎'!$A$4,0,0,MAX(1,COUNTA('M_送迎'!$A$4:$A$15)),1)")
defname("状態一覧", "'M_状態'!$A$4:$A$8")
defname("送付先一覧", "OFFSET('M_FAX送付先'!$A$4,0,0,MAX(1,COUNTA('M_FAX送付先'!$A$4:$A$24)),1)")


# ══════════════════════════════════════════════════════════════════
# 01_予約入力
# ══════════════════════════════════════════════════════════════════
ws = sheet("01_予約入力")
title(ws, "予約入力（1滞在 = 1行）",
      "★日付は「年月」を最初の行に一度だけ入れれば、あとは日にちの数字だけ。"
      "時刻も「時」の数字だけ（16 → 16:00）。食事は空欄なら標準どおりです。"
      "黄色の列だけ入力します。Y列・Z列が赤くなったら、その行を直してください。")
headers = ["予約ID", "氏名", "区分", "部屋",
           "年月\n(最初の行だけ)", "入所\n(日)", "入所日\n(自動)", "入所\n(時)",
           "退所\n(日)", "退所予定日\n(自動)", "退所\n(時)",
           "朝食", "昼食", "夕食",
           "利用日数", "宿泊数", "予約状態", "送迎", "送迎方法", "送迎場所・時間帯",
           "外部サービス", "備考", "更新日", "更新者",
           "⚠重複・部屋チェック", "⚠支給量チェック", "支給量対象日数",
           "作業:実効年月", "作業:並び順", "作業:入所時刻", "作業:退所時刻",
           "作業:入所日数値", "作業:退所日数値"]
widths  = [10, 16, 8, 10, 12, 8, 12, 8, 8, 12, 8, 7, 7, 7,
           9, 8, 11, 9, 16, 20, 16, 30, 12, 12, 17, 15, 13, 12, 12, 10, 10, 12, 12]
head_row(ws, 4, headers, widths)
ws.row_dimensions[4].height = 40
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A4:AA{BK_LAST}"
AUTO_COLS = {"A", "G", "J", "O", "P", "Y", "Z", "AA",
             "AB", "AC", "AD", "AE", "AF", "AG"}
WORK_COLS = {"AB", "AC", "AD", "AE", "AF", "AG"}
MEAL_COLS = {"L", "M", "N"}

for d in range(5, BK_LAST + 1):
    prev = d - 1
    ws[f"A{d}"] = f'=IF($G{d}="","","SS-"&TEXT(ROW()-4,"0000"))'
    ws[f"AB{d}"] = (f'=IF($E{d}<>"",$E{d},"")' if d == 5
                    else f'=IF($E{d}<>"",$E{d},$AB{prev})')
    ws[f"G{d}"] = (f'=IF($F{d}="","",IF($F{d}>31,$F{d},'
                   f'IF($AB{d}="","",DATE(YEAR($AB{d}),MONTH($AB{d}),$F{d}))))')
    ws[f"J{d}"] = (f'=IF($I{d}="","",IF($I{d}>31,$I{d},IF($G{d}="","",'
                   f'IF(DATE(YEAR($G{d}),MONTH($G{d}),$I{d})>=$G{d},'
                   f'DATE(YEAR($G{d}),MONTH($G{d}),$I{d}),'
                   f'DATE(YEAR($G{d}),MONTH($G{d})+1,$I{d})))))')
    ws[f"O{d}"] = f'=IF(OR($G{d}="",$J{d}=""),"",$J{d}-$G{d}+1)'
    ws[f"P{d}"] = f'=IF(OR($G{d}="",$J{d}=""),"",$J{d}-$G{d})'
    # 時刻の表示用文字列。0〜24 の整数は「時」とみなして 16 → "16:00" に直す。
    # 16:30 のように入れた場合は時刻値（1未満）になるのでそのまま分まで出す。
    for src, dst in (("H", "AD"), ("K", "AE")):
        ws[f"{dst}{d}"] = (f'=IF(${src}{d}="","",IF(ISNUMBER(${src}{d}),'
                           f'IF(AND(${src}{d}=INT(${src}{d}),${src}{d}>=0,${src}{d}<=24),'
                           f'TEXT(INT(${src}{d}),"00")&":00",'
                           f'TEXT(${src}{d},"hh:mm")),${src}{d}))')
    ws[f"AF{d}"] = f'=IF($G{d}="",0,$G{d})'
    ws[f"AG{d}"] = f'=IF($J{d}="",0,$J{d})'
    ws[f"AC{d}"] = (f'=IF(OR($G{d}="",$J{d}=""),"",'
                    f'IF(AND($G{d}<=EOMONTH({M1},0),$J{d}>={M1}),$G{d}*10000+ROW(),""))')
    ws[f"Y{d}"] = (
        f'=IF($G{d}="","",'
        f'IF($D{d}="","⚠部屋が未選択",'
        f'IF(COUNTIF({RROW},$D{d})=0,"⚠部屋が未登録",'
        f'IF(OR($J{d}="",$Q{d}="キャンセル"),"",'
        f'IF(SUMPRODUCT(($D$5:$D${BK_LAST}=$D{d})*($G$5:$G${BK_LAST}<>"")*($J$5:$J${BK_LAST}<>"")'
        f'*($Q$5:$Q${BK_LAST}<>"キャンセル")*($G$5:$G${BK_LAST}<$J{d})*($J$5:$J${BK_LAST}>$G{d}))>1,'
        f'"⚠重複","")))))')
    ws[f"AA{d}"] = (
        f'=IF(OR($G{d}="",$J{d}="",$Q{d}="キャンセル"),0,'
        f'IF(IFERROR(INDEX({CCNT},MATCH($C{d},{CROW},0)),"○")="○",$J{d}-$G{d}+1,0))')
    ws[f"Z{d}"] = (
        f'=IF(OR($B{d}="",$G{d}=""),"",'
        f'IF(SUMPRODUCT(({UROW}=$B{d})*({UALW}<>""))=0,"",'
        f'IF(SUMIFS($AA$5:$AA${BK_LAST},$B$5:$B${BK_LAST},$B{d},'
        f'$G$5:$G${BK_LAST},">="&DATE(YEAR($G{d}),MONTH($G{d}),1),'
        f'$G$5:$G${BK_LAST},"<="&EOMONTH($G{d},0))'
        f'>SUMIFS({UALW},{UROW},$B{d}),"⚠支給量超過","")))')
    for col in range(1, 34):
        L = get_column_letter(col); c = ws[f"{L}{d}"]
        c.border = BORDER
        c.font, c.fill = (C_AUTO, F_AUTO) if L in AUTO_COLS else (C_BODY, F_INPUT)
        if L in MEAL_COLS:
            c.fill = PatternFill("solid", fgColor="FBF3E4", bgColor="FBF3E4")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name=FONT, size=11, bold=True, color="8A6D3B")
        if L in ("G", "J", "W"): c.number_format = "yyyy/mm/dd"
        if L == "E": c.number_format = "yyyy/mm"
        if L in ("F", "I"): c.number_format = "0"
        if L in ("H", "K"): c.number_format = "General"
        if L not in MEAL_COLS: c.alignment = Alignment(vertical="center")

samples = [
    # (氏名, 区分, 部屋, 年月, 入所日, 入所時, 退所日, 退所時, 朝, 昼, 夕, 状態, 送迎, 送迎方法, 場所, 外部, 備考)
    ("利用者A", "SS",  "部屋1", datetime.date(2026, 4, 1), 6, 16, 8, 9, "", "", "", "確定",
     "あり", "施設車", "自宅前 16:00", "", "サンプル行です。削除してお使いください"),
    ("利用者B", "契約", "部屋2", None, 6, "16:30", 7, 9, "", "○", "", "確定", "なし", "", "", "",
     "昼食を全日つける例（既定では中日だけ）"),
    ("利用者C", "無料", "部屋1", None, 11, 10, 11, 16, "", "", "○", "確定", "あり", "家族送迎", "", "",
     "日帰りだが夕食もつける例"),
    ("利用者A", "SS",  "部屋1", None, 20, 16, 24, 9, "×", "", "", "仮予約", "調整中", "", "", "",
     "朝食を出さない例。仮予約の例でもあります"),
]
for i, s in enumerate(samples):
    d = 5 + i
    for col, val in zip(["B","C","D","E","F","H","I","K","L","M","N","Q","R","S","T","U","V"], s):
        if val not in (None, ""):
            ws[f"{col}{d}"] = val

ws["E4"].comment = Comment("その月の日付（例: 2026/8）を最初の行に一度だけ入れてください。\n"
                           "以降の行は空欄のままで、上の行の年月を自動で引き継ぎます。\n"
                           "月が変わる行で新しい年月を入れると、そこから下が新しい月になります。\n"
                           "※ 行の並べ替えはしないでください（引き継ぎが崩れます）。",
                           "設計メモ", height=130, width=350)
ws["F4"].comment = Comment("日にちの数字だけ入れてください（1〜31）。\n"
                           "年月と組み合わせて G列に日付が自動で入ります。\n"
                           "別の月を直接入れたいときは、日付そのもの（2026/9/3 など）でも構いません。",
                           "設計メモ", height=110, width=350)
ws["H4"].comment = Comment("時の数字だけ入れてください（例: 16 → 16:00）。\n"
                           "9時なら 9 だけでかまいません（09:00 と表示されます）。\n"
                           "分まで指定したいときは 16:30 のように入れてください。\n"
                           "空欄のままでも予約は成立します（カレンダーに時刻が出ないだけです）。",
                           "設計メモ", height=120, width=350)
ws["K4"].comment = Comment("時の数字だけ入れてください（例: 9 → 09:00）。\n"
                           "分まで指定したいときは 9:30 のように入れてください。",
                           "設計メモ", height=90, width=350)
ws["I4"].comment = Comment("日にちの数字だけ入れてください（1〜31）。\n"
                           "入所日より前の日を入れた場合は「翌月のその日」とみなします。\n"
                           "（例: 入所31 → 退所1 なら翌月1日）\n"
                           "1か月を超える滞在のときは、日付そのものを入れてください。",
                           "設計メモ", height=120, width=350)
MEAL_HELP = ("空欄 … 標準どおり\n"
             "　 日帰り=昼のみ ／ 初日=夕 ／ 中日=朝昼夕 ／ 最終日=朝\n"
             "○ … この滞在の全ての利用日（入所日〜退所日）で出す\n"
             "× … この滞在では出さない\n\n"
             "ここで決めた内容は 02_カレンダー の食事行と 06_食数表 に反映されます。\n"
             "特定の1日だけ変えたいときは 02_カレンダー 下部の「食事の個別調整」を使ってください"
             "（そちらが優先されます）。")
for cc in ("L4", "M4", "N4"):
    ws[cc].comment = Comment(MEAL_HELP, "設計メモ", height=170, width=380)
ws["Y4"].comment = Comment("部屋が空欄だと「⚠部屋が未選択」。この予約はカレンダーに表示されません。\n"
                           "M_部屋 に無い部屋名なら「⚠部屋が未登録」。\n"
                           "同じ部屋・同じ夜に2件以上あれば「⚠重複」。\n"
                           "退所日の夜は空き扱いなので、前の方の退所日＝次の方の入所日は重複になりません。",
                           "設計メモ", height=140, width=350)
ws["Z4"].comment = Comment("利用者マスタ に支給量が入っている方だけ判定します。\n"
                           "入所日の属する月で合計し、M_区分 で「○」の区分だけを数えます。",
                           "設計メモ", height=100, width=350)

for rng, name in [(f"B5:B{BK_LAST}", "氏名一覧"), (f"C5:C{BK_LAST}", "区分一覧"),
                  (f"D5:D{BK_LAST}", "部屋一覧"), (f"Q5:Q{BK_LAST}", "状態一覧"),
                  (f"R5:R{BK_LAST}", "送迎一覧")]:
    dv = DataValidation(type="list", formula1=f"={name}", allow_blank=True, showDropDown=False)
    dv.error = "一覧にない値です。マスタに登録してから選んでください。"
    dv.errorTitle = "入力できません"
    ws.add_data_validation(dv); dv.add(rng)
dvm = DataValidation(type="list", formula1='"○,×"', allow_blank=True, showDropDown=False)
dvm.error = "○（全日出す）か ×（出さない）を選んでください。標準どおりでよければ空欄のままにします。"
dvm.errorTitle = "入力できません"
dvm.promptTitle = "食事"
dvm.prompt = "空欄=標準 ／ ○=全ての利用日で出す ／ ×=出さない"
dvm.showInputMessage = True
ws.add_data_validation(dvm)
dvm.add(f"L5:N{BK_LAST}")

rng_all = f"A5:AA{BK_LAST}"
ws.conditional_formatting.add(rng_all, FormulaRule(formula=['$Q5="キャンセル"'],
    font=Font(name=FONT, size=10, color="999999", strike=True),
    fill=PatternFill("solid", fgColor="EFEFEF", bgColor="EFEFEF"), stopIfTrue=True))
ws.conditional_formatting.add(rng_all, FormulaRule(formula=['$Y5<>""'], fill=F_WARN, stopIfTrue=True))
ws.conditional_formatting.add(rng_all, FormulaRule(formula=['$Z5<>""'], fill=F_WARN2, stopIfTrue=True))
ws.conditional_formatting.add(rng_all, FormulaRule(formula=['OR($Q5="仮予約",$Q5="調整中")'], fill=F_TENT))
for L in WORK_COLS:
    ws.column_dimensions[L].hidden = True




# ══════════════════════════════════════════════════════════════════
# 04_月間表  ―  計算のハブ。カレンダーと帯表はここを引くだけ
# ══════════════════════════════════════════════════════════════════
ws = sheet("04_月間表")
title(ws, "月間表（日別の一覧・自動計算）",
      "★対象月の変更は 02_カレンダー で行います。L列より右は計算用の作業列です（非表示にしてあります）。")
label_auto(ws, 3, "対象月", f"={M1}", "02_カレンダー の対象月に連動します", "yyyy/mm/dd")

hdr = ["日付", "曜日"] + [f"部屋{i}" for i in range(1, MAX_ROOMS + 1)] + \
      ["使用室数", "空室数", "仮予約", "空き記号", "日帰り"]
head_row(ws, 6, hdr, [12, 6] + [24] * MAX_ROOMS + [9, 9, 9, 11, 9])
for i in range(MAX_ROOMS):
    ws[f"{get_column_letter(3+i)}6"] = f'=IF(COUNTA({RROW})>={i+1},INDEX({RROW},{i+1}),"")'
USE_C = get_column_letter(3 + MAX_ROOMS)     # G
VAC_C = get_column_letter(4 + MAX_ROOMS)     # H
TEN_C = get_column_letter(5 + MAX_ROOMS)     # I
SYM_C = get_column_letter(6 + MAX_ROOMS)     # J
DAY_C = get_column_letter(7 + MAX_ROOMS)     # K
IDX0  = 8 + MAX_ROOMS                        # L = 12（作業列の先頭）
CATM0 = IDX0 + 3 * MAX_ROOMS                 # X = 24（区分ミラーの先頭）
MEAL0 = CATM0 + MAX_ROOMS                    # AB = 28（食事の作業列の先頭）
MB_A, MB_L, MB_D = (get_column_letter(MEAL0), get_column_letter(MEAL0 + 1),
                    get_column_letter(MEAL0 + 2))          # 朝昼夕（自動）
MB_X = get_column_letter(MEAL0 + 3)                        # 調整表の行番号
MC_A, MC_L, MC_D = (get_column_letter(MEAL0 + 4), get_column_letter(MEAL0 + 5),
                    get_column_letter(MEAL0 + 6))          # 朝昼夕（確定）
MC_F = get_column_letter(MEAL0 + 7)                        # 表示用文字列
MB_H = get_column_letter(MEAL0 + 8)                        # 実際に値が入った調整があるか
# 02_カレンダー 下部の「食事の個別調整」表（日付・朝・昼・夕）
ADJ_D  = f"{CAL}!$B${ADJ_TOP}:$B${ADJ_END}"
ADJ_A  = f"{CAL}!$C${ADJ_TOP}:$C${ADJ_END}"
ADJ_L  = f"{CAL}!$D${ADJ_TOP}:$D${ADJ_END}"
ADJ_D2 = f"{CAL}!$E${ADJ_TOP}:$E${ADJ_END}"

# 作業列の見出し
for j in range(MAX_ROOMS):
    for k, nm in enumerate(["退所", "宿泊", "日帰"]):
        c = ws.cell(row=6, column=IDX0 + 3 * j + k, value=f"作業{j+1}{nm}")
        c.font = C_NOTE; c.fill = F_SUB; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=6, column=CATM0 + j, value=f"色{j+1}")
    c.font = C_NOTE; c.fill = F_SUB; c.border = BORDER
    c.alignment = Alignment(horizontal="center")
for k, nm in enumerate(["朝自動", "昼自動", "夕自動", "調整行", "朝確定", "昼確定", "夕確定",
                        "食事表示", "調整有無"]):
    c = ws.cell(row=6, column=MEAL0 + k, value=nm)
    c.font = C_NOTE; c.fill = F_SUB; c.border = BORDER
    c.alignment = Alignment(horizontal="center")

for i in range(31):
    d = 7 + i
    ws[f"A{d}"] = f'=IF(MONTH({M1}+{i})<>MONTH({M1}),"",{M1}+{i})'
    ws[f"B{d}"] = f'=IF($A{d}="","",MID("日月火水木金土",WEEKDAY($A{d}),1))'
    for j in range(MAX_ROOMS):
        RN = f"{get_column_letter(3+j)}$6"                 # 部屋名の見出しセル
        DEP = f"${get_column_letter(IDX0+3*j)}{d}"
        NGT = f"${get_column_letter(IDX0+3*j+1)}{d}"
        TRP = f"${get_column_letter(IDX0+3*j+2)}{d}"
        base = f'({ROOM_}={RN})*({ST_}<>"キャンセル")*({IN_}<>"")'
        ws[f"{get_column_letter(IDX0+3*j)}{d}"] = (
            f'=IF(OR($A{d}="",{RN}=""),0,SUMPRODUCT(MAX({base}'
            f'*({OUT_}=$A{d})*({IN_}<$A{d})*{ROWID})))')
        ws[f"{get_column_letter(IDX0+3*j+1)}{d}"] = (
            f'=IF(OR($A{d}="",{RN}=""),0,SUMPRODUCT(MAX({base}'
            f'*({IN_}<=$A{d})*({OUT_}>$A{d})*{ROWID})))')
        ws[f"{get_column_letter(IDX0+3*j+2)}{d}"] = (
            f'=IF(OR($A{d}="",{RN}=""),0,SUMPRODUCT(MAX({base}'
            f'*({IN_}=$A{d})*({OUT_}=$A{d})*{ROWID})))')
        # 表示文字列: 退所者 ◀退所時刻 / ▶入所時刻 入所者 / 継続者 / ◆日帰り 時刻
        SHOW = f'({SHOWTIME}="○")'
        ws[f"{get_column_letter(3+j)}{d}"] = (
            f'=IF(OR($A{d}="",{RN}=""),"",'
            f'IF({DEP}=0,"",INDEX({NAME_},{DEP})&" ◀"&IF({SHOW},INDEX({TOUT_},{DEP}),""))'
            f'&IF(AND({DEP}>0,OR({NGT}>0,{TRP}>0)),"  ","")'
            f'&IF({NGT}=0,"",IF(INDEX({IN_},{NGT})=$A{d},'
            f'"▶"&IF({SHOW},INDEX({TIN_},{NGT})&" "," "),"")&INDEX({NAME_},{NGT}))'
            f'&IF(AND({NGT}>0,{TRP}>0),"  ","")'
            f'&IF({TRP}=0,"","◆ "&INDEX({NAME_},{TRP})'
            f'&IF({SHOW}," "&INDEX({TIN_},{TRP})&"-"&INDEX({TOUT_},{TRP}),"")))')
        # 色ミラー: 宿泊者 > 日帰り > 退所者 の優先で、区分名または「仮」
        def catof(ix):
            return (f'IF(OR(INDEX({ST_},{ix})="仮予約",INDEX({ST_},{ix})="調整中"),'
                    f'"仮",INDEX({CAT_},{ix}))')
        ws[f"{get_column_letter(CATM0+j)}{d}"] = (
            f'=IF($A{d}="","",IF({NGT}>0,{catof(NGT)},'
            f'IF({TRP}>0,{catof(TRP)},IF({DEP}>0,{catof(DEP)},""))))')
    ws[f"{USE_C}{d}"] = f'=IF($A{d}="","",SUMPRODUCT({VALID}*({IN_}<=$A{d})*({OUT_}>$A{d})))'
    ws[f"{VAC_C}{d}"] = f'=IF($A{d}="","",{CAP}-${USE_C}{d})'
    ws[f"{TEN_C}{d}"] = (f'=IF($A{d}="","",SUMPRODUCT({TENT}*({IN_}<>"")'
                         f'*({IN_}<=$A{d})*({OUT_}>$A{d})))')
    ws[f"{SYM_C}{d}"] = (f'=IF($A{d}="","",IF(${TEN_C}{d}>0,"仮",'
                         f'IF(${VAC_C}{d}<=0,"×",IF(${VAC_C}{d}<{CAP},"△","○"))))')
    ws[f"{DAY_C}{d}"] = f'=IF($A{d}="","",SUMPRODUCT({VALID}*({IN_}=$A{d})*({OUT_}=$A{d})))'
    # ── 食事 ──────────────────────────────────────────────
    # 既定: 日帰り=昼のみ／初日=夕／中日=朝昼夕／最終日=朝
    # 予約ごとの指定: ○=利用日すべて / ×=出さない / 空欄=標準
    USE = f'({IN_}<=$A{d})*({OUT_}>=$A{d})'
    ws[f"{MB_A}{d}"] = (f'=IF($A{d}="",0,SUMPRODUCT({VALID}*(({BF_}="○")*{USE}'
                        f'+({BF_}="")*({IN_}<$A{d})*({OUT_}>=$A{d}))))')
    ws[f"{MB_L}{d}"] = (f'=IF($A{d}="",0,SUMPRODUCT({VALID}*(({LF_}="○")*{USE}'
                        f'+({LF_}="")*((({IN_}<$A{d})*({OUT_}>$A{d}))'
                        f'+(({IN_}=$A{d})*({OUT_}=$A{d}))))))')
    ws[f"{MB_D}{d}"] = (f'=IF($A{d}="",0,SUMPRODUCT({VALID}*(({DF_}="○")*{USE}'
                        f'+({DF_}="")*({IN_}<=$A{d})*({OUT_}>$A{d}))))')
    # 調整表（02_カレンダー の下部）を日付で引く。無ければ既定のまま。
    ws[f"{MB_X}{d}"] = f'=IF($A{d}="",0,IFERROR(MATCH($A{d},{ADJ_D},0),0))'
    for k, (auto, conf) in enumerate([(MB_A, MC_A), (MB_L, MC_L), (MB_D, MC_D)]):
        src = [ADJ_A, ADJ_L, ADJ_D2][k]
        ws[f"{conf}{d}"] = (f'=IF($A{d}="",0,IF(${MB_X}{d}=0,${auto}{d},'
                            f'IF(INDEX({src},${MB_X}{d})="",${auto}{d},INDEX({src},${MB_X}{d}))))')
    # 日付だけ書いて値を入れていない行は「調整」とみなさない
    ws[f"{MB_H}{d}"] = (f'=IF(OR($A{d}="",${MB_X}{d}=0),0,'
                        f'IF(OR(INDEX({ADJ_A},${MB_X}{d})<>"",INDEX({ADJ_L},${MB_X}{d})<>"",'
                        f'INDEX({ADJ_D2},${MB_X}{d})<>""),1,0))')
    ws[f"{MC_F}{d}"] = (f'=IF($A{d}="","",IF(AND(${MC_A}{d}=0,${MC_L}{d}=0,${MC_D}{d}=0),"",'
                        f'"朝"&${MC_A}{d}&" 昼"&${MC_L}{d}&" 夕"&${MC_D}{d}'
                        f'&IF(${MB_H}{d}=0,"","　※調整")))')
    for col in range(1, MEAL0 + 9):
        c = ws.cell(row=d, column=col)
        c.border = BORDER; c.font = C_AUTO; c.fill = F_AUTO
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{d}"].number_format = "yyyy/mm/dd"
    for j in range(MAX_ROOMS):
        ws.cell(row=d, column=3 + j).alignment = Alignment(horizontal="left", vertical="center")

def mark_cf(worksheet, rng):
    top = rng.split(":")[0]
    for mark, fill in [("×", F_WARN), ("△", F_WARN2), ("仮", F_TENT), ("○", F_OK)]:
        worksheet.conditional_formatting.add(rng, FormulaRule(formula=[f'{top}="{mark}"'], fill=fill))

mark_cf(ws, f"{SYM_C}7:{SYM_C}37")
ws.conditional_formatting.add("A7:B37", FormulaRule(formula=['AND($A7<>"",WEEKDAY($A7,2)>=6)'], fill=F_WKND))
for col in range(IDX0, MEAL0 + 9):
    ws.column_dimensions[get_column_letter(col)].hidden = True
    ws.column_dimensions[get_column_letter(col)].width = 8
ws["A39"] = ("※ 「使用室数／空室数／空き記号」はその日の夜の宿泊で数えます。退所日の夜は空きです。"
             "部屋の欄に出る「◀」は朝に退所することを示します。")
ws["A39"].font = C_NOTE
ws["A40"] = "※ L列より右は計算用の作業列です。非表示にしてあります（表示しても構いませんが、編集しないでください）。"
ws["A40"].font = C_NOTE
ws.print_area = f"A1:{DAY_C}37"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True


# ══════════════════════════════════════════════════════════════════
# 02_カレンダー  ―  月グリッド。対象月の唯一の入力元
#   B..H = 日〜土の表示列 / J..P = 同じ形の隠しミラー(区分・状態)
#   条件付き書式は「8列右のミラー」を見て色を決める。
# ══════════════════════════════════════════════════════════════════
ws = sheet("02_カレンダー")
title(ws, "カレンダー",
      "★対象月を変えると、03_帯表・04_月間表・06_食数表・09_FAX空き表 も同じ月になります。")
label_input(ws, 3, "対象月", MONTH0, "", "yyyy/mm/dd")
ws["C3"] = ('="その月の日付ならどれでも可。　　【今日は "&YEAR(TODAY())&"年"&MONTH(TODAY())'
            '&"月"&DAY(TODAY())&"日】　"&IF(TEXT($B$3,"yyyymm")=TEXT(TODAY(),"yyyymm"),'
            '"（今月を表示しています）","← 今月に切り替えるときはここを直します")')
ws["C3"].font = C_NOTE
label_auto (ws, 4, "月初", '=DATE(YEAR($B$3),MONTH($B$3),1)', "自動", "yyyy/mm/dd")
label_input(ws, 5, "時刻を表示", "○", "○で入室・退室時刻を表示。狭く感じるときは空欄に")
label_input(ws, 6, "食事を表示", "○", "○で各週の下に朝・昼・夕の食数を表示。個別の調整はこのシートの下部で")
NOROOM = f'SUMPRODUCT(({ROOM_}="")*({IN_}<>""))'
ws["A7"] = "確認"; ws["A7"].font = C_SUB; ws["A7"].fill = F_SUB; ws["A7"].border = BORDER
ws["B7"] = (f'=IF({NOROOM}=0,"✓ 部屋の指定もれはありません",'
            f'"⚠ 部屋が未選択の予約が "&{NOROOM}&" 件あります。'
            f'部屋の行には出ませんが「空き」の判定には数えています。01_予約入力 のV列（赤）をご確認ください。")')
ws["B7"].font = Font(name=FONT, size=10, bold=True, color="B03A2E")
ws["B7"].alignment = Alignment(horizontal="left", vertical="center")
ws["A8"] = "記号"; ws["A8"].font = C_SUB
ws["B8"] = ("▶ 入所（時刻つき）　／　記号なし 連泊の中日　／　◀ 退所（朝に退所。その夜は空きます）"
            "　／　◆ 日帰り　／　色は区分（M_区分）、グレー斜体は仮予約")
ws["B8"].font = C_NOTE
ws["B8"].alignment = Alignment(horizontal="left")

WD = ["日", "月", "火", "水", "木", "金", "土"]
ws.column_dimensions["A"].width = 13
for c in range(7):
    L = get_column_letter(2 + c)
    ws.column_dimensions[L].width = 23
    cell = ws[f"{L}{CAL_START-1}"]; cell.value = WD[c]
    cell.font = Font(name=FONT, size=11, bold=True,
                     color="FFFFFF" if 0 < c < 6 else ("FFE0E0" if c == 0 else "DDEBFF"))
    cell.fill = PatternFill("solid", fgColor="2F6690", bgColor="2F6690") if 0 < c < 6 else \
                PatternFill("solid", fgColor=("A93226" if c == 0 else "1F618D"),
                            bgColor=("A93226" if c == 0 else "1F618D"))
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[CAL_START - 1].height = 24

BLOCK = 3 + MAX_ROOMS          # 日付 + 部屋 + 空き + 食事
MIRROR0 = 10                   # J列。表示列 B(2) からのオフセットは +8
CAL_LAST = CAL_START + 5 * CAL_STRIDE + BLOCK - 1
THICK = Side(style="medium", color="2F6690")

for w in range(6):
    base = CAL_START + w * CAL_STRIDE
    ws.cell(row=base, column=1, value=f"第{w+1}週").font = C_NOTE
    ws.cell(row=base, column=1).alignment = Alignment(horizontal="right", vertical="center")
    for j in range(MAX_ROOMS):
        c = ws.cell(row=base + 1 + j, column=1,
                    value=f'=IF(COUNTA({RROW})>={j+1},INDEX({RROW},{j+1}),"")')
        c.font = C_SUB; c.fill = F_SUB; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=base + 1 + MAX_ROOMS, column=1, value="空き")
    c.font = C_SUB; c.fill = F_SUB; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=base + 2 + MAX_ROOMS, column=1, value="食事")
    c.font = C_SUB; c.fill = PatternFill("solid", fgColor="EDE7DC", bgColor="EDE7DC"); c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

    for cidx in range(7):
        L = get_column_letter(2 + cidx)
        n = w * 7 + cidx
        # 日付行
        cell = ws[f"{L}{base}"]
        cell.value = (f'=IF(MONTH($B$4-WEEKDAY($B$4)+1+{n})<>MONTH($B$4),"",'
                      f'$B$4-WEEKDAY($B$4)+1+{n})')
        cell.number_format = "d"
        cell.font = Font(name=FONT, size=14, bold=True,
                         color="A93226" if cidx == 0 else ("1F618D" if cidx == 6 else "333333"))
        cell.fill = PatternFill("solid", fgColor="F4F7FA", bgColor="F4F7FA")
        cell.border = Border(left=THICK, right=THICK, top=THICK, bottom=THIN)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        # 部屋レーン
        for j in range(MAX_ROOMS):
            rc = get_column_letter(3 + j)
            mc = get_column_letter(CATM0 + j)
            cell = ws.cell(row=base + 1 + j, column=2 + cidx)
            cell.value = (f'=IF({L}{base}="","",IFERROR(INDEX({MON}!{rc}$7:{rc}$37,'
                          f'MATCH({L}{base},{MON}!$A$7:$A$37,0)),""))')
            cell.font = Font(name=FONT, size=10, bold=True, color="1F3B52")
            cell.border = Border(left=THICK, right=THICK, top=HAIR, bottom=HAIR)
            cell.alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
            m = ws.cell(row=base + 1 + j, column=MIRROR0 + cidx)
            m.value = (f'=IF({L}{base}="","",IFERROR(INDEX({MON}!{mc}$7:{mc}$37,'
                       f'MATCH({L}{base},{MON}!$A$7:$A$37,0)),""))')
        # 空き記号行
        cell = ws.cell(row=base + 1 + MAX_ROOMS, column=2 + cidx)
        cell.value = (f'=IF({L}{base}="","",IFERROR(INDEX({MON}!${SYM_C}$7:${SYM_C}$37,'
                      f'MATCH({L}{base},{MON}!$A$7:$A$37,0)),""))')
        cell.font = Font(name=FONT, size=14, bold=True)
        cell.border = Border(left=THICK, right=THICK, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # 食事行
        cell = ws.cell(row=base + 2 + MAX_ROOMS, column=2 + cidx)
        cell.value = (f'=IF(OR({L}{base}="",$B$6<>"○"),"",IFERROR(INDEX({MON}!${MC_F}$7:${MC_F}$37,'
                      f'MATCH({L}{base},{MON}!$A$7:$A$37,0)),""))')
        cell.font = Font(name=FONT, size=9, color="6B5B3E")
        cell.fill = PatternFill("solid", fgColor="FBF8F2", bgColor="FBF8F2")
        cell.border = Border(left=THICK, right=THICK, top=THIN, bottom=THICK)
        cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    ws.row_dimensions[base].height = 22
    for j in range(MAX_ROOMS):
        ws.row_dimensions[base + 1 + j].height = 21
    ws.row_dimensions[base + 1 + MAX_ROOMS].height = 22
    ws.row_dimensions[base + 2 + MAX_ROOMS].height = 17

# 色分け（ミラーは8列右）。仮予約を最優先、次に区分の色。
GRID = f"B{CAL_START}:H{CAL_LAST}"
ws.conditional_formatting.add(GRID, FormulaRule(
    formula=[f'AND(B{CAL_START}<>"",J{CAL_START}="仮")'],
    fill=F_TENT2, font=Font(name=FONT, size=10, italic=True, color="6E6E6E"), stopIfTrue=True))
for i in range(4):
    ws.conditional_formatting.add(GRID, FormulaRule(
        formula=[f'AND(B{CAL_START}<>"",{MCAT}!$A${4+i}<>"",J{CAL_START}={MCAT}!$A${4+i})'],
        fill=CAT_FILLS[i], font=Font(name=FONT, size=10, bold=True, color=CAT_FONTS[i]),
        stopIfTrue=True))
for w in range(6):
    r_ = CAL_START + w * CAL_STRIDE + 1 + MAX_ROOMS
    mark_cf(ws, f"B{r_}:H{r_}")

for c in range(MIRROR0, MIRROR0 + 7):
    ws.column_dimensions[get_column_letter(c)].hidden = True

# ── 食事の個別調整（日付で結び付けるので、月を切り替えてもずれない）──
hr = ADJ_TOP - 2
ws[f"A{hr}"] = "■ 食事の個別調整（既定と違う日だけ入れてください）"
ws[f"A{hr}"].font = C_SUB
ws[f"A{hr+1}"] = "使い方"
ws[f"A{hr+1}"].font = C_NOTE
ws[f"B{hr+1}"] = ("日付を入れ、朝・昼・夕に「その日の実際の食数」を入れます。空欄の欄は自動計算のままです。"
                  "0を入れれば「その食事なし」になります。既定は 日帰り=昼のみ／初日=夕／中日=朝昼夕／最終日=朝。")
ws[f"B{hr+1}"].font = C_NOTE
ws[f"B{hr+1}"].alignment = Alignment(horizontal="left")
for i, (lbl, wdt) in enumerate([("日付", 14), ("朝", 8), ("昼", 8), ("夕", 8), ("メモ", 40)]):
    c = ws.cell(row=ADJ_TOP - 1, column=2 + i, value=lbl)
    c.font = C_HEAD; c.fill = F_HEAD; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
for r_ in range(ADJ_TOP, ADJ_END + 1):
    for i in range(5):
        c = ws.cell(row=r_, column=2 + i)
        c.fill = F_INPUT; c.font = C_BODY; c.border = BORDER
        c.alignment = Alignment(horizontal="center" if i else "left", vertical="center")
    ws.cell(row=r_, column=2).number_format = "yyyy/mm/dd"
    ws.cell(row=r_, column=6).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=r_, column=1, value=(f'=IF($B{r_}="","",IF(MONTH($B{r_})<>MONTH($B$4),"（他の月）",'
                                     f'IF(COUNTIF($B${ADJ_TOP}:$B${ADJ_END},$B{r_})>1,"⚠日付が重複","✓")))'))
    ws.cell(row=r_, column=1).font = C_AUTO
    ws.cell(row=r_, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.conditional_formatting.add(f"A{ADJ_TOP}:F{ADJ_END}", FormulaRule(
    formula=[f'$A{ADJ_TOP}="⚠日付が重複"'], fill=F_WARN))

ws[f"A{ADJ_END+2}"] = ("※ 使わない部屋の行（部屋名が空欄の行）は、6週ぶんまとめて選んで 右クリック→非表示 にしてください。")
ws[f"A{ADJ_END+3}"] = ("※ J〜P列は色分け用の隠し列です。編集しないでください。")
ws[f"A{ADJ_END+4}"] = ("※ 退所日に名前が出ていても、その日の夜は空いています（「空き」行の記号が正です）。")
ws[f"A{ADJ_END+5}"] = ("※ 食事の調整は日付で結び付けているので、対象月を切り替えても他の月に影響しません。")
for k in range(2, 6):
    ws[f"A{ADJ_END+k}"].font = C_NOTE
ws.print_area = f"A1:H{CAL_LAST}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.freeze_panes = f"B{CAL_START}"



# ══════════════════════════════════════════════════════════════════
# 03_帯表  ―  1人1行 × 1〜31日のガント
#   B..AF = 表示 / AP..BT = 隠しミラー(状態コード)。オフセットは +40。
#   コード = 1(利用) + 2(入所日) + 4(退所日) + 8(仮予約)
# ══════════════════════════════════════════════════════════════════
ws = sheet("03_帯表")
title(ws, "帯表（だれが・いつから・いつまで）",
      "1人1行、横軸が1〜31日。▶入所　■連泊　◀退所　◆日帰り。白抜き（▷□◁◇）は仮予約・調整中です。")
# 対象月は列幅3.4の列だと "#" になるので、B3:F3 を結合して表示する
ws["A3"] = "対象月"; ws["A3"].font = C_SUB; ws["A3"].fill = F_SUB; ws["A3"].border = BORDER
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("B3:F3")
ws["B3"] = f"={M1}"; ws["B3"].number_format = "yyyy/mm/dd"
ws["B3"].font = C_AUTO; ws["B3"].fill = F_AUTO
ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
for cc in range(2, 7): ws.cell(row=3, column=cc).border = BORDER
ws.merge_cells("G3:T3")
ws["G3"] = "02_カレンダー の対象月に連動します"; ws["G3"].font = C_NOTE
ws["G3"].alignment = Alignment(horizontal="left", vertical="center")
ws.column_dimensions["A"].width = 18
BAND_LAST = 6 + BAND_ROWS
MIRROR_B = 42                   # AP列。表示列 B(2) からのオフセットは +40

ws["A4"] = "日付"; ws["A5"] = "日"; ws["A6"] = "曜日"
for r_ in (4, 5, 6):
    ws[f"A{r_}"].font = C_SUB; ws[f"A{r_}"].fill = F_SUB; ws[f"A{r_}"].border = BORDER
    ws[f"A{r_}"].alignment = Alignment(horizontal="center", vertical="center")
for i in range(31):
    L = get_column_letter(2 + i)
    ws.column_dimensions[L].width = 3.4
    ws[f"{L}4"] = f'=IF(MONTH({M1}+{i})<>MONTH({M1}),"",{M1}+{i})'
    ws[f"{L}5"] = f'=IF({L}4="","",DAY({L}4))'
    ws[f"{L}6"] = f'=IF({L}4="","",MID("日月火水木金土",WEEKDAY({L}4),1))'
    for r_ in (5, 6):
        c = ws[f"{L}{r_}"]; c.border = BORDER
        c.font = Font(name=FONT, size=9, bold=(r_ == 5))
        c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[4].hidden = True

for i in range(BAND_ROWS):
    r_ = 7 + i
    ws[f"A{r_}"] = f'=IF(COUNTA({UROW})>={i+1},INDEX({UROW},{i+1}),"")'
    ws[f"A{r_}"].font = Font(name=FONT, size=10, bold=True, color="1F3B52")
    ws[f"A{r_}"].fill = F_AUTO; ws[f"A{r_}"].border = BORDER
    ws.row_dimensions[r_].height = 19
    ws[f"A{r_}"].alignment = Alignment(horizontal="left", vertical="center")
    for i2 in range(31):
        L = get_column_letter(2 + i2)
        ML = get_column_letter(MIRROR_B + i2)
        ws[f"{ML}{r_}"] = (
            f'=IF(OR($A{r_}="",{L}$4=""),0,SUMPRODUCT({VALID}*({NAME_}=$A{r_})'
            f'*({IN_}<={L}$4)*({OUT_}>={L}$4)'
            f'*(1+2*({IN_}={L}$4)+4*({OUT_}={L}$4)+8*{TENT})))')
        m = f"${ML}{r_}"
        ws[f"{L}{r_}"] = (
            f'=IF({m}=0,"",'
            f'IF(MOD({m},8)=7,IF({m}>=8,"◇","◆"),'
            f'IF(MOD({m},8)=3,IF({m}>=8,"▷","▶"),'
            f'IF(MOD({m},8)=5,IF({m}>=8,"◁","◀"),'
            f'IF({m}>=8,"□","■")))))')
        c = ws[f"{L}{r_}"]; c.border = BORDER_L
        c.font = Font(name=FONT, size=11, bold=True, color="12395B")
        c.alignment = Alignment(horizontal="center", vertical="center")
    MB0 = get_column_letter(MIRROR_B); MB1 = get_column_letter(MIRROR_B + 30)
    ws[f"AG{r_}"] = f'=IF($A{r_}="","",COUNTIF(${MB0}{r_}:${MB1}{r_},">0"))'
    ws[f"AG{r_}"].font = C_AUTO; ws[f"AG{r_}"].fill = F_AUTO; ws[f"AG{r_}"].border = BORDER
    ws[f"AG{r_}"].alignment = Alignment(horizontal="center", vertical="center")
ws["AG5"] = "当月"; ws["AG6"] = "利用日数"
for r_ in (5, 6):
    c = ws[f"AG{r_}"]; c.font = C_SUB; c.fill = F_SUB; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.column_dimensions["AG"].width = 10

BGRID = f"B7:AF{BAND_LAST}"
ws.conditional_formatting.add(BGRID, FormulaRule(
    formula=['AND(B7<>"",AP7>=8)'], fill=F_BANDT,
    font=Font(name=FONT, size=11, bold=True, color="6B6B7B"), stopIfTrue=True))
ws.conditional_formatting.add(BGRID, FormulaRule(
    formula=['AND(B7<>"",AP7>0)'], fill=F_BAND,
    font=Font(name=FONT, size=11, bold=True, color="12395B"), stopIfTrue=True))
ws.conditional_formatting.add(BGRID, FormulaRule(
    formula=['OR(B$6="土",B$6="日")'], fill=F_WKND))
ws.conditional_formatting.add(f"B5:AF6", FormulaRule(
    formula=['OR(B$6="土",B$6="日")'], fill=F_WKND))
for c in range(MIRROR_B, MIRROR_B + 31):
    ws.column_dimensions[get_column_letter(c)].hidden = True
ws.freeze_panes = "B7"
ws[f"A{BAND_LAST+2}"] = (f"※ 帯に出るのは 利用者マスタ の上から{BAND_ROWS}名を表示します。並び順を変えたいときは 利用者マスタ の行を入れ替えてください。")
ws[f"A{BAND_LAST+3}"] = "※ 「当月利用日数」は入所日〜退所日を数えた日数です。支給量の残は 08_支給量管理 で確認してください。"
ws[f"A{BAND_LAST+4}"] = "※ AP列より右は色分け用の隠し列です。編集しないでください。"
for k in (2, 3, 4):
    ws[f"A{BAND_LAST+k}"].font = C_NOTE

# ── 当月の予約明細（01_予約入力 の内容がそのまま入所日順に並ぶ）──────
# 帯の列幅は3.1しかないので、セルを結合して読める幅にする。
DET_H = BAND_LAST + 6          # 見出し行
DET_ROWS = 40
DET = [("氏名", 1, 1), ("部屋", 2, 5), ("区分", 6, 8), ("入所", 9, 14),
       ("退所", 15, 20), ("泊数", 21, 23), ("食事", 24, 27), ("状態", 28, 30),
       ("備考", 31, 33)]
ws[f"A{DET_H-1}"] = "■ 当月の予約明細（入所日順・01_予約入力 から自動）"
ws[f"A{DET_H-1}"].font = C_SUB
for lbl, c0, c1 in DET:
    if c1 > c0: ws.merge_cells(start_row=DET_H, start_column=c0, end_row=DET_H, end_column=c1)
    c = ws.cell(row=DET_H, column=c0, value=lbl)
    c.font = C_HEAD; c.fill = F_HEAD
    c.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(c0, c1 + 1): ws.cell(row=DET_H, column=cc).border = BORDER
ws.row_dimensions[DET_H].height = 20

MD = lambda dt: f'MONTH({dt})&"/"&DAY({dt})&"("&MID("日月火水木金土",WEEKDAY({dt}),1)&")"'
for i in range(DET_ROWS):
    r_ = DET_H + 1 + i
    K = f"$AP{r_}"; X = f"$AQ{r_}"
    ws[f"AP{r_}"] = f'=IFERROR(SMALL({SORTK_},{i+1}),"")'
    ws[f"AQ{r_}"] = f'=IF({K}="",0,IFERROR(MATCH({K},{SORTK_},0),0))'
    din = f'INDEX({IN_},{X})'; dout = f'INDEX({OUT_},{X})'
    bf, lf, df = f'INDEX({BF_},{X})', f'INDEX({LF_},{X})', f'INDEX({DF_},{X})'
    vals = {
        1:  f'=IF({X}=0,"",INDEX({NAME_},{X}))',
        2:  f'=IF({X}=0,"",IF(INDEX({ROOM_},{X})="","⚠未選択",INDEX({ROOM_},{X})))',
        6:  f'=IF({X}=0,"",INDEX({CAT_},{X}))',
        9:  (f'=IF({X}=0,"",{MD(din)}&" "&INDEX({TIN_},{X}))'),
        15: (f'=IF({X}=0,"",{MD(dout)}&" "&INDEX({TOUT_},{X}))'),
        21: f'=IF({X}=0,"",IF({dout}={din},"日帰り",{dout}-{din}&"泊"))',
        24: (f'=IF({X}=0,"",IF(AND({bf}="",{lf}="",{df}=""),"標準",'
             f'"朝"&IF({bf}="","−",{bf})&" 昼"&IF({lf}="","−",{lf})&" 夕"&IF({df}="","−",{df})))'),
        28: f'=IF({X}=0,"",IF(INDEX({ST_},{X})="","確定",INDEX({ST_},{X})))',
        31: f'=IF({X}=0,"",INDEX({NOTE_},{X}))',
    }
    for lbl, c0, c1 in DET:
        if c1 > c0: ws.merge_cells(start_row=r_, start_column=c0, end_row=r_, end_column=c1)
        cell = ws.cell(row=r_, column=c0, value=vals[c0])
        cell.font = C_AUTO; cell.fill = F_AUTO
        cell.alignment = Alignment(horizontal="center" if c0 in (2, 6, 21, 24, 28) else "left",
                                   vertical="center", shrink_to_fit=True)
        for cc in range(c0, c1 + 1): ws.cell(row=r_, column=cc).border = BORDER
DET_LAST = DET_H + DET_ROWS
ws.conditional_formatting.add(f"A{DET_H+1}:AF{DET_LAST}", FormulaRule(
    formula=[f'$B{DET_H+1}="⚠未選択"'], fill=F_WARN))
ws.conditional_formatting.add(f"A{DET_H+1}:AF{DET_LAST}", FormulaRule(
    formula=[f'OR($AB{DET_H+1}="仮予約",$AB{DET_H+1}="調整中")'], fill=F_TENT))
ws[f"A{DET_LAST+2}"] = ("※ この一覧は 01_予約入力 をそのまま入所日順に並べたものです。"
                        "対象月にかかる予約だけが出ます（月をまたぐ滞在も含みます）。")
ws[f"A{DET_LAST+3}"] = "※ 部屋が「⚠未選択」の行は、カレンダーの部屋の行に表示されません。01_予約入力 で部屋を選んでください。"
for k in (2, 3):
    ws[f"A{DET_LAST+k}"].font = C_NOTE

ws.print_area = f"A1:AG{DET_LAST}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True


# ══════════════════════════════════════════════════════════════════
# 05_空室照会
# ══════════════════════════════════════════════════════════════════
ws = sheet("05_空室照会")
title(ws, "空室照会 ―「この期間空いていますか」に即答する", "日付を2つ入れるだけです。対象月の指定は要りません。")
label_input(ws, 3, "開始日（入所日）", datetime.date(2026, 4, 6), "", "yyyy/mm/dd")
label_input(ws, 4, "終了日（退所予定日）", datetime.date(2026, 4, 9),
            "★退所日の夜は数えません（その夜は次の方が入れるため）", "yyyy/mm/dd")
label_auto (ws, 5, "泊数", '=IF(OR($B$3="",$B$4=""),"",MAX(0,$B$4-$B$3))',
            "0泊（同日）なら日帰り・体験利用です")
S_, E_ = "$B$3", "$B$4"
BOOKED = (f'((({OUTN_}<={E_})*{OUTN_}+({OUTN_}>{E_})*{E_})'
          f'-(({INN_}>={S_})*{INN_}+({INN_}<{S_})*{S_}))')
head_row(ws, 7, ["部屋", "空いている夜数", "ふさがっている夜数", "判定"], [18, 16, 20, 30])
for i in range(MAX_ROOMS):
    d = 8 + i
    ws[f"A{d}"] = f'=IF(COUNTA({RROW})>={i+1},INDEX({RROW},{i+1}),"")'
    ws[f"C{d}"] = (f'=IF(OR($A{d}="",{E_}<={S_}),"",'
                   f'SUMPRODUCT({VALID}*({ROOM_}=$A{d})*({BOOKED}>0)*{BOOKED}))')
    ws[f"B{d}"] = f'=IF($C{d}="","",MAX(0,{E_}-{S_}-$C{d}))'
    ws[f"D{d}"] = (f'=IF($A{d}="","",IF({E_}<={S_},"日付を確認してください",'
                   f'IF($B{d}={E_}-{S_},"○ 全日空き",IF($B{d}=0,"× 空きなし","△ 一部空き"))))')
    for col in range(1, 5):
        c = ws.cell(row=d, column=col); c.border = BORDER; c.font = C_AUTO; c.fill = F_AUTO
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{d}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"D{d}"].alignment = Alignment(horizontal="left", vertical="center")
for mark, fill in [("○", F_OK), ("△", F_WARN2), ("×", F_WARN)]:
    ws.conditional_formatting.add("D8:D11", FormulaRule(formula=[f'LEFT($D8,1)="{mark}"'], fill=fill))
ws["A13"] = "※ 仮予約・調整中も「ふさがっている」として数えます。"
ws["A14"] = "※ 日帰り（0泊）は夜をふさがないため、この照会には出ません。02_カレンダー の ◆ で確認してください。"
ws["A15"] = "※ 「一部空き」のときは 02_カレンダー を開けば、どの日が埋まっているかが一目で分かります。"
for r_ in (13, 14, 15): ws[f"A{r_}"].font = C_NOTE


# ══════════════════════════════════════════════════════════════════
# 06_食数表（厨房用）
# ══════════════════════════════════════════════════════════════════
ws = sheet("06_食数表")
title(ws, "食数表（厨房用）",
      "既定は 日帰り=昼のみ／初日=夕／中日=朝昼夕／最終日=朝。"
      "★個別の調整は 02_カレンダー の下部「食事の個別調整」で日付ごとに入れてください。")
label_auto(ws, 3, "対象月", f"={M1}", "★月の変更は 02_カレンダー で行ってください", "yyyy/mm/dd")
hdr = ["日付", "曜日", "朝食", "昼食", "夕食", "調整"] + [f"部屋{i}" for i in range(1, MAX_ROOMS + 1)]
head_row(ws, 5, hdr, [12, 6, 10, 10, 10, 8] + [24] * MAX_ROOMS)
for i in range(MAX_ROOMS):
    ws[f"{get_column_letter(7+i)}5"] = f'={MON}!{get_column_letter(3+i)}$6'
for i in range(31):
    d = 6 + i
    src = 7 + i
    ws[f"A{d}"] = f'=IF({MON}!$A{src}="","",{MON}!$A{src})'
    ws[f"B{d}"] = f'=IF($A{d}="","",MID("日月火水木金土",WEEKDAY($A{d}),1))'
    for k, col in enumerate(("C", "D", "E")):
        srccol = [MC_A, MC_L, MC_D][k]
        ws[f"{col}{d}"] = f'=IF($A{d}="","",{MON}!${srccol}{src})'
    ws[f"F{d}"] = f'=IF($A{d}="","",IF({MON}!${MB_H}{src}=0,"","※"))'
    for j in range(MAX_ROOMS):
        ws[f"{get_column_letter(7+j)}{d}"] = f'=IF($A{d}="","",{MON}!{get_column_letter(3+j)}{src})'
    for col in range(1, 7 + MAX_ROOMS):
        c = ws.cell(row=d, column=col); c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font, c.fill = C_AUTO, F_AUTO
    for col in ("C", "D", "E"):
        ws[f"{col}{d}"].font = Font(name=FONT, size=11, bold=True, color="1F3B52")
    ws[f"F{d}"].font = Font(name=FONT, size=10, bold=True, color="B03A2E")
    ws[f"A{d}"].number_format = "yyyy/mm/dd"
    for j in range(MAX_ROOMS):
        ws.cell(row=d, column=7 + j).alignment = Alignment(horizontal="left", vertical="center",
                                                            shrink_to_fit=True)
ws["A38"] = "月合計"; ws["A38"].font = C_SUB; ws["A38"].fill = F_SUB; ws["A38"].border = BORDER
for col in ("C", "D", "E"):
    c = ws[f"{col}38"]; c.value = f"=SUM({col}6:{col}36)"
    c.font = Font(name=FONT, size=11, bold=True); c.fill = F_SUB; c.border = BORDER
    c.alignment = Alignment(horizontal="center")
ws["G38"] = "「調整」に ※ が付いた日は、02_カレンダー で個別に指定した食数です"
ws["G38"].font = C_NOTE
ws.conditional_formatting.add("A6:B36", FormulaRule(formula=['AND($A6<>"",WEEKDAY($A6,2)>=6)'], fill=F_WKND))
ws.conditional_formatting.add(f"A6:{get_column_letter(6+MAX_ROOMS)}36",
    FormulaRule(formula=['$F6="※"'], fill=PatternFill("solid", fgColor="FBF3E4", bgColor="FBF3E4")))
ws.print_area = "A1:F38"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True


# ══════════════════════════════════════════════════════════════════
# 07_月次集計
# ══════════════════════════════════════════════════════════════════
ws = sheet("07_月次集計")
title(ws, "期間集計・稼働率（上司報告・請求突合用）",
      "期間を入れると稼働率・区分別・利用者別が出ます。キャンセルは除外。"
      "期間にかかった滞在は、期間内に重なる日数だけ数えます。")
label_input(ws, 3, "開始日", datetime.date(2026, 8, 1), "", "yyyy/mm/dd")
label_input(ws, 4, "終了日", datetime.date(2026, 8, 31), "", "yyyy/mm/dd")
label_auto (ws, 5, "期間の日数", '=IF(OR($B$3="",$B$4="",$B$4<$B$3),"",$B$4-$B$3+1)',
            "月単位でも年度単位でも指定できます（例 4/1〜3/31）")
S_, E_ = "$B$3", "$B$4"
DAYS_TERM = (f'((({OUTN_}<={E_})*{OUTN_}+({OUTN_}>{E_})*{E_})'
             f'-(({INN_}>={S_})*{INN_}+({INN_}<{S_})*{S_})+1)')
NIGHTS_TERM = (f'((({OUTN_}<={E_}+1)*{OUTN_}+({OUTN_}>{E_}+1)*({E_}+1))'
               f'-(({INN_}>={S_})*{INN_}+({INN_}<{S_})*{S_}))')
OVL_D = f'({IN_}<={E_})*({OUT_}>={S_})'
OVL_N = f'({IN_}<={E_})*({OUT_}>{S_})'
NIGHTS = f'{OVL_N}*({NIGHTS_TERM}>0)*{NIGHTS_TERM}'
CONF   = f'({ST_}<>"仮予約")*({ST_}<>"調整中")'          # 確定・利用済み・未記入
TENTV  = f'(({ST_}="仮予約")+({ST_}="調整中"))'

# ── 稼働の概要 ────────────────────────────────────────────
ws["A7"] = "■ 稼働の概要"; ws["A7"].font = C_SUB
head_row(ws, 8, ["期間の日数", "定員（室）", "提供できる\n延べ室数",
                 "延べ宿泊数\n（確定）", "稼働率\n（確定）",
                 "延べ宿泊数\n（仮予約）", "稼働率\n（仮予約を含む）"],
         [18, 13, 14, 14, 14, 14, 16])
SLOT = f'($B$5*{CAP})'
ws["A9"] = '=IF($B$5="","",$B$5&" 日")'
ws["B9"] = f'={CAP}'
ws["C9"] = f'=IF($B$5="","",{SLOT})'
ws["D9"] = f'=IF($B$5="","",SUMPRODUCT({VALID}*{CONF}*{NIGHTS}))'
ws["E9"] = f'=IF(OR($B$5="",{SLOT}=0),"",$D$9/{SLOT})'
ws["F9"] = f'=IF($B$5="","",SUMPRODUCT({VALID}*{TENTV}*{NIGHTS}))'
ws["G9"] = f'=IF(OR($B$5="",{SLOT}=0),"",($D$9+$F$9)/{SLOT})'
for col in range(1, 8):
    c = ws.cell(row=9, column=col); c.border = BORDER; c.fill = F_AUTO
    c.font = Font(name=FONT, size=12, bold=True, color="1F3B52")
    c.alignment = Alignment(horizontal="center", vertical="center")
ws["E9"].number_format = "0.0%"; ws["G9"].number_format = "0.0%"
ws.row_dimensions[9].height = 24
ws.conditional_formatting.add("E9:E9", FormulaRule(formula=['$E$9>=0.9'], fill=F_WARN))
ws.conditional_formatting.add("E9:E9", FormulaRule(formula=['$E$9>=0.7'], fill=F_WARN2))
ws["A10"] = ('※ 稼働率 ＝ 延べ宿泊数 ÷（定員 × 期間の日数）。「泊」で数えるので、'
             '退所日の夜は次の方が使える前提です。日帰り（0泊）は稼働率に入りません。')
ws["A11"] = '※ 「確定」は 確定・利用済み・状態未記入 の予約です。仮予約・調整中は右の2列で見込みとして示します。'
for r_ in (10, 11): ws[f"A{r_}"].font = C_NOTE

# ── 区分別 ────────────────────────────────────────────────
CAT_TOP = 14
ws[f"A{CAT_TOP-1}"] = "■ 区分別"; ws[f"A{CAT_TOP-1}"].font = C_SUB
head_row(ws, CAT_TOP, ["区分", "利用者数", "利用回数", "延べ利用日数", "延べ宿泊数", "稼働率"], None)
for i in range(CAT_ROWS):
    d = CAT_TOP + 1 + i
    ws[f"A{d}"] = f'=IF(COUNTA({CROW})>={i+1},INDEX({CROW},{i+1}),"")'
    ws[f"B{d}"] = (f'=IF($A{d}="","",SUMPRODUCT(--(COUNTIFS({NAME_},{UROW},{CAT_},$A{d},'
                   f'{IN_},"<="&{E_},{OUT_},">="&{S_},{ST_},"<>キャンセル")>0)*({UROW}<>"")))')
    ws[f"C{d}"] = (f'=IF($A{d}="","",COUNTIFS({CAT_},$A{d},{IN_},"<="&{E_},'
                   f'{OUT_},">="&{S_},{ST_},"<>キャンセル"))')
    ws[f"D{d}"] = f'=IF($A{d}="","",SUMPRODUCT({VALID}*({CAT_}=$A{d})*{OVL_D}*{DAYS_TERM}))'
    ws[f"E{d}"] = f'=IF($A{d}="","",SUMPRODUCT({VALID}*({CAT_}=$A{d})*{NIGHTS}))'
    ws[f"F{d}"] = f'=IF(OR($A{d}="",{SLOT}=0),"",$E{d}/{SLOT})'
    for col in range(1, 7):
        c = ws.cell(row=d, column=col); c.border = BORDER; c.font = C_AUTO; c.fill = F_AUTO
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{d}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"F{d}"].number_format = "0.0%"
d = CAT_TOP + 1 + CAT_ROWS
ws[f"A{d}"] = "合計"
ws[f"B{d}"] = (f'=SUMPRODUCT(--(COUNTIFS({NAME_},{UROW},{IN_},"<="&{E_},'
               f'{OUT_},">="&{S_},{ST_},"<>キャンセル")>0)*({UROW}<>""))')
for col in ("C", "D", "E"):
    ws[f"{col}{d}"] = f"=SUM({col}{CAT_TOP+1}:{col}{d-1})"
ws[f"F{d}"] = f'=IF({SLOT}=0,"",$E{d}/{SLOT})'
for col in range(1, 7):
    c = ws.cell(row=d, column=col); c.border = BORDER; c.font = C_BOLD; c.fill = F_SUB
    c.alignment = Alignment(horizontal="center", vertical="center")
ws[f"F{d}"].number_format = "0.0%"
ws[f"A{d+1}"] = ("※ 区分別の「延べ宿泊数」「稼働率」には仮予約・調整中も含みます。"
                 "確定だけの稼働率は上の「稼働の概要」をご覧ください。")
ws[f"A{d+2}"] = ("※ 「利用者数」は実人数（同じ方が何回利用しても1人）。"
                 "区分をまたぐ方がいる場合、区分別の合計と合計行は一致しません。")
for k in (1, 2): ws[f"A{d+k}"].font = C_NOTE

# ── 利用者別 ──────────────────────────────────────────────
USR_TOP = d + 4
ws[f"A{USR_TOP-1}"] = "■ 利用者別"; ws[f"A{USR_TOP-1}"].font = C_SUB
head_row(ws, USR_TOP, ["氏名", "利用回数", "延べ利用日数", "延べ宿泊数",
                       "支給量対象日数（滞在全体）"], None)
for i in range(USER_ROWS):
    d2 = USR_TOP + 1 + i
    ws[f"A{d2}"] = f'=IF(COUNTA({UROW})>={i+1},INDEX({UROW},{i+1}),"")'
    ws[f"B{d2}"] = (f'=IF($A{d2}="","",COUNTIFS({NAME_},$A{d2},{IN_},"<="&{E_},'
                    f'{OUT_},">="&{S_},{ST_},"<>キャンセル"))')
    ws[f"C{d2}"] = f'=IF($A{d2}="","",SUMPRODUCT({VALID}*({NAME_}=$A{d2})*{OVL_D}*{DAYS_TERM}))'
    ws[f"D{d2}"] = f'=IF($A{d2}="","",SUMPRODUCT({VALID}*({NAME_}=$A{d2})*{NIGHTS}))'
    ws[f"E{d2}"] = (f'=IF($A{d2}="","",SUMIFS({ALW_},{NAME_},$A{d2},'
                    f'{IN_},">="&{S_},{IN_},"<="&{E_}))')
    for col in range(1, 6):
        c = ws.cell(row=d2, column=col); c.border = BORDER; c.font = C_AUTO; c.fill = F_AUTO
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{d2}"].alignment = Alignment(horizontal="left", vertical="center")
for i, w in enumerate([20, 13, 14, 15, 15, 16, 17]):
    ws.column_dimensions[get_column_letter(1 + i)].width = w
ws.print_area = f"A1:G{USR_TOP + USER_ROWS}"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True



# ══════════════════════════════════════════════════════════════════
# 08_支給量管理
# ══════════════════════════════════════════════════════════════════
ws = sheet("08_支給量管理")
title(ws, "支給量の残日数（受給者証の月あたり上限）",
      "入所日の属する月で集計します。M_区分 で「○」を付けた区分だけを数えます。")
label_input(ws, 3, "対象月", MONTH0, "その月の日付ならどれでも可（02_カレンダー とは独立です）", "yyyy/mm/dd")
head_row(ws, 5, ["氏名", "支給量（日／月）", "当月の対象日数", "残り", "状態"], [22, 16, 16, 12, 18])
for i in range(USER_ROWS):
    d = 6 + i
    ws[f"A{d}"] = f'=IF(COUNTA({UROW})>={i+1},INDEX({UROW},{i+1}),"")'
    ws[f"B{d}"] = (f'=IF($A{d}="","",IF(SUMPRODUCT(({UROW}=$A{d})*({UALW}<>""))=0,"",'
                   f'SUMIFS({UALW},{UROW},$A{d})))')
    ws[f"C{d}"] = (f'=IF($A{d}="","",SUMIFS({ALW_},{NAME_},$A{d},'
                   f'{IN_},">="&DATE(YEAR($B$3),MONTH($B$3),1),{IN_},"<="&EOMONTH($B$3,0)))')
    ws[f"D{d}"] = f'=IF(OR($A{d}="",$B{d}=""),"",$B{d}-$C{d})'
    ws[f"E{d}"] = (f'=IF($A{d}="","",IF($B{d}="","（管理しない）",'
                   f'IF($D{d}<0,"⚠ 超過",IF($D{d}=0,"上限ちょうど",IF($D{d}<=1,"残りわずか","OK")))))')
    for col in range(1, 6):
        c = ws.cell(row=d, column=col); c.border = BORDER; c.font = C_AUTO; c.fill = F_AUTO
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{d}"].alignment = Alignment(horizontal="left", vertical="center")
last_u = 6 + USER_ROWS - 1
ws.conditional_formatting.add(f"A6:E{last_u}", FormulaRule(formula=['$E6="⚠ 超過"'], fill=F_WARN))
ws.conditional_formatting.add(f"A6:E{last_u}", FormulaRule(formula=['$E6="残りわずか"'], fill=F_WARN2))
ws[f"A{last_u+2}"] = "※ 月をまたぐ滞在は、入所日の属する月にまとめて計上します。月割りが必要な場合は 07_月次集計 を使ってください。"
ws[f"A{last_u+2}"].font = C_NOTE


# ══════════════════════════════════════════════════════════════════
# 09_FAX空き表（印刷用・氏名は載せない）
# ══════════════════════════════════════════════════════════════════
ws = sheet("09_FAX空き表")
ws.sheet_view.showGridLines = False
FX_LAST = 44
ws.column_dimensions["A"].width = 11
for c in range(2, 9):
    ws.column_dimensions[get_column_letter(c)].width = 10.4

MED  = Side(style="medium", color="333333")
THN  = Side(style="thin",   color="808080")
DOT  = Side(style="hair",   color="AAAAAA")
def box(r0, c0, r1, c1, outer=MED, inner=None):
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=outer if c == c0 else (inner or Border().left),
                right=outer if c == c1 else (inner or Border().right),
                top=outer if r == r0 else (inner or Border().top),
                bottom=outer if r == r1 else (inner or Border().bottom))
def merge(r, c0, c1, value=None, font=None, align=None, fill=None, fmt=None):
    if c1 > c0: ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
    cell = ws.cell(row=r, column=c0)
    if value is not None: cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    cell.alignment = align or Alignment(horizontal="left", vertical="center")
    return cell

F_TITLE = Font(name=FONT, size=16, bold=True, color="1F3B52")
F_LBL   = Font(name=FONT, size=10, bold=True, color="FFFFFF")
F_TXT   = Font(name=FONT, size=10)
F_BIG   = Font(name=FONT, size=10.5)
F_SMALL = Font(name=FONT, size=8.5, color="555555")
F_SEC   = Font(name=FONT, size=11, bold=True, color="1F3B52")
FILL_LBL = PatternFill("solid", fgColor="2F6690", bgColor="2F6690")
FILL_BOX = PatternFill("solid", fgColor="F7F9FB", bgColor="F7F9FB")

# ── 表題 ───────────────────────────────────────────────────
merge(2, 1, 8, "Ｆ Ａ Ｘ　送　信　票", F_TITLE,
      Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[2].height = 30
merge(3, 1, 8, '=' + YM + '&"　短期入所（ショートステイ）空き状況のご案内"',
      Font(name=FONT, size=11, bold=True),
      Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[3].height = 20
for c in range(1, 9):
    ws.cell(row=3, column=c).border = Border(bottom=MED)

# ── 送信先 ─────────────────────────────────────────────────
merge(5, 1, 1, "送 信 先", F_LBL, Alignment(horizontal="center", vertical="center"), FILL_LBL)
ws.merge_cells(start_row=5, start_column=1, end_row=8, end_column=1)
RCP  = f"'M_FAX送付先'!$A$4:$A$24"
RCP2 = f"'M_FAX送付先'!$B$4:$B$24"
RCP3 = f"'M_FAX送付先'!$C$4:$C$24"
IDX  = f'IFERROR(MATCH($B$5,{RCP},0),0)'
merge(5, 2, 5, None, F_TXT, Alignment(horizontal="left", vertical="center"), F_INPUT)
ws["B5"].comment = Comment("M_FAX送付先 に登録した事業所を選ぶと、御中・ご担当・FAX番号が自動で入ります。\n"
                           "空欄のままにすると、手書き用の下線が印刷されます。", "設計メモ",
                           height=90, width=340)
merge(6, 2, 5, f'=IF($B$5="","＿＿＿＿＿＿＿＿＿＿＿＿＿＿＿　御中",$B$5&"　御中")',
      Font(name=FONT, size=12, bold=True), Alignment(horizontal="left", vertical="center"))
merge(7, 2, 5, (f'=IF($B$5="","ご担当　＿＿＿＿＿＿＿＿　様",'
                f'IF(IFERROR(INDEX({RCP2},{IDX}),"")="","","ご担当　"&INDEX({RCP2},{IDX})&"　様"))'),
      F_TXT, Alignment(horizontal="left", vertical="center"))
merge(8, 2, 5, (f'=IF($B$5="","ＦＡＸ　＿＿＿＿＿＿＿＿＿＿",'
                f'"ＦＡＸ　"&IFERROR(INDEX({RCP3},{IDX}),""))'),
      F_TXT, Alignment(horizontal="left", vertical="center"))
merge(5, 6, 6, "送信日", F_TXT, Alignment(horizontal="center", vertical="center"), FILL_BOX)
merge(5, 7, 8, '=YEAR(TODAY())&"年"&MONTH(TODAY())&"月"&DAY(TODAY())&"日"', F_TXT,
      Alignment(horizontal="center", vertical="center"))
merge(6, 6, 6, "枚　数", F_TXT, Alignment(horizontal="center", vertical="center"), FILL_BOX)
merge(6, 7, 8, "本票を含め　1 枚", F_TXT, Alignment(horizontal="center", vertical="center"))
merge(7, 6, 8, '="全"&' + CAP + '&"室の施設です"', F_SMALL,
      Alignment(horizontal="center", vertical="center"))
merge(8, 6, 8, "利用者名は記載しておりません", F_SMALL,
      Alignment(horizontal="center", vertical="center"))
box(5, 1, 8, 8, MED, THN)
for r in range(5, 9): ws.row_dimensions[r].height = 19

# ── 送信元 ─────────────────────────────────────────────────
merge(10, 1, 1, "送 信 元", F_LBL, Alignment(horizontal="center", vertical="center"), FILL_LBL)
ws.merge_cells(start_row=10, start_column=1, end_row=13, end_column=1)
merge(10, 2, 8, f'={MSET}!$B$4&"　"&{MSET}!$B$5',
      Font(name=FONT, size=12, bold=True), Alignment(horizontal="left", vertical="center"))
UL = "＿＿＿＿＿＿＿＿＿＿"
merge(11, 2, 8, (f'=IF({MSET}!$B$11="","{UL}{UL}",'
                 f'IF({MSET}!$B$10="","","〒"&{MSET}!$B$10&"　")&{MSET}!$B$11)'),
      F_TXT, Alignment(horizontal="left", vertical="center"))
merge(12, 2, 8, (f'="ＴＥＬ　"&IF({TEL}="","{UL}",{TEL})'
                 f'&"　　ＦＡＸ　"&IF({FAXNO}="","{UL}",{FAXNO})'),
      F_TXT, Alignment(horizontal="left", vertical="center"))
merge(13, 2, 8, (f'="担当　"&IF({MSET}!$B$14="","{UL}",{MSET}!$B$14)'),
      F_TXT, Alignment(horizontal="left", vertical="center"))
box(10, 1, 13, 8, MED, THN)
for r in range(10, 14): ws.row_dimensions[r].height = 19

# ── 本文 ───────────────────────────────────────────────────
merge(15, 1, 8, "平素より大変お世話になっております。", F_BIG,
      Alignment(horizontal="left", vertical="center"))
merge(16, 1, 8,
      '=' + YM + '&"の短期入所の空き状況は下記のとおりです。'
      'ご確認のほど、よろしくお願いいたします。"',
      F_BIG, Alignment(horizontal="left", vertical="center"))
for r in (15, 16): ws.row_dimensions[r].height = 18

# ── 空き状況カレンダー ────────────────────────────────────────
merge(18, 1, 8, '="■　"&' + YM + '&"　空き状況"', F_SEC,
      Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[18].height = 20
WDF = ["日", "月", "火", "水", "木", "金", "土"]
for c in range(7):
    L = get_column_letter(2 + c)
    cell = ws[f"{L}19"]; cell.value = WDF[c]
    cell.font = Font(name=FONT, size=10, bold=True,
                     color="FFFFFF" if 0 < c < 6 else "FFFFFF")
    cell.fill = PatternFill("solid",
                            fgColor="A93226" if c == 0 else ("1F618D" if c == 6 else "2F6690"),
                            bgColor="A93226" if c == 0 else ("1F618D" if c == 6 else "2F6690"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
merge(19, 1, 1, "", None, Alignment(horizontal="center", vertical="center"), FILL_BOX)
ws.row_dimensions[19].height = 19
for w in range(6):
    rd, rm = 20 + w * 2, 21 + w * 2
    merge(rd, 1, 1, f"第{w+1}週", F_SMALL, Alignment(horizontal="center", vertical="center"), FILL_BOX)
    ws.merge_cells(start_row=rd, start_column=1, end_row=rm, end_column=1)
    for c in range(7):
        L = get_column_letter(2 + c)
        n = w * 7 + c
        d = ws[f"{L}{rd}"]
        d.value = (f'=IF(MONTH({M1}-WEEKDAY({M1})+1+{n})<>MONTH({M1}),"",'
                   f'{M1}-WEEKDAY({M1})+1+{n})')
        d.number_format = "d"
        d.font = Font(name=FONT, size=9, bold=True,
                      color="A93226" if c == 0 else ("1F618D" if c == 6 else "444444"))
        d.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        m = ws[f"{L}{rm}"]
        m.value = (f'=IF({L}{rd}="","",IFERROR(INDEX({MON}!${SYM_C}$7:${SYM_C}$37,'
                   f'MATCH({L}{rd},{MON}!$A$7:$A$37,0)),""))')
        m.font = Font(name=FONT, size=16, bold=True, color="222222")
        m.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rd].height = 14
    ws.row_dimensions[rm].height = 24
box(19, 1, 31, 8, MED, THN)
mark_cf(ws, "B21:H31")

# ── 凡例・注意書き・集計 ──────────────────────────────────────
LEG = [("○", "空室あり"), ("△", "残りわずか"), ("仮", "仮予約で調整中"), ("×", "満室")]
merge(32, 1, 1, "凡　例", Font(name=FONT, size=9, bold=True),
      Alignment(horizontal="center", vertical="center"), FILL_BOX)
for i, (mk, tx) in enumerate(LEG):
    c0 = 2 + i * 2
    cm = merge(32, c0, c0, mk, Font(name=FONT, size=12, bold=True),
               Alignment(horizontal="center", vertical="center"))
    merge(32, c0 + 1, c0 + 1, tx, F_SMALL, Alignment(horizontal="left", vertical="center"))
box(32, 1, 32, 8, THN, THN)
ws.row_dimensions[32].height = 20
NOTES = [
    "・各日の欄は「その日の夜（宿泊）」の空き状況です。退所日の当日は、次の方がご利用いただけます。",
    "・空き状況は日々変動いたします。ご予約の前に、必ずお電話でご確認くださいますようお願いいたします。",
    '="・全"&' + CAP + '&'
    '"室の施設です。本用紙に利用者様のお名前は記載しておりません。"',
]
for i, tx in enumerate(NOTES):
    merge(33 + i, 1, 8, tx, F_SMALL, Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[33 + i].height = 14
merge(36, 1, 8,
      '="【月内の集計】　空室あり "&COUNTIF($B$21:$H$31,"○")&" 日　／　残りわずか "'
      '&COUNTIF($B$21:$H$31,"△")&" 日　／　仮予約 "&COUNTIF($B$21:$H$31,"仮")'
      '&" 日　／　満室 "&COUNTIF($B$21:$H$31,"×")&" 日"',
      Font(name=FONT, size=9.5, bold=True), Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[36].height = 18

# ── 連絡事項（自由記入）──────────────────────────────────────
merge(38, 1, 8, "■　連絡事項", F_SEC, Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[38].height = 20
ws.merge_cells(start_row=39, start_column=1, end_row=42, end_column=8)
cell = ws.cell(row=39, column=1)
cell.fill = F_INPUT; cell.font = F_TXT
cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
cell.comment = Comment("送付先へのひとこと（任意）。空欄のままでも印刷できます。",
                       "設計メモ", height=60, width=300)
box(39, 1, 42, 8, MED, None)
for r in range(39, 43): ws.row_dimensions[r].height = 16
merge(44, 1, 8,
      "※ 本状が誤って届いた場合は、お手数ですが上記の連絡先までご一報のうえ、ご破棄くださいますようお願いいたします。",
      F_SMALL, Alignment(horizontal="left", vertical="center"))

dvf = DataValidation(type="list", formula1="=送付先一覧", allow_blank=True, showDropDown=False)
dvf.promptTitle = "送信先"; dvf.prompt = "M_FAX送付先 から選びます。空欄なら手書き用の下線が出ます。"
dvf.showInputMessage = True
ws.add_data_validation(dvf); dvf.add("B5")

ws.print_area = f"A1:H{FX_LAST}"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.5
ws.page_margins.top = ws.page_margins.bottom = 0.5
ws.print_options.horizontalCentered = True



order = ["00_使い方", "01_予約入力", "02_カレンダー", "03_帯表", "04_月間表", "05_空室照会",
         "06_食数表", "07_月次集計", "08_支給量管理", "09_FAX空き表",
         "M_施設設定", "M_部屋", "利用者マスタ", "M_区分", "M_送迎", "M_状態", "M_FAX送付先"]
wb._sheets = [wb[n] for n in order]
wb.active = wb.index(wb["02_カレンダー"])
wb.save(OUT)
print("saved:", OUT)
