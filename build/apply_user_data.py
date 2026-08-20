# -*- coding: utf-8 -*-
"""生成したテンプレートに、お客様が入力済みの内容を移し替える。

移す対象:
  M_部屋 / 利用者マスタ / 01_予約入力 の予約 / 各シートの対象月
日付は「年月＋日にち」形式へ変換する。月をまたぐ退所日は日にちだけで表現できる
（退所日が入所日より前なら翌月と解釈される）ことを利用する。
"""
import openpyxl, datetime
from openpyxl.styles import Alignment

SRC = "ショートステイ管理台帳.xlsx"
DST = "ショートステイ管理台帳.xlsx"

ROOMS  = ["１Ｆ"]
USERS  = [("◎◎様", 7), ("木村様", 5), ("伊藤メイ様", None), ("小澤様", None), ("小林様", None)]
MONTH  = datetime.date(2026, 8, 1)

# (氏名, 区分, 部屋, 入所日, 入所時, 退所日, 退所時, 状態, 送迎, 送迎方法, 場所, 備考)
# 食事は全件「標準」（空欄）のまま。必要になったら L/M/N 列で ○/× を選ぶ。
# 時刻は「時」の数字だけ。分が必要なものは "16:30" のように文字で残す。
BOOKINGS = [
    ("◎◎様",    "SS", "１Ｆ", datetime.date(2026,8,1),  16, datetime.date(2026,8,2),  9, "確定", "あり", "施設車", "自宅前 16:00", ""),
    ("木村様",    "SS", "１Ｆ", datetime.date(2026,8,5),  "16:30", datetime.date(2026,8,7),  9, "確定", "なし", "", "", ""),
    ("◎◎様",    "SS", "１Ｆ", datetime.date(2026,8,8),  10, datetime.date(2026,8,8),  16, "確定", "あり", "家族送迎", "", "日帰り（体験利用）"),
    ("小澤様",    "SS", "１Ｆ", datetime.date(2026,8,14), 16, datetime.date(2026,8,16), 9, "確定", "調整中", "", "", ""),
    ("木村様",    "SS", "１Ｆ", datetime.date(2026,8,19), 15, datetime.date(2026,8,21), "", "", "", "", "", ""),
    ("伊藤メイ様", "SS", "１Ｆ", datetime.date(2026,8,21), 15, datetime.date(2026,8,22), "", "", "", "", "", ""),
    ("小澤様",    "SS", "１Ｆ", datetime.date(2026,8,22), 15, datetime.date(2026,8,23), "", "", "", "", "", ""),
    ("小林様",    "SS", "１Ｆ", datetime.date(2026,8,29), 15, datetime.date(2026,8,30), "", "", "", "", "", ""),
    ("木村様",    "SS", "１Ｆ", datetime.date(2026,8,31), 15, datetime.date(2026,9,1),  "", "", "", "", "", ""),
]

wb = openpyxl.load_workbook(SRC)

# ── M_部屋 ────────────────────────────────────────────────
ws = wb["M_部屋"]
for r in range(4, 21):
    ws.cell(row=r, column=1).value = None
    ws.cell(row=r, column=2).value = None
for i, nm in enumerate(ROOMS):
    ws.cell(row=4 + i, column=1, value=nm)

# ── 利用者マスタ ──────────────────────────────────────────
ws = wb["利用者マスタ"]
for r in range(4, 61):
    for c in (1, 2, 3): ws.cell(row=r, column=c).value = None
for i, (nm, alw) in enumerate(USERS):
    ws.cell(row=4 + i, column=1, value=nm)
    if alw is not None: ws.cell(row=4 + i, column=2, value=alw)

# ── 01_予約入力 ───────────────────────────────────────────
ws = wb["01_予約入力"]
INPUT_COLS = ["B","C","D","E","F","H","I","K","L","M","N","Q","R","S","T","U","V","W","X"]
for r in range(5, 205):
    for c in INPUT_COLS: ws[f"{c}{r}"].value = None

for i, bk in enumerate(BOOKINGS):
    r = 5 + i
    name, cat, room, din, tin, dout, tout, st, pk, pkm, pkl, note = bk
    ws[f"B{r}"] = name
    ws[f"C{r}"] = cat
    if room: ws[f"D{r}"] = room
    if i == 0:                       # 年月は最初の行だけ。以降は引き継ぎ
        ws[f"E{r}"] = MONTH
    ws[f"F{r}"] = din.day            # 入所は日にちだけ
    ws[f"I{r}"] = dout.day           # 退所も日にちだけ（翌月は自動判定）
    if tin != "":  ws[f"H{r}"] = tin
    if tout != "": ws[f"K{r}"] = tout
    if st:   ws[f"Q{r}"] = st
    if pk:   ws[f"R{r}"] = pk
    if pkm:  ws[f"S{r}"] = pkm
    if pkl:  ws[f"T{r}"] = pkl
    if note: ws[f"V{r}"] = note

# ── 各シートの対象月・期間をお客様の月に合わせる ─────────────
wb["02_カレンダー"]["B3"] = MONTH
# 07・08 は空欄のままにして、02_カレンダー の対象月に連動させる
wb["08_支給量管理"]["B3"] = None
wb["07_月次集計"]["B3"] = None
wb["07_月次集計"]["B4"] = None
wb["05_空室照会"]["B3"] = datetime.date(2026, 8, 24)
wb["05_空室照会"]["B4"] = datetime.date(2026, 8, 27)

wb.save(DST)
print("applied user data ->", DST)
