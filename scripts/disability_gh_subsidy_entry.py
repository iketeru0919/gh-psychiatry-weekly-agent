# Google Colab用: 障がいGH 加算算定状況を施設別原本へ入力するコード
#
# 使い方:
# 1. Google Driveに施設別の利用状況ファイル（例: 宇都宮⑤.xlsx）を同じフォルダへ入れる
# 2. TEMPLATE_FILE に「障がいGH_事業所名(原本).xlsx」のパスを指定する
# 3. FOLDER_PATH / OUTPUT_FOLDER を環境に合わせて指定して実行する
#
# 入力対象:
# - 福祉専門職員配置等加算 I/II/III
# - 人員配置体制加算
# - 夜勤職員加配加算
# - 重度障害者支援加算 I/II/III
# - 医療連携体制加算 VII
#
# 夜勤・医療の算定数は「取得欄に●がある人数」としてカウントします。
# 感染対策向上加算は入力しません。
#
# 出力:
# - 施設別に入力済みの原本コピー（障がいGH_<施設名>.xlsx）
# - 処理結果一覧.xlsx
# - 上記すべてを1つにまとめたzipファイル（障がいGH_出力一式_YYYYMMDD_HHMMSS.zip）

import os
import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

try:
    from google.colab import drive
except ImportError:
    drive = None


# ============================================================
# 1. Google Driveをマウント
# ============================================================
if drive is not None:
    drive.mount("/content/drive")


# ============================================================
# 2. 設定
# ============================================================
FOLDER_PATH = Path(os.environ.get("FOLDER_PATH", "/content/drive/MyDrive/施設実績データ"))
TEMPLATE_FILE = Path(
    os.environ.get("TEMPLATE_FILE", "/content/drive/MyDrive/障がいGH_事業所名(原本).xlsx")
)
OUTPUT_FOLDER = Path(os.environ.get("OUTPUT_FOLDER", "/content/drive/MyDrive/障がいGH_入力済"))

# 原本の年度ブロックに対応するシートだけ処理します。
# R8.4=2026年4月、R8.5=2026年5月、...、R9.3=2027年3月
MONTH_ROW_MAP = {
    "R8.4": (5, 6),
    "R8.5": (9, 10),
    "R8.6": (13, 14),
    "R8.7": (17, 18),
    "R8.8": (21, 22),
    "R8.9": (25, 26),
    "R8.10": (29, 30),
    "R8.11": (33, 34),
    "R8.12": (37, 38),
    "R9.1": (41, 42),
    "R9.2": (45, 46),
    "R9.3": (49, 50),
}

EXCLUDE_SHEETS = {"【入力例】", "入力例", "テンプレート", "template"}
MARK = "●"
ROMAN_I = "Ⅰ"
ROMAN_II = "Ⅱ"
ROMAN_III = "Ⅲ"
ROMAN_V = "Ⅴ"
ROMAN_VII = "Ⅶ"


# ============================================================
# 3. 共通関数
# ============================================================
def clean_facility_name(file_path: Path) -> str:
    """出力ファイル名に使う施設名。例: 宇都宮⑤.xlsx -> 宇都宮"""
    name = file_path.stem
    name = re.sub(r"[①-⑳ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ0-9]+$", "", name)
    return name.strip() or file_path.stem


def build_header_map(ws, header_row=3):
    """3行目の見出しから列番号を作る。同じ「取得」「単位」は使わず、固有見出しだけ使う。"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        headers[str(value).strip()] = col
    return headers


def find_total_row(ws):
    """A列に「合計」がある行を探す。見つからない場合は最終行+1を返す。"""
    for row in range(ws.max_row, 1, -1):
        if "合計" in str(ws.cell(row, 1).value or ""):
            return row
    return ws.max_row + 1


def get_person_rows(ws):
    """利用者名が入っている行を対象にする。"""
    total_row = find_total_row(ws)
    rows = []
    for row in range(4, total_row):
        name = str(ws.cell(row, 2).value or "").strip()
        if name:
            rows.append(row)
    return rows


def cell_text(ws, row, col):
    if col is None:
        return ""
    return str(ws.cell(row, col).value or "").strip()


def count_equal(ws, rows, col, value):
    if col is None:
        return 0
    return sum(1 for row in rows if cell_text(ws, row, col) == value)


def count_mark(ws, rows, col):
    return count_equal(ws, rows, col, MARK)


def extract_month_data(ws):
    """原本に必要な項目だけ抽出する。"""
    headers = build_header_map(ws)
    rows = get_person_rows(ws)

    col_welfare = headers.get("福祉専門")
    col_staff_x = headers.get("人員配置X")
    col_staff_y = headers.get("人員配置Y")
    col_night = headers.get("夜勤加配")
    col_heavy = headers.get("重度")
    col_medical = headers.get("医療Ⅶ")

    return {
        "利用者数": len(rows),
        "福祉専門I": count_equal(ws, rows, col_welfare, ROMAN_I),
        "福祉専門II": count_equal(ws, rows, col_welfare, ROMAN_II),
        "福祉専門III": count_equal(ws, rows, col_welfare, ROMAN_III),
        "人員配置X_V": count_equal(ws, rows, col_staff_x, ROMAN_V),
        "人員配置Y_VII": count_equal(ws, rows, col_staff_y, ROMAN_VII),
        "夜勤人数": count_mark(ws, rows, col_night),
        "重度I": count_equal(ws, rows, col_heavy, ROMAN_I),
        "重度II": count_equal(ws, rows, col_heavy, ROMAN_II),
        "重度III": count_equal(ws, rows, col_heavy, ROMAN_III),
        "医療VII人数": count_mark(ws, rows, col_medical),
    }


def fill_template(template_file, facility_name, extracted_by_month):
    """施設ごとの原本コピーに入力する。"""
    wb = openpyxl.load_workbook(template_file)
    ws = wb["障がいGH"] if "障がいGH" in wb.sheetnames else wb.active

    # シート名とA1は施設名にしておく
    ws.title = facility_name[:31]

    for sheet_name, data in extracted_by_month.items():
        if sheet_name not in MONTH_ROW_MAP:
            continue

        count_row, user_row = MONTH_ROW_MAP[sheet_name]

        # 算定数（実数）回
        ws.cell(count_row, 3, data["福祉専門I"])
        ws.cell(count_row, 4, data["福祉専門II"])
        ws.cell(count_row, 5, data["福祉専門III"])
        ws.cell(count_row, 6, data["人員配置X_V"])
        ws.cell(count_row, 7, data["人員配置Y_VII"])
        ws.cell(count_row, 8, data["夜勤人数"])
        ws.cell(count_row, 9, data["重度I"])
        ws.cell(count_row, 10, data["重度II"])
        ws.cell(count_row, 11, data["重度III"])
        ws.cell(count_row, 12, data["医療VII人数"])

        # 契約（利用）数 人
        # 感染対策向上加算(M/N)は入力しない
        for col in range(3, 13):
            ws.cell(user_row, col, data["利用者数"])

    return wb


def process_facility_file(file_path):
    facility_name = clean_facility_name(file_path)
    src_wb = openpyxl.load_workbook(file_path, data_only=True)

    extracted_by_month = {}
    for sheet_name in src_wb.sheetnames:
        if sheet_name in EXCLUDE_SHEETS:
            continue
        if sheet_name not in MONTH_ROW_MAP:
            continue
        extracted_by_month[sheet_name] = extract_month_data(src_wb[sheet_name])

    if not extracted_by_month:
        return None, {
            "施設名": facility_name,
            "ファイル": file_path.name,
            "結果": "対象シートなし",
        }

    out_wb = fill_template(TEMPLATE_FILE, facility_name, extracted_by_month)
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_FOLDER / f"障がいGH_{facility_name}.xlsx"
    out_wb.save(output_file)

    return output_file, {
        "施設名": facility_name,
        "ファイル": file_path.name,
        "結果": "作成済",
        "対象月": ", ".join(extracted_by_month.keys()),
        "出力先": str(output_file),
    }


def create_output_zip(output_files):
    """出力済みファイル一式を1つのzipにまとめる。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_path = OUTPUT_FOLDER / f"障がいGH_出力一式_{timestamp}.zip"

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in output_files:
            if file_path is None or not Path(file_path).exists():
                continue
            zf.write(file_path, arcname=Path(file_path).name)

    return zip_path


def main():
    results = []
    output_files = []

    for file_path in sorted(FOLDER_PATH.glob("*.xlsx")):
        if file_path.name.startswith("~$"):
            continue
        if file_path.resolve() == TEMPLATE_FILE.resolve():
            continue
        try:
            output_file, result = process_facility_file(file_path)
            if output_file is not None:
                output_files.append(output_file)
            results.append(result)
            print(f"{result['施設名']}: {result['結果']}")
        except Exception as exc:
            results.append(
                {
                    "施設名": clean_facility_name(file_path),
                    "ファイル": file_path.name,
                    "結果": "エラー",
                    "エラー内容": str(exc),
                }
            )
            print(f"{file_path.name}: エラー - {exc}")

    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    summary_file = OUTPUT_FOLDER / "処理結果一覧.xlsx"
    pd.DataFrame(results).to_excel(summary_file, index=False)
    output_files.append(summary_file)

    zip_path = create_output_zip(output_files)

    print("\n完了")
    print(f"処理結果一覧: {summary_file}")
    print(f"圧縮フォルダ: {zip_path}")


# ============================================================
# 4. 実行
# ============================================================
if __name__ == "__main__":
    main()
