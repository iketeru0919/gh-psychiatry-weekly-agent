"""
内部監査チェックリスト 改善スクリプト
改善内容:
1. ヘッダー拡充（監査実施者・前回監査日・是正確認欄）
2. 判定コード凡例の追加
3. 列追加：指摘内容・是正期限・担当者・是正確認日
4. 既存項目の修正（表記ゆれ・GH等→GH・大規模住居の定義具体化等）
5. 不足項目の追加
6. サマリー欄の追加
"""
import copy
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = '/root/.claude/uploads/1d33ad44-761d-53a1-bc4c-9772bc5d1911/eac2e4ef-_____________.xlsx'
DST = '/home/user/gh-psychiatry-weekly-agent/内部監査チェックリスト_改善版.xlsx'

# ── カラーパレット ──────────────────────────────────────────
COL_HEADER   = 'FF1F4E79'   # 濃紺（タイトル行背景）
COL_SUBHEADER= 'FF2E75B6'   # 青（列ヘッダー背景）
COL_S        = 'FFFCE4D6'   # 薄オレンジ（S優先度）
COL_A        = 'FFE2EFDA'   # 薄緑（A優先度）
COL_B        = 'FFFFF2CC'   # 薄黄（B優先度）
COL_NEW_COLS = 'FFF2F2F2'   # 薄グレー（追加列背景）
COL_LEGEND   = 'FFDAE3F3'   # 薄青（凡例行）
COL_SUMMARY  = 'FFEAF4FB'   # 淡青（サマリー欄）

WHITE        = 'FFFFFFFF'
BLACK        = 'FF000000'

def fill(rgb):
    return PatternFill('solid', fgColor=rgb)

def font(bold=False, color=BLACK, size=10):
    return Font(bold=bold, color=color, size=size, name='游ゴシック')

def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style='thin', color='FFB0B0B0')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, bg=None, h_align='left',
             v_align='center', wrap=True, font_size=10, font_color=BLACK):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, size=font_size, name='游ゴシック')
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h_align, v_align, wrap)
    c.border = thin_border()
    return c

# ── 優先度に対応する背景色 ──────────────────────────────────
PRIORITY_COLOR = {'S': COL_S, 'A': COL_A, 'B': COL_B}

# ── 新規追加チェック項目 ────────────────────────────────────
NEW_ITEMS = [
    # (優先度, サービス, 分類, 監査項目, 確認内容, 確認資料, 確認方法)
    ('S', '共通', '人員配置',
     '人員欠如発生時の行政報告・減算処理確認',
     '人員欠如が発生した場合、速やかに都道府県・市町村へ報告し、翌月以降の請求に減算を反映しているか。欠如解消後の回復届も適切に行われているか。',
     '人員欠如報告書・変更届・国保連請求書',
     'Dropbox'),
    ('A', '共通', '事故・ヒヤリハット',
     '事故報告書・ヒヤリハット記録の管理・行政報告',
     '事故発生時に速報・詳細報告書を作成し、市町村・都道府県への報告が行われているか。ヒヤリハット記録を蓄積し再発防止策を講じているか。',
     '事故報告書・ヒヤリハット記録・行政への報告控え',
     'Dropbox+現地'),
    ('A', '共通', '緊急時対応',
     '緊急連絡体制・夜間対応手順書の整備確認',
     '緊急連絡網・夜間緊急対応マニュアルが最新化されているか。職員全員に周知されているか。',
     '緊急連絡網・夜間対応マニュアル・周知記録',
     'Dropbox+現地'),
    ('A', '共通', '労務',
     '健康診断の実施記録確認',
     '雇い入れ時健康診断・定期健康診断（年1回）が全職員に実施されているか。結果記録を保管しているか。',
     '健康診断結果票・受診記録',
     'Dropbox'),
    ('A', '共通', '個人情報',
     '退職者の個人情報管理（返却・廃棄）確認',
     '退職時に貸与物（PC・カード等）の返却、業務上知り得た個人情報の廃棄・消去が記録されているか。',
     '退職時チェックリスト・返却記録・廃棄証明書',
     'Dropbox+現地'),
    ('A', 'GH', '加算・減算',
     '2024年改定加算の算定要件確認（強度行動障害・集中的支援等）',
     '2024年障害福祉サービス等報酬改定で新設・変更された加算（強度行動障害支援者養成研修加算・集中的支援加算・重度障害者支援加算等）を算定している場合、各加算の要件（研修修了・アセスメント・支援計画等）を満たしているか。',
     '研修修了証・アセスメント記録・加算別支援計画',
     'Dropbox'),
    ('A', 'SS', '加算・減算',
     '緊急短期入所加算の要件確認',
     '緊急短期入所加算を算定している場合、緊急受入体制（空床確保）・緊急連絡先の整備・受入記録が揃っているか。',
     '緊急受入記録・緊急連絡体制・国保連請求書',
     'Dropbox'),
    ('B', '共通', '契約・重説',
     '重要事項説明書の改訂日・法改正追従確認',
     '直近の法改正・報酬改定を反映した重説の改訂がなされているか。改訂履歴と利用者への説明記録があるか。',
     '重要事項説明書（最新版）・改訂履歴・説明記録',
     'Dropbox+原本'),
    ('B', '共通', '内部統制',
     '前回監査指摘事項の是正確認',
     '前回の内部監査・行政指導で指摘された事項について、是正措置が完了しているか。未了の場合、対応状況を記録しているか。',
     '前回監査報告書・是正報告書',
     'Dropbox'),
    ('B', 'GH', '設備基準',
     '消防設備点検・防火管理者の確認',
     '消防設備点検（年2回：機器点検・総合点検）の記録があるか。防火管理者が選任・届出されているか。消防計画が最新化されているか。',
     '消防設備点検報告書・防火管理者選任届・消防計画',
     'Dropbox+現地'),
]

def get_row_color(priority):
    return PRIORITY_COLOR.get(str(priority).strip(), 'FFFFFFFF')

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = copy.copy(src_cell.font)
        dst_cell.fill      = copy.copy(src_cell.fill)
        dst_cell.border    = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)

def main():
    wb = openpyxl.load_workbook(SRC)

    # ── シート1: 内部監査チェックリスト を改修 ────────────────
    ws = wb['内部監査チェックリスト']

    # ---------- 既存データを読み出して修正 ----------
    data_rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            prio, svc, cat, item, content, docs, method, verdict = row
            # 表記ゆれ修正
            if svc == 'GH等':
                svc = 'GH'
            if isinstance(docs, str):
                docs = docs.replace('Dropbox＋', 'Dropbox+')
            # 大規模住居等減算の説明を具体化
            if item == '大規模住居等減算':
                item = '大規模住居等減算確認'
                content = ('共同生活住居の利用定員合計が21名以上、またはユニット入居定員が11名以上の場合に'
                           '減算対象となる。該当する場合、請求に正しく反映されているか。'
                           '（指定申請書・平面図・定員表で実態確認）')
            # サビ管配置数の計算式を明確化
            if item == 'サビ管の配置数確認':
                content = ('サービス利用者数÷30（端数切り上げ）の人数以上のサービス管理責任者が配置されているか。'
                           '（利用者30人以下：1人以上、31〜60人：2人以上）')
            # 処遇改善加算の項目名を実績まで含む形に修正
            if item == '処遇改善加算の届出・計画':
                item = '処遇改善加算の届出・計画・実績確認'
            # BCP項目を感染と自然災害に分割
            if item == 'BCP研修・訓練記録：感染・自然災害それぞれ':
                data_rows.append((prio, svc, 'BCP',
                    '自然災害BCP研修・訓練記録確認（年2回）',
                    '自然災害を想定したBCPの内容が整備されているか。年2回以上の研修・訓練が実施され、記録が保管されているか。',
                    'BCP（自然災害）・研修記録・訓練記録', method, '未'))
                data_rows.append((prio, svc, '感染症BCP',
                    '感染症BCP研修・訓練記録確認（年2回）',
                    '感染症を想定したBCPの内容が整備されているか。年2回以上の研修・訓練が実施され、記録が保管されているか。',
                    'BCP（感染症）・研修記録・訓練記録', method, '未'))
                continue
            # WAM NET 確認内容に期限明記
            if item == 'WAM NET情報公表報告':
                content = ('障害福祉サービス等情報公表制度への報告が完了しているか（未報告は5%減算）。'
                           '報告期限（毎年度末）を確認し、報告済みであることをWAM NETの画面またはスクリーンショットで証跡として保管しているか。')
            data_rows.append((prio, svc, cat, item, content, docs, method, verdict))

    # 新規項目を追加
    for item_tuple in NEW_ITEMS:
        data_rows.append((*item_tuple, '未'))

    # 優先度順にソート（S→A→B）
    order = {'S': 0, 'A': 1, 'B': 2}
    data_rows.sort(key=lambda r: order.get(str(r[0]).strip(), 9))

    # ---------- シートをクリアして再構築 ----------
    ws.delete_rows(1, ws.max_row)
    for mc in list(ws.merged_cells):
        pass  # delete_rowsで解除される

    # ── 行1: タイトル ──────────────────────────────────────
    ws.merge_cells('A1:L1')
    c = ws.cell(row=1, column=1,
                value='内部監査チェックリスト　【施設名・監査日・サービス種別を入力して使用】')
    c.font      = Font(bold=True, color=WHITE, size=13, name='游ゴシック')
    c.fill      = fill(COL_HEADER)
    c.alignment = align('center', 'center')
    ws.row_dimensions[1].height = 22

    # ── 行2: 基本情報（施設名・監査日） ────────────────────
    ws.merge_cells('A2:D2')
    c = ws.cell(row=2, column=1, value='施設名')
    c.font = font(bold=True, color=WHITE); c.fill = fill(COL_SUBHEADER)
    c.alignment = align('center')
    ws.merge_cells('E2:G2')
    ws.cell(row=2, column=5, value='').alignment = align('left')

    ws.merge_cells('H2:I2')
    c = ws.cell(row=2, column=8, value='監査日')
    c.font = font(bold=True, color=WHITE); c.fill = fill(COL_SUBHEADER)
    c.alignment = align('center')
    ws.merge_cells('J2:L2')
    ws.cell(row=2, column=10, value='').alignment = align('left')
    ws.row_dimensions[2].height = 20

    # ── 行3: 監査員・前回監査 ─────────────────────────────
    ws.merge_cells('A3:D3')
    c = ws.cell(row=3, column=1, value='監査実施者')
    c.font = font(bold=True, color=WHITE); c.fill = fill(COL_SUBHEADER)
    c.alignment = align('center')
    ws.merge_cells('E3:G3')
    ws.cell(row=3, column=5, value='').alignment = align('left')

    ws.merge_cells('H3:I3')
    c = ws.cell(row=3, column=8, value='前回監査日')
    c.font = font(bold=True, color=WHITE); c.fill = fill(COL_SUBHEADER)
    c.alignment = align('center')
    ws.merge_cells('J3:L3')
    ws.cell(row=3, column=10, value='').alignment = align('left')
    ws.row_dimensions[3].height = 20

    # ── 行4: 前回指摘是正確認 ────────────────────────────
    ws.merge_cells('A4:D4')
    c = ws.cell(row=4, column=1, value='前回指摘事項の是正状況')
    c.font = font(bold=True, color=WHITE); c.fill = fill(COL_SUBHEADER)
    c.alignment = align('center')
    ws.merge_cells('E4:L4')
    c = ws.cell(row=4, column=5,
                value='□ 全て完了　□ 一部未了（詳細は指摘内容欄参照）　□ 未対応（理由：　　　　　　　　）')
    c.font = font(); c.alignment = align('left')
    ws.row_dimensions[4].height = 20

    # ── 行5: 判定コード凡例 ────────────────────────────────
    ws.merge_cells('A5:L5')
    legend = ('【判定コード】　OK＝適合　　NG＝不適合（是正要）　　N/A＝対象外（理由を指摘内容欄へ）'
              '　　是正中＝指摘済み対応中　　未＝未確認')
    c = ws.cell(row=5, column=1, value=legend)
    c.font = Font(bold=False, color='FF1F4E79', size=9, name='游ゴシック')
    c.fill = fill(COL_LEGEND)
    c.alignment = align('left', 'center')
    ws.row_dimensions[5].height = 18

    # ── 行6: 列ヘッダー ───────────────────────────────────
    headers = ['優先度', 'サービス', '分類', '監査項目', '確認内容',
               '確認資料', '確認方法', '判定', '指摘内容', '是正期限', '担当者', '是正確認日']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font      = Font(bold=True, color=WHITE, size=10, name='游ゴシック')
        c.fill      = fill(COL_SUBHEADER)
        c.alignment = align('center', 'center')
        c.border    = thin_border()
    ws.row_dimensions[6].height = 24

    # ── 行7以降: チェック項目 ─────────────────────────────
    ROW_START = 7
    for i, row_data in enumerate(data_rows):
        r = ROW_START + i
        prio, svc, cat, item, content, docs, method, verdict = row_data
        bg = get_row_color(prio)
        values = [prio, svc, cat, item, content, docs, method, verdict,
                  '', '', '', '']  # 追加列は空欄
        for col_idx, val in enumerate(values, start=1):
            cell_bg = bg if col_idx <= 8 else COL_NEW_COLS
            bold = col_idx in (1, 2, 3, 4)
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font      = Font(bold=bold, size=9, name='游ゴシック')
            c.fill      = fill(cell_bg)
            c.alignment = align('center' if col_idx in (1,2,7,8,10,11,12) else 'left',
                                'center', True)
            c.border    = thin_border()
        ws.row_dimensions[r].height = 34

    # ── 判定欄にドロップダウン ─────────────────────────────
    last_data_row = ROW_START + len(data_rows) - 1
    dv = DataValidation(
        type='list',
        formula1='"未,OK,NG,N/A,是正中"',
        allow_blank=True,
        showDropDown=False
    )
    dv.sqref = f'H{ROW_START}:H{last_data_row}'
    ws.add_data_validation(dv)

    # ── サマリー欄 ────────────────────────────────────────
    summary_row = last_data_row + 2
    ws.merge_cells(f'A{summary_row}:L{summary_row}')
    c = ws.cell(row=summary_row, column=1, value='■ 監査結果サマリー')
    c.font = Font(bold=True, color=WHITE, size=11, name='游ゴシック')
    c.fill = fill(COL_HEADER)
    c.alignment = align('left', 'center')
    ws.row_dimensions[summary_row].height = 20

    summary_row += 1
    summary_headers = ['区分', 'OK', 'NG', 'N/A', '是正中', '未確認', '合計', '所見・総括']
    merge_cols = [1, 2, 3, 4, 5, 6, 7]
    ws.merge_cells(f'H{summary_row}:L{summary_row}')
    for ci, hh in enumerate(summary_headers, start=1):
        c = ws.cell(row=summary_row, column=ci, value=hh)
        c.font = Font(bold=True, color=WHITE, size=9, name='游ゴシック')
        c.fill = fill(COL_SUBHEADER)
        c.alignment = align('center', 'center')
        c.border = thin_border()
    ws.row_dimensions[summary_row].height = 18

    for label in ['S優先度項目', 'A優先度項目', 'B優先度項目', '合計']:
        summary_row += 1
        ws.merge_cells(f'H{summary_row}:L{summary_row}')
        ws.cell(row=summary_row, column=1, value=label).border = thin_border()
        for ci in range(2, 9):
            c = ws.cell(row=summary_row, column=ci, value='')
            c.fill = fill(COL_SUMMARY)
            c.border = thin_border()
            c.alignment = align('center')
        ws.row_dimensions[summary_row].height = 18

    # ── 列幅設定 ──────────────────────────────────────────
    col_widths = {
        'A': 5.5,   # 優先度
        'B': 8,     # サービス
        'C': 13,    # 分類
        'D': 28,    # 監査項目
        'E': 52,    # 確認内容
        'F': 36,    # 確認資料
        'G': 13,    # 確認方法
        'H': 9,     # 判定
        'I': 28,    # 指摘内容
        'J': 11,    # 是正期限
        'K': 10,    # 担当者
        'L': 12,    # 是正確認日
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── シート2: 確認資料リスト は表記ゆれ修正のみ ─────────
    ws2 = wb['確認資料リスト']
    for row in ws2.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.replace('Dropbox＋', 'Dropbox+')

    wb.save(DST)
    print(f'保存完了: {DST}')
    print(f'チェック項目数: {len(data_rows)} 件（新規追加: {len(NEW_ITEMS)} 件）')

if __name__ == '__main__':
    main()
