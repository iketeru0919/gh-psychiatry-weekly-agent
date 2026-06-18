#!/usr/bin/env python3
"""
週次チェック（Python版）
VBAの週次チェック_軽量版をopenpyxlで再実装。
ExcelをCOM経由で開かないため、VBAより大幅に高速。

使い方:
    python 週次チェック.py マスターファイル.xlsm
    python 週次チェック.py  # カレントディレクトリの *.xlsm を自動検出

依存パッケージ:
    pip install openpyxl
"""

import sys
import os
import re
from pathlib import Path
from datetime import datetime
import unicodedata

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import column_index_from_string, coordinate_to_tuple
except ImportError:
    print("エラー: openpyxlがインストールされていません。")
    print("  pip install openpyxl")
    sys.exit(1)


# ============================================================
# セルアドレス定義（様式タイプ別）
# ============================================================

CELL_MAP = {
    "標準": {
        "支援員基準":         "FE79",
        "世話人基準":         "FE82",
        "追加必要人員":       "EV79",
        "加算判定":           "EV78",
        "支援員日中配置換算": "EA19",
        "支援員日中配置超過": "EA22",
        "支援員実配置":       "DY24",
        "世話人日中配置換算": "EA30",
        "世話人日中配置超過": "EA36",
        "世話人実配置":       "DY40",
        "基準人員751":        "DU75",
        "現在の基準値":       "BU73",
        "人員過不足":         "BU75",
    },
    "西浅田ⅰ用": {
        "支援員基準":         "FE80",
        "世話人基準":         "FE83",
        "追加必要人員":       "EV80",
        "加算判定":           "EV79",
        "支援員日中配置換算": "EA19",
        "支援員日中配置超過": "EA22",
        "支援員実配置":       "DY24",
        "世話人日中配置換算": "EA30",
        "世話人日中配置超過": "EA36",
        "世話人実配置":       "DY40",
        "基準人員751":        "DU76",
        "現在の基準値":       "BU74",
        "人員過不足":         "BU76",
    },
    "舞多聞阪南用": {
        "支援員基準":         "FE89",
        "世話人基準":         "FE92",
        "追加必要人員":       "EV89",
        "加算判定":           "EV78",
        "支援員日中配置換算": "EA19",
        "支援員日中配置超過": "EA22",
        "支援員実配置":       "DY24",
        "世話人日中配置換算": "EA30",
        "世話人日中配置超過": "EA36",
        "世話人実配置":       "DY40",
        "基準人員751":        "DU85",
        "現在の基準値":       "BU83",
        "人員過不足":         "BU85",
    },
}

NUMERIC_KEYS = [
    "支援員基準", "世話人基準", "追加必要人員",
    "支援員日中配置換算", "支援員日中配置超過", "支援員実配置",
    "世話人日中配置換算", "世話人日中配置超過", "世話人実配置",
    "基準人員751", "現在の基準値", "人員過不足",
]

# 出力列ヘッダー
SUMMARY_HEADERS = [
    ("A", "判定"),
    ("B", "施設名"),
    ("C", "対象月"),
    ("D", "支援員数(法令基準)(前年度実績)"),
    ("E", "世話人数(法令基準)(前年度実績)"),
    ("F", "追加で必要な人員数(人員配置体制加算)"),
    ("G", "人員配置体制加算達成・未達判定"),
    ("H", "支援員 日中配置換算数"),
    ("I", "支援員 日中配置超過換算"),
    ("J", "支援員 実配置数"),
    ("K", "世話人 日中配置換算数"),
    ("L", "世話人 日中配置超過換算"),
    ("M", "世話人 実配置数"),
    ("N", "7.5:1 配置基準人員数"),
    ("O", "現在の基準値"),
    ("P", "実配置合計"),
    ("Q", "人員超過・過不足数 結果"),
    ("R", "更新日時"),
    ("S", "ファイル名"),
    ("T", "採用シート名"),
    ("U", "ファイルパス"),
    ("V", "備考"),
]

HEADER_COLORS = {
    "A": "DDEAF7", "B": "DDEAF7", "C": "DDEAF7",
    "D": "E2EFDA", "E": "E2EFDA", "F": "E2EFDA", "G": "E2EFDA",
    "H": "FFF2CC", "I": "FFF2CC", "J": "FFF2CC",
    "K": "FCE4D6", "L": "FCE4D6", "M": "FCE4D6",
    "N": "EAD1DC", "O": "EAD1DC", "P": "EAD1DC", "Q": "EAD1DC",
    "R": "D9D9D9", "S": "D9D9D9", "T": "D9D9D9", "U": "D9D9D9", "V": "D9D9D9",
}

COL_WIDTHS = {
    "A": 10, "B": 18, "C": 10,
    "D": 16, "E": 16, "F": 16, "G": 20,
    "H": 14, "I": 14, "J": 14,
    "K": 14, "L": 14, "M": 14,
    "N": 14, "O": 14, "P": 14, "Q": 16,
    "R": 18, "S": 38, "T": 16, "U": 70, "V": 60,
}


# ============================================================
# テキスト正規化
# ============================================================

def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace(" ", "").replace("　", "").replace("\t", "")
    s = s.replace("Ⅰ", "ⅰ").replace("Ⅱ", "ⅱ").replace("Ⅲ", "ⅲ")
    return s.lower()


def normalize_filename(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("　", "").replace(" ", "")
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("－", "-").replace("ー", "-")
    s = s.replace("Ⅰ", "ⅰ").replace("Ⅱ", "ⅱ").replace("Ⅲ", "ⅲ")
    return s.lower()


# ============================================================
# 設定読み込み
# ============================================================

def read_settings(ws_set) -> dict:
    settings = {}
    for row in ws_set.iter_rows(min_row=1, values_only=True):
        if row[0] is not None and row[1] is not None:
            key = normalize_text(str(row[0]))
            settings[key] = row[1]
    return settings


def get_setting(settings: dict, key_ja: str, default):
    return settings.get(normalize_text(key_ja), default)


# ============================================================
# 施設一覧読み込み
# ============================================================

def read_facility_list(ws_fac) -> tuple[list[str], list[dict]]:
    headers = []
    for cell in ws_fac[1]:
        headers.append(str(cell.value) if cell.value is not None else "")

    def find_col(name):
        nt = normalize_text(name)
        for i, h in enumerate(headers):
            if normalize_text(h) == nt:
                return i
        return -1

    col_target   = find_col("対象")
    col_manager  = find_col("担当者")
    col_facility = find_col("施設名")
    col_key      = find_col("検索フォルダ/検索キー")
    col_style    = find_col("様式タイプ")

    if col_facility < 0:
        print("エラー: 施設一覧シートに「施設名」列がありません。")
        sys.exit(1)

    facilities = []
    for row in ws_fac.iter_rows(min_row=2, values_only=True):
        name = str(row[col_facility]).strip() if row[col_facility] is not None else ""
        if not name:
            continue

        search_key = name
        if col_key >= 0 and row[col_key] is not None:
            k = str(row[col_key]).strip()
            if k:
                search_key = k

        style = "標準"
        if col_style >= 0 and row[col_style] is not None:
            s = str(row[col_style]).strip()
            if s:
                style = s

        target_val = str(row[col_target]).strip() if col_target >= 0 and row[col_target] is not None else ""
        manager = str(row[col_manager]).strip() if col_manager >= 0 and row[col_manager] is not None else ""

        facilities.append({
            "name":       name,
            "search_key": search_key,
            "style":      style,
            "target_val": target_val,
            "manager":    manager,
        })

    return facilities


def is_target(fac: dict, extract_mode: str, target_manager: str) -> bool:
    mode = normalize_text(extract_mode)
    if "全施設" in mode:
        return True
    if "担当者" in mode:
        return normalize_text(fac["manager"]) == normalize_text(target_manager) and bool(target_manager.strip())
    v = normalize_text(fac["target_val"])
    return v in ("1", "true", "yes", "対象", "?", "○")


# ============================================================
# 様式タイプ推定
# ============================================================

def infer_style_type(facility_name: str, style: str) -> str:
    f = normalize_text(facility_name)
    s = normalize_text(style)
    if not s or s == "標準":
        if "西浅田" in f and "ⅰ" in f:
            return "西浅田ⅰ用"
        if "舞多聞" in f or "阪南" in f:
            return "舞多聞阪南用"
    return style if style else "標準"


def resolve_cell_map(style: str) -> dict:
    s = normalize_text(style)
    if "西浅田" in s and "ⅰ" in s:
        return CELL_MAP["西浅田ⅰ用"]
    if "舞多聞" in s or "阪南" in s:
        return CELL_MAP["舞多聞阪南用"]
    return CELL_MAP["標準"]


# ============================================================
# ファイル検索
# ============================================================

def month_patterns(target_month: str) -> list[str]:
    target_month = target_month.strip()
    if len(target_month) == 4 and target_month.isdigit():
        yy = target_month[:2]
        mm = target_month[2:]
        yyyy = "20" + yy
        m1 = str(int(mm))
        m2 = mm.zfill(2)
        return [
            yy + mm,
            yyyy + mm,
            yyyy + "." + m1,
            yyyy + "." + m2,
            yyyy + "年" + m1 + "月",
            yyyy + "年" + m2 + "月",
            m1 + "月",
            m2 + "月",
        ]
    return [target_month]


def file_matches(path: Path, target_month: str) -> bool:
    nm = normalize_filename(path.name)
    if nm.startswith("~$"):
        return False
    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
        return False
    if "予定" in nm:
        return False
    if "実績" not in nm:
        return False
    patterns = month_patterns(target_month)
    for p in patterns:
        if normalize_filename(p) in nm:
            return True
    return False


def find_latest_file(search_root: Path, target_month: str) -> tuple[Path | None, int]:
    latest_path = None
    latest_mtime = 0.0
    count = 0

    try:
        for p in search_root.rglob("*"):
            if p.is_file() and file_matches(p, target_month):
                count += 1
                mtime = p.stat().st_mtime
                if mtime > latest_mtime:
                    latest_mtime = mtime
                    latest_path = p
    except PermissionError:
        pass

    return latest_path, count


def resolve_search_root(root_path: Path, search_key: str) -> Path | None:
    search_key = search_key.strip()
    if not search_key:
        return root_path

    direct = root_path / search_key
    if direct.is_dir():
        return direct

    return find_folder_recursive(root_path, search_key)


def find_folder_recursive(root: Path, search_key: str) -> Path | None:
    nt_key = normalize_text(search_key)
    try:
        for d in root.iterdir():
            if d.is_dir() and nt_key in normalize_text(d.name):
                return d
        for d in root.iterdir():
            if d.is_dir():
                result = find_folder_recursive(d, search_key)
                if result:
                    return result
    except PermissionError:
        pass
    return None


# ============================================================
# シート検索
# ============================================================

def find_target_sheet(wb, target_sheet_name: str):
    nt = normalize_text(target_sheet_name)
    for ws in wb.worksheets:
        if normalize_text(ws.title) == nt:
            return ws
    for ws in wb.worksheets:
        if "管理" in normalize_text(ws.title):
            return ws
    return None


# ============================================================
# セル値取得
# ============================================================

def get_cell_value(ws, cell_addr: str):
    try:
        return ws[cell_addr].value
    except Exception:
        return None


def read_managed_sheet(ws, style: str) -> dict:
    cmap = resolve_cell_map(style)
    vals = {}
    for key, addr in cmap.items():
        vals[key] = get_cell_value(ws, addr)

    v1 = vals.get("支援員実配置")
    v2 = vals.get("世話人実配置")
    try:
        vals["実配置合計"] = float(v1) + float(v2)
    except (TypeError, ValueError):
        vals["実配置合計"] = None

    return vals


# ============================================================
# バリデーション・判定
# ============================================================

def validate_values(vals: dict, facility_name: str, file_name: str, file_path: str) -> tuple[str, list[dict]]:
    note_parts = []
    errors = []

    def add_err(category, detail):
        errors.append({
            "施設名": facility_name,
            "エラー区分": category,
            "内容": detail,
            "ファイル名": file_name,
            "ファイルパス": file_path,
        })

    for key in NUMERIC_KEYS:
        v = vals.get(key)
        sv = str(v).strip() if v is not None else ""
        if sv == "" or sv == "None":
            note_parts.append(f"{key}が空欄")
            add_err("セル空欄", f"{key}が空欄です。")
        else:
            try:
                float(sv)
            except ValueError:
                note_parts.append(f"{key}が数値以外")
                add_err("数値セルに文字列", f"{key}に数値以外が入っています。値：{sv}。セル位置ズレの可能性があります。")

    v_judge = str(vals.get("加算判定", "")).strip()
    if not v_judge or v_judge == "None":
        note_parts.append("加算判定が空欄")
        add_err("セル空欄", "人員配置体制加算達成・未達判定が空欄です。")

    note = " / ".join(note_parts)
    return note, errors


def determine_status(vals: dict, note: str) -> str:
    if "空欄" in note or "数値以外" in note:
        return "要確認"

    judge = str(vals.get("加算判定", ""))
    if "未達" in judge or "基準未達" in judge:
        return "要確認"

    try:
        if float(vals.get("追加必要人員", 0) or 0) > 0:
            return "要確認"
    except (TypeError, ValueError):
        pass

    try:
        if float(vals.get("人員過不足", 0) or 0) != 0:
            return "要確認"
    except (TypeError, ValueError):
        pass

    return "OK"


# ============================================================
# 出力Excelシート作成
# ============================================================

def setup_summary_sheet(ws):
    for col_letter, header in SUMMARY_HEADERS:
        cell = ws[col_letter + "1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        color = HEADER_COLORS.get(col_letter, "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"


def write_summary_row(ws, row: int, status: str, facility: str, month: str,
                      vals: dict | None, file_date, file_name: str,
                      sheet_name: str, file_path: str, note: str):
    status_color = "FF0000" if status == "要確認" else "00B050"
    ws[f"A{row}"].value = status
    ws[f"A{row}"].font = Font(bold=True, color=status_color)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    ws[f"B{row}"].value = facility
    ws[f"C{row}"].value = month

    if vals:
        ws[f"D{row}"].value = vals.get("支援員基準")
        ws[f"E{row}"].value = vals.get("世話人基準")
        ws[f"F{row}"].value = vals.get("追加必要人員")
        ws[f"G{row}"].value = vals.get("加算判定")
        ws[f"H{row}"].value = vals.get("支援員日中配置換算")
        ws[f"I{row}"].value = vals.get("支援員日中配置超過")
        ws[f"J{row}"].value = vals.get("支援員実配置")
        ws[f"K{row}"].value = vals.get("世話人日中配置換算")
        ws[f"L{row}"].value = vals.get("世話人日中配置超過")
        ws[f"M{row}"].value = vals.get("世話人実配置")
        ws[f"N{row}"].value = vals.get("基準人員751")
        ws[f"O{row}"].value = vals.get("現在の基準値")
        ws[f"P{row}"].value = vals.get("実配置合計")
        ws[f"Q{row}"].value = vals.get("人員過不足")

    if file_date:
        ws[f"R{row}"].value = file_date
        ws[f"R{row}"].number_format = "yyyy/m/d h:mm"

    ws[f"S{row}"].value = file_name
    ws[f"T{row}"].value = sheet_name
    ws[f"U{row}"].value = file_path
    ws[f"V{row}"].value = note

    if file_path:
        ws[f"S{row}"].hyperlink = file_path
        ws[f"S{row}"].font = Font(color="0563C1", underline="single")

    if status == "要確認":
        for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:22]:
            ws[f"{col}{row}"].fill = PatternFill("solid", fgColor="FFEB9C")

    for col in "BCDEFGHIJKLMNOPQ":
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")


def setup_adopted_sheet(ws):
    headers = ["施設名", "対象月", "検索フォルダ", "採用ファイル名", "ファイルパス", "候補数", "備考"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [20, 10, 30, 36, 80, 8, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def setup_error_sheet(ws):
    headers = ["発生日時", "施設名", "エラー区分", "内容", "ファイル名", "ファイルパス"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 20, 16, 80, 36, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ============================================================
# メイン処理
# ============================================================

def main():
    # ---- マスターファイルの特定 ----
    if len(sys.argv) >= 2:
        master_path = Path(sys.argv[1])
    else:
        candidates = list(Path(".").glob("*.xlsm"))
        if not candidates:
            print("エラー: .xlsmファイルが見つかりません。")
            print("使い方: python 週次チェック.py マスターファイル.xlsm")
            sys.exit(1)
        master_path = candidates[0]
        print(f"マスターファイル: {master_path}")

    if not master_path.exists():
        print(f"エラー: ファイルが見つかりません: {master_path}")
        sys.exit(1)

    # ---- マスターファイル読み込み ----
    print("設定・施設一覧を読み込んでいます...")
    try:
        master_wb = openpyxl.load_workbook(str(master_path), data_only=True)
    except Exception as e:
        print(f"エラー: マスターファイルを開けません: {e}")
        sys.exit(1)

    ws_set = master_wb["設定"]
    ws_fac = master_wb["施設一覧"]

    settings = read_settings(ws_set)
    root_path_str = str(get_setting(settings, "親フォルダ", ""))
    target_month  = str(get_setting(settings, "対象月", ""))
    target_sheet  = str(get_setting(settings, "対象シート名", "管理用シート"))
    extract_mode  = str(get_setting(settings, "抽出方法", "手動選択"))
    target_mgr    = str(get_setting(settings, "対象マネージャー", "") or "")
    old_days      = int(get_setting(settings, "更新日時チェック日数", 7) or 7)

    if not root_path_str:
        print("エラー: 設定シートの「親フォルダ」が空欄です。")
        sys.exit(1)

    root_path = Path(root_path_str)
    if not root_path.is_dir():
        print(f"エラー: 親フォルダが見つかりません: {root_path}")
        sys.exit(1)

    facilities = read_facility_list(ws_fac)
    master_wb.close()

    targets = [f for f in facilities if is_target(f, extract_mode, target_mgr)]
    print(f"対象施設数: {len(targets)} 施設 / 全{len(facilities)} 施設")
    print(f"対象月: {target_month}  親フォルダ: {root_path}")
    print()

    # ---- 出力ブック準備 ----
    out_wb = openpyxl.Workbook()
    ws_sum = out_wb.active
    ws_sum.title = "施設別サマリ"
    ws_pick = out_wb.create_sheet("採用ファイル一覧")
    ws_err  = out_wb.create_sheet("エラー一覧")

    setup_summary_sheet(ws_sum)
    setup_adopted_sheet(ws_pick)
    setup_error_sheet(ws_err)

    sum_row  = 2
    pick_row = 2
    err_row  = 2
    now_str  = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    # ---- 施設ループ ----
    for idx, fac in enumerate(targets, 1):
        name       = fac["name"]
        search_key = fac["search_key"]
        style      = infer_style_type(name, fac["style"])

        print(f"[{idx:2d}/{len(targets)}] {name} ...", end=" ", flush=True)

        note   = ""
        status = "OK"
        vals   = None
        file_date   = None
        file_name   = ""
        file_path_s = ""
        adopted_sheet = ""

        # フォルダ探索
        search_root = resolve_search_root(root_path, search_key)
        if search_root is None:
            note = "検索フォルダなし"
            status = "要確認"
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="検索フォルダなし")
            ws_err.cell(row=err_row, column=4, value=f"検索キーに一致するフォルダが見つかりません。検索キー：{search_key}")
            err_row += 1
            write_summary_row(ws_sum, sum_row, status, name, target_month, None, None, "", "", "", note)
            sum_row += 1
            print(f"→ {status} ({note})")
            continue

        # ファイル探索
        found_path, cand_count = find_latest_file(search_root, target_month)

        if found_path is None:
            note = "対象月の実績ファイルなし"
            status = "要確認"
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="ファイルなし")
            ws_err.cell(row=err_row, column=4, value=f"対象月の実績ファイルなし。検索フォルダ：{search_root}")
            err_row += 1
            ws_pick.cell(row=pick_row, column=1, value=name)
            ws_pick.cell(row=pick_row, column=2, value=target_month)
            ws_pick.cell(row=pick_row, column=3, value=str(search_root))
            ws_pick.cell(row=pick_row, column=7, value=note)
            pick_row += 1
            write_summary_row(ws_sum, sum_row, status, name, target_month, None, None, "", "", "", note)
            sum_row += 1
            print(f"→ {status} ({note})")
            continue

        file_name   = found_path.name
        file_path_s = str(found_path)
        file_stat   = found_path.stat()
        file_date   = datetime.fromtimestamp(file_stat.st_mtime)

        if cand_count > 1:
            note_dup = f"重複あり・最新更新ファイルを採用"
            note = note_dup if not note else note + " / " + note_dup
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="重複あり")
            ws_err.cell(row=err_row, column=4, value=f"候補ファイルが複数あります。最新更新ファイルを採用しました。候補数：{cand_count}")
            ws_err.cell(row=err_row, column=5, value=file_name)
            ws_err.cell(row=err_row, column=6, value=file_path_s)
            err_row += 1

        days_old = (datetime.now() - file_date).days
        if days_old >= old_days:
            note_old = f"更新日時が{old_days}日以上前"
            note = note_old if not note else note + " / " + note_old
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="更新古い")
            ws_err.cell(row=err_row, column=4, value=f"更新日時が{old_days}日以上前です。")
            ws_err.cell(row=err_row, column=5, value=file_name)
            ws_err.cell(row=err_row, column=6, value=file_path_s)
            err_row += 1

        ws_pick.cell(row=pick_row, column=1, value=name)
        ws_pick.cell(row=pick_row, column=2, value=target_month)
        ws_pick.cell(row=pick_row, column=3, value=str(search_root))
        ws_pick.cell(row=pick_row, column=4, value=file_name)
        ws_pick.cell(row=pick_row, column=5, value=file_path_s)
        ws_pick.cell(row=pick_row, column=6, value=cand_count)
        ws_pick.cell(row=pick_row, column=7, value=note)
        pick_row += 1

        # Excelファイル読み込み（openpyxl: Excelを起動しない）
        try:
            src_wb = openpyxl.load_workbook(str(found_path), data_only=True)
        except Exception as e:
            note_open = "ファイル読込エラー"
            note = note_open if not note else note + " / " + note_open
            status = "要確認"
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="ファイル読込エラー")
            ws_err.cell(row=err_row, column=4, value=str(e))
            ws_err.cell(row=err_row, column=5, value=file_name)
            ws_err.cell(row=err_row, column=6, value=file_path_s)
            err_row += 1
            write_summary_row(ws_sum, sum_row, status, name, target_month, None, file_date, file_name, "", file_path_s, note)
            sum_row += 1
            print(f"→ {status} ({note_open}: {e})")
            continue

        src_ws = find_target_sheet(src_wb, target_sheet)
        if src_ws is None:
            src_wb.close()
            note_ns = "管理用シートなし"
            note = note_ns if not note else note + " / " + note_ns
            status = "要確認"
            sheets = ", ".join(ws.title for ws in src_wb.worksheets)
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=name)
            ws_err.cell(row=err_row, column=3, value="シートなし")
            ws_err.cell(row=err_row, column=4, value=f"対象シートが見つかりません。存在シート：{sheets}")
            ws_err.cell(row=err_row, column=5, value=file_name)
            ws_err.cell(row=err_row, column=6, value=file_path_s)
            err_row += 1
            write_summary_row(ws_sum, sum_row, status, name, target_month, None, file_date, file_name, "", file_path_s, note)
            sum_row += 1
            print(f"→ {status} ({note_ns})")
            continue

        adopted_sheet = src_ws.title
        vals = read_managed_sheet(src_ws, style)
        src_wb.close()

        val_note, val_errors = validate_values(vals, name, file_name, file_path_s)
        if val_note:
            note = val_note if not note else note + " / " + val_note
        for e in val_errors:
            ws_err.cell(row=err_row, column=1, value=now_str)
            ws_err.cell(row=err_row, column=2, value=e["施設名"])
            ws_err.cell(row=err_row, column=3, value=e["エラー区分"])
            ws_err.cell(row=err_row, column=4, value=e["内容"])
            ws_err.cell(row=err_row, column=5, value=e["ファイル名"])
            ws_err.cell(row=err_row, column=6, value=e["ファイルパス"])
            err_row += 1

        status = determine_status(vals, note)
        write_summary_row(ws_sum, sum_row, status, name, target_month, vals, file_date, file_name, adopted_sheet, file_path_s, note)
        sum_row += 1

        print(f"→ {status}" + (f"  ({note})" if note else ""))

    # ---- 出力ファイル保存 ----
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"週次チェック結果_{target_month}_{ts}.xlsx"
    out_path = Path(out_name)

    print()
    print(f"結果を保存中: {out_name} ...")
    out_wb.save(str(out_path))
    print(f"完了！  → {out_path.resolve()}")

    ok_count  = sum(1 for r in range(2, sum_row) if ws_sum.cell(row=r, column=1).value == "OK")
    req_count = sum(1 for r in range(2, sum_row) if ws_sum.cell(row=r, column=1).value == "要確認")
    print()
    print(f"  OK    : {ok_count} 施設")
    print(f"  要確認: {req_count} 施設")


if __name__ == "__main__":
    main()
