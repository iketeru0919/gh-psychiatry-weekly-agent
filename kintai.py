# ============================================================
# RASIEL グループホーム 勤怠自動処理スクリプト（ローカル版）
# ============================================================
#
# 【毎月の作業手順】
#   1. 以下の3つのフォルダにファイルを置く
#
#      📁 dckintai\   ← DC勤怠CSVをここに置く
#      📁 タイミー\   ← タイミー利用明細CSVをここに置く
#      📁 シフト\     ← 勤務表Excelをここに置く
#
#   2. 下の YEAR・MONTH を今月の値に変更する
#   3. スクリプトを実行する
#
# 【実行方法】
#   PowerShellまたはコマンドプロンプトで：
#   python kintai.py
#
# 【出力先】
#   勤怠自動化\output\YYYYMM\
#     ・勤怠一覧_整合性チェック_全施設_YYYYMM.xlsx
#     ・実行ログ_YYYYMM.txt
#
# ============================================================

# ▼▼▼ 毎月ここだけ変更する ▼▼▼
YEAR  = 2026
MONTH = 5
# ▲▲▲ ここまで ▲▲▲

# ============================================================
# 施設設定
# ============================================================
FACILITIES = {
    '春日部': {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '南中丸': {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '宇都宮': {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '三蔵子': {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '上西郷': {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '気賀':   {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '高丘':   {'has_ii': False, 'timee_i': True,  'timee_ii': False},
    '萩丘':   {'has_ii': True,  'timee_i': True,  'timee_ii': False},
    '西浅田': {'has_ii': True,  'timee_i': True,  'timee_ii': True },
}

# ============================================================
# フォルダパス設定（ここは変更不要）
# ============================================================
import os

BASE_DIR  = r'C:\Users\k-takayama\OneDrive\デスクトップ\勤怠自動化'
DC_DIR    = os.path.join(BASE_DIR, 'dckintai')
TIMEE_DIR = os.path.join(BASE_DIR, 'タイミー')
SHIFT_DIR = os.path.join(BASE_DIR, 'シフト')
OUT_DIR   = os.path.join(BASE_DIR, 'output', f'{YEAR}{MONTH:02d}')
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# 以下は変更不要（処理本体）
# ============================================================

import glob
import datetime
from collections import defaultdict, Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS   = list(range(1, 32))
YM_STR = f'{YEAR}{MONTH:02d}'
DOW    = ['月', '火', '水', '木', '金', '土', '日']


# ============================================================
# ファイル検索
# ============================================================
def find_dc_file(施設名):
    hits = [
        f for f in glob.glob(os.path.join(DC_DIR, '*.csv'))
        if 施設名 in os.path.basename(f)
    ]
    return hits[0] if hits else None


def find_timee_files(施設名, use_timee_i, use_timee_ii):
    hits = [
        f for f in glob.glob(os.path.join(TIMEE_DIR, '*.csv'))
        if 施設名 in os.path.basename(f)
    ]
    file_i  = next((f for f in hits if 'ⅱ' not in os.path.basename(f)), None) if use_timee_i  else None
    file_ii = next((f for f in hits if 'ⅱ'     in os.path.basename(f)), None) if use_timee_ii else None
    return file_i, file_ii


def find_shift_files(施設名, has_ii):
    hits = [
        f for f in glob.glob(os.path.join(SHIFT_DIR, '*.xlsx'))
        if 施設名 in os.path.basename(f)
    ]
    if not has_ii:
        return (hits[0] if hits else None), None
    file_i  = next((f for f in hits if 'ⅱ' not in os.path.basename(f)), None)
    file_ii = next((f for f in hits if 'ⅱ'     in os.path.basename(f)), None)
    return file_i, file_ii


# ============================================================
# スタイル定義
# ============================================================
def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=9):
    return Font(name='Arial', bold=bold, color=color, size=size)

def _border():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)

BDR = _border()

COLOR = {
    'd':  ('DBEAFE', '1D4ED8'),
    'ni': ('FEF3C7', '92400E'),
    'na': ('DCFCE7', '166534'),
    'kk': ('F1F5F9', '64748B'),
    'nk': ('FFE4E1', '9B1C1C'),
    None: ('FFFFFF', 'CCCCCC'),
}
CODE_LABEL = {
    'd': '日勤', 'ni': '夜勤入り', 'na': '夜勤明け',
    'kk': '公休', 'nk': '年休', None: '―',
}


# ============================================================
# DC勤怠CSV読み込み
# ============================================================
def load_dc_csv(path):
    with open(path, 'rb') as f:
        content = f.read().decode('shift-jis', errors='replace')

    staff_data  = {}
    staff_order = []
    seen        = set()

    for line in content.splitlines():
        cols = line.split(',')
        if len(cols) < 15 or cols[0] == 'カード番号':
            continue

        name     = cols[2].strip()
        date_str = cols[4].strip()
        syukkin  = cols[8].strip()
        taikin   = cols[14].strip()
        fuzai    = cols[7].strip()
        heikyuu  = cols[6].strip()
        comment  = cols[20].strip() if len(cols) > 20 else ''

        try:
            day = int(date_str.split('/')[2])
        except Exception:
            continue

        if name not in seen:
            seen.add(name)
            staff_order.append(name)
            staff_data[name] = {}

        code = _classify_dc(syukkin, taikin, fuzai, heikyuu)
        staff_data[name][day] = (code, syukkin, taikin, comment)

    return staff_data, staff_order


def _classify_dc(syukkin, taikin, fuzai, heikyuu):
    if fuzai:
        return 'nk'
    if heikyuu == '01' and not syukkin and not taikin:
        return 'kk'
    if syukkin and ':' in syukkin:
        s_h = int(syukkin.split(':')[0])
        return 'ni' if s_h >= 14 else 'd'
    if not syukkin and taikin and ':' in taikin:
        t_h = int(taikin.split(':')[0])
        if t_h <= 12:
            return 'na'
    return None


def _parse_break_minutes(comment):
    import re
    if not comment: return 0
    comment = comment.replace('１','1').replace('２','2').replace('３','3') \
                     .replace('４','4').replace('５','5')
    m = re.search(r'休憩.*?(\d+)時間', comment)
    if m: return int(m.group(1)) * 60
    return 0


def calc_staff_minutes(dc_data):
    result = {}
    for name, daily in dc_data.items():
        events       = []
        break_by_day = {}
        for day, info in daily.items():
            if not info: continue
            syukkin = info[1] if len(info) > 1 else ''
            taikin  = info[2] if len(info) > 2 else ''
            comment = info[3] if len(info) > 3 else ''
            break_by_day[day] = _parse_break_minutes(comment)
            day_base = (day - 1) * 24 * 60
            if syukkin and ':' in syukkin:
                try:
                    h, m = map(int, syukkin.split(':'))
                    events.append((day_base + h*60 + m, 'in', day))
                except Exception:
                    pass
            if taikin and ':' in taikin:
                try:
                    h, m = map(int, taikin.split(':'))
                    events.append((day_base + h*60 + m, 'out', day))
                except Exception:
                    pass

        events.sort(key=lambda x: (x[0], 0 if x[1] == 'in' else 1))

        total_min = 0
        start_abs = None
        start_day = None
        for abs_min, kind, day in events:
            if kind == 'in':
                start_abs = abs_min
                start_day = day
            elif kind == 'out' and start_abs is not None:
                duration = abs_min - start_abs
                if duration < 0:
                    duration += 24 * 60
                if 1 <= duration <= 24 * 60:
                    rest    = break_by_day.get(start_day, 0)
                    net     = max(0, duration - rest)
                    total_min += net
                start_abs = None
                start_day = None

        result[name] = total_min
    return result


# ============================================================
# タイミーCSV読み込み
# ============================================================
def load_timee_csv(path):
    with open(path, 'rb') as f:
        content = f.read().decode('utf-8', errors='replace')

    lines = content.splitlines()
    if not lines:
        return {}, []

    header = lines[0].lstrip('﻿').split(',')
    if len(header) <= 10:
        IDX_DATE, IDX_TIN, IDX_TOUT, IDX_SEI, IDX_MEI = 0, 1, 2, 7, 8
        MIN_COLS = 9
    else:
        IDX_DATE, IDX_TIN, IDX_TOUT, IDX_SEI, IDX_MEI = 4, 5, 6, 13, 14
        MIN_COLS = 15

    data  = defaultdict(dict)
    names = []
    seen  = set()

    for line in lines[1:]:
        cols = line.split(',')
        if len(cols) < MIN_COLS:
            continue

        date_str = cols[IDX_DATE].strip()
        tin      = cols[IDX_TIN].strip()
        tout     = cols[IDX_TOUT].strip()
        sei      = cols[IDX_SEI].strip()
        mei      = cols[IDX_MEI].strip()
        name     = sei + mei

        try:
            day = int(date_str.split('/')[2])
        except Exception:
            continue

        if name not in seen:
            seen.add(name)
            names.append(name)

        tin_h = int(tin.split(':')[0]) if ':' in tin else 0
        if tin_h >= 17:
            data[name][day] = ('ni', tin, '')
            if day + 1 <= 31:
                data[name][day + 1] = ('na', '', tout)
        else:
            data[name][day] = ('d', tin, tout)

    return dict(data), names


def load_timee_slots(path):
    with open(path, 'rb') as f:
        content = f.read().decode('utf-8', errors='replace')

    lines = content.splitlines()
    if not lines:
        return {}

    header = lines[0].lstrip('﻿').split(',')
    if len(header) <= 10:
        IDX_DATE, IDX_TIN, IDX_TOUT = 0, 1, 2
        MIN_COLS = 9
    else:
        IDX_DATE, IDX_TIN, IDX_TOUT = 4, 5, 6
        MIN_COLS = 15

    by_day = defaultdict(lambda: {'日勤': [], '夜勤': [], '深夜': []})

    for line in lines[1:]:
        cols = line.split(',')
        if len(cols) < MIN_COLS: continue
        date_str = cols[IDX_DATE].strip()
        tin      = cols[IDX_TIN].strip()
        tout     = cols[IDX_TOUT].strip()
        try:
            day   = int(date_str.split('/')[2])
            tin_h = int(tin.split(':')[0]) if ':' in tin else 0
        except Exception:
            continue

        if tin_h >= 20:   kubun = '深夜'
        elif tin_h >= 14: kubun = '夜勤'
        else:              kubun = '日勤'

        by_day[day][kubun].append((tin, tout))

    return dict(by_day)


# ============================================================
# 勤務表Excel読み込み
# ============================================================
def load_kinmuhyo(path):
    SKIP_NAMES = {
        'タイミー日勤①', 'タイミー日勤②',
        'タイミー夜勤①', 'タイミー夜勤②',
        'タイミー深夜①', 'タイミー深夜②', 'タイミー深夜③',
    }
    REST_EXACT      = {'休', '欠', '希望休', '振休', '振8', '振代'}
    REST_STARTSWITH = ('非有', '有', '常有', '時有')

    def is_rest(val):
        if val is None: return True
        v = str(val).strip()
        if not v: return True
        if v in REST_EXACT: return True
        if v.startswith(REST_STARTSWITH): return True
        return False

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if '入力シート' in s), None)
    if not sheet_name:
        raise ValueError(f'入力シートが見つかりません: {path}')

    ws = wb[sheet_name]
    date_col = {}
    for c in range(1, 50):
        v = ws.cell(row=4, column=c).value
        if v and hasattr(v, 'day'):
            date_col[v.day] = c

    data  = {}
    names = []
    seen  = set()

    for r in range(5, 60):
        name_v = ws.cell(row=r, column=3).value
        if not name_v: continue
        name = str(name_v).strip()
        if not name or name in ('0', 'None') or name in SKIP_NAMES: continue

        if name not in seen:
            seen.add(name)
            names.append(name)
            data[name] = {}

        for day, col in date_col.items():
            v       = ws.cell(row=r, column=col).value
            raw     = str(v).strip() if v else ''
            working = not is_rest(v)
            if day in data[name]:
                if not data[name][day][0] and working:
                    data[name][day] = (working, raw)
            else:
                data[name][day] = (working, raw)

    return data, names


def load_km_timee_slots(path):
    REST_EXACT      = {'休', '欠', '希望休', '振休', '振8', '振代'}
    REST_STARTSWITH = ('非有', '有', '常有', '時有')

    def is_rest(val):
        if val is None: return True
        v = str(val).strip()
        if not v: return True
        if v in REST_EXACT: return True
        if v.startswith(REST_STARTSWITH): return True
        return False

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if '入力シート' in s), None)
    if not sheet_name: return {}

    ws = wb[sheet_name]
    date_col = {}
    for c in range(1, 50):
        v = ws.cell(row=4, column=c).value
        if v and hasattr(v, 'day'):
            date_col[v.day] = c

    slots = defaultdict(int)
    for r in range(5, 60):
        name_v = ws.cell(row=r, column=3).value
        if not name_v: continue
        name = str(name_v).strip()
        if 'タイミー' not in name: continue
        for day, col in date_col.items():
            v = ws.cell(row=r, column=col).value
            if not is_rest(v):
                slots[day] += 1

    return dict(slots)


def load_km_timee_rows(path):
    REST_EXACT      = {'休', '欠', '希望休', '振休', '振8', '振代'}
    REST_STARTSWITH = ('非有', '有', '常有', '時有')

    def is_rest(val):
        if val is None: return True
        v = str(val).strip()
        if not v: return True
        if v in REST_EXACT: return True
        if v.startswith(REST_STARTSWITH): return True
        return False

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if '入力シート' in s), None)
    if not sheet_name: return []

    ws = wb[sheet_name]
    date_col = {}
    for c in range(1, 50):
        v = ws.cell(row=4, column=c).value
        if v and hasattr(v, 'day'):
            date_col[v.day] = c

    result = []
    for r in range(5, 60):
        name_v = ws.cell(row=r, column=3).value
        if not name_v: continue
        name = str(name_v).strip()
        if 'タイミー' not in name: continue

        row_data = {}
        for day, col in date_col.items():
            v = ws.cell(row=r, column=col).value
            if not is_rest(v):
                row_data[day] = str(v).strip()
        if row_data:
            result.append((name, row_data))

    return result


def parse_yukyu(symbol):
    import re
    if not symbol: return False, 0
    s = str(symbol).strip()
    if '有' not in s: return False, 0
    nums = re.findall(r'\d+', s)
    hours = int(nums[-1]) if nums else 8
    return True, hours * 60


def calc_yukyu_from_km(km_i_data, km_ii_data):
    result = {}
    for km_data in [km_i_data or {}, km_ii_data or {}]:
        for name, daily in km_data.items():
            base = name.rstrip('①②③').strip()
            if base not in result:
                result[base] = {'days': 0, 'minutes': 0}
            for day, (working, raw) in daily.items():
                is_y, mins = parse_yukyu(raw)
                if is_y:
                    if mins >= 8 * 60:
                        result[base]['days']    += 1
                        result[base]['minutes'] += mins
                    else:
                        result[base]['days']    += round(mins / (8 * 60), 1)
                        result[base]['minutes'] += mins
    return result


def classify_km_timee_symbol(symbol):
    import re
    if not symbol: return None
    s = str(symbol).strip()
    nums = re.findall(r'\d+', s)
    hour = int(nums[0]) if nums else None
    if '明' in s:
        return f'out_{hour:02d}' if hour is not None else 'out_xx'
    if '入' in s and hour is not None:
        return f'in_{hour:02d}'
    if '深夜' in s or ('深' in s and '明' not in s and '入' not in s):
        return 'shinyakan'
    if '日勤' in s or '日' in s or 'F' in s or 'f' in s:
        return 'nikkin'
    if '夜勤' in s or '夜' in s or '非夜' in s:
        return 'yakan'
    return None


def match_timee_symbol_dc(km_class, tin_h, tout_h):
    if km_class is None: return None
    if km_class.startswith('in_'):
        target = int(km_class[3:])
        return tin_h is not None and abs(tin_h - target) <= 1
    if km_class == 'out_xx':
        return tout_h is not None
    if km_class.startswith('out_'):
        target = int(km_class[4:])
        return tout_h is not None and abs(tout_h - target) <= 1
    if km_class == 'nikkin':
        return tin_h is not None and tin_h < 14
    if km_class == 'yakan':
        return tin_h is not None and 14 <= tin_h < 20
    if km_class == 'shinyakan':
        return tin_h is not None and tin_h >= 20
    return None


# ============================================================
# 整合性チェック
# ============================================================
def check_consistency(dc_data, km_data):
    def normalize(name):
        return name.replace('　', '').replace(' ', '') \
                   .rstrip('①②③④⑤').strip()

    dc_norm = {normalize(n): (n, d) for n, d in dc_data.items()}
    km_norm = {normalize(n): (n, d) for n, d in km_data.items()}

    discrepancies = []
    match_count   = 0

    for norm_name, (dc_name, dc_daily) in dc_norm.items():
        if norm_name not in km_norm:
            continue
        _, km_daily = km_norm[norm_name]

        for day in DAYS:
            dc_code, dc_in, dc_out = (dc_daily.get(day) or (None,'',''))[:3]
            km_working, km_raw     = km_daily.get(day, (False, ''))
            dc_work = dc_code in ('d', 'ni', 'na')

            if dc_work and not km_working:
                discrepancies.append({
                    'dc_name': dc_name, 'day': day,
                    'type': 'DC出勤 / 勤務表休み',
                    'dc_code': dc_code, 'dc_in': dc_in, 'dc_out': dc_out,
                    'km_raw': km_raw,
                })
            elif not dc_work and km_working:
                discrepancies.append({
                    'dc_name': dc_name, 'day': day,
                    'type': 'DC休み / 勤務表出勤',
                    'dc_code': dc_code, 'dc_in': dc_in, 'dc_out': dc_out,
                    'km_raw': km_raw,
                })
            else:
                match_count += 1

    return discrepancies, match_count


# ============================================================
# Excel出力
# ============================================================
def write_kintai_excel(wb_check, 施設名, dc_data, dc_order,
                       timee_i_data, timee_i_names,
                       timee_ii_data, timee_ii_names,
                       km_i_data=None, km_ii_data=None,
                       km_timee_slots_i=None, km_timee_slots_ii=None,
                       km_order_i=None, km_order_ii=None,
                       dc_minutes=None,
                       timee_slots_i=None, timee_slots_ii=None,
                       km_timee_rows_i=None, km_timee_rows_ii=None,
                       yukyu_data=None):

    km_merged = {}
    for src_km in [km_i_data or {}, km_ii_data or {}]:
        for name, daily in src_km.items():
            base = name.rstrip('①②③').strip()
            if base not in km_merged:
                km_merged[base] = {}
            for day, (working, raw) in daily.items():
                if day not in km_merged[base] or \
                   (working and not km_merged[base][day][0]):
                    km_merged[base][day] = (working, raw)

    def _norm(n):
        return n.replace('　','').replace(' ','').lstrip('★').rstrip('①②③④⑤').strip()

    km_order_combined = []
    seen_km = set()
    for km_name in (km_order_i or []) + (km_order_ii or []):
        base = km_name.rstrip('①②③').strip()
        norm = _norm(base)
        if norm not in seen_km:
            seen_km.add(norm)
            km_order_combined.append(base)

    if km_order_combined:
        def _sort_key(dc_name):
            norm = _norm(dc_name)
            for i, km_name in enumerate(km_order_combined):
                if _norm(km_name) == norm:
                    return i
            return len(km_order_combined)
        sorted_order = sorted(dc_order, key=_sort_key)
    else:
        sorted_order = dc_order

    km_timee_slots = {}
    for slots in [km_timee_slots_i or {}, km_timee_slots_ii or {}]:
        for day, cnt in slots.items():
            km_timee_slots[day] = km_timee_slots.get(day, 0) + cnt

    ws = wb_check.create_sheet(f'{施設名}_勤怠一覧')

    timee_slots_merged = {}
    for slots in [timee_slots_i or {}, timee_slots_ii or {}]:
        for day, kubun_dict in slots.items():
            if day not in timee_slots_merged:
                timee_slots_merged[day] = {'日勤':[],'夜勤':[],'深夜':[]}
            for kubun, entries in kubun_dict.items():
                timee_slots_merged[day][kubun].extend(entries)

    km_timee_rows_merged = list(km_timee_rows_i or []) + list(km_timee_rows_ii or [])

    _write_kintai_sheet(ws, 施設名, dc_data, sorted_order,
                        timee_i_data, timee_i_names,
                        timee_ii_data, timee_ii_names,
                        km_merged, km_timee_slots,
                        timee_slots_merged, km_timee_rows_merged)

    ws2 = wb_check.create_sheet(f'{施設名}_職員別集計')
    _write_summary_sheet(ws2, dc_data, sorted_order,
                         timee_i_data, timee_i_names,
                         timee_ii_data, timee_ii_names,
                         dc_minutes or {},
                         yukyu_data or {})


def _write_kintai_sheet(ws, 施設名, dc_data, dc_order,
                        timee_i_data, timee_i_names,
                        timee_ii_data, timee_ii_names,
                        km_merged=None, km_timee_slots=None,
                        timee_slots=None, km_timee_rows=None):
    CH       = '2D5F8A'
    last_col = get_column_letter(37)

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = f'{YEAR}年{MONTH}月 職員勤怠一覧（RASIEL{施設名}）'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = _fill(CH)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws['A2'] = '職員名'
    ws['A2'].font      = _font(bold=True, color='FFFFFF')
    ws['A2'].fill      = _fill(CH)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].border    = BDR
    ws.column_dimensions['A'].width = 16

    for si, sc in enumerate(['日勤', '夜勤入', '夜勤明', '公休', '年休']):
        cl          = get_column_letter(33 + si)
        c           = ws[f'{cl}2']
        c.value     = sc
        c.font      = _font(bold=True, color='FFFFFF', size=8)
        c.fill      = _fill(CH)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = BDR
        ws.column_dimensions[cl].width = 5

    ws.row_dimensions[2].height = 28

    for day in DAYS:
        try:
            wd = datetime.date(YEAR, MONTH, day).weekday()
        except ValueError:
            continue
        cl      = get_column_letter(day + 1)
        cell    = ws[f'{cl}2']
        cell.value     = f'{day}\n({DOW[wd]})'
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = BDR
        ws.column_dimensions[cl].width = 5.5
        if wd == 6:
            cell.fill = _fill('C0392B')
            cell.font = _font(bold=True, color='FFFFFF', size=8)
        elif wd == 5:
            cell.fill = _fill('2471A3')
            cell.font = _font(bold=True, color='FFFFFF', size=8)
        else:
            cell.fill = _fill('4A90C4')
            cell.font = _font(bold=True, color='FFFFFF', size=8)

    def write_sep(row, label, color):
        ws.row_dimensions[row].height = 13
        ws.merge_cells(f'A{row}:{last_col}{row}')
        c           = ws.cell(row=row, column=1, value=label)
        c.font      = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        c.fill      = _fill(color)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    def normalize_name(name):
        return name.replace('　','').replace(' ','').lstrip('★').rstrip('①②③④⑤').strip()

    def get_km(name):
        if not km_merged: return {}
        norm = normalize_name(name)
        for km_name, km_daily in km_merged.items():
            if normalize_name(km_name) == norm:
                return km_daily
        return {}

    def dc_code_to_text(code, t_in, t_out):
        if code == 'd':  return f'日勤\n{t_in}-{t_out}' if t_in else '日勤'
        if code == 'ni': return f'夜入\n{t_in}〜'       if t_in else '夜入'
        if code == 'na': return f'夜明\n〜{t_out}'      if t_out else '夜明'
        if code == 'kk': return '公休'
        if code == 'nk': return '年休'
        return ''

    def is_dc_working(code): return code in ('d', 'ni', 'na')

    def write_row(row, name, daily, ri, bg_e, bg_o, fg):
        bg = bg_e if ri % 2 == 0 else bg_o
        km_daily = get_km(name)

        ws.row_dimensions[row].height   = 20
        ws.row_dimensions[row+1].height = 14

        nc           = ws.cell(row=row, column=1, value=name)
        nc.font      = _font(bold=True, size=9, color=fg)
        nc.fill      = _fill(bg)
        nc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        nc.border    = BDR

        nc2           = ws.cell(row=row+1, column=1, value='勤務表↓')
        nc2.font      = Font(name='Arial', size=7, color='888888')
        nc2.fill      = _fill('F8F8F8')
        nc2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        nc2.border    = BDR

        counts = {'d': 0, 'ni': 0, 'na': 0, 'kk': 0, 'nk': 0}

        for day in DAYS:
            col = day + 1
            info  = daily.get(day)
            code  = info[0] if info else None
            t_in  = info[1] if info and len(info) > 1 else ''
            t_out = info[2] if info and len(info) > 2 else ''

            bg_c, fg_c = COLOR.get(code, COLOR[None])
            dc_text = dc_code_to_text(code, t_in, t_out)
            if code == 'd':  counts['d']  += 1
            elif code == 'ni': counts['ni'] += 1
            elif code == 'na': counts['na'] += 1
            elif code == 'kk': counts['kk'] += 1
            elif code == 'nk': counts['nk'] += 1

            km_info    = km_daily.get(day, (False, ''))
            km_working = km_info[0]
            km_raw     = km_info[1] if len(km_info) > 1 else ''
            dc_working = is_dc_working(code)
            mismatch   = km_merged and (dc_working != km_working)

            c_dc           = ws.cell(row=row, column=col, value=dc_text)
            c_dc.font      = Font(name='Arial', size=7,
                                  color='C0392B' if mismatch else fg_c)
            c_dc.fill      = _fill('FFD0D0' if mismatch else bg_c)
            c_dc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c_dc.border    = BDR

            km_text        = km_raw if km_raw else ('（打刻のみ）' if dc_working and not km_working else '')
            c_km           = ws.cell(row=row+1, column=col,
                                     value=('⚠要確認\n' + km_text) if mismatch else km_text)
            c_km.font      = Font(name='Arial', size=7,
                                  color='C0392B' if mismatch else '555555')
            c_km.fill      = _fill('FFD0D0' if mismatch else 'F8F8F8')
            c_km.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c_km.border    = BDR

        for si, key in enumerate(['d', 'ni', 'na', 'kk', 'nk']):
            c           = ws.cell(row=row, column=33 + si, value=counts[key])
            c.font      = _font(size=9, bold=True)
            c.fill      = _fill(bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = BDR
            c2          = ws.cell(row=row+1, column=33 + si, value='')
            c2.fill     = _fill('F8F8F8')
            c2.border   = BDR

        return counts

    r = 3
    write_sep(r, '▼ 固定職員　（上段:DC勤怠打刻　下段:勤務表記号　赤:不一致）', '2D5F8A'); r += 1
    for ri, name in enumerate(dc_order):
        write_row(r, name, dc_data[name], ri, 'EBF3FB', 'F4F8FD', '000000')
        r += 2

    if timee_slots or km_timee_rows:
        write_sep(r, '▼ タイミー（上段:DC勤怠 打刻時間帯　下段:勤務表記号　赤:不一致）', '4A3580'); r += 1

        dc_pool_by_day = {}
        if timee_slots:
            for day, kubun_dict in timee_slots.items():
                entries = []
                for kubun, pairs in kubun_dict.items():
                    for tin, tout in pairs:
                        tin_h = int(tin.split(':')[0]) if ':' in tin else 0
                        entries.append((tin_h, tin, tout, kubun))
                entries.sort()
                dc_pool_by_day[day] = entries

        def pop_matching_dc(day, km_class, used_indices):
            pool = dc_pool_by_day.get(day, [])
            for i, (tin_h, tin, tout, kubun) in enumerate(pool):
                if i in used_indices: continue
                match = match_timee_symbol_dc(km_class, tin_h, None)
                if match or match is None:
                    used_indices.add(i)
                    return tin, tout, kubun
            return None

        used_dc_by_day  = {}
        n_km_rows       = len(km_timee_rows or [])
        max_dc_per_day  = max((len(v) for v in dc_pool_by_day.values()), default=0)
        n_display_rows  = max(n_km_rows, max_dc_per_day, 1)

        ROW_COLORS = [
            ('EDE9FF', '3D2B8A'),
            ('FEF3C7', '92400E'),
            ('E0F2FE', '0369A1'),
            ('DCFCE7', '166534'),
            ('FFF0F8', '7B1A4A'),
        ]

        for row_idx in range(n_display_rows):
            bg_c, fg_c = ROW_COLORS[row_idx % len(ROW_COLORS)]

            if row_idx < n_km_rows:
                km_row_name, km_row_data = km_timee_rows[row_idx]
                label = km_row_name
            else:
                km_row_name = ''
                km_row_data = {}
                label = f'タイミー追加枠{row_idx+1}'

            ws.row_dimensions[r].height = 20
            nc = ws.cell(row=r, column=1, value=label)
            nc.font = _font(bold=True, size=8, color=fg_c)
            nc.fill = _fill(bg_c); nc.border = BDR
            nc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            ws.row_dimensions[r+1].height = 14
            nc2 = ws.cell(row=r+1, column=1, value='勤務表↓')
            nc2.font = Font(name='Arial', size=7, color='888888')
            nc2.fill = _fill('F8F8F8'); nc2.border = BDR
            nc2.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            km_class_map = {}
            for day in DAYS:
                km_raw = km_row_data.get(day, '')
                km_class_map[day] = classify_km_timee_symbol(km_raw) if km_raw else None

            for day in DAYS:
                col = day + 1
                if day not in used_dc_by_day:
                    used_dc_by_day[day] = set()

                km_raw     = km_row_data.get(day, '')
                km_working = bool(km_raw)
                km_class   = km_class_map[day]

                dc_entry = pop_matching_dc(day, km_class, used_dc_by_day[day]) if km_working else None

                if not km_working:
                    pool = dc_pool_by_day.get(day, [])
                    remaining = [(i,e) for i,e in enumerate(pool) if i not in used_dc_by_day[day]]
                    if row_idx < len(remaining):
                        idx, entry = remaining[row_idx]
                        used_dc_by_day[day].add(idx)
                        dc_entry = (entry[1], entry[2], entry[3])

                if dc_entry:
                    tin, tout, kubun = dc_entry
                    tin_h  = int(tin.split(':')[0])  if tin  and ':' in tin  else None
                    tout_h = int(tout.split(':')[0]) if tout and ':' in tout else None
                    if kubun == '夜勤':   dc_text = f'夜入\n{tin}〜'
                    elif kubun == '深夜': dc_text = f'深夜\n{tin}〜'
                    else:                 dc_text = f'日勤\n{tin}-{tout}'
                    dc_working = True
                else:
                    tin_h = tout_h = None
                    dc_text    = ''
                    dc_working = False

                if not km_timee_rows:
                    mismatch = False
                elif dc_working and km_working:
                    result   = match_timee_symbol_dc(km_class, tin_h, tout_h)
                    mismatch = (result is False)
                elif dc_working != km_working:
                    mismatch = True
                else:
                    mismatch = False

                c_dc = ws.cell(row=r, column=col, value=dc_text)
                c_dc.font = Font(name='Arial', size=7,
                                 color='C0392B' if mismatch else fg_c)
                c_dc.fill = _fill('FFD0D0' if mismatch else (bg_c if dc_working else 'F8F8F8'))
                c_dc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c_dc.border = BDR

                km_display = ('⚠要確認\n' + km_raw) if mismatch else km_raw
                c_km = ws.cell(row=r+1, column=col, value=km_display)
                c_km.font = Font(name='Arial', size=7,
                                 color='C0392B' if mismatch else '555555')
                c_km.fill = _fill('FFD0D0' if mismatch else 'F8F8F8')
                c_km.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c_km.border = BDR

            r += 2

    ws.freeze_panes = 'B3'


def _write_summary_sheet(ws2, dc_data, dc_order,
                         timee_i_data, timee_i_names,
                         timee_ii_data, timee_ii_names,
                         dc_minutes=None,
                         yukyu_data=None):
    CH   = '2D5F8A'
    hdrs = ['職員名', '区分', '日勤', '夜勤入り', '夜勤明け', '公休', '年休', '打刻なし', '勤務計', '月間勤務時間', '有休日数', '有休時間']
    for ci, h in enumerate(hdrs, start=1):
        c           = ws2.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = _fill(CH)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = BDR
    ws2.row_dimensions[1].height = 20
    for i, w in enumerate([16, 12, 8, 8, 8, 6, 6, 8, 8, 14, 8, 10]):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    all_entries = [(n, dc_data[n], '固定職員') for n in dc_order]

    r2 = 2
    for ri, (name, daily, kubun) in enumerate(all_entries):
        cnts = {'d': 0, 'ni': 0, 'na': 0, 'kk': 0, 'nk': 0, 'none': 0}
        for day in DAYS:
            info = daily.get(day)
            code = info[0] if info else None
            if code in cnts: cnts[code] += 1
            else:             cnts['none'] += 1

        bg   = 'EBF3FB' if ri % 2 == 0 else 'FFFFFF'
        vals = [name, kubun, cnts['d'], cnts['ni'], cnts['na'],
                cnts['kk'], cnts['nk'], cnts['none']]
        for ci, v in enumerate(vals, start=1):
            c           = ws2.cell(row=r2, column=ci, value=v)
            c.font      = _font(size=9)
            c.fill      = _fill(bg)
            c.alignment = Alignment(
                horizontal='center' if ci > 1 else 'left',
                vertical='center', indent=1 if ci == 1 else 0)
            c.border = BDR

        c           = ws2.cell(row=r2, column=9, value=f'=C{r2}+D{r2}')
        c.font      = _font(size=9, bold=True)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = BDR

        raw_name  = name.lstrip('★').strip()
        total_min = (dc_minutes or {}).get(raw_name, 0)
        time_text = f'{total_min // 60}時間{total_min % 60:02d}分' if total_min > 0 else '―'
        c10           = ws2.cell(row=r2, column=10, value=time_text)
        c10.font      = _font(size=9, bold=True,
                              color='1D4ED8' if total_min > 0 else '888888')
        c10.fill      = _fill(bg)
        c10.alignment = Alignment(horizontal='center', vertical='center')
        c10.border    = BDR

        yukyu   = (yukyu_data or {}).get(raw_name, {'days': 0, 'minutes': 0})
        y_days  = yukyu['days']
        y_mins  = yukyu['minutes']
        y_time  = f'{y_mins // 60}時間{y_mins % 60:02d}分' if y_mins > 0 else '―'

        c11           = ws2.cell(row=r2, column=11, value=y_days if y_days > 0 else '―')
        c11.font      = _font(size=9, bold=True, color='166534' if y_days > 0 else '888888')
        c11.fill      = _fill(bg)
        c11.alignment = Alignment(horizontal='center', vertical='center')
        c11.border    = BDR

        c12           = ws2.cell(row=r2, column=12, value=y_time)
        c12.font      = _font(size=9, bold=True, color='166534' if y_mins > 0 else '888888')
        c12.fill      = _fill(bg)
        c12.alignment = Alignment(horizontal='center', vertical='center')
        c12.border    = BDR

        ws2.row_dimensions[r2].height = 16
        r2 += 1

    for ci in range(1, 13):
        cl      = get_column_letter(ci)
        c       = ws2.cell(row=r2, column=ci)
        c.value = '合計' if ci == 1 else (
            '' if ci in (2, 10, 11, 12) else
            f'=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{r2 - 1})'
        )
        c.font      = _font(bold=True, size=10, color='FFFFFF')
        c.fill      = _fill('2D5F8A')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = BDR
    ws2.row_dimensions[r2].height = 18
    ws2.freeze_panes = 'C2'


def write_timee_summary_sheet(wb_check, 施設名, timee_slots_i, timee_slots_ii):
    CH    = '2D5F8A'
    KUBUN = ['日勤', '夜勤', '深夜']
    KUBUN_COLOR = {
        '日勤': ('DBEAFE', '1D4ED8'),
        '夜勤': ('FEF3C7', '92400E'),
        '深夜': ('F3E8FF', '5B21B6'),
    }

    by_day = defaultdict(lambda: {k: {'count': 0, 'minutes': 0} for k in KUBUN})
    for slots in [timee_slots_i or {}, timee_slots_ii or {}]:
        for day, kubun_dict in slots.items():
            for kubun, entries in kubun_dict.items():
                for tin, tout in entries:
                    by_day[day][kubun]['count'] += 1
                    if tin and tout and ':' in tin and ':' in tout:
                        try:
                            sh, sm = map(int, tin.split(':'))
                            eh, em = map(int, tout.split(':'))
                            s_min = sh*60+sm; e_min = eh*60+em
                            if e_min < s_min: e_min += 24*60
                            by_day[day][kubun]['minutes'] += e_min - s_min
                        except Exception:
                            pass

    ws = wb_check.create_sheet(f'{施設名}_タイミー集計')
    last_col = get_column_letter(34)

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = f'{YEAR}年{MONTH}月 タイミー勤務集計（RASIEL{施設名}）'
    ws['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = _fill(CH)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value='区分').font = _font(bold=True, color='FFFFFF')
    ws.cell(row=2, column=1).fill = _fill(CH); ws.cell(row=2, column=1).border = BDR
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = 14
    ws.cell(row=3, column=1, value='').fill = _fill(CH)
    ws.cell(row=3, column=1).border = BDR

    for day in DAYS:
        try:
            wd = datetime.date(YEAR, MONTH, day).weekday()
        except ValueError:
            continue
        cl = get_column_letter(day + 1)
        c  = ws.cell(row=2, column=day+1, value=f'{day}\n({DOW[wd]})')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BDR
        ws.column_dimensions[cl].width = 5.5
        if wd == 6:   c.fill = _fill('C0392B'); c.font = _font(bold=True, color='FFFFFF', size=8)
        elif wd == 5: c.fill = _fill('2471A3'); c.font = _font(bold=True, color='FFFFFF', size=8)
        else:         c.fill = _fill('4A90C4'); c.font = _font(bold=True, color='FFFFFF', size=8)

    for si, label in enumerate(['月計人数', '月計時間'], start=33):
        c = ws.cell(row=2, column=si, value=label)
        c.font = _font(bold=True, color='FFFFFF', size=9)
        c.fill = _fill(CH); c.border = BDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(si)].width = 9

    for day in DAYS:
        c = ws.cell(row=3, column=day+1, value='人数/時間')
        c.font = _font(size=7, color='888888')
        c.fill = _fill('F0F0F0'); c.border = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 12

    r = 4
    grand_count = 0; grand_min = 0

    for kubun in KUBUN:
        bg_c, fg_c = KUBUN_COLOR[kubun]
        ws.row_dimensions[r].height = 16
        nc = ws.cell(row=r, column=1, value=f'{kubun}　人数')
        nc.font = _font(bold=True, size=9, color=fg_c)
        nc.fill = _fill(bg_c); nc.border = BDR
        nc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ws.row_dimensions[r+1].height = 14
        nc2 = ws.cell(row=r+1, column=1, value=f'{kubun}　時間')
        nc2.font = Font(name='Arial', size=8, color=fg_c)
        nc2.fill = _fill(bg_c); nc2.border = BDR
        nc2.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        month_count = 0; month_min = 0
        for day in DAYS:
            col = day + 1
            d   = by_day.get(day, {}).get(kubun, {'count': 0, 'minutes': 0})
            cnt = d['count']; mins = d['minutes']
            h, m = divmod(mins, 60)

            c1 = ws.cell(row=r, column=col, value=cnt if cnt > 0 else '')
            c1.font = Font(name='Arial', size=9, bold=cnt > 0, color=fg_c if cnt > 0 else 'CCCCCC')
            c1.fill = _fill(bg_c if cnt > 0 else 'FAFAFA')
            c1.alignment = Alignment(horizontal='center', vertical='center')
            c1.border = BDR

            time_str = f'{h}h{m:02d}m' if mins > 0 else ''
            c2 = ws.cell(row=r+1, column=col, value=time_str)
            c2.font = Font(name='Arial', size=8, color=fg_c if mins > 0 else 'CCCCCC')
            c2.fill = _fill(bg_c if mins > 0 else 'FAFAFA')
            c2.alignment = Alignment(horizontal='center', vertical='center')
            c2.border = BDR

            month_count += cnt; month_min += mins

        mh, mm = divmod(month_min, 60)
        c_mc = ws.cell(row=r, column=33, value=month_count if month_count > 0 else '―')
        c_mc.font = _font(bold=True, size=10, color=fg_c)
        c_mc.fill = _fill(bg_c); c_mc.border = BDR
        c_mc.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r+1, column=33, value='').fill = _fill(bg_c)
        ws.cell(row=r+1, column=33).border = BDR

        c_mh = ws.cell(row=r, column=34, value=f'{mh}時間{mm:02d}分' if month_min > 0 else '―')
        c_mh.font = _font(bold=True, size=10, color=fg_c)
        c_mh.fill = _fill(bg_c); c_mh.border = BDR
        c_mh.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=r+1, column=34, value='').fill = _fill(bg_c)
        ws.cell(row=r+1, column=34).border = BDR

        grand_count += month_count; grand_min += month_min
        r += 2

    ws.row_dimensions[r].height = 20
    nc_g = ws.cell(row=r, column=1, value='合計')
    nc_g.font = _font(bold=True, size=11, color='FFFFFF')
    nc_g.fill = _fill(CH); nc_g.border = BDR
    nc_g.alignment = Alignment(horizontal='center', vertical='center')

    for day in DAYS:
        col = day + 1
        day_total = sum(by_day.get(day, {}).get(k, {}).get('count', 0) for k in KUBUN)
        c = ws.cell(row=r, column=col, value=day_total if day_total > 0 else '')
        c.font = _font(bold=True, size=9, color='FFFFFF')
        c.fill = _fill(CH); c.border = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')

    gh, gm = divmod(grand_min, 60)
    for col, val in [(33, grand_count), (34, f'{gh}時間{gm:02d}分')]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = _font(bold=True, size=11, color='FFFFFF')
        c.fill = _fill(CH); c.border = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'B4'


def write_check_sheet_to_wb(wb_check, 施設名, dc_data, km_i_data, km_ii_data):
    km_merged = {}
    for src in [km_i_data, km_ii_data]:
        if not src: continue
        for name, daily in src.items():
            base = name.rstrip('①②③').strip()
            if base not in km_merged:
                km_merged[base] = {}
            for day, (working, raw) in daily.items():
                if day not in km_merged[base] or \
                   (working and not km_merged[base][day][0]):
                    km_merged[base][day] = (working, raw)

    discs, match_cnt = check_consistency(dc_data, km_merged)
    by_person        = defaultdict(list)
    for d in discs:
        by_person[d['dc_name']].append(d)
    by_type = Counter(d['type'] for d in discs)

    CH = '2D5F8A'

    ws_s = wb_check.create_sheet(f'{施設名}_サマリー')
    ws_s.merge_cells('A1:F1')
    ws_s['A1'] = f'整合性チェック（{施設名}）　{YEAR}年{MONTH}月'
    ws_s['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws_s['A1'].fill      = _fill(CH)
    ws_s['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_s.row_dimensions[1].height = 22

    ws_s.merge_cells('A2:F2')
    ws_s['A2'] = ('【このレポートについて】'
                  'DC勤怠（打刻実績）を基準に勤務表との差異を抽出。'
                  '勤務表は実績ベースですが修正未完了の場合があります。'
                  '不一致箇所は担当者が確認してください。')
    ws_s['A2'].font      = _font(size=9, color='5B5B5B')
    ws_s['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    ws_s.row_dimensions[2].height = 28

    overview = [
        ('チェック対象職員数',   f'{len(dc_data)}名'),
        ('不一致件数　合計',      f'{len(discs)}件'),
        ('  DC出勤 / 勤務表休み', f'{by_type.get("DC出勤 / 勤務表休み", 0)}件'),
        ('  DC休み / 勤務表出勤', f'{by_type.get("DC休み / 勤務表出勤", 0)}件'),
        ('一致件数',              f'{match_cnt}件'),
    ]
    for ri, (label, val) in enumerate(overview, start=3):
        ws_s.row_dimensions[ri].height = 16
        c           = ws_s.cell(row=ri, column=1, value=label)
        c.font      = _font(bold=True, size=10)
        c.fill      = _fill('EBF3FB' if ri % 2 else 'F4F8FD')
        c.border    = BDR
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c           = ws_s.cell(row=ri, column=2, value=val)
        c.font      = Font(name='Arial', bold=True, size=10,
                           color='C0392B' if '不一致' in label and label.strip().startswith('不') else '000000')
        c.fill      = _fill('EBF3FB' if ri % 2 else 'F4F8FD')
        c.border    = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')

    r9 = 9
    ws_s.cell(row=r9, column=1, value='職員別不一致件数').font = \
        Font(name='Arial', bold=True, size=10, color='FFFFFF')
    ws_s.cell(row=r9, column=1).fill      = _fill(CH)
    ws_s.cell(row=r9, column=1).border    = BDR
    ws_s.cell(row=r9, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_s.row_dimensions[r9].height = 18

    for ci, h in enumerate(['職員名', '不一致合計', 'DC出勤/勤務表休み', 'DC休み/勤務表出勤'], start=1):
        c           = ws_s.cell(row=r9+1, column=ci, value=h)
        c.font      = _font(bold=True, color='FFFFFF', size=9)
        c.fill      = _fill('4A90C4')
        c.border    = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_s.row_dimensions[r9+1].height = 16

    ri2 = r9 + 2
    for name, items in sorted(by_person.items(), key=lambda x: -len(x[1])):
        tc = Counter(i['type'] for i in items)
        bg = 'FFF0F0' if len(items) > 3 else ('FFFDE7' if len(items) > 0 else 'FFFFFF')
        for ci, v in enumerate(
            [name, len(items),
             tc.get('DC出勤 / 勤務表休み', 0),
             tc.get('DC休み / 勤務表出勤', 0)], start=1
        ):
            c      = ws_s.cell(row=ri2, column=ci, value=v)
            c.font = Font(name='Arial', bold=(ci == 2 and len(items) > 0), size=9,
                          color='C0392B' if ci == 2 and len(items) > 3 else '000000')
            c.fill      = _fill(bg)
            c.border    = BDR
            c.alignment = Alignment(
                horizontal='center' if ci > 1 else 'left',
                vertical='center', indent=1 if ci == 1 else 0)
        ws_s.row_dimensions[ri2].height = 16
        ri2 += 1

    for i, w in enumerate([20, 10, 16, 16]):
        ws_s.column_dimensions[get_column_letter(i+1)].width = w

    ws_d = wb_check.create_sheet(f'{施設名}_詳細')
    hdrs = ['職員名', '日', '不一致種別', '勤務表の記号', 'DC勤怠区分', 'DC出勤打刻', 'DC退勤打刻']
    for ci, h in enumerate(hdrs, start=1):
        c           = ws_d.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color='FFFFFF', size=9)
        c.fill      = _fill(CH)
        c.border    = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_d.row_dimensions[1].height = 18

    for ri, disc in enumerate(sorted(discs, key=lambda x: (x['dc_name'], x['day'])), start=2):
        bg   = 'FFF0F0' if disc['type'] == 'DC出勤 / 勤務表休み' else 'FFFDE7'
        vals = [disc['dc_name'], f"{MONTH}/{disc['day']}", disc['type'],
                disc['km_raw'], CODE_LABEL.get(disc['dc_code'], ''),
                disc['dc_in'], disc['dc_out']]
        for ci, v in enumerate(vals, start=1):
            c           = ws_d.cell(row=ri, column=ci, value=v)
            c.font      = _font(size=9)
            c.fill      = _fill(bg)
            c.border    = BDR
            c.alignment = Alignment(
                horizontal='center' if ci > 1 else 'left',
                vertical='center', indent=1 if ci == 1 else 0)
        ws_d.row_dimensions[ri].height = 15

    for i, w in enumerate([20, 8, 20, 14, 12, 10, 10]):
        ws_d.column_dimensions[get_column_letter(i+1)].width = w
    ws_d.freeze_panes = 'A2'

    return len(discs), match_cnt


# ============================================================
# メイン処理
# ============================================================
def main():
    log         = []
    total_discs = 0

    wb_check = Workbook()
    ws_top   = wb_check.active
    ws_top.title = '全施設サマリー'
    CH = '2D5F8A'

    ws_top.merge_cells('A1:G1')
    ws_top['A1'] = f'整合性チェック　全施設サマリー　{YEAR}年{MONTH}月'
    ws_top['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws_top['A1'].fill      = _fill(CH)
    ws_top['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_top.row_dimensions[1].height = 22

    for ci, h in enumerate(
        ['施設名', '不一致合計', 'DC出勤/勤務表休み', 'DC休み/勤務表出勤', '一致件数', '状態', '備考'], start=1
    ):
        c           = ws_top.cell(row=2, column=ci, value=h)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = _fill('4A90C4')
        c.border    = BDR
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_top.row_dimensions[2].height = 18
    for i, w in enumerate([12, 10, 16, 16, 10, 14, 30]):
        ws_top.column_dimensions[get_column_letter(i+1)].width = w

    top_row = 3

    for 施設名, config in FACILITIES.items():
        print(f'\n処理中: {施設名} ...')

        dc_path                     = find_dc_file(施設名)
        timee_i_path, timee_ii_path = find_timee_files(施設名, config['timee_i'], config['timee_ii'])
        shift_i_path, shift_ii_path = find_shift_files(施設名, config['has_ii'])

        print(f'  DC勤怠   : {os.path.basename(dc_path)       if dc_path       else "❌ 見つからず"}')
        if config['timee_i']:
            print(f'  タイミーⅰ: {os.path.basename(timee_i_path)  if timee_i_path  else "❌ 見つからず"}')
        if config['timee_ii']:
            print(f'  タイミーⅱ: {os.path.basename(timee_ii_path) if timee_ii_path else "❌ 見つからず"}')
        print(f'  勤務表ⅰ  : {os.path.basename(shift_i_path)  if shift_i_path  else "❌ 見つからず"}')
        if config['has_ii']:
            print(f'  勤務表ⅱ  : {os.path.basename(shift_ii_path) if shift_ii_path else "❌ 見つからず"}')

        missing = []
        if not dc_path:                                    missing.append('DC勤怠CSV')
        if config['timee_i']  and not timee_i_path:        missing.append('タイミーⅰCSV')
        if config['timee_ii'] and not timee_ii_path:       missing.append('タイミーⅱCSV')
        if not shift_i_path:                               missing.append('勤務表ⅰ')
        if config['has_ii']   and not shift_ii_path:       missing.append('勤務表ⅱ')

        if missing:
            msg = f'⚠️  {施設名}：{missing} が見つかりません → スキップ'
            print(msg); log.append(msg)
            note = ' / '.join(missing) + ' が見つからず'
            for ci, v in enumerate([施設名, '-', '-', '-', '-', '⚠️ スキップ', note], start=1):
                c           = ws_top.cell(row=top_row, column=ci, value=v)
                c.font      = _font(size=10, color='856404')
                c.fill      = _fill('FFF3CD')
                c.border    = BDR
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws_top.row_dimensions[top_row].height = 16
            top_row += 1
            continue

        try:
            dc_data, dc_order = load_dc_csv(dc_path)
            dc_minutes        = calc_staff_minutes(dc_data)

            timee_i_data,  timee_i_names  = {}, []
            timee_ii_data, timee_ii_names = {}, []
            if config['timee_i']  and timee_i_path:
                timee_i_data,  timee_i_names  = load_timee_csv(timee_i_path)
            if config['timee_ii'] and timee_ii_path:
                timee_ii_data, timee_ii_names = load_timee_csv(timee_ii_path)

            km_i_data,  km_order_i  = load_kinmuhyo(shift_i_path)
            km_ii_data, km_order_ii = ({}, []) if not shift_ii_path else load_kinmuhyo(shift_ii_path)

            km_timee_slots_i  = load_km_timee_slots(shift_i_path)
            km_timee_slots_ii = load_km_timee_slots(shift_ii_path) if shift_ii_path else {}

            yukyu_data = calc_yukyu_from_km(km_i_data, km_ii_data)

            timee_slots_i  = load_timee_slots(timee_i_path)  if timee_i_path  else {}
            timee_slots_ii = load_timee_slots(timee_ii_path) if timee_ii_path else {}

            km_timee_rows_i  = load_km_timee_rows(shift_i_path)
            km_timee_rows_ii = load_km_timee_rows(shift_ii_path) if shift_ii_path else []

            write_timee_summary_sheet(wb_check, 施設名, timee_slots_i, timee_slots_ii)

            write_kintai_excel(
                wb_check, 施設名,
                dc_data, dc_order,
                timee_i_data, timee_i_names,
                timee_ii_data, timee_ii_names,
                km_i_data, km_ii_data,
                km_timee_slots_i, km_timee_slots_ii,
                km_order_i, km_order_ii,
                dc_minutes,
                timee_slots_i, timee_slots_ii,
                km_timee_rows_i, km_timee_rows_ii,
                yukyu_data,
            )

            n_disc, n_match = write_check_sheet_to_wb(wb_check, 施設名, dc_data, km_i_data, km_ii_data)
            total_discs += n_disc

            bg_top = 'FFF0F0' if n_disc > 5 else ('FFFDE7' if n_disc > 0 else 'F0FFF4')
            status = '✅ 不一致なし' if n_disc == 0 else f'⚠️ {n_disc}件 要確認'
            for ci, v in enumerate([施設名, n_disc, '-', '-', n_match, status, ''], start=1):
                c      = ws_top.cell(row=top_row, column=ci, value=v)
                c.font = Font(name='Arial', size=10,
                              bold=(ci == 2 and n_disc > 0),
                              color='C0392B' if ci == 2 and n_disc > 5 else
                                    '166534' if ci == 6 and n_disc == 0 else '000000')
                c.fill      = _fill(bg_top)
                c.border    = BDR
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws_top.row_dimensions[top_row].height = 16
            top_row += 1

            msg = f'✅  {施設名}：完了（不一致 {n_disc}件）'
            print(msg); log.append(msg)

        except Exception as e:
            import traceback
            msg = f'❌  {施設名}：エラー → {str(e)}'
            print(msg)
            traceback.print_exc()
            log.append(msg)
            for ci, v in enumerate([施設名, '-', '-', '-', '-', '❌ エラー', str(e)], start=1):
                c           = ws_top.cell(row=top_row, column=ci, value=v)
                c.font      = _font(size=10, color='C0392B')
                c.fill      = _fill('FFE0E0')
                c.border    = BDR
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws_top.row_dimensions[top_row].height = 16
            top_row += 1

    out_check = os.path.join(OUT_DIR, f'勤怠一覧_整合性チェック_全施設_{YM_STR}.xlsx')
    wb_check.save(out_check)
    print(f'\n✅ Excel保存完了: {out_check}')

    log_path = os.path.join(OUT_DIR, f'実行ログ_{YM_STR}.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write(f'実行日時: {datetime.datetime.now()}\n')
        f.write(f'対象: {YEAR}年{MONTH}月\n\n')
        f.write('\n'.join(log))

    print(f'\n{"=" * 50}')
    print(f'全処理完了　不一致合計: {total_discs}件')
    print(f'出力先: {OUT_DIR}')
    print(f'{"=" * 50}')


if __name__ == '__main__':
    main()
