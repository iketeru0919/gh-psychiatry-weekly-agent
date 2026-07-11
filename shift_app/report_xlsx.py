"""取り込みチェック結果の Excel 出力。"""
import io


def report_to_xlsx(facility_name, report) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    warn_fill = PatternFill('solid', fgColor='FFF2CC')
    ng_font = Font(bold=True, color='B3261E')
    ok_font = Font(bold=True, color='1A7D3C')

    # --- シート1: チェック結果 ---
    ws = wb.active
    ws.title = 'チェック結果'
    ws['A1'] = f'{facility_name} {report["year"]}年{report["month"]}月 取り込みチェック結果'
    ws['A1'].font = Font(bold=True, size=14)
    r = 3
    ws.cell(row=r, column=1, value='■ サマリ').font = bold
    r += 1
    for k, v in report['summary'].items():
        ws.cell(row=r, column=1, value=k)
        c = ws.cell(row=r, column=2, value=v)
        if k == '警告件数' and isinstance(v, (int, float)) and v > 0:
            c.font = ng_font
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ ブック健全性').font = bold
    r += 1
    fm = report['format']
    items = []
    if fm.get('duplicate_import'):
        items.append('同じ施設・同じ月のデータが既に取り込まれています(二重取り込みの可能性)')
    if fm.get('broken_formulas'):
        items.append(f"壊れた数式(#REF!)が {fm['broken_formulas']} 個あります(元Excel由来)")
    items += fm.get('masters_warnings', [])
    for u in fm.get('unknown_codes', []):
        items.append(f"マスタ未登録の記号「{u['記号']}」使用日: {u['日付']}")
    if not items:
        ws.cell(row=r, column=1, value='問題なし').font = ok_font
        r += 1
    for it in items:
        c = ws.cell(row=r, column=1, value='⚠ ' + it)
        c.fill = warn_fill
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='■ 業務チェック').font = bold
    r += 1
    b = report['business']
    ha = b['haichi']
    for label, val in [('人員配置 基準判定', ha.get('基準判定')),
                       ('人員超過・過不足数', ha.get('人員超過・過不足数')),
                       ('④判定', ha.get('④判定')), ('④過不足', ha.get('④過不足')),
                       ('割増金発生件数', b.get('割増金発生件数')),
                       ('シフト入力率(%)', b.get('シフト入力率'))]:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val)
        if '未達' in str(val):
            c.font = ng_font
        elif '到達' in str(val):
            c.font = ok_font
        r += 1
    for label, lst in [('夜勤0名の日', b.get('夜勤0名の日', [])),
                       ('夜勤1名の日', b.get('夜勤1名の日', [])),
                       ('法定休日NG(割増金発生)', b.get('法定休日NG', [])),
                       ('異常値', b.get('異常値', []))]:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value='、'.join(str(x) for x in lst) or 'なし')
        if lst:
            c.fill = warn_fill
        r += 1

    if report.get('prev'):
        r += 1
        ws.cell(row=r, column=1, value=f"■ 前月比較(前月 {report['prev'].get('年月', '')})").font = bold
        r += 1
        for col, h in enumerate(['項目', '前月', '今月', '増減'], start=1):
            ws.cell(row=r, column=col, value=h).font = bold
        r += 1
        for k, v in report['prev'].items():
            if k == '年月':
                continue
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v.get('前月'))
            ws.cell(row=r, column=3, value=v.get('今月'))
            ws.cell(row=r, column=4, value=v.get('増減'))
            r += 1
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 30

    # --- シート2: 職員別月次集計 ---
    ws2 = wb.create_sheet('職員別月次集計')
    rows = report.get('staff_monthly', [])
    if rows:
        cols = list(rows[0].keys())
        for j, h in enumerate(cols, start=1):
            ws2.cell(row=1, column=j, value=h).font = bold
        for i, row in enumerate(rows, start=2):
            for j, h in enumerate(cols, start=1):
                ws2.cell(row=i, column=j, value=row.get(h))
        ws2.column_dimensions['A'].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
