"""Excel数式評価エンジン。

openpyxl でブックを読み込み、数式セルは AST 評価、入力セルはリテラル値として
参照する。日付・時刻は Excel シリアル値(1900日付システム)に正規化する。
"""
import datetime as _dt
import fnmatch
import sys

from .parser import parse_formula, num_to_col, col_to_num, tokenize, Parser

_EPOCH = _dt.date(1899, 12, 30)
_JP_WEEKDAYS = ['土', '日', '月', '火', '水', '木', '金']  # serial % 7 で引く


class XLErr:
    __slots__ = ('code',)

    def __init__(self, code):
        self.code = code

    def __repr__(self):
        return f'XLErr({self.code})'

    def __eq__(self, other):
        return isinstance(other, XLErr) and other.code == self.code

    def __hash__(self):
        return hash(self.code)


ERR_VALUE = XLErr('#VALUE!')
ERR_NA = XLErr('#N/A')
ERR_DIV0 = XLErr('#DIV/0!')
ERR_REF = XLErr('#REF!')
ERR_NAME = XLErr('#NAME?')
ERR_NUM = XLErr('#NUM!')


_LEAP_BUG_CUTOFF = _dt.date(1900, 3, 1)


def _date_serial(d: _dt.date) -> int:
    # Excel 1900 日付システム: 1900/2/29(実在しない)を挟むため
    # 1900/3/1 より前は基準日が1日ずれる
    if d < _LEAP_BUG_CUTOFF:
        return (d - _dt.date(1899, 12, 31)).days
    return (d - _EPOCH).days


def date_to_serial(d) -> float:
    if isinstance(d, _dt.datetime):
        frac = (d.hour * 3600 + d.minute * 60 + d.second) / 86400.0 + d.microsecond / 86400e6
        return _date_serial(d.date()) + frac
    if isinstance(d, _dt.date):
        return float(_date_serial(d))
    if isinstance(d, _dt.time):
        return (d.hour * 3600 + d.minute * 60 + d.second) / 86400.0 + d.microsecond / 86400e6
    if isinstance(d, _dt.timedelta):
        return d.total_seconds() / 86400.0
    raise TypeError(type(d))


def serial_to_date(s: float) -> _dt.date:
    return _EPOCH + _dt.timedelta(days=int(s))


class RangeVal:
    __slots__ = ('sheet', 'c1', 'r1', 'c2', 'r2', 'rows')

    def __init__(self, sheet, c1, r1, c2, r2, rows):
        self.sheet, self.c1, self.r1, self.c2, self.r2 = sheet, c1, r1, c2, r2
        self.rows = rows  # list[list[value]]

    def flat(self):
        for row in self.rows:
            yield from row


def _num(v, transition=False):
    """数値文脈への型強制。失敗時は XLErr を返す。

    transition=True は Excel の「Lotus 1-2-3 形式の計算」
    (sheetPr/@transitionEvaluation) 相当: 数値化できない文字列は 0 になる。
    """
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        try:
            return float(s) if s else (0.0 if transition else ERR_VALUE)
        except ValueError:
            return 0.0 if transition else ERR_VALUE
    if isinstance(v, XLErr):
        return v
    return ERR_VALUE


def _fmt_num(x: float) -> str:
    if x == int(x) and abs(x) < 1e15:
        return str(int(x))
    return repr(x)


def _text(v):
    """文字列文脈への型強制(& 連結など)。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float)):
        return _fmt_num(float(v))
    if isinstance(v, XLErr):
        return v
    return v


def _bool(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        u = v.upper()
        if u == 'TRUE':
            return True
        if u == 'FALSE':
            return False
        return ERR_VALUE
    if isinstance(v, XLErr):
        return v
    return ERR_VALUE


def _cmp(op, a, b, transition=False):
    """Excel の比較演算。型が異なる場合は 数値 < 文字列 < 論理値 の順。

    transition=True では文字列と数値の比較時に文字列を数値化(不能なら0)して
    数値同士で比較する(Lotus 互換)。
    """
    if isinstance(a, XLErr):
        return a
    if isinstance(b, XLErr):
        return b
    if transition:
        if isinstance(a, str) and isinstance(b, (int, float)) and not isinstance(b, bool):
            a = _num(a, True)
        elif isinstance(b, str) and isinstance(a, (int, float)) and not isinstance(a, bool):
            b = _num(b, True)
    # 空セルは相手の型に合わせる
    if a is None:
        a = ('' if isinstance(b, str) else False if isinstance(b, bool) else 0.0)
    if b is None:
        b = ('' if isinstance(a, str) else False if isinstance(a, bool) else 0.0)

    def rank(v):
        if isinstance(v, bool):
            return 2
        if isinstance(v, str):
            return 1
        return 0

    ra, rb = rank(a), rank(b)
    if ra != rb:
        c = -1 if ra < rb else 1
    else:
        if isinstance(a, str):
            aa, bb = a.casefold(), b.casefold()
            c = -1 if aa < bb else (0 if aa == bb else 1)
        else:
            fa, fb = float(a), float(b)
            c = -1 if fa < fb else (0 if fa == fb else 1)
    return {
        '=': c == 0, '<>': c != 0,
        '<': c < 0, '<=': c <= 0, '>': c > 0, '>=': c >= 0,
    }[op]


class _Criteria:
    """COUNTIF/SUMIF 系の条件。"""

    def __init__(self, crit):
        self.op = '='
        self.rhs = crit
        if isinstance(crit, str):
            for op in ('<>', '<=', '>=', '<', '>', '='):
                if crit.startswith(op):
                    self.op = op
                    rest = crit[len(op):]
                    try:
                        self.rhs = float(rest)
                    except ValueError:
                        self.rhs = rest
                    return
            self.rhs = crit
        elif crit is None:
            self.rhs = 0.0

    def match(self, v):
        op, rhs = self.op, self.rhs
        if isinstance(rhs, str):
            if op == '=':
                if rhs == '':
                    return v is None or v == ''
                if isinstance(v, XLErr) or v is None:
                    return False
                sv = _text(v) if not isinstance(v, str) else v
                pat = rhs.casefold()
                if '*' in pat or '?' in pat:
                    return fnmatch.fnmatchcase(sv.casefold(), pat)
                return sv.casefold() == pat
            if op == '<>':
                if rhs == '':
                    return not (v is None or v == '')
                if not isinstance(v, str):
                    return True
                return v.casefold() != rhs.casefold()
            if not isinstance(v, str):
                return False
            r = _cmp(op, v, rhs)
            return r if isinstance(r, bool) else False
        # 数値条件: 数値セルのみ対象
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return False
        r = _cmp(op, float(v), rhs)
        return r if isinstance(r, bool) else False


class Workbook:
    def __init__(self, path, today=None, keep_sheets=None, cache_dir=None):
        import openpyxl
        self.path = path
        self.today = today or _dt.date.today()
        self.cells = {}      # sheet -> {(col,row): ('f', src) | ('v', value)}
        self.cached = {}     # sheet -> {(col,row): cached value} (数式セルの計算済み値)
        self.formulas = {}   # sheet -> {(col,row): formula str}
        self._memo = {}
        self._stack = set()
        self.volatile_cells = set()  # TODAY 等に依存したセル

        self.transition_sheets = set()
        self.extent = {}  # sheet -> (max_row, max_col) 使用範囲(列全体参照の評価境界)
        if cache_dir:
            import hashlib
            import os
            import pickle
            h = hashlib.sha256(open(path, 'rb').read()).hexdigest()[:16]
            pkl = os.path.join(cache_dir, f'xlmodel-v3-{h}.pkl')
            if os.path.exists(pkl):
                with open(pkl, 'rb') as fh:
                    (self.cells, self.cached, self.formulas,
                     self.transition_sheets, self.extent) = pickle.load(fh)
                return
            self._load(path, keep_sheets)
            with open(pkl, 'wb') as fh:
                pickle.dump((self.cells, self.cached, self.formulas,
                             self.transition_sheets, self.extent), fh)
            return
        self._load(path, keep_sheets)

    def _load_transition_flags(self, path):
        """sheetPr/@transitionEvaluation='1' のシートを列挙する。"""
        import re
        import zipfile
        z = zipfile.ZipFile(path)
        wbxml = z.read('xl/workbook.xml').decode('utf-8')
        sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="rId(\d+)"', wbxml)
        rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        relmap = dict(re.findall(r'Id="rId(\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', rels))
        out = set()
        for name, rid in sheets:
            target = relmap.get(rid)
            if not target:
                continue
            head = z.read('xl/' + target)[:2000].decode('utf-8', 'ignore')
            m = re.search(r'<sheetPr[^>]*transitionEvaluation="(?:1|true)"', head)
            if m:
                out.add(name)
        z.close()
        return out

    def _load(self, path, keep_sheets):
        import openpyxl
        self.transition_sheets = self._load_transition_flags(path)
        wbf = openpyxl.load_workbook(path, data_only=False)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for name in wbf.sheetnames:
            if keep_sheets is not None and name not in keep_sheets:
                continue
            wsf, wsv = wbf[name], wbv[name]
            cmap, fmap, cachemap = {}, {}, {}
            for rowf, rowv in zip(wsf.iter_rows(), wsv.iter_rows()):
                for cf, cv in zip(rowf, rowv):
                    v = cf.value
                    if v is None:
                        continue
                    key = (cf.column, cf.row)
                    if hasattr(v, 'text'):  # ArrayFormula
                        v = str(v.text)
                    if isinstance(v, str) and v.startswith('='):
                        cmap[key] = ('f', v)
                        fmap[key] = v
                        cachemap[key] = self._norm_literal(cv.value)
                    else:
                        cmap[key] = ('v', self._norm_literal(v))
            self.cells[name] = cmap
            self.formulas[name] = fmap
            self.cached[name] = cachemap
            self.extent[name] = (
                max((k[1] for k in cmap), default=1),
                max((k[0] for k in cmap), default=1),
            )
        wbf.close()
        wbv.close()

    @staticmethod
    def _norm_literal(v):
        if isinstance(v, (_dt.datetime, _dt.date, _dt.time, _dt.timedelta)):
            return date_to_serial(v)
        if isinstance(v, str) and v.startswith('#') and v.endswith(('!', '?', 'A')):
            return XLErr(v)
        if isinstance(v, int) and not isinstance(v, bool):
            return float(v)
        return v

    # ---- セル値 ----
    def value(self, sheet, col, row):
        key = (sheet, col, row)
        if key in self._memo:
            return self._memo[key]
        cmap = self.cells.get(sheet)
        if cmap is None:
            return ERR_REF
        cell = cmap.get((col, row))
        if cell is None:
            return None
        kind, payload = cell
        if kind == 'v':
            return payload
        if key in self._stack:
            raise RuntimeError(f'circular ref at {sheet}!{num_to_col(col)}{row}')
        self._stack.add(key)
        try:
            ctx = _Ctx(sheet, col, row)
            ctx.transition = sheet in self.transition_sheets
            v = self.eval_node(parse_formula(payload), ctx)
            if isinstance(v, RangeVal):  # 動的配列: アンカーには左上値
                v = v.rows[0][0] if v.rows else None
            if v is None:
                v = 0.0  # 空白参照を返す数式の結果は 0 になる
            if ctx.volatile:
                self.volatile_cells.add(key)
                for dep in ctx.deps:
                    pass
            self._memo[key] = v
            return v
        finally:
            self._stack.discard(key)

    # ---- AST 評価 ----
    def eval_node(self, node, ctx):
        tag = node[0]
        if tag == 'num':
            return node[1]
        if tag == 'str':
            return node[1]
        if tag == 'bool':
            return node[1]
        if tag == 'err':
            return XLErr(node[1])
        if tag == 'blank':
            return None
        if tag == 'ref':
            sheet = node[1] or ctx.sheet
            v = self.value(sheet, node[2], node[3])
            if (sheet, node[2], node[3]) in self.volatile_cells:
                ctx.volatile = True
            return v
        if tag == 'range':
            return self._range(node, ctx)
        if tag == 'crange':
            sheet = node[1] or ctx.sheet
            maxr = self.extent.get(sheet, (1, 1))[0]
            return self._range(('range', node[1], node[2], 1, node[3], maxr), ctx)
        if tag == 'un':
            v = _num(self.eval_node(node[2], ctx), ctx.transition)
            if isinstance(v, XLErr):
                return v
            return -v if node[1] == '-' else v
        if tag == 'pct':
            v = _num(self.eval_node(node[1], ctx), ctx.transition)
            if isinstance(v, XLErr):
                return v
            return v / 100.0
        if tag == 'bin':
            return self._binop(node, ctx)
        if tag == 'call':
            return self._call(node[1], node[2], ctx)
        raise ValueError(f'bad node {node!r}')

    def _range(self, node, ctx):
        sheet = node[1] or ctx.sheet
        c1, r1, c2, r2 = node[2], node[3], node[4], node[5]
        rows = []
        vol = False
        for r in range(r1, r2 + 1):
            row = []
            for c in range(c1, c2 + 1):
                row.append(self.value(sheet, c, r))
                if (sheet, c, r) in self.volatile_cells:
                    vol = True
            rows.append(row)
        if vol:
            ctx.volatile = True
        return RangeVal(sheet, c1, r1, c2, r2, rows)

    def _binop(self, node, ctx):
        op = node[1]
        a = self.eval_node(node[2], ctx)
        b = self.eval_node(node[3], ctx)
        if isinstance(a, RangeVal) or isinstance(b, RangeVal):
            return ERR_VALUE
        if op in ('=', '<>', '<', '<=', '>', '>='):
            return _cmp(op, a, b, ctx.transition)
        if op == '&':
            ta, tb = _text(a), _text(b)
            if isinstance(ta, XLErr):
                return ta
            if isinstance(tb, XLErr):
                return tb
            return ta + tb
        fa, fb = _num(a, ctx.transition), _num(b, ctx.transition)
        if isinstance(fa, XLErr):
            return fa
        if isinstance(fb, XLErr):
            return fb
        if op == '+':
            return fa + fb
        if op == '-':
            return fa - fb
        if op == '*':
            return fa * fb
        if op == '/':
            if fb == 0:
                return ERR_DIV0
            return fa / fb
        if op == '^':
            try:
                return float(fa ** fb)
            except (OverflowError, ValueError):
                return ERR_NUM
        raise ValueError(op)

    # ---- 関数 ----
    def _call(self, name, args, ctx):
        # 遅延評価が必要なもの
        if name == 'IF':
            c = _bool(self.eval_node(args[0], ctx))
            if isinstance(c, XLErr):
                return c
            if c:
                return self.eval_node(args[1], ctx) if len(args) > 1 else True
            return self.eval_node(args[2], ctx) if len(args) > 2 else False
        if name == 'IFERROR':
            v = self.eval_node(args[0], ctx)
            if isinstance(v, XLErr):
                return self.eval_node(args[1], ctx)
            return v
        if name == 'IFS':
            for i in range(0, len(args) - 1, 2):
                c = _bool(self.eval_node(args[i], ctx))
                if isinstance(c, XLErr):
                    return c
                if c:
                    return self.eval_node(args[i + 1], ctx)
            return ERR_NA
        if name == 'INDIRECT':
            s = self.eval_node(args[0], ctx)
            if not isinstance(s, str):
                return ERR_REF
            return self._indirect(s, ctx)
        if name == 'ROW':
            if not args:
                return float(ctx.row)
            a = args[0]
            if a[0] == 'ref':
                return float(a[3])
            if a[0] == 'range':
                return float(a[3])
            return ERR_VALUE
        if name in ('TODAY', 'NOW'):
            ctx.volatile = True
            return date_to_serial(self.today)
        if name == 'INDEX' and args and args[0][0] in ('range', 'crange'):
            # 参照形式の INDEX は選択された行/列だけを評価する(Excel と同じ遅延性)。
            # 範囲全体を先に評価すると、選ばれない列の逆参照で疑似循環になる。
            a0 = args[0]
            sheet = a0[1] or ctx.sheet
            if a0[0] == 'crange':
                c1, r1, c2 = a0[2], 1, a0[3]
                r2 = self.extent.get(sheet, (1, 1))[0]
            else:
                c1, r1, c2, r2 = a0[2], a0[3], a0[4], a0[5]
            r = int(_num(self.eval_node(args[1], ctx))) if len(args) > 1 else 1
            c = int(_num(self.eval_node(args[2], ctx))) if len(args) > 2 else 1
            nrows, ncols = r2 - r1 + 1, c2 - c1 + 1
            if nrows == 1 and len(args) == 2:
                r, c = 1, r
            if ncols == 1 and len(args) == 2:
                c = 1
            if r == 0 and 1 <= c <= ncols:
                cc = c1 + c - 1
                return self._range(('range', a0[1], cc, r1, cc, r2), ctx)
            if c == 0 and 1 <= r <= nrows:
                rr = r1 + r - 1
                return self._range(('range', a0[1], c1, rr, c2, rr), ctx)
            if r < 1 or r > nrows or c < 1 or c > ncols:
                return ERR_REF
            return self.value(sheet, c1 + c - 1, r1 + r - 1)

        vals = [self.eval_node(a, ctx) for a in args]
        fn = getattr(self, '_fn_' + name.replace('.', '_'), None)
        if fn is None:
            raise NotImplementedError(f'function {name}')
        return fn(vals, ctx)

    def _indirect(self, s, ctx):
        try:
            toks = Parser(tokenize(s))
            node = toks.parse()
        except ValueError:
            return ERR_REF
        if node[0] in ('ref', 'range'):
            return self.eval_node(node, ctx)
        return ERR_REF

    # -- 集計 --
    @staticmethod
    def _iter_scalars(vals):
        for v in vals:
            if isinstance(v, RangeVal):
                yield from v.flat()
            else:
                yield v

    def _fn_SUM(self, vals, ctx):
        t = 0.0
        for v in self._iter_scalars(vals):
            if isinstance(v, XLErr):
                return v
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                t += v
        return t

    def _fn_SUBTOTAL(self, vals, ctx):
        fno = int(_num(vals[0]))
        rest = vals[1:]
        if fno in (9, 109):
            return self._fn_SUM(rest, ctx)
        if fno in (2, 102):
            return self._fn_COUNT(rest, ctx)
        if fno in (3, 103):
            return self._fn_COUNTA(rest, ctx)
        if fno in (4, 104):
            return self._fn_MAX(rest, ctx)
        if fno in (5, 105):
            return self._fn_MIN(rest, ctx)
        if fno in (1, 101):
            s = self._fn_SUM(rest, ctx)
            n = self._fn_COUNT(rest, ctx)
            return ERR_DIV0 if n == 0 else s / n
        raise NotImplementedError(f'SUBTOTAL({fno})')

    def _fn_COUNT(self, vals, ctx):
        n = 0
        for v in self._iter_scalars(vals):
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                n += 1
        return float(n)

    def _fn_COUNTA(self, vals, ctx):
        return float(sum(1 for v in self._iter_scalars(vals) if v is not None))

    def _fn_COUNTBLANK(self, vals, ctx):
        return float(sum(1 for v in self._iter_scalars(vals) if v is None or v == ''))

    def _fn_MAX(self, vals, ctx):
        best = None
        for v in self._iter_scalars(vals):
            if isinstance(v, XLErr):
                return v
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if best is None or v > best:
                    best = v
        return best if best is not None else 0.0

    def _fn_MIN(self, vals, ctx):
        best = None
        for v in self._iter_scalars(vals):
            if isinstance(v, XLErr):
                return v
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if best is None or v < best:
                    best = v
        return best if best is not None else 0.0

    def _fn_COUNTIF(self, vals, ctx):
        rng, crit = vals
        if isinstance(rng, XLErr):
            return rng
        c = _Criteria(crit)
        return float(sum(1 for v in rng.flat() if c.match(v)))

    def _fn_COUNTIFS(self, vals, ctx):
        for i in range(0, len(vals), 2):
            if isinstance(vals[i], XLErr):
                return vals[i]
        pairs = [(vals[i], _Criteria(vals[i + 1])) for i in range(0, len(vals), 2)]
        n = 0
        cells = [list(rng.flat()) for rng, _ in pairs]
        for i in range(len(cells[0])):
            if all(crit.match(cells[j][i]) for j, (_, crit) in enumerate(pairs)):
                n += 1
        return float(n)

    def _fn_SUMIFS(self, vals, ctx):
        if isinstance(vals[0], XLErr):
            return vals[0]
        sumrng = vals[0]
        tgt = list(sumrng.flat())
        pairs = [(list(vals[i].flat()), _Criteria(vals[i + 1])) for i in range(1, len(vals), 2)]
        t = 0.0
        for i, s in enumerate(tgt):
            if all(crit.match(cells[i]) for cells, crit in pairs):
                if isinstance(s, (int, float)) and not isinstance(s, bool):
                    t += s
        return t

    def _fn_ROWS(self, vals, ctx):
        rng = vals[0]
        if not isinstance(rng, RangeVal):
            return 1.0
        return float(len(rng.rows))

    def _fn_COLUMNS(self, vals, ctx):
        rng = vals[0]
        if not isinstance(rng, RangeVal):
            return 1.0
        return float(len(rng.rows[0]) if rng.rows else 0)

    def _fn_SUMIF(self, vals, ctx):
        if isinstance(vals[0], XLErr):
            return vals[0]
        rng, crit = vals[0], _Criteria(vals[1])
        sumrng = vals[2] if len(vals) > 2 else rng
        src = list(rng.flat())
        # 合計範囲は条件範囲と同じ形に拡張される
        tgt_rows = []
        for i in range(rng.r1, rng.r2 + 1):
            row = []
            for c in range(rng.c1, rng.c2 + 1):
                row.append(self.value(sumrng.sheet, sumrng.c1 + (c - rng.c1), sumrng.r1 + (i - rng.r1)))
            tgt_rows.append(row)
        tgt = [v for row in tgt_rows for v in row]
        t = 0.0
        for v, s in zip(src, tgt):
            if crit.match(v) and isinstance(s, (int, float)) and not isinstance(s, bool):
                t += s
        return t

    # -- 丸め --
    @staticmethod
    def _round_dir(vals, up):
        x = _num(vals[0])
        d = _num(vals[1]) if len(vals) > 1 else 0.0
        if isinstance(x, XLErr):
            return x
        if isinstance(d, XLErr):
            return d
        import math
        m = 10 ** int(d)
        y = x * m
        # 浮動小数の誤差で切り捨てが揺れないように微調整
        eps = 1e-9 * max(1.0, abs(y))
        if up:
            y = math.ceil(y - eps) if y >= 0 else -math.ceil(-y - eps)
        else:
            y = math.floor(y + eps) if y >= 0 else -math.floor(-y + eps)
        return y / m

    def _fn_ROUNDUP(self, vals, ctx):
        return self._round_dir(vals, True)

    def _fn_ROUNDDOWN(self, vals, ctx):
        return self._round_dir(vals, False)

    def _fn_ROUND(self, vals, ctx):
        x = _num(vals[0])
        d = _num(vals[1]) if len(vals) > 1 else 0.0
        if isinstance(x, XLErr):
            return x
        import math
        m = 10 ** int(d)
        return math.floor(abs(x) * m + 0.5) / m * (1 if x >= 0 else -1)

    # -- 検索 --
    def _fn_VLOOKUP(self, vals, ctx):
        v, table, col = vals[0], vals[1], vals[2]
        approx = _bool(vals[3]) if len(vals) > 3 else True
        ci = int(_num(col))
        if not isinstance(table, RangeVal):
            return ERR_VALUE
        if ci < 1 or ci > len(table.rows[0]):
            return ERR_REF
        return self._lookup_rows(v, table.rows, ci - 1, approx)

    def _fn_HLOOKUP(self, vals, ctx):
        v, table, row = vals[0], vals[1], vals[2]
        approx = _bool(vals[3]) if len(vals) > 3 else True
        ri = int(_num(row))
        if not isinstance(table, RangeVal):
            return ERR_VALUE
        cols = list(zip(*table.rows))
        if ri < 1 or ri > len(cols[0]):
            return ERR_REF
        return self._lookup_rows(v, cols, ri - 1, approx)

    @staticmethod
    def _lookup_rows(v, rows, idx, approx):
        def key_eq(a, b):
            r = _cmp('=', a, b)
            return r is True
        if not approx:
            for row in rows:
                if v is not None and key_eq(row[0], v):
                    return row[idx]
            return ERR_NA
        # 近似一致: 検索値と同じ型のキーのうち v 以下で最後のもの
        def same_type(k):
            if isinstance(v, str):
                return isinstance(k, str)
            if isinstance(v, bool):
                return isinstance(k, bool)
            return isinstance(k, (int, float)) and not isinstance(k, bool)
        best = None
        for row in rows:
            k = row[0]
            if k is None or not same_type(k):
                continue
            if _cmp('<=', k, v) is True:
                best = row
        if best is None:
            return ERR_NA
        return best[idx]

    def _fn_MATCH(self, vals, ctx):
        v, rng = vals[0], vals[1]
        mtype = int(_num(vals[2])) if len(vals) > 2 and vals[2] is not None else 1
        seq = list(rng.flat())
        if mtype == 0:
            crit = _Criteria(v if isinstance(v, str) else v)
            for i, x in enumerate(seq):
                if crit.match(x) if isinstance(v, str) else _cmp('=', x, v) is True:
                    return float(i + 1)
            return ERR_NA
        best = None
        for i, x in enumerate(seq):
            if x is None:
                continue
            if _cmp('<=', x, v) is True:
                best = i
        if best is None:
            return ERR_NA
        return float(best + 1)

    def _fn_INDEX(self, vals, ctx):
        rng = vals[0]
        r = int(_num(vals[1])) if len(vals) > 1 else 1
        c = int(_num(vals[2])) if len(vals) > 2 else 1
        if not isinstance(rng, RangeVal):
            return ERR_VALUE
        rows = rng.rows
        if len(rows) == 1 and len(vals) == 2:  # 横ベクトルに INDEX(v, n)
            r, c = 1, int(_num(vals[1]))
        if len(rows[0]) == 1 and len(vals) == 2:
            c = 1
        if r == 0 and 1 <= c <= len(rows[0]):  # INDEX(範囲, 0, 列) → 列全体
            col = [[row[c - 1]] for row in rows]
            cc = rng.c1 + c - 1
            return RangeVal(rng.sheet, cc, rng.r1, cc, rng.r2, col)
        if c == 0 and 1 <= r <= len(rows):     # INDEX(範囲, 行, 0) → 行全体
            rr = rng.r1 + r - 1
            return RangeVal(rng.sheet, rng.c1, rr, rng.c2, rr, [list(rows[r - 1])])
        if r < 1 or r > len(rows) or c < 1 or c > len(rows[0]):
            return ERR_REF
        return rows[r - 1][c - 1]

    # -- 論理 --
    def _fn_AND(self, vals, ctx):
        res = True
        for v in self._iter_scalars(vals):
            if isinstance(v, XLErr):
                return v
            if v is None or isinstance(v, str):
                continue
            b = _bool(v)
            if isinstance(b, XLErr):
                return b
            res = res and b
        return res

    def _fn_OR(self, vals, ctx):
        res = False
        for v in self._iter_scalars(vals):
            if isinstance(v, XLErr):
                return v
            if v is None or isinstance(v, str):
                continue
            b = _bool(v)
            if isinstance(b, XLErr):
                return b
            res = res or b
        return res

    def _fn_NOT(self, vals, ctx):
        b = _bool(vals[0])
        if isinstance(b, XLErr):
            return b
        return not b

    # -- 文字列 --
    def _fn_TEXT(self, vals, ctx):
        v, fmt = vals[0], vals[1]
        n = _num(v)
        if fmt == '@':
            t = _text(v)
            return t if not isinstance(t, XLErr) else t
        if isinstance(n, XLErr):
            return n
        if fmt in ('aaa', 'aaaa'):
            wd = _JP_WEEKDAYS[int(n) % 7]
            return wd if fmt == 'aaa' else wd + '曜日'
        if fmt == '0':
            return _fmt_num(float(int(round(n))))
        if fmt == 'd':
            return str(serial_to_date(n).day)
        if fmt == 'm':
            return str(serial_to_date(n).month)
        raise NotImplementedError(f'TEXT format {fmt!r}')

    def _fn_SUBSTITUTE(self, vals, ctx):
        t = _text(vals[0])
        old = _text(vals[1])
        new = _text(vals[2])
        if any(isinstance(x, XLErr) for x in (t, old, new)):
            return ERR_VALUE
        if len(vals) > 3:
            k = int(_num(vals[3]))
            parts = t.split(old)
            if len(parts) <= k:
                return t
            return old.join(parts[:k]) + new + old.join(parts[k:])
        return t.replace(old, new)

    def _fn_LEN(self, vals, ctx):
        t = _text(vals[0])
        return float(len(t)) if not isinstance(t, XLErr) else t

    # -- 日付 --
    def _fn_DATE(self, vals, ctx):
        y = int(_num(vals[0]))
        m = int(_num(vals[1]))
        d = int(_num(vals[2]))
        y2, m2 = y + (m - 1) // 12, (m - 1) % 12 + 1
        base = _dt.date(y2, m2, 1)
        return float((base - _EPOCH).days + d - 1)

    def _fn_DAY(self, vals, ctx):
        n = _num(vals[0])
        if isinstance(n, XLErr):
            return n
        return float(serial_to_date(n).day)

    def _fn_MONTH(self, vals, ctx):
        n = _num(vals[0])
        if isinstance(n, XLErr):
            return n
        return float(serial_to_date(n).month)

    def _fn_YEAR(self, vals, ctx):
        n = _num(vals[0])
        if isinstance(n, XLErr):
            return n
        return float(serial_to_date(n).year)

    @staticmethod
    def _add_months(d: _dt.date, m: int) -> _dt.date:
        y = d.year + (d.month - 1 + m) // 12
        mo = (d.month - 1 + m) % 12 + 1
        import calendar
        dd = min(d.day, calendar.monthrange(y, mo)[1])
        return _dt.date(y, mo, dd)

    def _fn_EDATE(self, vals, ctx):
        n = _num(vals[0])
        m = int(_num(vals[1]))
        if isinstance(n, XLErr):
            return n
        return float((self._add_months(serial_to_date(n), m) - _EPOCH).days)

    def _fn_EOMONTH(self, vals, ctx):
        n = _num(vals[0])
        m = int(_num(vals[1]))
        if isinstance(n, XLErr):
            return n
        import calendar
        d = self._add_months(serial_to_date(n), m)
        last = calendar.monthrange(d.year, d.month)[1]
        return float((_dt.date(d.year, d.month, last) - _EPOCH).days)

    def _fn_DATEDIF(self, vals, ctx):
        s = _num(vals[0])
        e = _num(vals[1])
        unit = str(vals[2]).upper()
        if isinstance(s, XLErr) or isinstance(e, XLErr):
            return ERR_VALUE
        ds, de = serial_to_date(s), serial_to_date(e)
        if unit == 'D':
            return float((de - ds).days)
        if unit == 'Y':
            y = de.year - ds.year
            if (de.month, de.day) < (ds.month, ds.day):
                y -= 1
            return float(y)
        if unit == 'M':
            m = (de.year - ds.year) * 12 + de.month - ds.month
            if de.day < ds.day:
                m -= 1
            return float(m)
        raise NotImplementedError(f'DATEDIF unit {unit}')

    # -- 動的配列(必要時のみ) --
    def _fn_UNIQUE(self, vals, ctx):
        rng = vals[0]
        seen, out = set(), []
        for row in rng.rows:
            key = tuple(row)
            if key not in seen:
                seen.add(key)
                out.append(list(row))
        return RangeVal(rng.sheet, rng.c1, rng.r1, rng.c2, rng.r1 + len(out) - 1, out)

    def _fn_FILTER(self, vals, ctx):
        rng, cond = vals[0], vals[1]
        out = []
        flags = [v for v in cond.flat()] if isinstance(cond, RangeVal) else [cond]
        for row, f in zip(rng.rows, flags):
            b = _bool(f)
            if b is True:
                out.append(list(row))
        if not out:
            return vals[2] if len(vals) > 2 else ERR_NA
        return RangeVal(rng.sheet, rng.c1, rng.r1, rng.c2, rng.r1 + len(out) - 1, out)


class _Ctx:
    __slots__ = ('sheet', 'col', 'row', 'volatile', 'deps', 'transition')

    def __init__(self, sheet, col, row):
        self.sheet = sheet
        self.col = col
        self.row = row
        self.volatile = False
        self.deps = ()
        self.transition = False


sys.setrecursionlimit(1_000_000)
