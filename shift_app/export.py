"""Excel への書き戻しと翌月ブック生成。

アプリ上の編集(上書きセル)を元のブックに適用した .xlsx を生成する。
数式はそのまま保持されるため、Excel で開けば従来どおり再計算される。
"""
import calendar
import io

from .model_manager import (GRID_DAY_COL0, GRID_FIRST_ROW, GRID_LAST_ROW,
                            CHECKLIST_COLS, CHECKLIST_ROWS, KINTAI_EDIT_COLS,
                            KINTAI_NOTE_CELL, KINTAI_ROWS, MASTER_SHEET,
                            PLACEMENT_SHEET, _FW)
import re


def apply_overrides(path, overrides):
    """元ブックに上書きセルを適用した openpyxl Workbook を返す。"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    for sheet, col, row, value in overrides:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        ws.cell(row=row, column=col).value = value
    return wb


def export_bytes(path, overrides) -> bytes:
    wb = apply_overrides(path, overrides)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def next_month_bytes(path, overrides):
    """翌月用ブックを生成する。

    - マスタ設定の年月を翌月に更新(マスタ・職員・利用者はそのまま引き継ぐ)
    - シフト入力の日別セルをクリア
    - 勤怠実績(1)〜(31)・日次チェックリストの記入をクリア
    - 人員配置【入力用】の「月の日数」を翌月の日数に更新
    戻り値: (bytes, 年, 月)
    """
    wb = apply_overrides(path, overrides)

    ms = wb[MASTER_SHEET]
    year = int(ms['A2'].value or 0)
    month = int(ms['J2'].value or 0)
    year2, month2 = (year + 1, 1) if month == 12 else (year, month + 1)
    ms['A2'] = year2
    ms['J2'] = month2

    # シフト入力の日別セルをクリア
    input_sheet = '入力シート【現場配布用】'
    if input_sheet in wb.sheetnames:
        ws = wb[input_sheet]
        for r in range(GRID_FIRST_ROW, GRID_LAST_ROW + 1):
            for c in range(GRID_DAY_COL0, GRID_DAY_COL0 + 31):
                if ws.cell(row=r, column=c).value is not None:
                    ws.cell(row=r, column=c).value = None

    # 人員配置【入力用】の月の日数
    if PLACEMENT_SHEET in wb.sheetnames:
        wb[PLACEMENT_SHEET]['E5'] = calendar.monthrange(year2, month2)[1]

    # 日別シートの記入クリア
    for name in wb.sheetnames:
        n = name.translate(_FW).strip()
        if re.fullmatch(r'\(\d+\)', n):
            ws = wb[name]
            for r in KINTAI_ROWS:
                for c in KINTAI_EDIT_COLS:
                    if ws.cell(row=r, column=c).value is not None:
                        ws.cell(row=r, column=c).value = None
            nc, nr = KINTAI_NOTE_CELL
            if ws.cell(row=nr, column=nc).value is not None:
                ws.cell(row=nr, column=nc).value = None
        elif re.fullmatch(r'日次\s*\(\d+\)', n):
            ws = wb[name]
            for r in CHECKLIST_ROWS:
                for c in CHECKLIST_COLS:
                    v = ws.cell(row=r, column=c).value
                    # 定型の選択肢文言(適切/不適切 等)は残し、記入(それ以外)は消す
                    if v is not None and not _is_template_text(v):
                        ws.cell(row=r, column=c).value = None

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), year2, month2


def _is_template_text(v):
    """未記入時の定型文言(選択肢のプレースホルダ)かどうか。数字入りは記入とみなす。"""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if any(ch.isdigit() for ch in s):
        return False
    if s in ('：', ':'):
        return True
    return ('適切' in s or ('あり' in s and 'なし' in s) or s in ('有　　　無', '有無')) and len(s) <= 20
