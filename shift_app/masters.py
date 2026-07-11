"""マスタ設定シートからのマスタ抽出(Excel取り込み機能のコア)。

シフト略記号マスタと職員マスタを、施設ブックの「マスタ設定」シートから
読み取って構造化する。6施設のブックで位置が共通であることを確認済み:
  - 略記号表: 見出し「略記号」= A8、データは 11〜115 行
      A=略記号 / E=開始 / K=終了 / O=休憩開始 / U=休憩終了
      Y=日中時間内勤務時間 / AF=全体勤務時間
  - 職員表: 見出し行 118(E=職種, U=従業員番号, AE=氏名, AO=資格, BV=雇用形態)

固定座標を直接使わず、見出しセルを探して起点を決める(数行のずれを吸収)。

使い方:
    python -m shift_app.masters ブック.xlsx           # 人間向け表示
    python -m shift_app.masters ブック.xlsx --json    # JSON出力
"""
import datetime as _dt
import json
import sys
from dataclasses import dataclass, asdict, field


@dataclass
class ShiftCode:
    code: str
    start: str | None = None       # "9:00" など。時刻なしの記号(有・欠等)は None
    end: str | None = None         # 日またぎは "24:00" 超え表現("翌0:00"→"24:00")
    break_start: str | None = None
    break_end: str | None = None
    daytime_hours: float | None = None   # 日中時間内勤務時間
    total_hours: float | None = None     # 全体勤務時間
    row: int = 0


@dataclass
class Staff:
    name: str
    role: str | None = None          # 職種(管理者 兼 生活支援員 など)
    employee_no: str | None = None   # 従業員番号(7桁)
    qualification: str | None = None  # 資格(上位1つ)
    qualification_others: str | None = None
    pay_type: str | None = None      # 月給制 / 時給制
    category: str | None = None      # C列の区分見出し(生活支援員/世話人/営業)
    row: int = 0


@dataclass
class Masters:
    facility_name: str
    year: int | None
    month: int | None
    service_type: str | None
    shift_codes: list[ShiftCode] = field(default_factory=list)
    staff: list[Staff] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _fmt_time(v):
    if v is None:
        return None
    if isinstance(v, _dt.time):
        return f'{v.hour}:{v.minute:02d}'
    if isinstance(v, _dt.datetime):
        return f'{v.hour}:{v.minute:02d}'
    if isinstance(v, _dt.timedelta):
        secs = int(v.total_seconds())
        return f'{secs // 3600}:{secs % 3600 // 60:02d}'
    return str(v)


def _find(ws, text, max_row=300, max_col=120, startswith=False):
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for c in row:
            v = c.value
            if isinstance(v, str):
                s = v.strip()
                if (s.startswith(text) if startswith else s == text):
                    return c.row, c.column
    return None


def extract_masters(path) -> Masters:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'マスタ設定' not in wb.sheetnames:
        raise ValueError('「マスタ設定」シートが見つかりません')
    ws = wb['マスタ設定']

    m = Masters(
        facility_name=str(ws['A1'].value or '').strip(),
        year=int(ws['A2'].value) if isinstance(ws['A2'].value, (int, float)) else None,
        month=int(ws['J2'].value) if isinstance(ws['J2'].value, (int, float)) else None,
        service_type=ws['AN2'].value if isinstance(ws['AN2'].value, str) else None,
    )

    # --- シフト略記号 ---
    hdr = _find(ws, '略記号')
    if hdr is None:
        m.warnings.append('「略記号」見出しが見つからないため既定位置(A8)を使用')
        hdr = (8, 1)
    # 見出し行の下、「開始/終了」ヘッダ行(+2)を飛ばしてデータ開始
    r = hdr[0] + 3
    seen = {}
    blank_run = 0
    while r < hdr[0] + 200 and blank_run < 40:
        code = ws.cell(row=r, column=1).value
        if code is None or str(code).strip() == '':
            blank_run += 1
            r += 1
            continue
        blank_run = 0
        code = str(code).strip()
        sc = ShiftCode(
            code=code,
            start=_fmt_time(ws.cell(row=r, column=5).value),    # E
            end=_fmt_time(ws.cell(row=r, column=11).value),     # K
            break_start=_fmt_time(ws.cell(row=r, column=15).value),  # O
            break_end=_fmt_time(ws.cell(row=r, column=21).value),    # U
            daytime_hours=_numf(ws.cell(row=r, column=25).value),    # Y
            total_hours=_numf(ws.cell(row=r, column=32).value),      # AF
            row=r,
        )
        if code in seen:
            m.warnings.append(f'略記号「{code}」が重複しています(行{seen[code]}と行{r})')
        else:
            seen[code] = r
        m.shift_codes.append(sc)
        r += 1

    # --- 職員 ---
    nh = _find(ws, '氏名', startswith=True)
    if nh is None:
        raise ValueError('職員表の「氏名」見出しが見つかりません')
    hdr_row, name_col = nh
    role_col = _col_of(ws, hdr_row, '職種') or 5           # E
    no_col = _col_of(ws, hdr_row, '従業員番号', startswith=True) or 21   # U
    qual_col = _col_of(ws, hdr_row, '資格') or 41          # AO
    # 雇用形態列: 見出しがないため 月給制/時給制 の値が並ぶ列を検出
    pay_col = _detect_pay_col(ws, hdr_row)

    seen_names = {}
    category = None
    for r in range(hdr_row + 1, hdr_row + 80):
        cat = ws.cell(row=r, column=3).value  # C列
        if isinstance(cat, str) and cat.strip():
            category = cat.strip()
        name = ws.cell(row=r, column=name_col).value
        if name is None or str(name).strip() == '':
            continue
        name = str(name).strip()
        st = Staff(
            name=name,
            role=_strip(ws.cell(row=r, column=role_col).value),
            employee_no=_empno(ws.cell(row=r, column=no_col).value),
            qualification=_strip(ws.cell(row=r, column=qual_col).value),
            qualification_others=_strip(ws.cell(row=r, column=52).value),  # AZ
            pay_type=_strip(ws.cell(row=r, column=pay_col).value) if pay_col else None,
            category=category,
            row=r,
        )
        if name in seen_names:
            m.warnings.append(f'氏名「{name}」が複数行にあります(行{seen_names[name]}と行{r})')
        else:
            seen_names[name] = r
        m.staff.append(st)

    if not m.shift_codes:
        m.warnings.append('シフト略記号が1件も読み取れませんでした(様式が想定と異なる可能性)')
    if not m.staff:
        m.warnings.append('職員が1件も読み取れませんでした(様式が想定と異なる可能性)')
    wb.close()
    return m


def _numf(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _strip(v):
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None


def _empno(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s or None


def _col_of(ws, hdr_row, text, startswith=False):
    for c in range(1, 120):
        v = ws.cell(row=hdr_row, column=c).value
        if isinstance(v, str):
            s = v.strip()
            if (s.startswith(text) if startswith else s == text):
                return c
    return None


def _detect_pay_col(ws, hdr_row):
    from collections import Counter
    hits = Counter()
    for r in range(hdr_row + 1, hdr_row + 80):
        for c in range(1, 120):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in ('月給制', '時給制'):
                hits[c] += 1
    if not hits:
        return None
    return hits.most_common(1)[0][0]


def main():
    path = sys.argv[1]
    as_json = '--json' in sys.argv
    m = extract_masters(path)
    if as_json:
        print(json.dumps(asdict(m), ensure_ascii=False, indent=1))
        return
    print(f'施設: {m.facility_name}  ({m.year}年{m.month}月) サービス種類: {m.service_type}')
    print(f'\n■ シフト略記号: {len(m.shift_codes)}件')
    for sc in m.shift_codes:
        t = f'{sc.start or "--"}〜{sc.end or "--"}'
        br = f' 休憩{sc.break_start}〜{sc.break_end}' if sc.break_start else ''
        print(f'  {sc.code:10s} {t:15s}{br}  日中h={sc.daytime_hours} 全体h={sc.total_hours}')
    print(f'\n■ 職員: {len(m.staff)}名')
    for st in m.staff:
        print(f'  [{st.category or "-"}] {st.name:14s} {st.role or "-"} '
              f'番号={st.employee_no or "-"} 資格={st.qualification or "-"} {st.pay_type or "-"}')
    if m.warnings:
        print(f'\n⚠ 警告 {len(m.warnings)}件:')
        for w in m.warnings:
            print('  -', w)


if __name__ == '__main__':
    main()
