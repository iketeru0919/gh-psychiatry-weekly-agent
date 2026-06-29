"""
Excel読み込みモジュール
- 1ファイル・施設ごとにシートが分かれている構成を想定
- セル位置はcell_map.jsonで管理（コードを触らずに変更可能）
"""

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

CELL_MAP_PATH = Path(__file__).parent / "cell_map.json"


# ── ユーティリティ ────────────────────────────────────────

def col_letter_to_index(letter: str) -> int:
    return column_index_from_string(letter)


def get_cell_value(ws, row: int, col_letter: str) -> Any:
    return ws.cell(row=row, column=col_letter_to_index(col_letter)).value


def advance_col(col_letter: str, steps: int, stride: int = 1) -> str:
    idx = col_letter_to_index(col_letter) + steps * stride
    return get_column_letter(idx)


def cell_str_to_row_col(cell_str: str) -> tuple[int, int]:
    match = re.match(r"([A-Za-z]+)(\d+)", cell_str)
    if not match:
        raise ValueError(f"不正なセル番地: {cell_str}")
    col = col_letter_to_index(match.group(1))
    row = int(match.group(2))
    return row, col


def safe_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("%", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(v) -> int | None:
    f = safe_float(v)
    return int(f) if f is not None else None


# ── メイン読み込み関数 ────────────────────────────────────

def load_cell_map() -> dict:
    with open(CELL_MAP_PATH, encoding="utf-8") as f:
        return json.load(f)


def month_label_to_iso(label: str, fiscal_year_start: int = 2026) -> str:
    """
    '3月' → '2026-03'
    会計年度: 4月始まりを想定。4〜12月→fiscal_year_start、1〜3月→fiscal_year_start+1
    """
    num = int(re.sub(r"[^\d]", "", label))
    year = fiscal_year_start if num >= 4 else fiscal_year_start + 1
    return f"{year}-{num:02d}"


def read_monthly_table(ws, table_cfg: dict, months_list: list[str]) -> list[dict]:
    """
    月次テーブルを読み込む。
    col_stride: 月ごとの列間隔（例: 3=3列おき）
    """
    rows_cfg = table_cfg["rows"]
    col_start = table_cfg["data_col_start"]
    col_stride = table_cfg.get("col_stride", 1)
    defined_months = table_cfg.get("months", months_list)

    records = []
    for i, month_label in enumerate(defined_months):
        col = advance_col(col_start, i, stride=col_stride)
        row_data: dict = {}
        any_value = False

        for field, row_num in rows_cfg.items():
            if field.startswith("_"):
                continue
            val = get_cell_value(ws, row_num, col)
            if isinstance(val, float) and val < 2.0 and "rate" in field:
                val = round(val * 100, 1)
            row_data[field] = val
            if val is not None and isinstance(val, (int, float)):
                any_value = True

        if not any_value:
            continue

        row_data["month"] = month_label_to_iso(month_label)
        records.append(row_data)

    return records


def read_sheet(ws, cmap: dict, sheet_name: str) -> dict:
    """1シート（1施設）を読み込んでdictで返す"""
    result: dict = {
        "facility_id":    sheet_name,
        "facility_name":  sheet_name,
        "group_company":  "",
        "capacity":       0,
        "monthly":        [],
        "weekly_current": [],
    }

    # 基本情報
    finfo = cmap.get("facility_info", {})
    if "company" in finfo:
        row, col = cell_str_to_row_col(finfo["company"])
        result["group_company"] = str(ws.cell(row=row, column=col).value or "")
    if "facility_name" in finfo:
        row, col = cell_str_to_row_col(finfo["facility_name"])
        name = ws.cell(row=row, column=col).value
        if name:
            result["facility_name"] = str(name)
    if "capacity" in finfo:
        row, col = cell_str_to_row_col(finfo["capacity"])
        result["capacity"] = safe_int(ws.cell(row=row, column=col).value) or 0

    # 月次テーブル（重要指標推移）
    if "monthly_table" in cmap:
        months_list = cmap["monthly_table"].get("months", [])
        result["monthly"] = read_monthly_table(ws, cmap["monthly_table"], months_list)
    else:
        result["monthly"] = []
        months_list = []

    monthly_by_month = {m["month"]: m for m in result["monthly"]}

    # スタッフィングテーブルをマージ
    if "staffing_table" in cmap:
        staffing = read_monthly_table(ws, cmap["staffing_table"], months_list)
        for s in staffing:
            month = s.pop("month")
            if month in monthly_by_month:
                monthly_by_month[month].update(s)

    # ブランディング（GH問合せ・営業）テーブルをマージ
    if "branding_table" in cmap:
        branding = read_monthly_table(ws, cmap["branding_table"], months_list)
        for b in branding:
            month = b.pop("month")
            if month in monthly_by_month:
                monthly_by_month[month].update(b)

    # ブランディング（SS問合せ・見学）テーブルをマージ
    if "branding_ss_table" in cmap:
        branding_ss = read_monthly_table(ws, cmap["branding_ss_table"], months_list)
        for b in branding_ss:
            month = b.pop("month")
            if month in monthly_by_month:
                for k, v in b.items():
                    if k not in monthly_by_month[month] or monthly_by_month[month][k] is None:
                        monthly_by_month[month][k] = v

    # タイミーテーブルをマージ
    if "timee_table" in cmap:
        timee = read_monthly_table(ws, cmap["timee_table"], months_list)
        for t in timee:
            month = t.pop("month")
            if month in monthly_by_month:
                monthly_by_month[month].update(t)

    # 運営評価テーブルをマージ
    if "operations_table" in cmap:
        ops = read_monthly_table(ws, cmap["operations_table"], months_list)
        for o in ops:
            month = o.pop("month")
            if month in monthly_by_month:
                monthly_by_month[month].update(o)

    result["monthly"] = list(monthly_by_month.values())

    # 週次テーブル
    if "weekly_table" in cmap:
        result["weekly_current"] = read_weekly_table(ws, cmap["weekly_table"])

    return result


def read_weekly_table(ws, table_cfg: dict) -> list[dict]:
    rows_cfg = table_cfg["rows"]
    col_start = table_cfg["data_col_start"]
    col_stride = table_cfg.get("col_stride", 1)
    weeks = table_cfg["weeks"]

    records = []
    for i, week_label in enumerate(weeks):
        col = advance_col(col_start, i, stride=col_stride)
        row_data: dict = {"week": i + 1}
        any_value = False
        for field, row_num in rows_cfg.items():
            if field.startswith("_"):
                continue
            val = get_cell_value(ws, row_num, col)
            if isinstance(val, float) and val < 2.0 and "rate" in field:
                val = round(val * 100, 1)
            row_data[field] = val
            if val is not None:
                any_value = True
        if any_value:
            records.append(row_data)
    return records


# ── 公開インターフェース ──────────────────────────────────

def load_from_excel(excel_path: str | Path) -> list[dict]:
    """
    Excelファイルから全施設データを読み込む。
    戻り値: [facility_dict, ...]  （app.pyのload_all_facilities()と同じ形式）
    """
    cmap = load_cell_map()
    skip_sheets = set(cmap.get("_skip_sheets", []))

    wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)

    facilities = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        try:
            facility = read_sheet(ws, cmap, sheet_name)
            facilities.append(facility)
        except Exception as e:
            print(f"[警告] シート '{sheet_name}' の読み込みに失敗しました: {e}")

    wb.close()
    return facilities
